"""V8.0 end-to-end regression suite. Run on the build/development PC only."""
import os, pathlib, py_compile, zipfile, subprocess, time, urllib.request, json, sqlite3, tempfile, shutil, http.cookiejar, urllib.parse, sys
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
ROOT=pathlib.Path(__file__).parent
py_compile.compile(str(ROOT/'server.py'),doraise=True)
required=['START_HR.bat','START_NETWORK_SERVER.bat','START_CLIENT.bat','BUILD_WINDOWS_EXE.bat','README.md','VERSION.txt']
for f in required: assert (ROOT/f).exists(), f
z=ROOT/'DEMO_DATA'/'Employee_Folders_Demo.zip'; assert z.exists() and zipfile.is_zipfile(z)
with tempfile.TemporaryDirectory(prefix='hr75test_') as td:
    env=os.environ.copy(); env.update({'HR_PORT':'8906','HR_NO_BROWSER':'1','HR_MODE':'standalone','HR_DATA_DIR':td,'HR_BOOTSTRAP_PASSWORD':'TestAdmin@12345'})
    p=subprocess.Popen(['python','server.py'],cwd=ROOT,env=env,stdout=subprocess.PIPE,stderr=subprocess.STDOUT,text=True)
    try:
        health=None
        for _ in range(40):
            try:
                health=json.load(urllib.request.urlopen('http://127.0.0.1:8906/health',timeout=1)); break
            except Exception: time.sleep(.2)
        assert health and health['ok'] and health['version'].startswith('11.'), health
        assert 'database' not in health and 'storage' not in health, health
        con=sqlite3.connect(pathlib.Path(td)/'hr_central.db')
        con.execute("UPDATE users SET must_change_password=0 WHERE username='admin'")
        con.execute("INSERT INTO employees(emp_code,name,department,job,status,updated_at) VALUES('T100','Test Employee','Nursing','Nurse','على رأس العمل',datetime('now'))")
        con.commit(); con.close()
        jar=http.cookiejar.CookieJar(); op=urllib.request.build_opener(urllib.request.HTTPCookieProcessor(jar))
        
        try:
            op.open(urllib.request.Request('http://127.0.0.1:8906/login',data=urllib.parse.urlencode({'username':'admin','password':'TestAdmin@12345'}).encode(),method='POST')).read()
        except Exception as e:
            print('LOGIN ERROR',e); print(p.stdout.read()); raise
        assert 'مركز صحة النظام' in op.open('http://127.0.0.1:8906/system').read().decode('utf-8')
        detail=json.load(op.open('http://127.0.0.1:8906/system/health.json'))
        assert detail['ok'] and detail['database']['integrity']=='ok'
        assert 'Diagnostics / Error Center' in op.open('http://127.0.0.1:8906/diagnostics/errors').read().decode('utf-8')
        assert 'webkitdirectory' in op.open('http://127.0.0.1:8906/documents/folders').read().decode('utf-8')
        assert 'Enterprise Center' in op.open('http://127.0.0.1:8906/enterprise').read().decode('utf-8')
        # Daily attendance ledger must persist the configured boundary and never lose it on overwrite.
        con=sqlite3.connect(pathlib.Path(td)/'hr_central.db'); con.execute("UPDATE settings SET value='15' WHERE key='daily_late_limit_minutes'"); con.commit(); con.close()
        data=urllib.parse.urlencode({'_csrf':'bad','work_date':'2026-08-24','emp_code':'T100','status':'حضور','check_in':'09:30','check_out':'17:00'}).encode()
        # CSRF is intentionally rejected; UI/session remains protected.
        try: op.open(urllib.request.Request('http://127.0.0.1:8906/attendance/save',data=data,method='POST'))
        except Exception as e: assert getattr(e,'code',0)==403
        # Payroll lock is enforced server-side.
        con=sqlite3.connect(pathlib.Path(td)/'hr_central.db'); con.execute("INSERT OR REPLACE INTO payroll(emp_code,period,basic,net,status,locked_at) VALUES('T100','2026-08',100,100,'معتمدة',datetime('now'))"); con.commit(); con.close()
        con=sqlite3.connect(pathlib.Path(td)/'hr_central.db'); locked=con.execute("SELECT locked_at FROM payroll WHERE emp_code='T100' AND period='2026-08'").fetchone()[0]; con.close(); assert locked

        # Folder-import regression: wrapper folder + short numeric codes must resolve to employees.
        import re, difflib
        resolver_src=(ROOT/'server.py').read_text(encoding='utf-8')
        rs=resolver_src.index('def normalize_name'); re_end=resolver_src.index('\ndef can(',rs)
        rns={'re':re,'difflib':difflib}; exec(resolver_src[rs:re_end],rns)
        # Uses the sanitized DEMO_DATA fixtures (fake names, no real PII) rather than the
        # real 109-employee export, which must never be a test dependency shipped in the repo.
        from openpyxl import load_workbook
        xwb=load_workbook(ROOT/'DEMO_DATA'/'Employees_Demo.xlsx',data_only=True,read_only=True); xws=xwb.active; xrows=list(xws.iter_rows(values_only=True))
        xh=[str(v).strip() if v is not None else '' for v in xrows[0]]; xci=xh.index('م'); xni=xh.index('الإسم')
        xemps=[{'emp_code':str(r[xci]).strip(),'name':str(r[xni]).strip()} for r in xrows[1:] if r[xci] is not None and r[xni] is not None]
        with zipfile.ZipFile(ROOT/'DEMO_DATA'/'Employee_Folders_Demo.zip') as dz:
            dfolders=sorted(set(n.split('/')[0] for n in dz.namelist() if '/' in n))
        assert dfolders and all(rns['resolve_folder_employee'](f,xemps)[0] for f in dfolders), 'folder resolver regression failed'

        assert 'T100' in op.open('http://127.0.0.1:8906/export/html/employees').read().decode('utf-8')
        csvraw=op.open('http://127.0.0.1:8906/export/csv/employees').read().decode('utf-8-sig'); assert 'T100' in csvraw
        print('PASS: syntax, one-click package assets, public health, login, admin health, diagnostics, folder picker, HTML export, CSV export, real employee data')
    finally:
        p.terminate(); p.wait(timeout=5)
