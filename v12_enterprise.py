"""HR Enterprise 11.2 enterprise UX/data completion layer.
Adds universal Excel mapping, leave-balance import/adjustment history, QR identity UX,
per-employee account provisioning, modern navigation, branding controls, ID-card designer,
and PostgreSQL-ready network configuration without disturbing the stable core routes.
"""
import os, io, re, json, secrets, hashlib, zipfile, csv, mimetypes
from urllib.parse import urlparse, parse_qs, quote
from datetime import datetime, date
from openpyxl import load_workbook, Workbook

ALIASES = {
    'emp_code':['employee id','employee code','emp id','emp code','code','id','رقم الموظف','كود الموظف','كود','الرقم الوظيفي','م','م/كود الموظف'],
    'name':['name','employee name','full name','الاسم','اسم الموظف','الاسم بالكامل','اسم الموظف بالكامل'],
    'leave_type':['leave type','leave','type','نوع الاجازة','نوع الإجازة','الاجازة','الإجازة','نوع'],
    'annual':['annual','allocated','allocation','entitled','balance','المخصص','المستحق','الرصيد السنوي','الرصيد','المتاح'],
    'used':['used','taken','consumed','utilized','المستخدم','المستهلك','المأخوذ','المستخدم من الرصيد'],
    'remaining':['remaining','remain','available','left','المتبقي','المتبقى','المتاح المتبقي'],
    'department':['department','dept','الإدارة','الادارة','القسم'],
    'job':['job','title','position','الوظيفة','المسمى الوظيفي','المسمى'],
    'branch':['branch','location','الفرع','الموقع'],
}

def norm(v):
    s=str(v or '').strip().lower()
    s=re.sub(r'[\u064B-\u065F\u0670]', '', s)
    s=s.replace('أ','ا').replace('إ','ا').replace('آ','ا').replace('ة','ه').replace('ى','ي')
    s=re.sub(r'[^\w\u0600-\u06FF]+',' ',s)
    return re.sub(r'\s+',' ',s).strip()

def best_field(header, candidates):
    h=norm(header)
    if not h: return None,0
    for c in candidates:
        n=norm(c)
        if h==n: return c,1.0
    for c in candidates:
        n=norm(c)
        if n and (n in h or h in n): return c,.88
    return None,0

def detect_mapping(headers, kind='employees'):
    out={}; confidence={}
    for field,cands in ALIASES.items():
        if kind=='employees' and field in ('leave_type','annual','used','remaining'): continue
        if kind=='leave' and field in ('department','job','branch'): continue
        best=(None,0)
        for h in headers:
            c,score=best_field(h,cands)
            if score>best[1]: best=(h,score)
        if best[0]: out[field]=best[0]; confidence[field]=best[1]
    return out,confidence

def read_rows(data, filename):
    ext=os.path.splitext(filename)[1].lower()
    if ext=='.csv':
        text=data.decode('utf-8-sig') if isinstance(data,(bytes,bytearray)) else str(data)
        first=text.splitlines()[0] if text.splitlines() else ''
        delim='\t' if first.count('\t')>=1 else ','
        rows=list(csv.reader(io.StringIO(text),delimiter=delim)); return rows
    wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True)
    ws=wb.active
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]

def clean_num(v):
    if v is None or v=='': return 0.0
    try: return float(str(v).replace(',','').strip())
    except: return 0.0

def install(g):
    H=g['H']; db=g['db']; now=g['now']; page0=g['page']; esc=g['esc']; csrf=g['csrf_field']; can=g['can']; audit=g['audit']; DATA=g['DATA']; BRAND=g['BRAND']
    clear_settings_cache=g.get('clear_settings_cache', lambda: None)
    try: qrcode=g.get('qrcode')
    except: qrcode=None

    def upgrade():
        c=db(); c.executescript('''
        CREATE TABLE IF NOT EXISTS leave_balance_history(
            id INTEGER PRIMARY KEY, emp_code TEXT NOT NULL, leave_type TEXT NOT NULL,
            action TEXT NOT NULL, amount REAL DEFAULT 0, before_annual REAL DEFAULT 0,
            before_used REAL DEFAULT 0, after_annual REAL DEFAULT 0, after_used REAL DEFAULT 0,
            reason TEXT, source TEXT, actor TEXT, created_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_lbh_emp ON leave_balance_history(emp_code,leave_type,created_at);
        CREATE TABLE IF NOT EXISTS qr_scan_history(
            id INTEGER PRIMARY KEY, emp_code TEXT, token_hash TEXT, scanned_by TEXT,
            ip TEXT, result TEXT, scanned_at TEXT, details TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_qr_scan_emp ON qr_scan_history(emp_code,scanned_at);
        ''')
        # Link employee accounts to their employee records.
        cols={r['name'] for r in c.execute('PRAGMA table_info(users)').fetchall()}
        if 'employee_code' not in cols: c.execute('ALTER TABLE users ADD COLUMN employee_code TEXT')
        if 'temporary_password_expires_at' not in cols: c.execute('ALTER TABLE users ADD COLUMN temporary_password_expires_at TEXT')
        for k,v in {'ui_font':'Segoe UI','login_tagline':'Enterprise HR Management','login_background':'','login_card_style':'glass'}.items():
            c.execute('INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)',(k,v))
        c.commit(); c.close()
    old_init = g['init']
    def init_v12():
        old_init()
        upgrade()
    g['init'] = init_v12

    def page(title,body,user,active='dashboard'):
        out=page0(title,body,user,active)
        if not user: return out
        # Add a cleaner grouped navigation without deleting the stable legacy links.
        # Hide duplicate low-level links on the visual layer; routes remain intact.
        css='''<style>
        .side{width:292px;padding:16px 12px}.main{margin-right:292px;width:calc(100% - 292px);padding:26px 34px}
        .brand{padding:10px 12px 18px}.brand-logo{width:52px;height:52px;object-fit:contain;border-radius:14px;background:#fff;padding:6px;box-shadow:0 8px 22px rgba(0,0,0,.16)}
        .nav-section{margin:16px 8px 6px;color:#7dd3fc;font-size:10px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}.nav a{margin:3px 0;padding:11px 12px}
        .profile-chip{display:flex;align-items:center;gap:9px;margin:14px 5px;padding:10px;border:1px solid rgba(255,255,255,.1);border-radius:13px;background:rgba(255,255,255,.04)}
        .profile-chip b{font-size:12px}.profile-chip small{display:block;color:#98a2b3;margin-top:3px}
        .table .qr-mini{width:58px;height:58px;object-fit:contain;background:#fff;border-radius:9px;padding:4px;border:1px solid #e4e7ec}
        .hero-card{border:1px solid rgba(132,173,255,.28);background:linear-gradient(135deg,#ffffff,#f7fbff);box-shadow:0 18px 55px rgba(16,24,40,.08)}
        .identity-strip{display:flex;gap:16px;align-items:center;flex-wrap:wrap}.identity-strip .identity-photo{width:88px;height:106px;border-radius:18px;object-fit:cover;border:1px solid #e4e7ec;background:#f2f4f7}
        @media(max-width:900px){.side{width:280px}.main{margin:0;width:100%;padding:18px}}
        </style>'''
        out=out.replace('</head>',css+'</head>',1)
        return out
    g['page']=page

    # -------- Leave balances --------
    def leave_page(self,u):
        qs=parse_qs(urlparse(self.path).query); q=qs.get('q',[''])[0].strip(); typ=qs.get('leave_type',[''])[0].strip()
        c=db(); types=[r['name'] for r in c.execute('SELECT name FROM leave_types ORDER BY name').fetchall()]
        rows=c.execute('SELECT lb.emp_code,e.name,lb.leave_type,lb.annual,lb.used,(lb.annual-lb.used) remaining FROM leave_balances lb JOIN employees e ON e.emp_code=lb.emp_code ORDER BY e.name,lb.leave_type LIMIT 5000').fetchall(); c.close()
        rows=[r for r in rows if g['emp_allowed'](u,r['emp_code']) and (not q or q.lower() in (r['emp_code']+' '+r['name']).lower()) and (not typ or r['leave_type']==typ)]
        trs=''.join(f'''<tr><td><b>{esc(r['emp_code'])}</b></td><td>{esc(r['name'])}</td><td><span class="badge b-blue">{esc(r['leave_type'])}</span></td><td>{float(r['annual'] or 0):g}</td><td>{float(r['used'] or 0):g}</td><td><b>{float(r['remaining'] or 0):g}</b></td><td><button class="btn gray" onclick="adj('{esc(r['emp_code'])}','{esc(r['leave_type'])}',{float(r['annual'] or 0)},{float(r['used'] or 0)})">تعديل</button></td></tr>''' for r in rows)
        opts=''.join(f'<option>{esc(x)}</option>' for x in types)
        body=f'''<div class="top"><div class="title"><h1>🏖 إدارة أرصدة الإجازات</h1><p>استيراد ذكي من Excel، لصق مباشر، تعديل يدوي، وسجل تاريخي لكل تغيير.</p></div><div class="actions"><a class="btn" href="/leave-balances/import">📥 استيراد Excel</a><a class="btn gray" href="/leave-balances/history">🧾 History</a></div></div>
        <div class="grid g3"><div class="card"><h3>⚡ تعديل سريع</h3><form method="post" action="/leave-balances/adjust" class="form">{csrf(u)}<div class="field"><label>الموظف</label><input name="emp_code" id="adj_emp" required></div><div class="field"><label>نوع الإجازة</label><select name="leave_type" id="adj_type">{opts}</select></div><div class="field"><label>المخصص الجديد</label><input name="annual" id="adj_annual" type="number" step="0.01"></div><div class="field"><label>المستخدم الجديد</label><input name="used" id="adj_used" type="number" step="0.01"></div><div class="field full"><label>سبب التعديل</label><input name="reason" required></div><div class="full"><button class="btn">حفظ التعديل + تسجيل History</button></div></form></div>
        <div class="card"><h3>📊 كيف يعمل</h3><p>Employee ID → Leave Type → Annual/Used/Remaining.</p><p>النظام يكتشف أسماء الأعمدة بالعربي والإنجليزي ويعرض Preview قبل الحفظ.</p><span class="badge b-ok">Import-safe</span></div>
        <div class="card"><h3>📋 Paste</h3><p>انسخ من Excel والصق مباشرة في صفحة الاستيراد بدون رفع ملف.</p><a class="btn gray" href="/leave-balances/import?mode=paste">فتح Paste Center</a></div></div>
        <div class="card table-wrap" style="margin-top:16px"><form class="toolbar"><input name="q" value="{esc(q)}" placeholder="بحث موظف / كود"><select name="leave_type"><option value="">كل الأنواع</option>{''.join('<option '+('selected' if typ==x else '')+'>'+esc(x)+'</option>' for x in types)}</select><button class="btn gray">تطبيق</button></form><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>النوع</th><th>المخصص</th><th>المستخدم</th><th>المتبقي</th><th></th></tr></thead><tbody>{trs or '<tr><td colspan="7"><div class="alert">لا توجد بيانات.</div></td></tr>'}</tbody></table></div>
        <script>function adj(e,t,a,u){{document.getElementById('adj_emp').value=e;document.getElementById('adj_type').value=t;document.getElementById('adj_annual').value=a;document.getElementById('adj_used').value=u;window.scrollTo({{top:0,behavior:'smooth'}})}}</script>'''
        H.send(self, page('أرصدة الإجازات',body,u,'leave-balances'))

    def leave_import(self,u):
        if not can(u,'import.validate'): return H.forbid(self,u)
        mode=parse_qs(urlparse(self.path).query).get('mode',[''])[0]
        body=f'''<div class="top"><div class="title"><h1>📥 Leave Balance Import Center</h1><p>Smart mapping: Arabic/English columns + Preview + Commit.</p></div><a class="btn gray" href="/leave-balances">رجوع للأرصدة</a></div>
        <div class="grid g2"><div class="card"><h3>رفع Excel / CSV</h3><form method="post" action="/leave-balances/import" enctype="multipart/form-data">{csrf(u)}<input type="file" name="file" accept=".xlsx,.xls,.csv" required><div class="actions" style="margin-top:12px"><button class="btn">Analyze & Preview</button></div></form><div class="alert" style="margin-top:12px">يدعم Long format: Employee ID | Leave Type | Annual | Used | Remaining. وسيحاول أيضًا اكتشاف الأعمدة العربية والإنجليزية تلقائيًا.</div></div>
        <div class="card"><h3>Paste from Excel</h3><form method="post" action="/leave-balances/paste">{csrf(u)}<textarea name="data" style="width:100%;min-height:260px;font-family:Consolas,monospace" placeholder="Employee ID\tLeave Type\tAnnual\tUsed\tRemaining\nEMP001\tAnnual\t21\t8\t13"></textarea><button class="btn" style="margin-top:10px">Analyze & Preview</button></form></div></div>'''
        H.send(self, page('Leave Import',body,u,'leave-balances'))

    def parse_leave_preview(self,u, data, filename, pasted=False):
        try: rows=read_rows(data,filename)
        except Exception as e: return H.send(self, page('Import Error',f'<div class="card"><div class="alert">{esc(e)}</div></div>',u),400)
        if not rows or not rows[0]: return H.send(self, page('Import Error','<div class="card"><div class="alert">الملف فارغ.</div></div>',u),400)
        headers=[str(x or '').strip() for x in rows[0]]; mapping,conf=detect_mapping(headers,'leave')
        preview=[]; errors=[]
        for i,row in enumerate(rows[1:],2):
            vals={h:(row[j] if j<len(row) else '') for j,h in enumerate(headers)}
            rec={k:vals.get(col,'') for k,col in mapping.items()}
            if not rec.get('emp_code'): errors.append((i,'emp_code','لم يتم العثور على Employee ID')) ; continue
            lt=rec.get('leave_type') or 'اعتيادي'
            annual=clean_num(rec.get('annual')); used=clean_num(rec.get('used')); remaining=rec.get('remaining')
            if remaining not in (None,'') and not rec.get('used'): used=max(0,annual-clean_num(remaining))
            preview.append({'row':i,'emp_code':str(rec['emp_code']).strip(),'leave_type':str(lt).strip(),'annual':annual,'used':used,'remaining':max(0,annual-used)})
        token=secrets.token_urlsafe(18); g.setdefault('V12_PREVIEWS',{})[token]={'records':preview,'errors':errors,'mapping':mapping,'confidence':conf,'source':filename}
        trs=''.join(f"<tr><td>{r['row']}</td><td>{esc(r['emp_code'])}</td><td>{esc(r['leave_type'])}</td><td>{r['annual']:g}</td><td>{r['used']:g}</td><td><b>{r['remaining']:g}</b></td></tr>" for r in preview[:300])
        maprows=''.join(f'<tr><td>{esc(k)}</td><td>{esc(v)}</td><td>{conf.get(k,0):.0%}</td></tr>' for k,v in mapping.items())
        err='<div class="alert" style="margin-top:12px">Rows needing attention: '+str(len(errors))+'</div>' if errors else ''
        body=f'''<div class="top"><div class="title"><h1>Leave Balance Preview</h1><p>{len(preview)} valid rows · {len(errors)} attention rows. Import Valid Rows: {len(preview)}</p></div></div><div class="grid g2"><div class="card"><h3>🧠 Detected Mapping</h3><table class="table"><tr><th>System Field</th><th>Your Column</th><th>Confidence</th></tr>{maprows}</table>{err}</div><div class="card"><h3>Commit</h3><p>لن يتم تغيير أي رصيد قبل الضغط على Commit.</p><form method="post" action="/leave-balances/import/commit">{csrf(u)}<input type="hidden" name="token" value="{token}"><button class="btn ok">✅ Commit {len(preview)} rows</button></form></div></div><div class="card table-wrap" style="margin-top:16px"><h3>Preview</h3><table class="table"><tr><th>Row</th><th>Employee</th><th>Leave Type</th><th>Annual</th><th>Used</th><th>Remaining</th></tr>{trs}</table></div>'''
        H.send(self, page('Leave Import Preview',body,u,'leave-balances'))

    def leave_import_post(self,u):
        if not can(u,'import.validate'): return H.forbid(self,u)
        f,up=H.parse_upload(self)
        if f.get('_csrf')!=u.get('csrf'): return H.send(self, page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
        if not up:return leave_import(self,u)
        _,data,fn=up; return parse_leave_preview(self,u,data,fn)

    def leave_paste(self,u):
        if not can(u,'import.validate'): return H.forbid(self,u)
        f=H.form(self)
        if f.get('_csrf')!=u.get('csrf'): return H.send(self, page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
        text=f.get('data',f.get('paste_data','')); return parse_leave_preview(self,u,text.encode('utf-8-sig'),'pasted.csv',True)

    def leave_commit(self,u):
        if not can(u,'import.validate'): return H.forbid(self,u)
        f=H.form(self); token=f.get('token',''); data=g.get('V12_PREVIEWS',{}).pop(token,None)
        if f.get('_csrf')!=u.get('csrf') or not data:return H.send(self, page('Import','<div class="card"><div class="alert">Preview expired or invalid.</div></div>',u),400)
        c=db(); new=upd=0
        for r in data['records']:
            if not g['emp_allowed'](u,r['emp_code']): continue
            old=c.execute('SELECT * FROM leave_balances WHERE emp_code=? AND leave_type=?',(r['emp_code'],r['leave_type'])).fetchone()
            if old:
                c.execute('UPDATE leave_balances SET annual=?,used=? WHERE emp_code=? AND leave_type=?',(r['annual'],r['used'],r['emp_code'],r['leave_type'])); upd+=1
            else:
                c.execute('INSERT INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,?)',(r['emp_code'],r['leave_type'],r['annual'],r['used'])); new+=1
            c.execute('INSERT INTO leave_balance_history(emp_code,leave_type,action,amount,before_annual,before_used,after_annual,after_used,reason,source,actor,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(r['emp_code'],r['leave_type'],'import',r['annual']-r['used'],float(old['annual']) if old else 0,float(old['used']) if old else 0,r['annual'],r['used'],'Excel import','Excel',u['username'],now()))
        c.commit(); c.close(); audit(u['username'],u['role'],'LEAVE_BALANCE_IMPORT','Leave Balances','Excel',f'new={new},updated={upd}'); self.redirect('/leave-balances')

    def leave_adjust(self,u):
        if not can(u,'leave.approve'): return H.forbid(self,u)
        f=H.form(self)
        if f.get('_csrf')!=u.get('csrf'): return H.send(self, page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
        emp=f.get('emp_code','').strip(); lt=f.get('leave_type','').strip(); annual=clean_num(f.get('annual')); used=clean_num(f.get('used'))
        if not emp or not lt or not g['emp_allowed'](u,emp): return H.send(self, page('Leave','<div class="card"><div class="alert">الموظف غير مسموح أو البيانات ناقصة.</div></div>',u),403)
        c=db(); old=c.execute('SELECT * FROM leave_balances WHERE emp_code=? AND leave_type=?',(emp,lt)).fetchone()
        c.execute('INSERT INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,?) ON CONFLICT(emp_code,leave_type) DO UPDATE SET annual=excluded.annual,used=excluded.used',(emp,lt,annual,used))
        c.execute('INSERT INTO leave_balance_history(emp_code,leave_type,action,amount,before_annual,before_used,after_annual,after_used,reason,source,actor,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(emp,lt,'manual_adjust',annual-used,float(old['annual']) if old else 0,float(old['used']) if old else 0,annual,used,f.get('reason',''),'manual',u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'LEAVE_BALANCE_ADJUST','Leave Balances',emp,f'{lt}: {annual}/{used}'); self.redirect('/leave-balances')

    def leave_history(self,u):
        c=db(); rows=c.execute('SELECT * FROM leave_balance_history ORDER BY id DESC LIMIT 1000').fetchall(); c.close(); rows=[r for r in rows if g['emp_allowed'](u,r['emp_code'])]
        trs=''.join(f"<tr><td>{esc(r['created_at'])}</td><td>{esc(r['emp_code'])}</td><td>{esc(r['leave_type'])}</td><td>{esc(r['action'])}</td><td>{r['before_annual']:g}/{r['before_used']:g}</td><td>{r['after_annual']:g}/{r['after_used']:g}</td><td>{esc(r['reason'] or '')}</td><td>{esc(r['actor'])}</td></tr>" for r in rows)
        H.send(self, page('Leave Balance History',f'<div class="top"><div class="title"><h1>🧾 Leave Balance History</h1><p>كل استيراد وتعديل محفوظ مع المستخدم والسبب.</p></div></div><div class="card table-wrap"><table class="table"><tr><th>التاريخ</th><th>الكود</th><th>النوع</th><th>العملية</th><th>قبل</th><th>بعد</th><th>السبب</th><th>بواسطة</th></tr>{trs}</table></div>',u,'leave-balances'))

    # -------- Universal employee mapping page --------
    def mapping_page(self,u):
        if not can(u,'import.mapping'): return H.forbid(self,u)
        c=db(); rows=c.execute('SELECT * FROM import_mappings ORDER BY name').fetchall(); c.close()
        trs=''.join(f"<tr><td>{esc(r['name'])}</td><td>{esc(r['kind'])}</td><td><code>{esc(r['mapping_json'])}</code></td><td>{esc(r['created_at'])}</td></tr>" for r in rows)
        body=f'''<div class="top"><div class="title"><h1>🧠 Universal Excel Mapping</h1><p>مش Template واحد: النظام يفهم العربي والإنجليزي ويحتفظ بالـmapping لكل شركة/شيت.</p></div></div>
        <div class="grid g2"><div class="card"><h3>Smart detection</h3><form method="post" action="/import/mapping/auto">{csrf(u)}<textarea name="headers" placeholder="Employee ID\tName\tDepartment\tJob\nأو الصق أول صف من Excel" style="width:100%;min-height:180px"></textarea><button class="btn" style="margin-top:10px">Detect Columns</button></form></div><div class="card"><h3>Saved mappings</h3><table class="table"><tr><th>Name</th><th>Kind</th><th>Mapping</th><th>Updated</th></tr>{trs or '<tr><td colspan="4">لا توجد mappings محفوظة.</td></tr>'}</table></div></div>'''
        H.send(self, page('Universal Mapping',body,u,'import'))
    def mapping_auto(self,u):
        f=H.form(self); headers=[x for x in re.split(r'[\t,\n]+',f.get('headers','')) if x.strip()][:100]; mapping,conf=detect_mapping(headers,'employees'); rows=''.join(f'<tr><td>{esc(k)}</td><td>{esc(v)}</td><td>{conf.get(k,0):.0%}</td></tr>' for k,v in mapping.items())
        H.send(self, page('Mapping Preview',f'<div class="top"><div class="title"><h1>Detected Mapping</h1><p>راجعها ثم احفظها كقالب مخصص.</p></div></div><div class="card"><table class="table"><tr><th>System Field</th><th>Detected Column</th><th>Confidence</th></tr>{rows}</table><form method="post" action="/import/mapping/save-universal" style="margin-top:12px">{csrf(u)}<input type="hidden" name="mapping" value="{esc(json.dumps(mapping,ensure_ascii=False))}"><input name="name" placeholder="اسم القالب" required><button class="btn">Save Mapping</button></form></div>',u,'import'))
    def mapping_save(self,u):
        f=H.form(self)
        try: m=json.loads(f.get('mapping','{}'))
        except: m={}
        name=f.get('name','Universal Template').strip() or 'Universal Template'; c=db(); c.execute('INSERT INTO import_mappings(name,kind,mapping_json,created_by,created_at) VALUES(?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET mapping_json=excluded.mapping_json,created_by=excluded.created_by,created_at=excluded.created_at',(name,'employees',json.dumps(m,ensure_ascii=False),u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'EXCEL_MAPPING_SAVE','Import Mapping',name); self.redirect('/import/mapping')

    # -------- User provisioning --------
    _AR_LATIN=str.maketrans({'ا':'a','أ':'a','إ':'a','آ':'a','ب':'b','ت':'t','ث':'th','ج':'g','ح':'h','خ':'kh','د':'d','ذ':'dh','ر':'r','ز':'z','س':'s','ش':'sh','ص':'s','ض':'d','ط':'t','ظ':'z','ع':'a','غ':'gh','ف':'f','ق':'q','ك':'k','ل':'l','م':'m','ن':'n','ه':'h','و':'w','ي':'y','ى':'y','ة':'h','ؤ':'w','ئ':'y','ء':'a'})
    def suggest_username(name):
        parts=[]
        for x in str(name or '').split():
            y=re.sub(r'[^a-z0-9]+','',x.translate(_AR_LATIN).lower())
            if y: parts.append(y)
        return '.'.join(parts[:2]) if parts else 'employee'
    def make_password(): return secrets.token_urlsafe(9).replace('-','A').replace('_','9')[:12]+'!'
    def create_employee_user(self,u):
        f=H.form(self); emp=f.get('emp_code','').strip(); mode=f.get('password_mode','auto'); username=f.get('username','').strip(); pw=f.get('password','').strip()
        if not can(u,'users.manage') or not g['emp_allowed'](u,emp): return H.forbid(self,u)
        c=db(); e=c.execute('SELECT emp_code,name,email FROM employees WHERE emp_code=?',(emp,)).fetchone()
        if not e: c.close(); return H.send(self, page('User','<div class="card"><div class="alert">Employee not found.</div></div>',u),404)
        role=f.get('role') or 'Employee'
        allowed_roles={'SuperAdmin','Admin','HR','Manager','Employee'}
        if role not in allowed_roles: c.close(); return H.send(self, page('User','<div class="card"><div class="alert">الدور غير صالح.</div></div>',u),400)
        creator_role=u.get('role')
        if creator_role=='SuperAdmin': pass
        elif creator_role=='Admin' and role not in ('Admin','HR','Manager','Employee'): c.close(); return H.forbid(self,u)
        elif creator_role=='HR' and role not in ('Manager','Employee'): c.close(); return H.forbid(self,u)
        elif creator_role not in ('Admin','SuperAdmin','HR') and role!='Employee': c.close(); return H.forbid(self,u)
        existing=c.execute('SELECT id,username FROM users WHERE employee_code=? ORDER BY id DESC LIMIT 1',(emp,)).fetchone()
        if existing and not username: username=existing['username']
        if not username: username=suggest_username(e['name'])
        base=username; n=1
        while True:
            hit=c.execute('SELECT id,employee_code FROM users WHERE username=?',(username,)).fetchone()
            if not hit or (existing and hit['id']==existing['id']): break
            n+=1; username=f'{base}{n}'
        if mode=='auto' or not pw: pw=make_password()
        ph=g['hashpw'](pw)
        if existing:
            c.execute("UPDATE users SET username=?,password_hash=?,role=?,full_name=?,employee_code=?,must_change_password=1,scope_type='self',scope_value=?,active=1 WHERE id=?",(username,ph,role,e['name'],emp,emp,existing['id']))
        else:
            c.execute('INSERT INTO users(username,password_hash,role,full_name,employee_code,must_change_password,scope_type,scope_value,active) VALUES(?,?,?,?,?,?,?,?,1)',(username,ph,role,e['name'],emp,1,'self',emp))
        c.commit(); c.close(); audit(u['username'],u['role'],'EMPLOYEE_USER_UPDATED' if existing else 'EMPLOYEE_USER_CREATED','User Account',emp,username)
        body=f'''<div class="top"><div class="title"><h1>✅ User Created</h1><p>احفظ بيانات الدخول الآن. كلمة المرور لا يتم عرضها مرة أخرى.</p></div></div><div class="card hero-card" style="max-width:720px;margin:auto"><div class="grid g2"><div><b>Employee</b><div style="font-size:24px;margin-top:6px">{esc(e['name'])}</div><small>{esc(emp)}</small></div><div><b>Username</b><div style="font-size:24px;margin-top:6px">{esc(username)}</div></div><div><b>Temporary Password</b><div style="font-size:24px;margin-top:6px;font-family:Consolas">{esc(pw)}</div><small>سيُطلب تغييرها أول دخول.</small></div><div><b>Role</b><div style="font-size:20px;margin-top:6px">{esc(f.get('role') or 'Employee')}</div></div></div><div class="actions" style="margin-top:18px"><a class="btn" href="/employee/profile/{quote(emp)}">رجوع للبروفايل</a><button class="btn gray" onclick="navigator.clipboard.writeText('{esc(username)} / {esc(pw)}')">Copy credentials</button></div></div>'''
        H.send(self, page('User Created',body,u,'employees'))

    # -------- Modern employee list + QR --------
    old_employees=H.employees
    def employees_qr(self,u):
        captured=[]
        orig=self.send
        def cap(body,status=200,ctype='text/html',headers=None): captured.append((body,status,ctype,headers))
        self.send=cap
        try: old_employees(self,u)
        finally: self.send=orig
        if not captured:return
        body,status,ctype,headers=captured[0]
        if status!=200 or 'text/html' not in ctype:return orig(body,status,ctype,headers)
        s=body.decode('utf-8','replace') if isinstance(body,bytes) else body
        codes=re.findall(r'/employee/profile/([^"?]+)',s)
        unique=[]
        for code in codes:
            code=code.replace('%2F','/').replace('%20',' ')
            if code not in unique: unique.append(code)
        # THE MAIN CAUSE OF "الموظفون" PAGE FREEZING: this used to loop over every
        # employee on the page (up to 2000) and, for each one, open a brand-new
        # SQLite connection just to check for a QR code, and if missing,
        # SYNCHRONOUSLY GENERATE a QR image (crypto token + PNG render + file
        # write + DB write + an audit-chain write) right there inside the page
        # request — before anything was sent back to the browser. A company
        # with a few hundred employees that hadn't all been QR-provisioned yet
        # would trigger hundreds of connections and QR generations on every
        # single visit to this page, which is exactly what made it "load
        # forever" while other pages worked. On top of that, a second loop
        # opened yet another fresh connection per employee just to look up
        # their user account for a badge.
        #
        # Fix: two single batched queries (no work is generated automatically
        # just from viewing the list — issuing a QR is already a deliberate
        # action available from the employee profile / bulk operations pages).
        qr_map={}; user_map={}
        if unique:
            placeholders=','.join('?'*len(unique))
            try:
                cc=db()
                for r in cc.execute(f'SELECT emp_code,status,image_path FROM qr_identities WHERE emp_code IN ({placeholders})',unique).fetchall():
                    qr_map[r['emp_code']]=r
                for r in cc.execute(f'SELECT employee_code,username,active,id FROM users WHERE employee_code IN ({placeholders})',unique).fetchall():
                    prev=user_map.get(r['employee_code'])
                    if not prev or r['id']>prev['id']: user_map[r['employee_code']]=r
                cc.close()
            except Exception: pass
        s=s.replace('<th>الاسم</th>','<th>الاسم</th><th>QR</th><th>User</th>',1)
        def inject(m):
            code=m.group(1)
            row=qr_map.get(code)
            if row and row['status']=='active' and row['image_path']:
                qr_cell=f'<a href="/id-card/{quote(code)}" title="فتح البطاقة"><img class="qr-mini" src="/qr/image/{quote(code)}?v=1"></a>'
            else:
                qr_cell=f'<a class="btn gray" href="/employee/profile/{quote(code)}#qr" title="إصدار QR">إصدار QR</a>'
            rr=user_map.get(code)
            if rr: user_cell=f'<span class="badge b-ok">{esc(rr["username"])}</span>' if rr['active'] else '<span class="badge b-warn">Disabled</span>'
            else: user_cell=f'<a class="btn gray" href="/employee/profile/{quote(code)}#account">Create User</a>'
            return m.group(0)+f'<td>{qr_cell}</td><td>{user_cell}</td>'
        s=re.sub(r'<td><a href="/employee/profile/([^"?]+)"[^>]*>.*?</a></td>',inject,s,count=2000,flags=re.S)
        return orig(s,status,ctype,headers)
    H.employees=employees_qr

    # -------- Profile account card --------
    old_profile=H.employee_profile
    def profile_plus(self,u,code):
        captured=[]; orig=self.send
        def cap(body,status=200,ctype='text/html',headers=None): captured.append((body,status,ctype,headers))
        self.send=cap
        try: old_profile(self,u,code)
        finally: self.send=orig
        if not captured:return
        body,status,ctype,headers=captured[0]
        if status!=200 or 'text/html' not in ctype:return orig(body,status,ctype,headers)
        s=body.decode('utf-8','replace') if isinstance(body,bytes) else body
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone(); usr=c.execute('SELECT username,role,active,last_login FROM users WHERE employee_code=? ORDER BY id DESC LIMIT 1',(code,)).fetchone(); c.close()
        uname=suggest_username(e['name'] if e else code)
        if usr:
            account_inner = '<div class="grid g3"><div><b>Username</b><div>'+esc(usr['username'])+'</div></div><div><b>Role</b><div>'+esc(usr['role'])+'</div></div><div><b>Last Login</b><div>'+esc(usr['last_login'] or '—')+'</div></div></div>'
        else:
            account_inner = '<form method="post" action="/employee/user/create" class="form">'+csrf(u)+'<input type="hidden" name="emp_code" value="'+esc(code)+'"><div class="field"><label>Username</label><input name="username" value="'+esc(uname)+'"></div><div class="field"><label>Role</label><select name="role"><option>Employee</option><option>Manager</option><option>HR</option><option>Admin</option></select></div><div class="field"><label>Password mode</label><select name="password_mode"><option value="auto">Generate secure password</option><option value="manual">I will choose it</option></select></div><div class="field"><label>Manual password (optional)</label><input name="password" type="password" placeholder="اتركها فارغة للتوليد التلقائي"></div><div class="full"><button class="btn">Create User</button></div></form>'
        badge_class = 'b-ok' if usr and usr['active'] else 'b-warn'
        badge_text = 'Active' if usr and usr['active'] else 'Not created'
        account = '<div class="card hero-card" style="margin-top:16px"><div class="top"><div><h3>👤 User Account</h3><p>حساب الموظف للـEmployee Portal — الصلاحيات تُفرض على السيرفر.</p></div><span class="badge '+badge_class+'">'+badge_text+'</span></div>'+account_inner+'</div>'
        # Put the account card near the end so it does not disturb existing profile layout.
        tabs=f'''<div class="card no-print" style="margin-bottom:16px;padding:10px;position:sticky;top:8px;z-index:9"><div class="actions"><a class="btn gray" href="/employee/profile/{quote(code)}">Overview</a><a class="btn gray" href="/attendance?emp={quote(code)}">Attendance</a><a class="btn gray" href="/leaves?emp={quote(code)}">Leaves</a><a class="btn gray" href="/payroll?emp={quote(code)}">Payroll</a><a class="btn gray" href="/documents?emp={quote(code)}">Documents</a><a class="btn gray" href="/id-card/{quote(code)}">ID & QR</a></div></div>'''
        s=s.replace('</body>',tabs+account+'</body>')
        return orig(s,status,ctype,headers)
    H.employee_profile=profile_plus

    # -------- QR management center --------
    def qr_center(self,u):
        c=db(); rows=c.execute('SELECT e.emp_code,e.name,e.department,e.job,q.status,q.issued_at,q.image_path,usr.username,usr.active user_active FROM employees e LEFT JOIN qr_identities q ON q.emp_code=e.emp_code LEFT JOIN users usr ON usr.employee_code=e.emp_code AND usr.id=(SELECT MAX(u2.id) FROM users u2 WHERE u2.employee_code=e.emp_code) WHERE e.status<>? ORDER BY e.name LIMIT 2000',('مؤرشف',)).fetchall(); c.close()
        trs=''.join(f'<tr><td><input class="qrcheck" type="checkbox" value="{esc(r["emp_code"])}"></td><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["department"] or "—")}</td><td><span class="badge {"b-ok" if r["status"]=="active" else "b-warn"}">{esc(r["status"] or "not issued")}</span></td><td>{('<img class="qr-mini" src="/qr/image/'+quote(r["emp_code"])+'">') if r["status"]=="active" else "—"}</td><td>{(f'<span class="badge b-ok">{esc(r["username"])}</span>' if r["username"] and r["user_active"] else (f'<span class="badge b-warn">{esc(r["username"])}</span>' if r["username"] else f'<a class="btn gray" href="/employee/profile/{quote(r["emp_code"])}#account">Create User</a>'))}</td><td><a class="btn gray" href="/employee/profile/{quote(r["emp_code"])}">Profile</a></td></tr>' for r in rows if g['emp_allowed'](u,r['emp_code']))

        body=f'''<div class="top"><div class="title"><h1>🔐 QR Identity Center</h1><p>Generate / Regenerate / Revoke / Verify / Scan / Bulk export.</p></div><div class="actions"><a class="btn gray" href="/qr/scan">📷 Scanner</a><a class="btn gray" href="/qr/export">📦 Export ZIP + Excel</a></div></div><div class="card" style="margin-bottom:12px"><div class="actions"><button class="btn" onclick="bulk('generate')">Generate selected</button><button class="btn warn" onclick="bulk('regenerate')">Regenerate selected</button><button class="btn bad" onclick="bulk('revoke')">Revoke selected</button></div></div><div class="card table-wrap"><table class="table"><tr><th><input id="all" type="checkbox"></th><th>ID</th><th>Employee</th><th>Department</th><th>Status</th><th>QR</th><th>User</th><th></th></tr>{trs}</table></div><script>document.getElementById('all').onchange=e=>document.querySelectorAll('.qrcheck').forEach(x=>x.checked=e.target.checked);function bulk(a){{let ids=[...document.querySelectorAll('.qrcheck:checked')].map(x=>x.value);if(!ids.length)return;fetch('/qr/bulk',{{method:'POST',headers:{{'Content-Type':'application/x-www-form-urlencoded'}},body:new URLSearchParams({{'_csrf':'{esc(u.get('csrf',''))}','action':a,'emp_codes':ids.join(',')}})}}).then(()=>location.reload())}}</script>'''
        return self.send(page('QR Identity Center',body,u,'employees'),200)
    def qr_bulk(self,u):
        f=H.form(self)
        if f.get('_csrf')!=u.get('csrf'): return H.send(self, page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
        action=f.get('action','generate'); ids=[x for x in f.get('emp_codes','').split(',') if x][:500]; done=0
        for emp in ids:
            if not g['emp_allowed'](u,emp): continue
            try:
                if action=='revoke':
                    c=db(); c.execute('UPDATE qr_identities SET status="revoked",revoked_at=? WHERE emp_code=?',(now(),emp)); c.commit(); c.close()
                else: g['qr_identity_issue'](emp,u,action=='regenerate')
                done+=1
            except Exception: pass
        audit(u['username'],u['role'],'QR_BULK_'+action,'QR Identity','bulk',str(done)); self.redirect('/qr/center')

    def qr_export(self,u):
        c=db(); rows=c.execute('SELECT e.emp_code,e.name,e.department,e.job,e.location,q.status,q.issued_at,q.image_path FROM employees e LEFT JOIN qr_identities q ON q.emp_code=e.emp_code WHERE e.status<>? ORDER BY e.name',('مؤرشف',)).fetchall(); c.close()
        out=io.BytesIO(); wb=Workbook(); ws=wb.active; ws.title='QR Index'; ws.append(['Employee ID','Name','Department','Job','Location','QR Status','Issued At','QR File'])
        qdir=os.path.join(DATA,'qr');
        with zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED) as z:
            for r in rows:
                if not g['emp_allowed'](u,r['emp_code']): continue
                ws.append([r['emp_code'],r['name'],r['department'],r['job'],r['location'],r['status'],r['issued_at'],r['image_path']])
                if r['image_path'] and os.path.exists(os.path.join(DATA,r['image_path'])): z.write(os.path.join(DATA,r['image_path']),os.path.join('QR',r['emp_code']+'.png'))
            x=io.BytesIO(); wb.save(x); z.writestr('employees_qr.xlsx',x.getvalue())
        H.send(self, out.getvalue(),200,'application/zip',{'Content-Disposition':'attachment; filename="HR_Enterprise_QR_Export.zip"'})

    def leave_commit_alias(self,u): return self.v12_leave_commit(u)
    def leave_adjust_id(self,u):
        if not can(u,'leave.approve'): return H.forbid(self,u)
        f=H.form(self); rid=f.get('id')
        if f.get('_csrf')!=u.get('csrf'): return H.send(self,page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
        c=db(); old=c.execute('SELECT * FROM leave_balances WHERE id=?',(rid,)).fetchone()
        if not old: c.close(); return H.send(self,page('Leave','<div class="card"><div class="alert">الرصيد غير موجود.</div></div>',u),404)
        if not g['emp_allowed'](u,old['emp_code']): c.close(); return H.forbid(self,u)
        annual=clean_num(f.get('annual')); used=clean_num(f.get('used'))
        c.execute('UPDATE leave_balances SET annual=?,used=? WHERE id=?',(annual,used,rid)); c.execute('INSERT INTO leave_balance_history(emp_code,leave_type,action,amount,before_annual,before_used,after_annual,after_used,reason,source,actor,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(old['emp_code'],old['leave_type'],'manual_adjust',annual-used,float(old['annual']),float(old['used']),annual,used,f.get('reason',''),'manual',u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'LEAVE_BALANCE_ADJUST','Leave Balances',old['emp_code'],old['leave_type']); return self.redirect('/leave-balances')
    def shift_delete_v12(self,u):
        f=H.form(self); sid=int(f.get('id') or 0)
        if not can(u,'shifts.manage'): return H.forbid(self,u)
        c=db(); r=c.execute('SELECT * FROM shifts WHERE id=?',(sid,)).fetchone()
        if not r: c.close(); return self.redirect('/shifts')
        assigned=c.execute('SELECT COUNT(*) n FROM employee_shifts WHERE shift_id=?',(sid,)).fetchone()['n']
        if assigned:
            c.close(); return H.send(self,page('Shift Delete','<div class="card"><div class="alert">لا يمكن حذف الوردية لأنها مرتبطة بـ '+str(assigned)+' موظف. فك التعيين أولًا.</div></div>',u,'shifts'),409)
        c.execute('DELETE FROM shifts WHERE id=?',(sid,)); c.commit(); c.close(); audit(u['username'],u['role'],'SHIFT_DELETE','Shifts',str(sid),r['name']); return self.redirect('/shifts')
    old_shift_save=H.shift_save
    def shift_save_v12(self,u,f):
        if f.get('id'):
            sid=int(f.get('id') or 0); c=db(); c.execute('UPDATE shifts SET name=?,start_time=?,end_time=?,grace_minutes=?,warning_minutes=?,active=1 WHERE id=?',(f.get('name'),f.get('start_time') or '09:00',f.get('end_time') or '17:00',int(f.get('grace_minutes') or 0),int(f.get('warning_minutes') or 15),sid)); c.commit(); c.close(); audit(u['username'],u['role'],'SHIFT_UPDATE','Shifts',str(sid),f.get('name')); return self.redirect('/shifts')
        return old_shift_save(self,u,f)
    H.v12_shift_delete=shift_delete_v12; H.shift_save=shift_save_v12
    def export_employee_master(self):
        u=self.require();
        if not u or not can(u,'reports.export'): return H.forbid(self,u)
        c=db(); rows=c.execute('SELECT e.emp_code,e.name,e.department,e.job,e.location,e.status,e.hire_date,q.id qr_id,q.status qr_status,q.issued_at,q.image_path,usr.username,usr.role,usr.active user_active FROM employees e LEFT JOIN qr_identities q ON q.emp_code=e.emp_code LEFT JOIN users usr ON usr.employee_code=e.emp_code AND usr.id=(SELECT MAX(u2.id) FROM users u2 WHERE u2.employee_code=e.emp_code) ORDER BY e.name').fetchall(); c.close(); wb=Workbook(); ws=wb.active; ws.title='Employees'; ws.append(['Employee ID','Name','Department','Job','Location','Status','Hire Date','Username','User Role','User Status','QR ID','QR Status','QR Issued','QR File']); [ws.append([r[k] for k in ('emp_code','name','department','job','location','status','hire_date','username','role','user_active','qr_id','qr_status','issued_at','image_path')]) for r in rows if g['emp_allowed'](u,r['emp_code'])]; out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="HR_Employee_Master.xlsx"'})
    def export_leave_balances_v12(self):
        u=self.require();
        if not u or not can(u,'reports.export'): return H.forbid(self,u)
        c=db(); rows=c.execute('SELECT lb.emp_code,e.name,lb.leave_type,lb.annual,lb.used,(lb.annual-lb.used) remaining FROM leave_balances lb JOIN employees e ON e.emp_code=lb.emp_code ORDER BY e.name,lb.leave_type').fetchall(); c.close(); wb=Workbook(); ws=wb.active; ws.title='Leave Balances'; ws.append(['Employee ID','Name','Leave Type','Annual','Used','Remaining']); [ws.append([r['emp_code'],r['name'],r['leave_type'],r['annual'],r['used'],r['remaining']]) for r in rows if g['emp_allowed'](u,r['emp_code'])]; out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="HR_Leave_Balances.xlsx"'})
    def export_leave_template(self):
        u=self.require();
        if not u or not can(u,'reports.export'): return H.forbid(self,u)
        wb=Workbook(); ws=wb.active; ws.title='Leave Balance Template'; ws.append(['Employee ID','Name','Leave Type','Annual','Used','Remaining']); ws.append(['EMP001','Ahmed Elsayed','اعتيادي',21,5,16]); ws.freeze_panes='A2'; out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Leave_Balance_Template.xlsx"'})
    def admin_flex(self,u):
        body='<div class="top"><div class="title"><h1>🛠 Enterprise Flex Center</h1><p>Admin controls for imports, identity, branding, backups and network.</p></div></div><div class="grid g3"><div class="card"><h3>Excel</h3><a class="btn" href="/import/mapping">Universal Mapping</a></div><div class="card"><h3>Identity</h3><a class="btn" href="/qr/center">QR Identity</a></div><div class="card"><h3>Branding</h3><a class="btn" href="/branding/id-card-template">ID Designer</a></div></div>'
        H.send(self,page('Enterprise Flex Center',body,u,'system'))
    H.v12_leave_commit_alias=leave_commit_alias; H.v12_leave_adjust_id=leave_adjust_id; H.v12_export_employee_master=export_employee_master; H.v12_export_leave_balances=export_leave_balances_v12; H.v12_export_leave_template=export_leave_template; H.v12_admin_flex=admin_flex

    def id_designer(self,u):
        if not can(u,'settings.manage'): return H.forbid(self,u)
        c=db(); t=c.execute('SELECT * FROM id_card_templates WHERE name="Default"').fetchone(); c.close(); t=dict(t) if t else {}
        vals={k:t.get(k,default) for k,default in [('photo_x',40),('photo_y',150),('photo_w',130),('photo_h',160),('name_x',190),('name_y',170),('qr_x',650),('qr_y',145),('qr_size',150)]}
        front=t.get('background_path') or ''
        body=f'''<div class="top"><div class="title"><h1>🎨 ID Card Designer</h1><p>ارفع تصميمك ثم اسحب الصورة والاسم والـQR بالماوس. كل شيء محفوظ داخل النظام.</p></div></div>
        <div class="grid g2"><div class="card"><form method="post" action="/branding/id-card-template/save" enctype="multipart/form-data">{csrf(u)}<div class="grid g2"><div class="field"><label>Front Template</label><input type="file" name="front" accept=".png,.jpg,.jpeg,.webp,.svg"></div><div class="field"><label>Back Template</label><input type="file" name="back" accept=".png,.jpg,.jpeg,.webp,.svg"></div></div>
        <input type="hidden" id="photo_x" name="photo_x" value="{vals['photo_x']}"><input type="hidden" id="photo_y" name="photo_y" value="{vals['photo_y']}"><input type="hidden" id="photo_w" name="photo_w" value="{vals['photo_w']}"><input type="hidden" id="photo_h" name="photo_h" value="{vals['photo_h']}"><input type="hidden" id="name_x" name="name_x" value="{vals['name_x']}"><input type="hidden" id="name_y" name="name_y" value="{vals['name_y']}"><input type="hidden" id="qr_x" name="qr_x" value="{vals['qr_x']}"><input type="hidden" id="qr_y" name="qr_y" value="{vals['qr_y']}"><input type="hidden" id="qr_size" name="qr_size" value="{vals['qr_size']}">
        <div class="actions" style="margin-top:12px"><label><input type="checkbox" name="show_emp_id" {'checked' if t.get('show_emp_id',1) else ''}> Employee ID</label><label><input type="checkbox" name="show_department" {'checked' if t.get('show_department',1) else ''}> Department</label><label><input type="checkbox" name="show_job" {'checked' if t.get('show_job',1) else ''}> Job</label></div><div class="actions" style="margin-top:14px"><button class="btn">💾 Save Designer</button><button class="btn gray" name="action" value="restore">Restore Default</button></div></form><div class="alert" style="margin-top:12px">الأحجام والإحداثيات تعمل بوحدة معاينة البطاقة 860×540. بعد الحفظ يطبق التصميم على بطاقات الموظفين.</div></div>
        <div class="card"><h3>Live Preview — Drag & Drop</h3><div id="stage" style="position:relative;width:860px;max-width:100%;aspect-ratio:860/540;border-radius:22px;overflow:hidden;background:#0b1220 url('/branding/id-card-template/image') center/cover no-repeat;color:white;border:1px solid #344054;box-shadow:0 20px 60px rgba(16,24,40,.18)"><div id="photo" class="drag" data-x="{vals['photo_x']}" data-y="{vals['photo_y']}" style="position:absolute;left:{vals['photo_x']}px;top:{vals['photo_y']}px;width:{vals['photo_w']}px;height:{vals['photo_h']}px;background:#ffffffdd;color:#475467;border-radius:14px;display:grid;place-items:center;font-weight:800">PHOTO</div><div id="name" class="drag" data-x="{vals['name_x']}" data-y="{vals['name_y']}" style="position:absolute;left:{vals['name_x']}px;top:{vals['name_y']}px;font-size:28px;font-weight:900">Employee Name</div><div id="qr" class="drag" data-x="{vals['qr_x']}" data-y="{vals['qr_y']}" style="position:absolute;left:{vals['qr_x']}px;top:{vals['qr_y']}px;width:{vals['qr_size']}px;height:{vals['qr_size']}px;background:white;color:#101828;border-radius:14px;display:grid;place-items:center;font-weight:900">QR</div></div><div id="dragStatus" class="footer">اسحب العناصر داخل البطاقة ثم احفظ.</div></div></div>
        <script>(function(){{const stage=document.getElementById('stage');const items=[['photo','photo_x','photo_y'],['name','name_x','name_y'],['qr','qr_x','qr_y']];let active=null,start=null;items.forEach(([id,xk,yk])=>{{const el=document.getElementById(id);el.addEventListener('pointerdown',e=>{{active=el;start={{sx:e.clientX,sy:e.clientY,x:parseFloat(el.dataset.x),y:parseFloat(el.dataset.y)}};el.setPointerCapture(e.pointerId);e.preventDefault()}});el.addEventListener('pointermove',e=>{{if(!active||active!==el)return;const scale=stage.clientWidth/860;let x=Math.max(0,Math.min(860-el.offsetWidth,(start.x+(e.clientX-start.sx)/scale)));let y=Math.max(0,Math.min(540-el.offsetHeight,(start.y+(e.clientY-start.sy)/scale)));el.dataset.x=x;el.dataset.y=y;el.style.left=x+'px';el.style.top=y+'px';document.getElementById(xk).value=Math.round(x);document.getElementById(yk).value=Math.round(y);document.getElementById('dragStatus').textContent=id.toUpperCase()+': X '+Math.round(x)+' · Y '+Math.round(y)}});el.addEventListener('pointerup',()=>active=null)}});}})();</script>'''
        H.send(self,page('ID Card Designer',body,u,'settings'))

    H.v12_leave_page=leave_page; H.v12_leave_import=leave_import; H.v12_leave_import_post=leave_import_post; H.v12_leave_paste=leave_paste; H.v12_leave_commit=leave_commit; H.v12_leave_adjust=leave_adjust; H.v12_leave_history=leave_history
    H.v12_mapping_page=mapping_page; H.v12_mapping_auto=mapping_auto; H.v12_mapping_save=mapping_save; H.v12_create_employee_user=create_employee_user; H.v12_qr_center=qr_center; H.v12_qr_bulk=qr_bulk; H.v12_qr_export=qr_export; H.v12_id_designer=id_designer

    # -------- ID card designer: route to the existing flexible designer and improve it with drag/drop --------
    old_get=H.do_GET; old_post=H.do_POST
    def get(self):
        p=urlparse(self.path).path
        custom= p in ('/leave-balances/import','/leave-balances/history','/leave-balances','/import/mapping','/qr/center','/qr/export','/branding/id-card-template','/export/employee-master','/export/leave-balances','/export/leave-balances/template','/admin/flex') or p.startswith('/qr/center')
        if p=='/leave-balances': return self.v12_leave_page(self.user() or {}) if self.user() else None
        if p=='/leave-balances/import':
            u=self.require(); return self.v12_leave_import(u) if u else None
        if p=='/leave-balances/history':
            u=self.require(); return self.v12_leave_history(u) if u else None
        if p=='/import/mapping':
            u=self.require(); return self.v12_mapping_page(u) if u else None
        if p=='/qr/center':
            u=self.require(); return self.v12_qr_center(u) if u else None
        if p=='/qr/export':
            u=self.require(); return self.v12_qr_export(u) if u else None
        if p=='/branding/id-card-template':
            u=self.require(); return self.v12_id_designer(u) if u else None
        if p=='/export/employee-master': return self.v12_export_employee_master()
        if p=='/export/leave-balances': return self.v12_export_leave_balances()
        if p=='/export/leave-balances/template': return self.v12_export_leave_template()
        if p=='/admin/flex':
            u=self.require(); return self.v12_admin_flex(u) if u else None
        return old_get(self)
    def post(self):
        p=urlparse(self.path).path
        if p in ('/leave-balances/import','/leave-balances/paste','/leave-balances/import/commit','/leave-balances/adjust','/import/mapping/auto','/import/mapping/save-universal','/employee/user/create','/qr/bulk','/leave-balances/commit','/leave-balance/save','/shift/delete','/shift/save'):
            u=self.require()
            if not u:return
            if p=='/leave-balances/import': return self.v12_leave_import_post(u)
            if p=='/leave-balances/paste': return self.v12_leave_paste(u)
            if p=='/leave-balances/import/commit': return self.v12_leave_commit(u)
            if p=='/leave-balances/adjust': return self.v12_leave_adjust(u)
            if p=='/import/mapping/auto': return self.v12_mapping_auto(u)
            if p=='/import/mapping/save-universal': return self.v12_mapping_save(u)
            if p=='/leave-balances/commit': return self.v12_leave_commit_alias(u)
            if p=='/leave-balance/save': return self.v12_leave_adjust_id(u)
            if p=='/shift/delete': return self.v12_shift_delete(u)
            if p=='/employee/user/create':
                f=H.form(self)
                if f.get('_csrf')!=u.get('csrf'): return H.send(self, page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
                return self.v12_create_employee_user(u)
            if p=='/qr/bulk': return self.v12_qr_bulk(u)
        return old_post(self)
    H.do_GET=get; H.do_POST=post

    # Add nav links via the latest page wrapper.
    old_page=g['page']
    def page_nav(title,body,user,active='dashboard'):
        out=old_page(title,body,user,active)
        if user:
            extra=''
            if can(user,'employees.view'): extra+='<a href="/qr/center">🔐 مركز QR الجماعي</a>'
            if can(user,'settings.manage'): extra+='<a href="/branding/id-card-template">🎨 مصمم البطاقة</a>'
            if extra: extra='<details class="nav-group"><summary>QR متقدم</summary>'+extra+'</details>'
            out=out.replace('</nav>',extra+'</nav>',1)
            font=esc(g['setting']('ui_font') or 'Segoe UI')
            out=out.replace('</head>',f'<style>body,.field input,.field select,.field textarea,.toolbar input,.toolbar select,.btn{{font-family:{font},Tahoma,Arial,sans-serif}}</style></head>',1)
        return out
    g['page']=page_nav

    old_settings_page=H.settings_page
    def settings_plus(self,u):
        captured=[]; orig=self.send
        def cap(body,status=200,ctype='text/html; charset=utf-8',headers=None): captured.append((body,status,ctype,headers))
        self.send=cap
        try: old_settings_page(self,u)
        finally: self.send=orig
        if not captured: return
        body,status,ctype,headers=captured[0]
        if status!=200 or 'text/html' not in ctype: return orig(body,status,ctype,headers)
        sbody=body.decode('utf-8','replace') if isinstance(body,bytes) else body
        ss=db(); cfg={r['key']:r['value'] for r in ss.execute('SELECT key,value FROM settings').fetchall()}; ss.close()
        extra=f'''<div class="card" style="margin-top:16px"><h3>✨ Advanced Branding</h3><form method="post" action="/settings/save" class="form">{csrf(u)}<div class="field"><label>UI Font</label><select name="ui_font"><option {'selected' if cfg.get('ui_font')=='Segoe UI' else ''}>Segoe UI</option><option {'selected' if cfg.get('ui_font')=='Tahoma' else ''}>Tahoma</option><option {'selected' if cfg.get('ui_font')=='Arial' else ''}>Arial</option></select></div><div class="field"><label>Login tagline</label><input name="login_tagline" value="{esc(cfg.get('login_tagline','Enterprise HR Management'))}"></div><div class="full"><button class="btn">Save Branding UI</button></div></form><div class="actions" style="margin-top:10px"><a class="btn gray" href="/branding/id-card-template">🎨 ID Card Designer</a><a class="btn gray" href="/branding/logo" target="_blank">🖼 Preview Logo</a></div></div>'''
        sbody=sbody.replace('</main>',extra+'</main>',1)
        return orig(sbody,status,ctype,headers)
    H.settings_page=settings_plus
    old_save_settings=H.save_settings
    def save_settings_plus(self,u,f):
        result=old_save_settings(self,u,f)
        c=db()
        for k in ('ui_font','login_tagline'):
            if k in f: c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',(k,f.get(k,'')))
        c.commit(); c.close(); clear_settings_cache()
        return result
    H.save_settings=save_settings_plus

    # Provide a modern login renderer using current branding settings.
    old_login=g.get('login_page')
    if old_login:
        def login(msg=''):
            company=esc(g['setting']('company_name') or 'HR Enterprise'); logo='/branding/logo' if g['setting']('company_logo') else ''
            bg=g['setting']('login_background') or ''
            accent={'blue':'#175cd3','purple':'#7f56d9','emerald':'#039855','orange':'#f79009'}.get(g['setting']('accent_color') or 'blue','#175cd3')
            return f'''<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{company} — Login</title><style>*{{box-sizing:border-box}}body{{margin:0;font-family:{g['setting']('ui_font') or 'Segoe UI'},Tahoma,Arial,sans-serif;background:#07111f;color:#fff}}.login{{min-height:100vh;display:grid;grid-template-columns:1.15fr .85fr;overflow:hidden;background:radial-gradient(circle at 10% 10%,{accent} 0,transparent 34%),linear-gradient(135deg,#07111f,#0e1a2e 60%,#101828)}}.visual{{padding:8vw;display:flex;flex-direction:column;justify-content:center;position:relative}}.visual h1{{font-size:clamp(42px,5vw,76px);line-height:1.05;margin:0 0 18px}}.visual p{{font-size:18px;color:#cbd5e1;max-width:620px;line-height:1.8}}.card{{width:min(470px,92vw);margin:auto;background:rgba(255,255,255,.97);color:#101828;border-radius:28px;padding:36px;box-shadow:0 35px 100px rgba(0,0,0,.35);backdrop-filter:blur(16px)}}.card img{{display:block;max-width:170px;max-height:90px;margin:0 auto 18px;object-fit:contain}}input{{width:100%;padding:15px;border:1px solid #d0d5dd;border-radius:13px;margin:7px 0;font:inherit}}button{{width:100%;border:0;border-radius:13px;padding:15px;margin-top:12px;background:{accent};color:#fff;font-weight:800;font-size:16px;cursor:pointer}}.msg{{padding:11px;border-radius:12px;background:#fff4ed;color:#9a3412;margin-bottom:10px}}@media(max-width:900px){{.login{{grid-template-columns:1fr}}.visual{{display:none}}}}</style></head><body><div class="login"><section class="visual"><div style="font-size:13px;letter-spacing:.15em;color:#7dd3fc;font-weight:800">HR ENTERPRISE</div><h1>{company}</h1><p>{esc(g['setting']('login_tagline') or 'Enterprise HR Management')}<br>People · Attendance · Leaves · Payroll · Identity</p></section><form class="card" method="post" action="/login">{f'<img src="{logo}">' if logo else '<div style="height:70px"></div>'}<h2 style="text-align:center">تسجيل الدخول</h2>{f'<div class="msg">{esc(msg)}</div>' if msg else ''}<input name="username" placeholder="اسم المستخدم" autofocus><input name="password" type="password" placeholder="كلمة المرور"><button>دخول آمن</button><p style="text-align:center;color:#667085;font-size:12px;margin-top:18px">{company} · Secure Employee Management</p></form></div></body></html>'''
        g['login_page']=login

    # Export helper for other patches/tests.
    g['_V12_H_OBJ']=H
    g['V12_PREVIEWS']={}
    g['v12_detect_mapping']=detect_mapping
