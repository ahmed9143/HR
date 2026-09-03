import os, io, json, base64, secrets, hashlib, mimetypes, socket
from urllib.parse import quote, urlparse, parse_qs
from datetime import date, timedelta


def install_enterprise(g):
    db=g['db']; now=g['now']; page=g['page']; esc=g['esc']; csrf_field=g['csrf_field']; can=g['can']; H=g['H']; DATA=g['DATA']; BRAND=g['BRAND']
    clear_settings_cache=g.get('clear_settings_cache', lambda: None)
    try:
        import qrcode
    except Exception:
        qrcode=None

    def upgrade():
        c=db()
        c.executescript('''
        CREATE TABLE IF NOT EXISTS qr_identities(
            id INTEGER PRIMARY KEY, emp_code TEXT UNIQUE NOT NULL, token_hash TEXT UNIQUE NOT NULL,
            issued_at TEXT NOT NULL, revoked_at TEXT, status TEXT DEFAULT 'active',
            created_by TEXT, regenerated_from TEXT, image_path TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_qr_status ON qr_identities(status);
        CREATE TABLE IF NOT EXISTS contracts(
            id INTEGER PRIMARY KEY, emp_code TEXT NOT NULL, contract_no TEXT, contract_type TEXT DEFAULT 'employment',
            start_date TEXT, end_date TEXT, amount REAL DEFAULT 0, status TEXT DEFAULT 'active',
            notes TEXT, attachment_path TEXT, created_by TEXT, created_at TEXT, updated_at TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_contracts_emp ON contracts(emp_code,start_date,end_date);
        CREATE TABLE IF NOT EXISTS training_programs(
            id INTEGER PRIMARY KEY, name TEXT UNIQUE, provider TEXT, description TEXT, active INTEGER DEFAULT 1, created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS training_enrollments(
            id INTEGER PRIMARY KEY, program_id INTEGER, emp_code TEXT NOT NULL, start_date TEXT, end_date TEXT,
            completion_date TEXT, score REAL DEFAULT 0, certificate_no TEXT, certificate_expiry TEXT,
            status TEXT DEFAULT 'enrolled', notes TEXT, created_by TEXT, created_at TEXT,
            FOREIGN KEY(program_id) REFERENCES training_programs(id)
        );
        CREATE INDEX IF NOT EXISTS idx_training_enroll_emp ON training_enrollments(emp_code);
        CREATE INDEX IF NOT EXISTS idx_training_enroll_expiry ON training_enrollments(certificate_expiry);
        CREATE TABLE IF NOT EXISTS id_card_templates(
            id INTEGER PRIMARY KEY, name TEXT UNIQUE, width_mm REAL DEFAULT 86, height_mm REAL DEFAULT 54,
            background_path TEXT, show_logo INTEGER DEFAULT 1, show_photo INTEGER DEFAULT 1,
            show_qr INTEGER DEFAULT 1, created_by TEXT, created_at TEXT
        );
        ''')
        cols={r['name'] for r in c.execute('PRAGMA table_info(qr_identities)').fetchall()}
        # Self-healing migration: qr_identities may exist from an older build that predates
        # one or more of these columns. CREATE TABLE IF NOT EXISTS above is a no-op on an
        # existing table, so every column that schema needs has to be checked individually,
        # not just image_path — otherwise bulk QR generation fails with
        # "table qr_identities has no column named token_hash" on any database created
        # before token_hash existed.
        qr_required_cols={
            'token_hash':'TEXT','issued_at':'TEXT',"status":"TEXT DEFAULT 'active'",
            'revoked_at':'TEXT','created_by':'TEXT','regenerated_from':'TEXT','image_path':'TEXT',
        }
        for col,coltype in qr_required_cols.items():
            if col not in cols:
                c.execute(f'ALTER TABLE qr_identities ADD COLUMN {col} {coltype}')
        # token_hash was declared UNIQUE in the CREATE TABLE above; when the column is added
        # later via ALTER TABLE that constraint doesn't come along automatically, so recreate
        # it as an index. SQLite allows multiple NULLs under a UNIQUE index, so this is safe
        # even for older rows that don't have a token_hash yet.
        try:
            c.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_qr_token_hash ON qr_identities(token_hash)')
        except Exception:
            pass  # pre-existing duplicate token_hash values on an old DB; log and fix manually rather than block startup
        c.execute("INSERT OR IGNORE INTO id_card_templates(name,created_at,created_by) VALUES('Default',?, 'system')",(now(),))
        c.commit(); c.close()

    old_init=g['init']
    def init_wrapped():
        old_init(); upgrade()
    g['init']=init_wrapped

    def token_hash(token):
        return hashlib.sha256(token.encode()).hexdigest()

    def issue_qr(emp_code,user,regenerate=False):
        if not qrcode: raise RuntimeError('qrcode package is required')
        c=db(); emp=c.execute('SELECT emp_code,name FROM employees WHERE emp_code=?',(emp_code,)).fetchone()
        if not emp:
            c.close(); raise ValueError('Employee not found')
        old=c.execute('SELECT * FROM qr_identities WHERE emp_code=?',(emp_code,)).fetchone()
        if old and not regenerate and old['status']=='active':
            c.close(); return None
        old_token=old['token_hash'] if old else None
        token=secrets.token_urlsafe(32)
        raw=qr_png(token)
        qdir=os.path.join(DATA,'qr'); os.makedirs(qdir,exist_ok=True)
        rel=os.path.join('qr',hashlib.sha256(str(emp_code).encode('utf-8')).hexdigest()[:24]+'.png').replace('\\','/')
        with open(os.path.join(DATA,rel),'wb') as fh: fh.write(raw)
        if old:
            c.execute('UPDATE qr_identities SET token_hash=?,issued_at=?,revoked_at=NULL,status="active",created_by=?,regenerated_from=?,image_path=? WHERE emp_code=?',(token_hash(token),now(),user['username'],old_token,rel,emp_code))
        else:
            c.execute('INSERT INTO qr_identities(emp_code,token_hash,issued_at,status,created_by,regenerated_from,image_path) VALUES(?,?,?,?,?,?,?)',(emp_code,token_hash(token),now(),'active',user['username'],None,rel))
        c.commit(); c.close()
        g['audit'](user['username'],user['role'],'QR_CREATED' if not old else 'QR_REGENERATED','QR Identity',emp_code,'active')
        return token

    def current_token(emp_code):
        c=db(); r=c.execute('SELECT * FROM qr_identities WHERE emp_code=? AND status="active"',(emp_code,)).fetchone(); c.close(); return r

    def qr_png(token):
        if not qrcode: return None
        # ERROR_CORRECT_H (~30% recovery) is required, not M (~15%), because a
        # company-logo mark is optionally composited into the center below —
        # M-level correction cannot reliably survive that much covered area.
        qr=qrcode.QRCode(version=None,error_correction=qrcode.constants.ERROR_CORRECT_H,box_size=8,border=4)
        base=(g['setting']('server_url') or '').strip().rstrip('/')
        if not base: base=f"http://{g['local_ip']()}:{g['PORT']}"
        qr.add_data(base+'/qr/verify/'+token); qr.make(fit=True)
        img=qr.make_image().convert('RGB')
        # Optional company-logo mark in the QR center. Keep a white quiet-zone around
        # the mark so printed cards remain reliably scannable.
        try:
            from PIL import Image
            logo_rel=g['setting']('company_logo') or ''
            if logo_rel:
                lp=os.path.join(DATA,logo_rel)
                if os.path.exists(lp):
                    logo=Image.open(lp).convert('RGBA')
                    # Keep the mark (logo + white pad) at roughly <=18% of the QR's
                    # width. Even with ERROR_CORRECT_H, a larger mark risks
                    # unscannable codes on lower-end phone cameras/printers.
                    side=max(36,int(min(img.size)*0.16)); logo.thumbnail((side,side),Image.LANCZOS)
                    pad=8; box=Image.new('RGBA',(logo.width+pad*2,logo.height+pad*2),'white')
                    box.alpha_composite(logo,(pad,pad)); x=(img.width-box.width)//2; y=(img.height-box.height)//2
                    img=img.convert('RGBA'); img.alpha_composite(box,(x,y)); img=img.convert('RGB')
        except Exception:
            pass
        out=io.BytesIO(); img.save(out,format='PNG'); return out.getvalue()

    def qr_image(self,u,emp_code):
        if not can(u,'employees.view') or not g['emp_allowed'](u,emp_code): return self.forbid(u)
        c=db(); row=c.execute('SELECT image_path,status FROM qr_identities WHERE emp_code=?',(emp_code,)).fetchone(); c.close()
        if not row or not row['image_path']: return self.send(b'QR not issued',404,'text/plain')
        p=os.path.join(DATA,row['image_path'])
        if not os.path.exists(p): return self.send(b'QR image unavailable; regenerate identity',409,'text/plain')
        self.send(open(p,'rb').read(),200,'image/png',{'Cache-Control':'no-store'})

    def ensure_token(emp_code,u,regenerate=False):
        row=current_token(emp_code)
        if row and not regenerate and row['status']=='active' and row['image_path']:
            return 'active'
        return issue_qr(emp_code,u,regenerate) or 'active'

    def qr_verify(self,u,token):
        c=db(); r=c.execute('SELECT q.*,e.name,e.department,e.job,e.status emp_status FROM qr_identities q JOIN employees e ON e.emp_code=q.emp_code WHERE q.token_hash=?',(token_hash(token),)).fetchone(); c.close()
        if not r: return self.send(page('QR Verification','<div class="card"><div class="alert">QR غير صالح أو غير معروف.</div></div>',u),404)
        valid=r['status']=='active'
        try:
            cc=db(); cc.execute('INSERT INTO qr_scan_history(emp_code,token_hash,scanned_by,ip,result,scanned_at,details) VALUES(?,?,?,?,?,?,?)',(r['emp_code'],token_hash(token),u.get('username','anonymous'),getattr(self,'client_address',[None])[0] if getattr(self,'client_address',None) else '', 'valid' if valid else 'revoked',now(),r['status'])); cc.commit(); cc.close()
        except Exception: pass
        g['audit'](u['username'],u['role'],'QR_VERIFIED','QR Identity',r['emp_code'],r['status'])
        badge='<span class="badge b-ok">VALID</span>' if valid else '<span class="badge b-bad">REVOKED</span>'
        body=f'<div class="top"><div class="title"><h1>🔐 QR Identity</h1><p>تحقق من هوية الموظف بدون كشف بيانات حساسة داخل الـQR.</p></div></div><div class="card"><h2>{esc(r["name"])}</h2><p>Employee ID: <b>{esc(r["emp_code"])}</b></p><p>Department: {esc(r["department"] or "—")}</p><p>Job: {esc(r["job"] or "—")}</p><p>Status: {badge}</p><p>QR status: <b>{esc(r["status"])}</b></p><a class="btn" href="/employee/profile/{esc(r["emp_code"])}">فتح الملف</a></div>'
        self.send(page('QR Verification',body,u,'employees'),200)

    def qr_scan(self,u):
        body='''<div class="top"><div class="title"><h1>📷 QR Scanner</h1><p>يدعم USB keyboard-wedge scanner، كاميرا الهاتف/الويب، أو لصق كود QR يدويًا.</p></div></div>
<div class="card">
  <div style="display:flex;gap:16px;flex-wrap:wrap;align-items:flex-end">
    <div style="flex:1;min-width:240px">
      <label>الصق كود QR هنا (Paste QR ID) — أو امسحه بجهاز USB</label>
      <input id="scan" autofocus style="width:100%;padding:14px;border:1px solid #d0d5dd;border-radius:10px;font-size:18px" placeholder="الصق أو امسح كود QR هنا">
    </div>
    <button class="btn gray" id="camBtn" type="button">📷 فتح الكاميرا</button>
    <button class="btn" id="verifyBtn" type="button">تحقق</button>
  </div>
  <div id="camWrap" style="display:none;margin-top:16px">
    <video id="camVideo" style="width:100%;max-width:420px;border-radius:12px;background:#000" playsinline muted></video>
    <canvas id="camCanvas" style="display:none"></canvas>
    <div style="margin-top:8px;color:#667085;font-size:13px">وجّه الكاميرا نحو كود QR الخاص بالموظف — سيتم التحقق تلقائيًا.</div>
  </div>
  <div id="result" style="margin-top:14px"></div>
</div>
<script>
(function(){
  // Note: no external CDN script here on purpose — this app serves a strict
  // Content-Security-Policy (script-src 'self'), so a 3rd-party QR library
  // would just get silently blocked by the browser. We use the browser's own
  // BarcodeDetector API instead (no external script needed); where it isn't
  // available (older Firefox/Safari) we say so plainly and point at the
  // paste field / USB scanner, which always work.
  const s=document.getElementById('scan'), res=document.getElementById('result');
  function go(){
    let v=s.value.trim(); if(!v) return;
    if(v.includes('/qr/verify/')) v=v.split('/qr/verify/').pop();
    v=v.split('/').filter(Boolean).pop();
    if(v) location.href='/qr/verify/'+encodeURIComponent(v);
  }
  s.addEventListener('keydown',function(e){ if(e.key==='Enter'){ e.preventDefault(); go(); } });
  document.getElementById('verifyBtn').addEventListener('click',go);

  const camBtn=document.getElementById('camBtn'), camWrap=document.getElementById('camWrap'),
        video=document.getElementById('camVideo'), canvas=document.getElementById('camCanvas');
  let stream=null, raf=null, detector=null, supported=false;
  if('BarcodeDetector' in window){ try{ detector=new BarcodeDetector({formats:['qr_code']}); supported=true; }catch(e){ detector=null; } }
  if(!supported){ camBtn.title='هذا المتصفح لا يدعم قراءة QR بالكاميرا تلقائيًا — استخدم اللصق أو ماسح USB'; }

  async function startCam(){
    if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia){
      res.innerHTML='<div class="alert">المتصفح لا يدعم فتح الكاميرا — استخدم اللصق أو ماسح USB.</div>'; return;
    }
    if(!supported){
      res.innerHTML='<div class="alert">هذا المتصفح لا يدعم قراءة QR تلقائيًا من الكاميرا. الرجاء استخدام حقل اللصق بالأعلى أو ماسح USB — كلاهما يعمل بشكل كامل.</div>'; return;
    }
    try{
      stream=await navigator.mediaDevices.getUserMedia({video:{facingMode:'environment'}});
      video.srcObject=stream; await video.play();
      camWrap.style.display='block'; camBtn.textContent='⏹ إيقاف الكاميرا'; res.innerHTML='';
      tick();
    }catch(e){ res.innerHTML='<div class="alert">تعذّر فتح الكاميرا: '+(e.message||e)+'. تأكد من السماح للمتصفح بالوصول للكاميرا.</div>'; }
  }
  function stopCam(){
    if(stream){ stream.getTracks().forEach(function(t){t.stop();}); stream=null; }
    if(raf) cancelAnimationFrame(raf);
    camWrap.style.display='none'; camBtn.textContent='📷 فتح الكاميرا';
  }
  async function tick(){
    if(!stream) return;
    try{
      const codes=await detector.detect(video);
      if(codes && codes.length){ s.value=codes[0].rawValue; stopCam(); go(); return; }
    }catch(e){}
    raf=requestAnimationFrame(tick);
  }
  camBtn.addEventListener('click',function(){ stream ? stopCam() : startCam(); });
})();
</script>'''
        self.send(page('QR Scanner',body,u,'employees'))

    def id_template_page(self,u):
        if not can(u,'settings.manage'): return self.forbid(u)
        d=os.path.join(BRAND,'id_card_template'); files=[]
        if os.path.isdir(d): files=[x for x in os.listdir(d) if os.path.splitext(x)[1].lower() in ('.png','.jpg','.jpeg','.svg')]
        cur=files[0] if files else ''
        body=f'''<div class="top"><div class="title"><h1>🎨 ID Card Designer</h1><p>ارفع خلفية بطاقة الهوية من داخل النظام بدون تعديل الكود.</p></div></div><div class="grid g2"><div class="card"><form method="post" action="/branding/id-card-template/save" enctype="multipart/form-data">{csrf_field(u)}<div class="field"><label>Card background / template</label><input type="file" name="file" accept=".png,.jpg,.jpeg,.svg" required></div><div class="actions" style="margin-top:12px"><button class="btn">Upload Template</button><button class="btn gray" name="action" value="restore">Restore Default</button></div></form><p class="footer">PNG/JPG/SVG · حد أقصى 5MB. يطبق على بطاقات الموظفين الجديدة والمطبوعة.</p></div><div class="card"><h3>Current Template</h3>{f'<img src="/branding/id-card-template/image" style="max-width:100%;border-radius:16px">' if cur else '<div class="alert">Default template active.</div>'}</div></div>'''
        self.send(page('ID Card Designer',body,u,'settings'))

    def id_template_save(self,u):
        if not can(u,'settings.manage'): return self.forbid(u)
        fields,files=self.parse_upload_all(); action=fields.get('action',''); d=os.path.join(BRAND,'id_card_template'); os.makedirs(d,exist_ok=True)
        if action=='restore':
            for f in os.listdir(d):
                try: os.remove(os.path.join(d,f))
                except Exception: pass
            return self.redirect('/branding/id-card-template')
        fp=files.get('file')
        if not fp: return self.send(page('ID Card Designer','<div class="card"><div class="alert">اختر صورة.</div></div>',u),400)
        _,data,fname=fp; ext=os.path.splitext(fname)[1].lower()
        if ext not in ('.png','.jpg','.jpeg','.svg') or len(data)>5*1024*1024: return self.send(page('ID Card Designer','<div class="card"><div class="alert">صيغة أو حجم الصورة غير صالح.</div></div>',u),400)
        for f in os.listdir(d):
            try: os.remove(os.path.join(d,f))
            except Exception: pass
        open(os.path.join(d,'template'+ext),'wb').write(data); g['audit'](u['username'],u['role'],'ID card template update','Branding','id_card_template',fname); self.redirect('/branding/id-card-template')

    def id_template_image(self,u):
        if not can(u,'employees.view'): return self.send(b'',404,'image/png')
        d=os.path.join(BRAND,'id_card_template')
        for f in os.listdir(d) if os.path.isdir(d) else []:
            if os.path.splitext(f)[1].lower() in ('.png','.jpg','.jpeg','.svg'):
                p=os.path.join(d,f); self.send(open(p,'rb').read(),200,mimetypes.guess_type(p)[0] or 'image/png',{'Cache-Control':'no-cache'}); return
        self.send(b'',404,'image/png')

    def id_card(self,u,emp_code):
        if not can(u,'employees.view') or not g['emp_allowed'](u,emp_code): return self.forbid(u)
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(emp_code,)).fetchone(); t=c.execute('SELECT * FROM id_card_templates WHERE name="Default"').fetchone(); photo=c.execute("SELECT id,storage_path,file_name FROM documents WHERE emp_code=? AND category='صورة' ORDER BY id DESC LIMIT 1",(emp_code,)).fetchone(); c.close()
        if not e: return self.send(page('ID Card','<div class="card">Employee not found</div>',u),404)
        try: token=ensure_token(emp_code,u)
        except Exception as ex: token=''
        qr_src=f'/qr/image/{quote(emp_code)}' if token else ''
        logo='/branding/logo' if g['setting']('company_logo') else ''
        photo_html=''
        if photo:
            pp=os.path.join(DATA,photo['storage_path']) if photo['storage_path'] else ''
            if pp and os.path.exists(pp): photo_html=f'<img src="/document/{photo["id"]}" style="width:86px;height:104px;object-fit:cover;border-radius:8px">'
        if not photo_html: photo_html='<div style="width:86px;height:104px;border-radius:8px;background:#eef2f6;display:grid;place-items:center;font-size:28px">👤</div>'
        body=f'''<div class="top no-print"><div class="title"><h1>🪪 Employee ID Card</h1><p>بطاقة قابلة للطباعة مع QR.</p></div><div class="actions"><button class="btn" onclick="window.print()">🖨 Print</button><a class="btn gray" href="/id-card-pdf/{quote(code)}">⬇ PDF</a><a class="btn gray" href="/employee/profile/{esc(emp_code)}">Profile</a></div></div><div style="display:grid;place-items:center"><div class="id-card" style="width:860px;max-width:100%;min-height:540px;border-radius:28px;padding:30px;background:linear-gradient(135deg,#0b1220,#175cd3);color:white;box-shadow:0 25px 70px rgba(0,0,0,.22);background-size:cover;background-position:center;background-image:url('/branding/id-card-template/image'),linear-gradient(135deg,#0b1220,#175cd3)"><div style="display:flex;justify-content:space-between;align-items:center"><div style="font-size:25px;font-weight:800">{esc(g['setting']('company_name') or 'HR Enterprise')}</div>{f'<img src="{logo}" style="height:48px;max-width:170px;object-fit:contain;background:white;border-radius:8px;padding:4px">' if logo else ''}</div><div style="display:flex;gap:28px;align-items:center;margin-top:48px">{photo_html}<div style="flex:1"><div style="font-size:32px;font-weight:800">{esc(e['name'])}</div><div style="margin-top:12px;font-size:18px">ID: {esc(e['emp_code'])}</div><div>{esc(e['department'] or '—')} · {esc(e['job'] or '—')}</div></div>{f'<img src="{qr_src}" style="width:180px;height:180px;background:white;padding:10px;border-radius:12px">' if qr_src else ''}</div><div style="margin-top:48px;opacity:.85">HR Enterprise · Employee Identity Card</div></div></div><style>@media print{{@page{{size:A4 landscape;margin:10mm}}.id-card{{break-inside:avoid;box-shadow:none!important}}}}</style>'''
        self.send(page('ID Card',body,u,'employees'))

    def id_cards_bulk(self,u):
        qs=parse_qs(urlparse(self.path).query); codes=[]
        raw=qs.get('codes',[''])[0]
        if raw: codes=[x.strip() for x in raw.split(',') if x.strip()][:50]
        codes=[x for x in codes if g['emp_allowed'](u,x)]
        if not codes:
            scope_sql,scope_params=g['visible_employee_sql'](u,'e')
            c=db(); codes=[r['emp_code'] for r in c.execute(f"SELECT e.emp_code FROM employees e WHERE e.status<>'مؤرشف'{scope_sql} ORDER BY e.name LIMIT 50",scope_params).fetchall()]; c.close()
        scope_sql,scope_params=g['visible_employee_sql'](u,'e')
        c=db(); missing_qr=c.execute(f"SELECT COUNT(*) n FROM employees e WHERE e.status<>'مؤرشف' AND NOT EXISTS (SELECT 1 FROM qr_identities q WHERE q.emp_code=e.emp_code AND q.status='active'){scope_sql}",scope_params).fetchone()['n']; c.close()
        gen_btn=f'<button class="btn ok" onclick="genAllQr(this)">⚡ توليد QR لكل الموظفين ({missing_qr} بدون QR)</button>' if missing_qr and u.get('role') in ('Admin','SuperAdmin') else ''
        links=''.join(f'<div style="page-break-inside:avoid;margin-bottom:24px"><iframe src="/id-card/{quote(code)}" style="width:100%;height:600px;border:1px solid #ddd"></iframe></div>' for code in codes)
        body=f'''<div class="top no-print"><div class="title"><h1>🪪 Bulk ID Cards</h1><p>{len(codes)} بطاقة · اطبع الصفحة مرة واحدة.</p></div><div class="actions">{gen_btn}<button class="btn" onclick="window.print()">🖨 Print All</button></div></div>{links}
        <script>function genAllQr(btn){{btn.disabled=true;btn.textContent="جاري التوليد…";var fd=new FormData();fd.append('_csrf','{esc(u.get("csrf",""))}');fetch('/qr/generate-all',{{method:'POST',body:fd}}).then(r=>r.json()).then(d=>{{alert('تم توليد '+d.created+' كود QR جديد.');location.reload();}}).catch(()=>{{alert('حدث خطأ.');btn.disabled=false;}});}}</script>'''
        self.send(page('Bulk ID Cards',body,u,'employees'))

    def qr_generate_all(self,u):
        if u.get('role') not in ('Admin','SuperAdmin'): return self.forbid(u)
        c=db(); codes=[r['emp_code'] for r in c.execute("SELECT emp_code FROM employees e WHERE e.status<>'مؤرشف' AND NOT EXISTS (SELECT 1 FROM qr_identities q WHERE q.emp_code=e.emp_code AND q.status='active')").fetchall()]; c.close()
        codes=[x for x in codes if g['emp_allowed'](u,x)]
        created=0
        for code in codes:
            try:
                issue_qr(code,u,False); created+=1
            except Exception as e: g['log_error']('qr_generate_all',e)
        g['audit'](u['username'],u['role'],'QR_BULK_GENERATE','QR Identity',str(created),f'{created} of {len(codes)}')
        self.send(json.dumps({'created':created,'total':len(codes)}).encode(),200,'application/json')

    def contracts(self,u):
        scope_sql,scope_params=g['visible_employee_sql'](u,'e'); c=db(); rows=c.execute(f'SELECT ct.*,e.name FROM contracts ct JOIN employees e ON e.emp_code=ct.emp_code WHERE 1=1{scope_sql} ORDER BY COALESCE(ct.end_date,"9999") ASC,ct.id DESC LIMIT 500',scope_params).fetchall(); emps=c.execute(f'SELECT emp_code,name FROM employees e WHERE e.status<>"مؤرشف"{scope_sql}',scope_params).fetchall(); c.close()
        trs=''.join(f'<tr><td>{esc(r["contract_no"] or r["id"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["start_date"] or "")}</td><td>{esc(r["end_date"] or "")}</td><td>{esc(r["status"])}</td><td>{float(r["amount"] or 0):g}</td><td><form method="post" action="/contracts/action" style="display:inline">{csrf_field(u)}<input type="hidden" name="id" value="{r["id"]}"><input type="hidden" name="action" value="terminate"><button class="btn bad">إنهاء</button></form><form method="post" action="/contracts/action" style="display:inline;margin-right:5px">{csrf_field(u)}<input type="hidden" name="id" value="{r["id"]}"><input type="hidden" name="action" value="renew"><input type="date" name="end_date" value="{esc(r["end_date"] or "")}"><button class="btn warn">تجديد</button></form></td></tr>' for r in rows)
        opts=''.join(f'<option value="{esc(e["emp_code"])}">{esc(e["emp_code"])} — {esc(e["name"])}</option>' for e in emps)
        body=f'''<div class="top"><div class="title"><h1>📑 Contracts</h1><p>عقود متعددة لكل موظف مع تجديد وإنهاء وتاريخ كامل.</p></div></div><div class="card"><form method="post" action="/contracts/save" class="form">{csrf_field(u)}<div class="field"><label>Employee</label><select name="emp_code" required>{opts}</select></div><div class="field"><label>Contract No</label><input name="contract_no"></div><div class="field"><label>Type</label><input name="contract_type" value="employment"></div><div class="field"><label>Start</label><input type="date" name="start_date"></div><div class="field"><label>End</label><input type="date" name="end_date"></div><div class="field"><label>Amount</label><input type="number" step="0.01" name="amount"></div><div class="field"><label>Status</label><select name="status"><option>active</option><option>renewed</option><option>terminated</option><option>expired</option></select></div><div class="field full"><label>Notes</label><textarea name="notes"></textarea></div><div class="full"><button class="btn">Save Contract</button></div></form></div><div class="card" style="margin-top:16px"><table class="table"><tr><th>No</th><th>Employee ID</th><th>Name</th><th>Start</th><th>End</th><th>Status</th><th>Amount</th><th>Actions</th></tr>{trs or '<tr><td colspan="8">No contracts</td></tr>'}</table></div>'''
        self.send(page('Contracts',body,u,'enterprise'))

    def contract_save(self,u,f):
        if not can(u,'employees.edit'): return self.forbid(u)
        emp=f.get('emp_code','').strip()
        if not emp or not g['emp_allowed'](u,emp): return self.forbid(u)
        c=db(); exists=c.execute('SELECT 1 FROM employees WHERE emp_code=?',(emp,)).fetchone()
        if not exists: c.close(); return self.send(page('Contracts','<div class="card"><div class="alert">Employee not found.</div></div>',u),400)
        c.execute('INSERT INTO contracts(emp_code,contract_no,contract_type,start_date,end_date,amount,status,notes,created_by,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,f.get('contract_no','').strip(),f.get('contract_type','employment'),f.get('start_date',''),f.get('end_date',''),float(f.get('amount','0') or 0),f.get('status','active'),f.get('notes',''),u['username'],now(),now())); c.commit(); c.close(); g['audit'](u['username'],u['role'],'Contract create','Contracts',emp); self.redirect('/contracts')

    def training(self,u):
        scope_sql,scope_params=g['visible_employee_sql'](u,'e'); c=db(); programs=c.execute('SELECT * FROM training_programs ORDER BY name').fetchall(); emps=c.execute(f'SELECT emp_code,name FROM employees e WHERE e.status<>"مؤرشف"{scope_sql} ORDER BY name',scope_params).fetchall(); rows=c.execute(f'SELECT te.*,tp.name program_name,e.name FROM training_enrollments te LEFT JOIN training_programs tp ON tp.id=te.program_id JOIN employees e ON e.emp_code=te.emp_code WHERE 1=1{scope_sql} ORDER BY te.id DESC LIMIT 500',scope_params).fetchall(); c.close()
        opts=''.join(f'<option value="{p["id"]}">{esc(p["name"])}</option>' for p in programs); eopts=''.join(f'<option value="{esc(e["emp_code"])}">{esc(e["emp_code"])} — {esc(e["name"])}</option>' for e in emps)
        trs=''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["program_name"] or "")}</td><td>{esc(r["start_date"] or "")}</td><td>{esc(r["completion_date"] or "")}</td><td>{esc(r["certificate_expiry"] or "")}</td><td>{esc(r["status"])}</td></tr>' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>🎓 Training Center</h1><p>برامج تدريب، تسجيل، إتمام، شهادات وتنبيهات انتهاء.</p></div></div><div class="grid g2"><div class="card"><h3>New Program</h3><form method="post" action="/training/program/save">{csrf_field(u)}<div class="field"><label>Name</label><input name="name" required></div><div class="field"><label>Provider</label><input name="provider"></div><div class="field full"><label>Description</label><textarea name="description"></textarea></div><button class="btn">Save Program</button></form></div><div class="card"><h3>Enroll Employee</h3><form method="post" action="/training/enroll">{csrf_field(u)}<div class="field"><label>Program</label><select name="program_id" required>{opts}</select></div><div class="field"><label>Employee</label><select name="emp_code" required>{eopts}</select></div><div class="field"><label>Start</label><input type="date" name="start_date"></div><div class="field"><label>End</label><input type="date" name="end_date"></div><div class="field"><label>Completion</label><input type="date" name="completion_date"></div><div class="field"><label>Certificate Expiry</label><input type="date" name="certificate_expiry"></div><div class="field"><label>Status</label><select name="status"><option>enrolled</option><option>completed</option><option>failed</option><option>cancelled</option></select></div><button class="btn">Save Enrollment</button></form></div></div><div class="card" style="margin-top:16px"><table class="table"><tr><th>ID</th><th>Name</th><th>Program</th><th>Start</th><th>Completion</th><th>Certificate Expiry</th><th>Status</th></tr>{trs or '<tr><td colspan="7">No training records</td></tr>'}</table></div>'''
        self.send(page('Training',body,u,'enterprise'))

    def training_program_save(self,u,f):
        c=db(); c.execute('INSERT OR IGNORE INTO training_programs(name,provider,description,created_at) VALUES(?,?,?,?)',(f.get('name','').strip(),f.get('provider',''),f.get('description',''),now())); c.commit(); c.close(); self.redirect('/training')
    def training_enroll(self,u,f):
        if not can(u,'employees.edit'): return self.forbid(u)
        emp=f.get('emp_code','').strip()
        if not emp or not g['emp_allowed'](u,emp): return self.forbid(u)
        c=db(); c.execute('INSERT INTO training_enrollments(program_id,emp_code,start_date,end_date,completion_date,certificate_expiry,status,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(int(f.get('program_id','0') or 0),emp,f.get('start_date',''),f.get('end_date',''),f.get('completion_date',''),f.get('certificate_expiry',''),f.get('status','enrolled'),u['username'],now())); c.commit(); c.close(); self.redirect('/training')


    def id_template_flex(self,u):
        if not can(u,'settings.manage'): return self.forbid(u)
        c=db(); t=c.execute('SELECT * FROM id_card_templates WHERE name="Default"').fetchone(); c.close(); t=dict(t) if t else {}
        front=t.get('background_path') or ''
        body=f'''<div class="top"><div class="title"><h1>🎨 ID Card Designer</h1><p>ارفع تصميم البطاقة وعدّل أماكن الصورة والاسم والـID والـQR بدون كود.</p></div></div>
<div class="grid g2"><div class="card"><form method="post" action="/branding/id-card-template/save" enctype="multipart/form-data">{csrf_field(u)}
<div class="field"><label>Front Template</label><input type="file" name="front" accept=".png,.jpg,.jpeg,.webp,.svg"></div>
<div class="field"><label>Back Template (optional)</label><input type="file" name="back" accept=".png,.jpg,.jpeg,.webp,.svg"></div>
<div class="grid g2"><div class="field"><label>Photo X</label><input name="photo_x" type="number" step="1" value="{t.get('photo_x',40)}"></div><div class="field"><label>Photo Y</label><input name="photo_y" type="number" step="1" value="{t.get('photo_y',150)}"></div><div class="field"><label>Photo Width</label><input name="photo_w" type="number" step="1" value="{t.get('photo_w',130)}"></div><div class="field"><label>Photo Height</label><input name="photo_h" type="number" step="1" value="{t.get('photo_h',160)}"></div><div class="field"><label>Name X</label><input name="name_x" type="number" step="1" value="{t.get('name_x',190)}"></div><div class="field"><label>Name Y</label><input name="name_y" type="number" step="1" value="{t.get('name_y',170)}"></div><div class="field"><label>QR X</label><input name="qr_x" type="number" step="1" value="{t.get('qr_x',650)}"></div><div class="field"><label>QR Y</label><input name="qr_y" type="number" step="1" value="{t.get('qr_y',145)}"></div><div class="field"><label>QR Size</label><input name="qr_size" type="number" step="1" value="{t.get('qr_size',150)}"></div></div>
<div class="actions" style="margin-top:12px"><label><input type="checkbox" name="show_emp_id" {'checked' if t.get('show_emp_id',1) else ''}> Employee ID</label><label><input type="checkbox" name="show_department" {'checked' if t.get('show_department',1) else ''}> Department</label><label><input type="checkbox" name="show_job" {'checked' if t.get('show_job',1) else ''}> Job</label></div>
<div class="actions" style="margin-top:16px"><button class="btn">Save Designer</button><button class="btn gray" name="action" value="restore">Restore Default</button></div></form></div>
<div class="card"><h3>Live Preview</h3><div style="aspect-ratio:86/54;border-radius:20px;background:#eef2f6;overflow:hidden;position:relative;background-size:cover;background-position:center;{'background-image:url(/branding/id-card-template/image);' if front else ''}"><div style="position:absolute;left:30px;top:30px;font-weight:800;font-size:22px">{esc(g['setting']('company_name') or 'HR Enterprise')}</div><div style="position:absolute;left:40px;top:150px;width:130px;height:160px;background:#fff8;border-radius:14px;display:grid;place-items:center">PHOTO</div><div style="position:absolute;left:190px;top:170px;font-weight:800;font-size:24px">Employee Name</div><div style="position:absolute;right:40px;top:145px;width:150px;height:150px;background:#fff;border-radius:12px;display:grid;place-items:center">QR</div></div><p class="footer">احفظ ثم افتح بطاقة موظف حقيقية للتجربة والطباعة.</p></div></div>'''
        self.send(page('ID Card Designer',body,u,'settings'))

    def id_template_save_flex(self,u,fields,files):
        if not can(u,'settings.manage'): return self.forbid(u)
        d=os.path.join(BRAND,'id_card_template'); os.makedirs(d,exist_ok=True); action=fields.get('action','')
        if action=='restore':
            for fn in os.listdir(d):
                try: os.remove(os.path.join(d,fn))
                except Exception: pass
            c=db(); c.execute('UPDATE id_card_templates SET background_path="",created_by=?,created_at=? WHERE name="Default"',(u['username'],now())); c.commit(); c.close(); g['audit'](u['username'],u['role'],'ID_CARD_TEMPLATE_RESTORE','Branding','Default'); return self.redirect('/branding/id-card-template')
        allowed={'.png','.jpg','.jpeg','.webp','.svg'}; paths={}
        for key in ('front','back'):
            fp=next((x for x in files if x[2] and key in x[0].decode('utf-8','ignore')),None)
            if fp:
                _,data,fname=fp; ext=os.path.splitext(fname)[1].lower()
                if ext not in allowed or len(data)>8*1024*1024:return self.send(page('ID Card Designer','<div class="card"><div class="alert">صيغة/حجم القالب غير صالح.</div></div>',u),400)
                out=os.path.join(d,key+ext); open(out,'wb').write(data); paths[key]=os.path.relpath(out,DATA).replace('\\','/')
        c=db(); cols_db={r['name'] for r in c.execute('PRAGMA table_info(id_card_templates)').fetchall()}
        if 'background_back_path' not in cols_db: c.execute('ALTER TABLE id_card_templates ADD COLUMN background_back_path TEXT')
        cols=['photo_x','photo_y','photo_w','photo_h','name_x','name_y','qr_x','qr_y','qr_size']; vals=[float(fields.get(k) or 0) for k in cols]
        c.execute('UPDATE id_card_templates SET background_path=COALESCE(NULLIF(?,""),background_path),background_back_path=COALESCE(NULLIF(?,""),background_back_path),created_by=?,created_at=?,photo_x=?,photo_y=?,photo_w=?,photo_h=?,name_x=?,name_y=?,qr_x=?,qr_y=?,qr_size=?,show_department=?,show_job=?,show_emp_id=? WHERE name="Default"',(paths.get('front',''),paths.get('back',''),u['username'],now(),*vals,1 if fields.get('show_department') else 0,1 if fields.get('show_job') else 0,1 if fields.get('show_emp_id') else 0)); c.commit(); c.close(); g['audit'](u['username'],u['role'],'ID_CARD_TEMPLATE_SAVE','Branding','Default'); self.redirect('/branding/id-card-template')

    def id_card_flex(self,u,code):
        if not can(u,'employees.view') or not g['emp_allowed'](u,code): return self.forbid(u)
        autoprint=parse_qs(urlparse(self.path).query).get('autoprint',[''])[0]=='1'
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone(); q=c.execute('SELECT image_path FROM qr_identities WHERE emp_code=? AND status="active"',(code,)).fetchone(); ph=c.execute("SELECT id,file_name FROM documents WHERE emp_code=? AND category='صورة' AND status='current' ORDER BY id DESC LIMIT 1",(code,)).fetchone(); t=c.execute('SELECT * FROM id_card_templates WHERE name="Default"').fetchone(); c.close()
        if not e:return self.send(page('ID Card','<div class="card"><div class="alert">الموظف غير موجود.</div></div>',u),404)
        if not q and u.get('role') in ('Admin','SuperAdmin'):
            try:
                issue_qr(code,u,False)
                c=db(); q=c.execute('SELECT image_path FROM qr_identities WHERE emp_code=? AND status="active"',(code,)).fetchone(); c.close()
            except Exception as ex: g['log_error']('id_card_flex_auto_qr',ex)
        t=dict(t) if t else {}; bg=t.get('background_path') or ''; bgurl='/branding/id-card-template/image' if bg else ''; photo=f'/employee/photo/{quote(code)}?v={ph["id"]}' if ph else ''; qrurl=f'/qr/image/{quote(code)}' if q else ''
        px=float(t.get('photo_x',40)); py=float(t.get('photo_y',150)); pw=float(t.get('photo_w',130)); phh=float(t.get('photo_h',160)); nx=float(t.get('name_x',190)); ny=float(t.get('name_y',170)); qx=float(t.get('qr_x',650)); qy=float(t.get('qr_y',145)); qs=float(t.get('qr_size',150))
        company=esc(g['setting']('company_name') or 'HR Enterprise'); name=esc(e['name']); dept=esc(e['department'] or ''); job=esc(e['job'] or '')
        photo_html=f'<img src="{photo}" style="position:absolute;left:{px}px;top:{py}px;width:{pw}px;height:{phh}px;object-fit:cover;border-radius:12px">' if photo else f'<div style="position:absolute;left:{px}px;top:{py}px;width:{pw}px;height:{phh}px;background:#eef2f6;border-radius:12px;display:grid;place-items:center">PHOTO</div>'
        qr_html=f'<img src="{qrurl}" style="position:absolute;left:{qx}px;top:{qy}px;width:{qs}px;height:{qs}px;background:#fff;padding:8px;border-radius:12px">' if qrurl else '<div style="position:absolute;right:40px;top:145px">QR not issued</div>'
        fields_html=(f'<div style="margin-top:8px">ID: {esc(code)}</div>' if t.get('show_emp_id',1) else '')+(f'<div>{dept}</div>' if t.get('show_department',1) else '')+(f'<div>{job}</div>' if t.get('show_job',1) else '')
        body=f'''<div class="top no-print"><div class="title"><h1>🪪 بطاقة هوية الموظف</h1><p>{esc(code)} · {name}</p></div><div class="actions"><button class="btn" onclick="window.print()">🖨 طباعة البطاقة</button><a class="btn gray" href="/employee/profile/{quote(code)}">الملف الشخصي</a></div></div><div style="display:grid;place-items:center"><div class="id-card" style="position:relative;width:860px;height:540px;max-width:100%;overflow:hidden;border-radius:28px;background:#fff;border:1px solid #d0d5dd;box-shadow:0 25px 70px rgba(0,0,0,.18);background-size:cover;background-position:center;{'background-image:url('+bgurl+');' if bgurl else 'background:linear-gradient(135deg,#0b1220,#175cd3);color:#fff;'}"><div style="position:absolute;left:30px;top:25px;font-size:24px;font-weight:800">{company}</div>{photo_html}<div style="position:absolute;left:{nx}px;top:{ny}px;right:250px"><div style="font-size:28px;font-weight:800">{name}</div>{fields_html}</div>{qr_html}<div style="position:absolute;left:30px;bottom:24px;font-size:12px;opacity:.8">HR Enterprise · Employee Identity Card</div></div></div><style>@media print{{@page{{size:A4 landscape;margin:10mm}}body{{background:white!important}}.id-card{{box-shadow:none!important;break-inside:avoid}}}}</style>{'<script>window.addEventListener("load",function(){setTimeout(function(){window.print();},250);});</script>' if autoprint else ''}'''
        self.send(page('Employee ID Card',body,u,'employees'))


    def id_card_pdf(self,u,code):
        if not can(u,'employees.view') or not g['emp_allowed'](u,code): return self.forbid(u)
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from PIL import Image
        except Exception as e:
            return self.send(page('ID Card PDF',f'<div class="card"><div class="alert">PDF engine unavailable: {esc(e)}</div></div>',u),500)
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone(); q=c.execute('SELECT image_path FROM qr_identities WHERE emp_code=? AND status="active"',(code,)).fetchone(); ph=c.execute("SELECT storage_path,file_name FROM documents WHERE emp_code=? AND category='صورة' AND status='current' ORDER BY id DESC LIMIT 1",(code,)).fetchone(); t=c.execute('SELECT * FROM id_card_templates WHERE name="Default"').fetchone(); c.close()
        if not e:return self.send(page('ID Card PDF','<div class="card"><div class="alert">الموظف غير موجود.</div></div>',u),404)
        out=io.BytesIO(); card_w,card_h=86*mm,54*mm; pdf=canvas.Canvas(out,pagesize=(card_w,card_h))
        font_path=os.path.join(g['BASE'],'fonts','DejaVuSans.ttf')
        try: pdfmetrics.registerFont(TTFont('HRDejaVu',font_path)); font='HRDejaVu'
        except Exception: font='Helvetica'
        pdf.setFont(font,10); pdf.drawString(8*mm,47*mm,g['setting']('company_name') or 'HR Enterprise')
        if ph:
            try:
                raw=g['secure_file_bytes'](ph['storage_path']); im=Image.open(io.BytesIO(raw)).convert('RGB'); tmp=io.BytesIO(); im.save(tmp,format='JPEG'); tmp.seek(0); pdf.drawImage(tmp,5*mm,18*mm,width=25*mm,height=25*mm,preserveAspectRatio=True,mask='auto')
            except Exception: pass
        pdf.setFont(font,11); pdf.drawString(33*mm,35*mm,str(e['name'] or ''))
        pdf.setFont(font,8); pdf.drawString(33*mm,30*mm,'ID: '+str(code))
        pdf.drawString(33*mm,26*mm,str(e['department'] or ''))
        pdf.drawString(33*mm,22*mm,str(e['job'] or ''))
        if q:
            try:
                raw=g['secure_file_bytes'](q['image_path']); tmp=io.BytesIO(raw); pdf.drawImage(tmp,61*mm,10*mm,width=20*mm,height=20*mm,preserveAspectRatio=True,mask='auto')
            except Exception: pass
        pdf.save(); self.send(out.getvalue(),200,'application/pdf',{'Content-Disposition':f'attachment; filename="ID_{g['safe_name'](code)}.pdf"'})

    old_get=H.do_GET; old_post=H.do_POST
    def get(self):
        p=urlparse(self.path).path
        custom = p=='/qr/scan' or p.startswith('/qr/image/') or p.startswith('/qr/verify/') or p.startswith('/employee/photo/') or (p.startswith('/id-card/') and p.count('/')==2) or p=='/id-cards' or p in ('/contracts','/training','/branding/id-card-template','/branding/id-card-template/image')
        if not custom: return old_get(self)
        if p.startswith('/qr/verify/'):
            token=p.split('/qr/verify/',1)[1]
            return qr_verify(self,{'username':'anonymous','role':'Public','full_name':'Public'},token)
        u=self.require()
        if not u: return
        if p=='/qr/scan': return qr_scan(self,u)
        if p.startswith('/qr/image/'): return qr_image(self,u,p.split('/')[-1])
        if p.startswith('/qr/verify/'):
            token=p.split('/qr/verify/',1)[1]; return qr_verify(self,u,token)
        if p.startswith('/employee/photo/'):
            code=p.split('/employee/photo/',1)[1]; c=db(); r=c.execute("SELECT id,storage_path,file_name FROM documents WHERE emp_code=? AND category='صورة' AND status='current' ORDER BY id DESC LIMIT 1",(code,)).fetchone(); c.close()
            if not r: return self.send(b'',404,'image/png')
            try: return self.send(g['secure_file_bytes'](r['storage_path']),200,mimetypes.guess_type(r['file_name'])[0] or 'image/jpeg',{'Cache-Control':'private,max-age=86400'})
            except Exception: return self.send(b'',404,'image/png')
        if p.startswith('/id-card/') and p.count('/')==2: return id_card(self,u,p.split('/')[-1])
        if p=='/id-cards': return id_cards_bulk(self,u)
        if p=='/contracts': return contracts(self,u)
        if p=='/training': return training(self,u)
        if p=='/branding/id-card-template': return id_template_page(self,u)
        if p=='/branding/id-card-template/image': return id_template_image(self,u)
    def post(self):
        p=urlparse(self.path).path
        if p in ('/qr/generate','/qr/regenerate','/qr/revoke','/qr/generate-all','/contracts/save','/contracts/action','/training/program/save','/employee/photo/upload','/training/enroll','/branding/id-card-template/save'):
            u=self.require();
            if not u:return
            upload_part=None
            if self.headers.get('Content-Type','').lower().startswith('multipart/form-data'):
                f,upload_part=self.parse_upload()
            else:
                f=self.form()
            if f.get('_csrf')!=u.get('csrf'): return self.send(page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
            if p=='/employee/photo/upload':
                emp=f.get('emp_code','').strip()
                if not emp or u.get('role') not in ('Admin','SuperAdmin') or not g['emp_allowed'](u,emp): return self.forbid(u)
                if not upload_part: return self.send(page('صورة الموظف','<div class="card"><div class="alert">ملف الصورة غير موجود.</div></div>',u),400)
                _,data,fname=upload_part
                ext=os.path.splitext(fname)[1].lower()
                if not data or ext not in ('.jpg','.jpeg','.png','.webp') or len(data)>5*1024*1024:
                    return self.send(page('صورة الموظف','<div class="card"><div class="alert">PNG/JPG/WEBP وبحد أقصى 5MB.</div></div>',u),400)
                if not g['validate_file_signature'](fname,data):
                    return self.send(page('صورة الموظف','<div class="card"><div class="alert">محتوى الملف لا يطابق امتداده أو الملف تالف.</div></div>',u),400)
                # Original phone-camera uploads (often several MB, thousands of px wide) were being
                # stored and served as-is, even though every table row shows this photo at 36x42px
                # and the profile page at 110x130px. Every employee list page load was quietly
                # downloading the full-resolution original per row — this is a primary cause of the
                # system feeling heavy. Re-encode to a sane max size once, at upload time.
                try:
                    from PIL import Image, ImageOps
                    im=Image.open(io.BytesIO(data)); im=ImageOps.exif_transpose(im); im=im.convert('RGB')
                    im.thumbnail((900,900))
                    buf=io.BytesIO(); im.save(buf,'JPEG',quality=85,optimize=True); data=buf.getvalue(); ext='.jpg'
                except Exception:
                    pass  # Pillow unavailable or unreadable image: keep the validated original as-is.
                rel=g['save_employee_file'](emp,'profile'+ext,data); c=db()
                c.execute("UPDATE documents SET status='superseded' WHERE emp_code=? AND category='صورة' AND status='current'",(emp,))
                c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,'profile'+ext,ext,u['username'],now(),None,rel,'صورة',1,hashlib.sha256(data).hexdigest(),'current'))
                c.commit(); c.close(); g['audit'](u['username'],u['role'],'PROFILE_PHOTO_UPDATED','Employees',emp,'profile photo'); return self.redirect('/employee/profile/'+quote(emp))
            if p in ('/qr/generate','/qr/regenerate'):
                emp=f.get('emp_code','').strip()
                if not emp or not g['emp_allowed'](u,emp): return self.forbid(u)
                try: ensure_token(emp,u,p.endswith('regenerate')); return self.redirect('/employee/profile/'+quote(emp))
                except Exception as e: return self.send(page('QR','<div class="card"><div class="alert">'+esc(str(e))+'</div></div>',u),400)
            if p=='/qr/generate-all': return qr_generate_all(self,u)
            if p=='/qr/revoke':
                emp=f.get('emp_code','').strip()
                if not emp or not g['emp_allowed'](u,emp): return self.forbid(u)
                c=db(); c.execute('UPDATE qr_identities SET status="revoked",revoked_at=? WHERE emp_code=? AND status="active"',(now(),emp)); c.commit(); c.close(); g['audit'](u['username'],u['role'],'QR_REVOKED','QR Identity',emp); return self.redirect('/employee/profile/'+quote(emp))
            if p=='/branding/id-card-template/save': return id_template_save(self,u)
            if p=='/contracts/save': return contract_save(self,u,f)
            if p=='/contracts/action':
                if not can(u,'employees.edit'): return self.forbid(u)
                cid=int(f.get('id') or 0); action=f.get('action',''); c=db(); r=c.execute('SELECT * FROM contracts WHERE id=?',(cid,)).fetchone()
                if not r: c.close(); return self.redirect('/contracts')
                if not g['emp_allowed'](u,r['emp_code']): c.close(); return self.forbid(u)
                if action=='terminate': c.execute("UPDATE contracts SET status='terminated',updated_at=? WHERE id=?",(now(),cid))
                elif action=='renew': c.execute("UPDATE contracts SET status='renewed',end_date=?,updated_at=? WHERE id=?",(f.get('end_date',''),now(),cid))
                c.commit(); c.close(); g['audit'](u['username'],u['role'],'Contract '+action,'Contracts',str(cid)); return self.redirect('/contracts')
            if p=='/training/program/save': return training_program_save(u,f)
            return training_enroll(self,u,f)
        return old_post(self)
    H.do_GET=get; H.do_POST=post
    g['ensure_token']=ensure_token
    g['issue_qr']=issue_qr

    # Extend employee profile without rewriting its existing handler.
    old_profile=H.employee_profile
    def profile(self,u,code):
        # Render existing profile first; then inject QR/card controls before </body>.
        import types
        captured=[]
        old_send=self_send=getattr(H,'send')
        def capture(self,body,status=200,ctype='text/html',headers=None):
            if isinstance(body,bytes): captured.append((body,status,ctype,headers)); return
            captured.append((body,status,ctype,headers))
        # Simpler: use a temporary monkey patch on instance method.
        original=self.send; self.send=types.MethodType(capture,self)
        try: old_profile(self,u,code)
        finally: self.send=original
        if not captured: return
        body,status,ctype,headers=captured[0]
        if status!=200 or 'text/html' not in ctype: return original(body,status,ctype,headers)
        row=current_token(code); active=bool(row and row['status']=='active')
        cqr=db(); qr_any=cqr.execute('SELECT status,revoked_at FROM qr_identities WHERE emp_code=? ORDER BY id DESC LIMIT 1',(code,)).fetchone(); cqr.close()
        qr_revoked=bool(qr_any and not active and qr_any['status']=='revoked')
        cphoto=db(); photo=cphoto.execute("SELECT id FROM documents WHERE emp_code=? AND category='صورة' AND status='current' ORDER BY id DESC LIMIT 1",(code,)).fetchone(); cphoto.close()
        photo_url=f'/employee/photo/{quote(code)}?v={photo["id"]}' if photo else ''
        # PHASE 2 (2026-08): the photo + QR identity now live directly inside the
        # Hero Header at the very top of the profile page (built in
        # server.py's employee_profile, right above the name/status/actions —
        # not a separate card lower down, and not appended near </body>). This
        # function's job is just to fill the two marked slots server.py left for
        # us (#heroPhotoSlot / #heroQrSlot) with the real photo/QR markup, since
        # the photo + QR data and their upload/generate/revoke actions live here.
        can_edit_photo=u.get('role') in ('Admin','SuperAdmin')
        photo_inner=(f'<img id="empPhotoImg" src="{photo_url}" class="ph-photo" alt="صورة الموظف">'
                     if photo else '<div id="empPhotoImg" class="ph-photo-empty">👤</div>')
        photo_edit=(f'''<label for="empPhotoFile" class="ph-photo-edit" title="اضغط لتغيير الصورة">📷<span style="font-size:11px">تغيير</span></label>
          <form id="empPhotoForm" method="post" action="/employee/photo/upload" enctype="multipart/form-data" style="display:none">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><input id="empPhotoFile" name="file" type="file" accept=".png,.jpg,.jpeg,.webp"></form>
          <div id="empPhotoPreviewModal" class="no-print" style="display:none;position:fixed;inset:0;background:rgba(16,24,40,.55);z-index:100;place-items:center">
            <div class="card" style="width:340px;text-align:center">
              <h3 style="margin-top:0">معاينة الصورة الجديدة</h3>
              <img id="empPhotoPreviewImg" style="width:180px;height:180px;object-fit:cover;border-radius:16px;border:1px solid var(--line)">
              <div class="actions" style="justify-content:center;margin-top:16px">
                <button type="button" class="btn ok" id="empPhotoConfirmBtn">✔ رفع الصورة</button>
                <button type="button" class="btn gray" id="empPhotoCancelBtn">إلغاء</button>
              </div>
            </div>
          </div>
          <script>(function(){{
            var fileInput=document.getElementById('empPhotoFile'), modal=document.getElementById('empPhotoPreviewModal'),
                img=document.getElementById('empPhotoPreviewImg'), okBtn=document.getElementById('empPhotoConfirmBtn'),
                cancelBtn=document.getElementById('empPhotoCancelBtn'), form=document.getElementById('empPhotoForm');
            function reset(){{modal.style.display='none'; fileInput.value='';}}
            fileInput.addEventListener('change',function(){{
              if(!fileInput.files || !fileInput.files[0]) return;
              var reader=new FileReader();
              reader.onload=function(e){{img.src=e.target.result; modal.style.display='grid';}};
              reader.readAsDataURL(fileInput.files[0]);
            }});
            okBtn.addEventListener('click',function(){{okBtn.disabled=true; okBtn.textContent='جارٍ الرفع…'; form.submit();}});
            cancelBtn.addEventListener('click',reset);
            modal.addEventListener('click',function(e){{if(e.target===modal) reset();}});
          }})();</script>'''
                    if can_edit_photo else '')
        delete_btn=(f'''<form method="post" action="/employee/photo/upload" style="display:inline" onsubmit="return confirm('حذف صورة الموظف؟')">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><input type="hidden" name="delete" value="1"><button class="btn gray" type="submit">🗑 حذف</button></form>'''
                     if photo and can_edit_photo else '')
        photo_fill=f'<div class="ph-photo-wrap" id="heroPhotoSlot">{photo_inner}{photo_edit}</div>'
        qr_status_badge=('<span class="badge b-ok">QR نشط</span>' if active
                          else (f'<span class="badge b-bad">QR مُلغى{" بتاريخ "+esc(qr_any["revoked_at"][:10]) if qr_any and qr_any["revoked_at"] else ""}</span>' if qr_revoked
                                else '<span class="badge b-warn">QR غير مُصدر</span>'))
        qr_img=(f'<img src="/qr/image/{quote(code)}" alt="QR">' if active
                else '<div class="ph-photo-empty" style="width:130px;height:130px;font-size:36px;border-radius:10px">🔲</div>')
        qr_forms=(f'''<form method="post" action="/qr/generate" style="display:inline">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><button class="btn" {'disabled' if active else ''}>إصدار QR</button></form><form method="post" action="/qr/regenerate" style="display:inline">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><button class="btn warn">إعادة إصدار</button></form><form method="post" action="/qr/revoke" style="display:inline">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><button class="btn bad">إلغاء</button></form>'''
                  if can_edit_photo else '')
        qr_fill=f'''<div class="ph-qr" id="heroQrSlot">
          <div class="qr-status">{qr_status_badge}</div>
          {qr_img}
          <div class="qr-actions">
            <a class="btn gray" href="/id-card/{quote(code)}">🪪 بطاقة الهوية</a>
            {f'<a class="btn gray" href="/qr/image/{quote(code)}" download="{esc(code)}-qr.png">⬇ تنزيل QR</a>' if active else ''}
            {qr_forms}
            <a class="btn gray" href="/qr/scan">📷 الماسح</a>
            {delete_btn}
          </div>
        </div>'''
        if isinstance(body,bytes): body=body.decode('utf-8','replace')
        placeholder_photo='<div class="ph-photo-wrap" id="heroPhotoSlot"><div class="ph-photo-empty">👤</div></div>'
        placeholder_qr='<div class="ph-qr" id="heroQrSlot"></div>'
        if placeholder_photo in body: body=body.replace(placeholder_photo,photo_fill,1)
        if placeholder_qr in body: body=body.replace(placeholder_qr,qr_fill,1)
        return original(body,status,ctype,headers)
    H.employee_profile=profile

    # Make the admin branding upload actually control the live company logo used by login/sidebar.
    old_branding_save=getattr(H,'branding_profile_save',None)
    if old_branding_save:
        def branding_save_wrapper(self,u):
            result=old_branding_save(self,u)
            try:
                c=db(); r=c.execute('SELECT path FROM branding_profiles WHERE profile_key="login_logo"').fetchone()
                c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',('company_logo',r['path'] if r and r['path'] else '')); c.commit(); c.close(); clear_settings_cache()
            except Exception: pass
            return result
        H.branding_profile_save=branding_save_wrapper

    # Add nav items through the latest page wrapper.
    old_page=g['page']
    def page_plus(title,body,user,active='dashboard'):
        out=old_page(title,body,user,active)
        if user:
            idn='<a href="/qr/scan">🔐 QR Scanner</a><a href="/id-cards">🪪 كل بطاقات الموظفين</a>'
            extra='<details class="nav-group"><summary>بطاقات سريعة</summary>'+idn+'</details>'
            extra+='<details class="nav-group"><summary>العقود والتدريب</summary><a href="/contracts">📑 العقود</a><a href="/training">🎓 التدريب</a></details>'
            out=out.replace('</nav>',extra+'</nav>',1)
            accent={'blue':'#175cd3','purple':'#7f56d9','emerald':'#039855','orange':'#f79009'}.get(g['setting']('accent_color') or 'blue','#175cd3')
            dark=(g['setting']('theme') or 'light')=='dark'
            vars_css=f'<style>:root{{--brand:{accent};--brand2:{accent};}}' + ('body{background:#0b1220!important;color:#f2f4f7!important}.card{background:#101828!important;border-color:#344054!important;color:#f2f4f7}.table th{background:#1d2939!important;color:#d0d5dd!important}.table td{border-color:#344054!important}.field input,.field select,.field textarea,.toolbar input,.toolbar select{background:#0b1220!important;color:#f2f4f7!important;border-color:#475467!important}.top{background:rgba(11,18,32,.94)!important}.title p,.muted,.sub{color:#98a2b3!important}</style>' if dark else '</style>')
            if vars_css=='</style>': vars_css=f'<style>:root{{--brand:{accent};--brand2:{accent};}}</style>'
            out=out.replace('</head>',vars_css+'</head>',1)
        return out
    g['page']=page_plus

    globals().update({'qr_identity_issue':issue_qr,'qr_token_hash':token_hash,'id_template_flex':id_template_flex,'id_template_save_flex':id_template_save_flex,'id_card_flex':id_card_flex})
    g.update({'qr_identity_issue':issue_qr,'qr_token_hash':token_hash,'id_template_flex':id_template_flex,'id_template_save_flex':id_template_save_flex,'id_card_flex':id_card_flex,'id_card_pdf':id_card_pdf})
    install_admin_flex(g)


# ---------------------------------------------------------------------------
# Enterprise Flex / Admin UX pack
# ---------------------------------------------------------------------------
def install_admin_flex(g):
    # Adds the non-destructive admin customization layer on top of 11.1.1.
    import csv, io, zipfile, time, re
    from openpyxl import Workbook, load_workbook
    from urllib.parse import urlparse, parse_qs, quote

    H=g['H']; db=g['db']; now=g['now']; esc=g['esc']; page=g['page']; csrf_field=g['csrf_field']
    can=g['can']; DATA=g['DATA']; BRAND=g['BRAND']; hashpw=g['hashpw']; audit=g['audit']; emp_allowed=g['emp_allowed']
    FLEX_PREVIEWS={}

    def upgrade_flex():
        c=db()
        c.executescript('''
        CREATE TABLE IF NOT EXISTS employee_user_links(
            emp_code TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL, created_at TEXT NOT NULL, created_by TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS leave_balance_adjustments(
            id INTEGER PRIMARY KEY, emp_code TEXT NOT NULL, leave_type TEXT NOT NULL, amount REAL NOT NULL,
            before_value REAL, after_value REAL, field_name TEXT DEFAULT 'annual', reason TEXT, created_by TEXT, created_at TEXT
        );
        ''')
        cols={r['name'] for r in c.execute('PRAGMA table_info(id_card_templates)').fetchall()}
        for col,typ in {
            'photo_x':'REAL DEFAULT 40','photo_y':'REAL DEFAULT 150','photo_w':'REAL DEFAULT 130','photo_h':'REAL DEFAULT 160',
            'name_x':'REAL DEFAULT 190','name_y':'REAL DEFAULT 170','qr_x':'REAL DEFAULT 650','qr_y':'REAL DEFAULT 145','qr_size':'REAL DEFAULT 150',
            'show_department':'INTEGER DEFAULT 1','show_job':'INTEGER DEFAULT 1','show_emp_id':'INTEGER DEFAULT 1'
        }.items():
            if col not in cols: c.execute(f'ALTER TABLE id_card_templates ADD COLUMN {col} {typ}')
        c.commit(); c.close()

    old_init=g['init']
    def init_flex():
        old_init(); upgrade_flex()
    g['init']=init_flex

    def norm(v):
        v=str(v or '').strip().lower()
        v=v.replace('أ','ا').replace('إ','ا').replace('آ','ا').replace('ى','ي').replace('ة','ه')
        return re.sub(r'[\s_\-./()]+','',v)

    aliases={
        'emp_code': {'employeeid','employee code','empid','empcode','رقمالموظف','كودالموظف','الكود','كود'},
        'name': {'name','employeename','fullname','اسم','اسمالموظف','الاسم','الاسمالكامل'},
        'leave_type': {'leavetype','type','نوعالاجازه','نوعالاجازة','النوع','الاجازه','الاجازة'},
        'annual': {'annual','allocated','allocation','entitlement','المخصص','المخصصه','السنوي','سنوي','الرصيد','رصيدالاجازة'},
        'used': {'used','taken','consumed','المستخدم','المستهلك','المستخدمه'},
        'remaining': {'remaining','balance','available','المتبقي','المتبقيه','المتاح'},
    }
    alias_norm={k:{norm(x) for x in vals} for k,vals in aliases.items()}

    def map_headers(headers):
        out={}
        for i,h in enumerate(headers):
            nh=norm(h)
            for field,vals in alias_norm.items():
                if nh in vals or any(v and v in nh for v in vals if len(v)>3):
                    out.setdefault(field,i); break
        return out

    def as_num(v):
        if v in (None,''): return 0.0
        try: return float(str(v).replace(',','').strip())
        except Exception: return 0.0

    def parse_balance_rows(rows):
        if not rows: return [], ['الملف لا يحتوي على بيانات.']
        headers=[str(x or '').strip() for x in rows[0]]; idx=map_headers(headers)
        if 'emp_code' not in idx: return [], ['لم يتم التعرف على Employee ID / رقم الموظف / كود الموظف.']
        records=[]
        if 'leave_type' in idx:
            for rn,row in enumerate(rows[1:],2):
                code=str(row[idx['emp_code']] if idx['emp_code']<len(row) else '').strip()
                if not code: continue
                lt=str(row[idx['leave_type']] if idx['leave_type']<len(row) else 'اعتيادي').strip() or 'اعتيادي'
                annual=as_num(row[idx['annual']] if 'annual' in idx and idx['annual']<len(row) else 0)
                used=as_num(row[idx['used']] if 'used' in idx and idx['used']<len(row) else 0)
                rem=as_num(row[idx['remaining']] if 'remaining' in idx and idx['remaining']<len(row) else annual-used)
                if 'used' not in idx and 'remaining' in idx: used=max(0,annual-rem)
                if annual==0 and 'remaining' in idx: annual=used+rem
                records.append({'row':rn,'emp_code':code,'name':str(row[idx['name']]).strip() if 'name' in idx and idx['name']<len(row) else '', 'leave_type':lt,'annual':annual,'used':used,'remaining':max(0,annual-used)})
        else:
            identity={idx.get('emp_code'),idx.get('name')}
            for col_i,h in enumerate(headers):
                if col_i in identity or not h: continue
                nh=norm(h)
                if nh in {'total','totals','used','remaining','المجموع','اجمالي','اجماليالاجازات'}: continue
                for rn,row in enumerate(rows[1:],2):
                    code=str(row[idx['emp_code']] if idx['emp_code']<len(row) else '').strip()
                    if not code: continue
                    val=as_num(row[col_i] if col_i<len(row) else 0)
                    if val:
                        records.append({'row':rn,'emp_code':code,'name':str(row[idx['name']]).strip() if 'name' in idx and idx['name']<len(row) else '', 'leave_type':h.strip(),'annual':val,'used':0.0,'remaining':val})
        c=db(); existing={r['emp_code'] for r in c.execute('SELECT emp_code FROM employees').fetchall()}; c.close()
        for r in records:
            if r['emp_code'] not in existing: r['warning']='الموظف غير موجود في قاعدة البيانات.'
        return records, []

    def read_xlsx(data):
        wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True); ws=wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def leave_balances(self,u):
        if not can(u,'leave.create'): return self.forbid(u)
        q=parse_qs(urlparse(self.path).query).get('q',[''])[0]
        c=db(); rows=c.execute('SELECT lb.*,e.name FROM leave_balances lb LEFT JOIN employees e ON e.emp_code=lb.emp_code ORDER BY e.name,lb.leave_type').fetchall(); c.close()
        rows=[r for r in rows if emp_allowed(u,r['emp_code']) and (not q or q.lower() in (r['emp_code']+' '+(r['name'] or '')+' '+r['leave_type']).lower())]
        trs=''.join(f'''<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"] or "")}</td><td>{esc(r["leave_type"])}</td><td><form method="post" action="/leave-balance/save" class="actions">{csrf_field(u)}<input type="hidden" name="id" value="{r['id']}"><input name="annual" value="{r['annual']}" type="number" step="0.25" style="width:95px"><input name="used" value="{r['used']}" type="number" step="0.25" style="width:95px"><input name="reason" placeholder="سبب التعديل" style="width:150px"><button class="btn">حفظ</button></form></td><td><b>{float(r['annual'] or 0)-float(r['used'] or 0):g}</b></td></tr>''' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>🏖 أرصدة الإجازات</h1><p>تعديل مباشر + استيراد Excel/Paste + سجل تعديلات + تصدير.</p></div><div class="actions"><a class="btn" href="/leave-balances/import">📥 Import / Paste</a><a class="btn gray" href="/export/leave-balances">📤 Excel Export</a><a class="btn gray" href="/export/leave-balances/template">📄 Template</a></div></div><div class="card toolbar"><input id="lbq" placeholder="بحث موظف / كود / نوع الإجازة" value="{esc(q)}" oninput="filterLB(this.value)"></div><div class="card table-wrap"><table class="table" id="lbtable"><thead><tr><th>الكود</th><th>الموظف</th><th>النوع</th><th>المخصص / المستخدم</th><th>المتبقي</th></tr></thead><tbody>{trs or '<tr><td colspan="5">لا توجد أرصدة.</td></tr>'}</tbody></table></div><script>function filterLB(v){{v=v.toLowerCase();document.querySelectorAll('#lbtable tbody tr').forEach(r=>r.style.display=r.innerText.toLowerCase().includes(v)?'':'none')}}</script>'''
        self.send(page('أرصدة الإجازات',body,u,'leaves'))

    def leave_balance_import(self,u):
        if not can(u,'import.validate'): return self.forbid(u)
        body=f'''<div class="top"><div class="title"><h1>📥 Leave Balance Import Center</h1><p>ارفع Excel أو الصق من Excel مباشرة. النظام يتعرف على الأعمدة العربية والإنجليزية ويعرض Preview قبل الحفظ.</p></div><a class="btn gray" href="/leave-balances">عودة</a></div><div class="grid g2"><div class="card"><h3>Upload Excel</h3><form method="post" action="/leave-balances/import" enctype="multipart/form-data">{csrf_field(u)}<input type="file" name="file" accept=".xlsx,.xls,.csv" required><button class="btn" style="margin-top:12px">Detect & Preview</button></form></div><div class="card"><h3>Paste from Excel</h3><form method="post" action="/leave-balances/paste">{csrf_field(u)}<textarea name="paste_data" style="width:100%;min-height:240px" placeholder="Ctrl+A → Ctrl+C من Excel ثم Ctrl+V هنا"></textarea><button class="btn" style="margin-top:12px">Detect & Preview</button></form></div></div><div class="card" style="margin-top:16px"><h3>Accepted columns</h3><p>Employee ID / رقم الموظف · Leave Type / نوع الإجازة · Annual / المخصص · Used / المستخدم · Remaining / المتبقي. ويمكن أيضًا استخدام أعمدة الإجازات بشكل أفقي.</p></div>'''
        self.send(page('Leave Import',body,u,'import'))

    def leave_import_preview(self,u,rows,source):
        records,errors=parse_balance_rows(rows); token=secrets.token_urlsafe(18); FLEX_PREVIEWS[token]={'user':u['username'],'created':time.time(),'records':records}
        valid=[r for r in records if not r.get('warning')]; warnings=[r for r in records if r.get('warning')]
        trs=''.join(f'<tr><td>{r["row"]}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["leave_type"])}</td><td>{r["annual"]:g}</td><td>{r["used"]:g}</td><td>{r["remaining"]:g}</td><td>{esc(r.get("warning", ""))}</td></tr>' for r in records[:500])
        body=f'''<div class="top"><div class="title"><h1>🔎 Leave Balance Preview</h1><p>المصدر: {esc(source)} · إجمالي {len(records)} سجل · صالح {len(valid)} · تحذيرات {len(warnings)}</p></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>Row</th><th>Employee ID</th><th>Name</th><th>Leave Type</th><th>Annual</th><th>Used</th><th>Remaining</th><th>Validation</th></tr></thead><tbody>{trs}</tbody></table></div><div class="card actions" style="margin-top:16px"><form method="post" action="/leave-balances/commit">{csrf_field(u)}<input type="hidden" name="token" value="{token}"><button class="btn ok">Import Valid Rows</button><a class="btn gray" href="/leave-balances/import">Cancel</a></form></div>'''
        self.send(page('Leave Import Preview',body,u,'import'))

    def leave_commit(self,u,f):
        tok=f.get('token',''); x=FLEX_PREVIEWS.get(tok)
        if not x or x['user']!=u['username'] or time.time()-x['created']>1800: return self.send(page('Import','<div class="card"><div class="alert">انتهت جلسة المعاينة.</div></div>',u),400)
        c=db(); imported=0; skipped=0
        for r in x['records']:
            if r.get('warning') or not emp_allowed(u,r['emp_code']): skipped+=1; continue
            old=c.execute('SELECT * FROM leave_balances WHERE emp_code=? AND leave_type=?',(r['emp_code'],r['leave_type'])).fetchone()
            c.execute('INSERT INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,?) ON CONFLICT(emp_code,leave_type) DO UPDATE SET annual=excluded.annual,used=excluded.used',(r['emp_code'],r['leave_type'],r['annual'],r['used']))
            if old:
                if float(old['annual'] or 0)!=float(r['annual']): c.execute('INSERT INTO leave_balance_adjustments(emp_code,leave_type,amount,before_value,after_value,field_name,reason,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(r['emp_code'],r['leave_type'],float(r['annual'] or 0)-float(old['annual'] or 0),old['annual'],r['annual'],'annual','Excel import',u['username'],now()))
                if float(old['used'] or 0)!=float(r['used']): c.execute('INSERT INTO leave_balance_adjustments(emp_code,leave_type,amount,before_value,after_value,field_name,reason,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(r['emp_code'],r['leave_type'],float(r['used'] or 0)-float(old['used'] or 0),old['used'],r['used'],'used','Excel import',u['username'],now()))
            imported+=1
        c.commit(); c.close(); del FLEX_PREVIEWS[tok]; audit(u['username'],u['role'],'LEAVE_BALANCE_IMPORT','Leave Balances',str(imported),f'skipped={skipped}')
        self.send(page('Import Complete',f'<div class="card"><h2>✅ تم الاستيراد</h2><p>تم: {imported} · تم تجاهله: {skipped}</p><a class="btn" href="/leave-balances">فتح الأرصدة</a></div>',u,'import'))

    def leave_balance_save(self,u,f):
        if not can(u,'leave.approve'): return self.forbid(u)
        rid=int(f.get('id') or 0); c=db(); old=c.execute('SELECT * FROM leave_balances WHERE id=?',(rid,)).fetchone()
        if not old: c.close(); return self.send(page('Balance','<div class="card"><div class="alert">الرصيد غير موجود.</div></div>',u),404)
        if not emp_allowed(u,old['emp_code']): c.close(); return self.forbid(u)
        annual=as_num(f.get('annual')); used=as_num(f.get('used')); reason=f.get('reason','')[:500]
        c.execute('UPDATE leave_balances SET annual=?,used=? WHERE id=?',(annual,used,rid))
        if float(old['annual'] or 0)!=annual: c.execute('INSERT INTO leave_balance_adjustments(emp_code,leave_type,amount,before_value,after_value,field_name,reason,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(old['emp_code'],old['leave_type'],annual-float(old['annual'] or 0),old['annual'],annual,'annual',reason,u['username'],now()))
        if float(old['used'] or 0)!=used: c.execute('INSERT INTO leave_balance_adjustments(emp_code,leave_type,amount,before_value,after_value,field_name,reason,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(old['emp_code'],old['leave_type'],used-float(old['used'] or 0),old['used'],used,'used',reason,u['username'],now()))
        c.commit(); c.close(); audit(u['username'],u['role'],'LEAVE_BALANCE_EDIT','Leave Balances',old['emp_code'],f'{old["leave_type"]}: annual={annual}, used={used}, reason={reason}'); self.redirect('/leave-balances')

    def leave_export(self,u):
        if not can(u,'reports.export'): return self.forbid(u)
        c=db(); rows=c.execute('SELECT lb.*,e.name,e.department,e.job FROM leave_balances lb LEFT JOIN employees e ON e.emp_code=lb.emp_code ORDER BY e.name,lb.leave_type').fetchall(); c.close()
        wb=Workbook(); ws=wb.active; ws.title='Leave Balances'; ws.append(['Employee ID','Employee Name','Department','Job Title','Leave Type','Allocated','Used','Remaining'])
        for r in rows:
            if emp_allowed(u,r['emp_code']): ws.append([r['emp_code'],r['name'],r['department'],r['job'],r['leave_type'],r['annual'],r['used'],float(r['annual'] or 0)-float(r['used'] or 0)])
        out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="leave_balances.xlsx"'})

    def leave_template(self,u):
        if not can(u,'reports.export'): return self.forbid(u)
        wb=Workbook(); ws=wb.active; ws.title='Leave Balances'; ws.append(['Employee ID','Employee Name','Leave Type','Annual','Used','Remaining']); ws.append(['EMP-00001','Ahmed Elsayed','اعتيادي',21,5,16]); out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Leave_Balances_Template.xlsx"'})

    def master_export(self,u):
        if not can(u,'reports.export'): return self.forbid(u)
        c=db(); emps=c.execute('SELECT * FROM employees ORDER BY name').fetchall(); links={r['emp_code']:r for r in c.execute('SELECT * FROM employee_user_links').fetchall()}; qrs={r['emp_code']:r for r in c.execute('SELECT id,status,issued_at,revoked_at,image_path FROM qr_identities').fetchall()}; c.close()
        wb=Workbook(); ws=wb.active; ws.title='Employees'; ws.append(['Employee ID','Name','Department','Unit','Job Title','Status','Hire Date','Username','User Role','User Status','QR ID','QR Status','QR Issued','QR Revoked','Photo URL'])
        for e in emps:
            if not emp_allowed(u,e['emp_code']): continue
            lk=links.get(e['emp_code']); qr=qrs.get(e['emp_code']); ws.append([e['emp_code'],e['name'],e['department'],e['unit'],e['job'],e['status'],e['contract_date'],lk['username'] if lk else '', '', 'linked' if lk else '', qr['id'] if qr else '', qr['status'] if qr else '', qr['issued_at'] if qr else '', qr['revoked_at'] if qr else '', f'/employee/photo/{quote(e["emp_code"])}'])
        out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="HR_Employee_Master.xlsx"'})

    def qr_bundle(self,u):
        if not can(u,'reports.export'): return self.forbid(u)
        c=db(); rows=c.execute('SELECT q.*,e.name,e.department,e.job FROM qr_identities q JOIN employees e ON e.emp_code=q.emp_code ORDER BY e.name').fetchall(); c.close(); z=io.BytesIO()
        with zipfile.ZipFile(z,'w',zipfile.ZIP_DEFLATED) as zz:
            wb=Workbook(); ws=wb.active; ws.title='QR Index'; ws.append(['Employee ID','Name','Department','Job','QR ID','Status','Issued At','Revoked At'])
            for r in rows:
                if not emp_allowed(u,r['emp_code']): continue
                p=os.path.join(DATA,r['image_path'] or '')
                ws.append([r['emp_code'],r['name'],r['department'],r['job'],r['id'],r['status'],r['issued_at'],r['revoked_at'] or ''])
                if os.path.exists(p):
                    with open(p,'rb') as fh: zz.writestr(f'QR/{safe_name(r["emp_code"])}.png',fh.read())
            x=io.BytesIO(); wb.save(x); zz.writestr('employees.xlsx',x.getvalue())
        self.send(z.getvalue(),200,'application/zip',{'Content-Disposition':'attachment; filename="HR_QR_Export.zip"'})

    def _gen_username_from_name(name,taken):
        words=re.findall(r'[\u0600-\u06FFA-Za-z0-9]+',name or '')
        base='_'.join(words[:2]) if words else 'موظف'
        candidate=base; n=2
        while candidate in taken:
            candidate=f'{base}{n}'; n+=1
        taken.add(candidate)
        return candidate

    def _gen_numeric_password(digits=6):
        return ''.join(secrets.choice('0123456789') for _ in range(digits))

    def bulk_users_confirm(self,u):
        if not can(u,'users.manage'): return self.forbid(u)
        c=db()
        rows=c.execute("SELECT emp_code,name FROM employees WHERE status='على رأس العمل' AND emp_code NOT IN (SELECT emp_code FROM employee_user_links) ORDER BY name").fetchall()
        c.close()
        if not rows:
            body='<div class="card"><div class="alert">كل الموظفين النشطين عندهم حسابات بالفعل — مفيش حد محتاج حساب جديد دلوقتي.</div></div><a class="btn gray" href="/employees">عودة</a>'
            return self.send(page('توليد حسابات',body,u,'employees'))
        trs=''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td></tr>' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>👤 توليد حسابات دخول تلقائيًا</h1><p>هيتولّد اسم مستخدم عربي (أول كلمتين من اسم الموظف) وكلمة مرور أرقام لكل موظف نشط لسه ماعندوش حساب. العدد الحالي: <b>{len(rows)}</b> موظف.</p></div><a class="btn gray" href="/employees">إلغاء</a></div>
<div class="card table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الاسم</th></tr></thead><tbody>{trs}</tbody></table></div>
<div class="card" style="margin-top:16px"><form method="post" action="/employees/bulk-users/create">{csrf_field(u)}<div class="field"><label>عدد أرقام كلمة المرور</label><select name="digits"><option value="4">4 أرقام</option><option value="6" selected>6 أرقام</option></select></div><button class="btn ok">✓ توليد {len(rows)} حساب الآن</button></form></div>'''
        return self.send(page('توليد حسابات',body,u,'employees'))

    def bulk_users_create(self,u,f):
        if not can(u,'users.manage'): return self.forbid(u)
        try: digits=int(f.get('digits') or 6)
        except ValueError: digits=6
        digits=6 if digits not in (4,6) else digits
        c=db()
        rows=c.execute("SELECT emp_code,name FROM employees WHERE status='على رأس العمل' AND emp_code NOT IN (SELECT emp_code FROM employee_user_links) ORDER BY name").fetchall()
        taken=set(x[0] for x in c.execute('SELECT username FROM users').fetchall())
        results=[]
        for r in rows:
            uname=_gen_username_from_name(r['name'],taken)
            pwd=_gen_numeric_password(digits)
            c.execute('INSERT INTO users(username,password_hash,role,full_name,must_change_password,active) VALUES(?,?,?,?,1,1)',(uname,hashpw(pwd),'Employee',r['name']))
            c.execute('INSERT INTO employee_user_links(emp_code,username,created_at,created_by) VALUES(?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET username=excluded.username,created_at=excluded.created_at,created_by=excluded.created_by',(r['emp_code'],uname,now(),u['username']))
            results.append((r['emp_code'],r['name'],uname,pwd))
        c.commit(); c.close()
        audit(u['username'],u['role'],'BULK_USER_CREATE','Users','-',f'{len(results)} accounts created')
        if not results:
            body='<div class="card"><div class="alert">مفيش موظفين محتاجين حساب جديد.</div></div><a class="btn gray" href="/employees">عودة</a>'
            return self.send(page('حسابات جديدة',body,u,'employees'))
        trs=''.join(f'<tr><td>{esc(ec)}</td><td>{esc(nm)}</td><td><b>{esc(un)}</b></td><td><code style="font-size:15px">{esc(pw)}</code></td></tr>' for ec,nm,un,pw in results)
        body=f'''<div class="top no-print"><div class="title"><h1>✅ تم توليد {len(results)} حساب جديد</h1><p>اطبع أو صوّر الجدول ده دلوقتي — كلمات المرور دي بتتعرض مرة واحدة بس وميتسجلش في أي سجل أو تصدير تاني.</p></div><div class="actions"><button class="btn gray" onclick="window.print()">🖨 طباعة</button><a class="btn gray" href="/employees">عودة للموظفين</a></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الاسم</th><th>اسم المستخدم</th><th>كلمة المرور المؤقتة</th></tr></thead><tbody>{trs}</tbody></table></div>'''
        return self.send(page('حسابات جديدة',body,u,'employees'))

    def admin_flex_page(self,u):
        if not can(u,'settings.manage'): return self.forbid(u)
        body='''<div class="top"><div class="title"><h1>⚙️ Enterprise Admin Center</h1><p>تخصيص النظام بدون كود: هوية، بيانات، صلاحيات، بطاقات، استيراد وتصدير.</p></div></div><div class="grid g3"><div class="card"><h3>👥 People & Identity</h3><div class="actions"><a class="btn" href="/employees">Employees</a><a class="btn gray" href="/id-cards">ID Cards</a><a class="btn gray" href="/qr/scan">QR Scanner</a><a class="btn gray" href="/employees/bulk-users">👤 توليد حسابات دخول</a></div></div><div class="card"><h3>🏖 Leave Data</h3><div class="actions"><a class="btn" href="/leave-balances">Balances</a><a class="btn gray" href="/leave-balances/import">Import / Paste</a></div></div><div class="card"><h3>📤 Exports</h3><div class="actions"><a class="btn gray" href="/export/employee-master">Employee Master</a><a class="btn gray" href="/export/qr-bundle">QR ZIP</a></div></div></div><div class="grid g2" style="margin-top:16px"><div class="card"><h3>🎨 Branding & Templates</h3><p>ارفع الشعار، خلفية الـID، ألوان النظام وقوالب التقارير من داخل البرنامج.</p><div class="actions"><a class="btn" href="/branding">Branding</a><a class="btn gray" href="/branding/id-card-template">ID Card Designer</a></div></div><div class="card"><h3>🔐 Permissions</h3><p>SuperAdmin / Admin / HR / Manager / Employee مع صلاحيات granular على السيرفر.</p><a class="btn" href="/roles">Open RBAC</a></div></div>'''
        self.send(page('Enterprise Admin Center',body,u,'settings'))

    def shifts_page_flex(self,u):
        if not can(u,'shifts.manage'): return self.forbid(u)
        c=db(); rows=c.execute('SELECT s.*,COUNT(es.emp_code) assigned FROM shifts s LEFT JOIN employee_shifts es ON es.shift_id=s.id GROUP BY s.id ORDER BY s.active DESC,s.id').fetchall(); c.close()
        trs=''.join(f'''<tr><td><form method="post" action="/shift/save" class="actions">{csrf_field(u)}<input type="hidden" name="id" value="{r['id']}"><input name="name" value="{esc(r['name'])}" style="width:130px"><input type="time" name="start_time" value="{esc(r['start_time'])}"><input type="time" name="end_time" value="{esc(r['end_time'])}"><input type="number" name="grace_minutes" value="{r['grace_minutes']}" style="width:75px"><input type="number" name="warning_minutes" value="{r['warning_minutes'] if 'warning_minutes' in r.keys() else 15}" style="width:75px"><button class="btn">حفظ</button></form></td><td>{r['assigned']}</td><td>{'نشطة' if r['active'] else 'معطلة'}</td><td>{'' if r['assigned'] else f'''<form method="post" action="/shift/delete" style="display:inline" onsubmit="return confirm('حذف الوردية؟')">{csrf_field(u)}<input type="hidden" name="id" value="{r['id']}"><button class="btn bad">حذف</button></form>'''}</td></tr>''' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>🕐 الورديات / الشيفتات</h1><p>إضافة وتعديل وحذف وتعيين. لا يمكن حذف وردية مرتبطة بموظفين حتى لا تضيع بيانات الحضور.</p></div></div><div class="card"><form class="form" method="post" action="/shift/save">{csrf_field(u)}<div class="field"><label>اسم الوردية</label><input name="name" required></div><div class="field"><label>بداية</label><input type="time" name="start_time" value="09:00"></div><div class="field"><label>نهاية</label><input type="time" name="end_time" value="17:00"></div><div class="field"><label>السماح بالدقائق</label><input type="number" name="grace_minutes" value="15"></div><div class="field"><label>الإنذار بعد</label><input type="number" name="warning_minutes" value="15"></div><div class="full"><button class="btn">➕ إضافة وردية</button></div></form></div><div class="card table-wrap" style="margin-top:16px"><table class="table"><thead><tr><th>الوردية / التعديل</th><th>موظفون</th><th>الحالة</th><th>حذف</th></tr></thead><tbody>{trs}</tbody></table></div>'''
        self.send(page('الورديات',body,u,'shifts'))

    old_profile=H.employee_profile
    def profile_flex(self,u,code):
        import types
        captured=[]; original=self.send
        def capture(self,body,status=200,ctype='text/html',headers=None): captured.append((body,status,ctype,headers))
        self.send=types.MethodType(capture,self)
        try: old_profile(self,u,code)
        finally: self.send=original
        if not captured: return
        body,status,ctype,headers=captured[0]
        if status!=200 or 'text/html' not in ctype: return original(body,status,ctype,headers)
        if isinstance(body,bytes): body=body.decode('utf-8','replace')
        c=db(); emp=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone(); link=c.execute('SELECT u.username,u.role,u.active,u.last_login FROM employee_user_links l JOIN users u ON u.username=l.username WHERE l.emp_code=?',(code,)).fetchone(); qr=c.execute('SELECT * FROM qr_identities WHERE emp_code=? ORDER BY id DESC LIMIT 1',(code,)).fetchone(); photo=c.execute("SELECT id FROM documents WHERE emp_code=? AND category='صورة' AND status='current' ORDER BY id DESC LIMIT 1",(code,)).fetchone(); c.close()
        if not emp: return original(body,status,ctype,headers)
        photo_url=f'/employee/photo/{quote(code)}?v={photo["id"]}' if photo else ''
        userbox=f'''<div class="card" style="margin-top:16px"><div class="top"><div><h3>👤 حساب الموظف</h3><p>إنشاء حساب دخول مرتبط بهذا الموظف مع دور وصلاحيات حقيقية.</p></div>{('<span class="badge b-ok">'+esc(link['username'])+' · '+esc(link['role'])+'</span>') if link else '<span class="badge b-gray">لا يوجد حساب</span>'}</div>'''
        if link:
            userbox+=f'''<p>الحالة: {'نشط' if link['active'] else 'موقوف'} · آخر دخول: {esc(link['last_login'] or '—')}</p><form method="post" action="/employee/user/create">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><div class="actions"><input name="username" value="{esc(link['username'])}" placeholder="Username"><select name="role"><option>Employee</option><option>Manager</option><option>HR</option></select><input name="password" placeholder="كلمة مرور جديدة (اختياري)"><button class="btn">حفظ / إعادة تعيين</button></div></form>'''
        else:
            userbox+=f'''<form method="post" action="/employee/user/create">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><div class="actions"><input name="username" value="{esc(code).lower()}" placeholder="Username"><select name="role"><option>Employee</option><option>Manager</option><option>HR</option></select><button class="btn">Create User</button></div></form>'''
        userbox+='</div>'
        identity=f'''<div class="card" style="margin-top:16px"><h3>🪪 الهوية الرقمية</h3><div class="grid g3"><div><b>كود الموظف</b><div>{esc(code)}</div></div><div><b>معرّف QR</b><div>{qr['id'] if qr else '—'}</div></div><div><b>حالة QR</b><div>{esc(qr['status']) if qr else 'غير مُصدر'}</div></div></div><div class="actions" style="margin-top:12px"><a class="btn" href="/id-card/{quote(code)}">عرض بطاقة الهوية</a><a class="btn gray" href="/id-card/{quote(code)}?autoprint=1">🖨 طباعة البطاقة</a>{f'<a class="btn gray" href="/qr/image/{quote(code)}">QR PNG</a>' if qr and qr['image_path'] else ''}</div></div>'''
        photo_actions=f'''<div class="card" style="margin-top:16px"><div class="top"><div><h3>📸 صورة الموظف</h3><p>JPG / PNG / WEBP حتى 5MB، وتستخدم تلقائيًا في الملف والـID والـPDF.</p></div>{f'<img src="{photo_url}" style="width:96px;height:116px;object-fit:cover;border-radius:14px;border:1px solid #e4e7ec">' if photo else '<div style="width:96px;height:116px;border-radius:14px;background:#f2f4f7;display:grid;place-items:center;font-size:38px">👤</div>'}</div><form method="post" action="/employee/photo/upload" enctype="multipart/form-data">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><div class="actions"><input type="file" name="file" accept=".png,.jpg,.jpeg,.webp" required><button class="btn">رفع / استبدال</button>{('<button class="btn bad" name="delete" value="1">حذف الصورة</button>') if photo else ''}</div></form></div>'''
        body=body.replace('</body>',photo_actions+identity+userbox+'</body>')
        return original(body,status,ctype,headers)
    H.employee_profile=profile_flex

    old_get=H.do_GET; old_post=H.do_POST
    def get_flex(self):
        p=urlparse(self.path).path
        if p=='/leave-balances': u=self.require(); return leave_balances(self,u) if u else None
        if p=='/leave-balances/import': u=self.require(); return leave_balance_import(self,u) if u else None
        if p=='/export/leave-balances': u=self.require(); return leave_export(self,u) if u else None
        if p=='/export/leave-balances/template': u=self.require(); return leave_template(self,u) if u else None
        if p=='/export/employee-master': u=self.require(); return master_export(self,u) if u else None
        if p=='/export/qr-bundle': u=self.require(); return qr_bundle(self,u) if u else None
        if p=='/admin/flex': u=self.require(); return admin_flex_page(self,u) if u else None
        if p=='/branding/id-card-template': u=self.require(); return g['id_template_flex'](self,u) if u else None
        if p.startswith('/id-card/') and p.count('/')==2: u=self.require(); return g['id_card_flex'](self,u,p.split('/')[-1]) if u else None
        if p.startswith('/id-card-pdf/') and p.count('/')==2: u=self.require(); return g['id_card_pdf'](self,u,p.split('/')[-1]) if u else None
        if p=='/shifts': u=self.require(); return shifts_page_flex(self,u) if u else None
        if p=='/employees/bulk-users': u=self.require(); return bulk_users_confirm(self,u) if u else None
        return old_get(self)

    def post_flex(self):
        p=urlparse(self.path).path
        if p=='/branding/id-card-template/save':
            u=self.require()
            if not u:return
            fields,files=self.parse_upload_all()
            if fields.get('_csrf')!=u.get('csrf'): return self.send(page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
            return g['id_template_save_flex'](self,u,fields,files)
        custom={'/leave-balances/import','/leave-balances/paste','/leave-balances/commit','/leave-balance/save','/shift/save','/shift/delete','/employee/user/create','/employee/photo/upload','/employees/bulk-users/create'}
        if p in custom:
            u=self.require()
            if not u:return
            ctype=self.headers.get('Content-Type','').lower()
            if ctype.startswith('multipart/form-data'):
                f,part=self.parse_upload(); files=[part] if part else []
            else:
                f=self.form(); files=[]
            if f.get('_csrf')!=u.get('csrf'): return self.send(page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
            if p=='/leave-balances/import':
                if not can(u,'import.validate'): return self.forbid(u)
                if not files:return self.send(page('Import','<div class="card"><div class="alert">اختر ملف Excel.</div></div>',u),400)
                _,data,fname=files[0]; ext=os.path.splitext(fname)[1].lower()
                try:
                    rows=read_xlsx(data) if ext in ('.xlsx','.xls') else list(csv.reader(io.StringIO(data.decode('utf-8-sig',errors='replace'))))
                    return leave_import_preview(self,u,rows,fname)
                except Exception as e:return self.send(page('Import',f'<div class="card"><div class="alert">{esc(e)}</div></div>',u),400)
            if p=='/leave-balances/paste':
                if not can(u,'import.validate'): return self.forbid(u)
                rows=list(csv.reader(io.StringIO(f.get('paste_data','')),delimiter='\t'))
                return leave_import_preview(self,u,rows,'Excel Paste')
            if p=='/leave-balances/commit': return leave_commit(self,u,f)
            if p=='/leave-balance/save': return leave_balance_save(self,u,f)
            if p=='/shift/save':
                if not can(u,'shifts.manage'): return self.forbid(u)
                sid=int(f.get('id') or 0); name=f.get('name','').strip()
                if not name:return self.send(page('Shift','<div class="card"><div class="alert">اسم الوردية مطلوب.</div></div>',u),400)
                c=db(); args=(name,f.get('start_time') or '09:00',f.get('end_time') or '17:00',int(f.get('grace_minutes') or 0),int(f.get('warning_minutes') or 15))
                if sid:c.execute('UPDATE shifts SET name=?,start_time=?,end_time=?,grace_minutes=?,warning_minutes=?,active=1 WHERE id=?',args+(sid,))
                else:c.execute('INSERT INTO shifts(name,start_time,end_time,grace_minutes,warning_minutes,active) VALUES(?,?,?,?,?,1) ON CONFLICT(name) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,grace_minutes=excluded.grace_minutes,warning_minutes=excluded.warning_minutes,active=1',args)
                c.commit(); c.close(); audit(u['username'],u['role'],'SHIFT_SAVE','Shifts',name); return self.redirect('/shifts')
            if p=='/shift/delete':
                if not can(u,'shifts.manage'): return self.forbid(u)
                sid=int(f.get('id') or 0); c=db(); assigned=c.execute('SELECT COUNT(*) n FROM employee_shifts WHERE shift_id=?',(sid,)).fetchone()['n']
                if assigned:c.close(); return self.send(page('Shift','<div class="card"><div class="alert">لا يمكن حذف وردية مرتبطة بموظفين. غيّر التعيين أولًا.</div></div>',u),409)
                c.execute('DELETE FROM shifts WHERE id=?',(sid,)); c.commit(); c.close(); audit(u['username'],u['role'],'SHIFT_DELETE','Shifts',str(sid)); return self.redirect('/shifts')
            if p=='/employee/user/create':
                if not can(u,'users.manage'): return self.forbid(u)
                emp=f.get('emp_code','').strip(); c=db(); e=c.execute('SELECT emp_code,name FROM employees WHERE emp_code=?',(emp,)).fetchone()
                if not e or not emp_allowed(u,emp): c.close(); return self.forbid(u)
                username=(f.get('username') or emp.lower()).strip(); role=f.get('role') or 'Employee'; password=f.get('password') or secrets.token_urlsafe(8)
                if not re.fullmatch(r'[A-Za-z0-9._\u0600-\u06FF-]{3,80}',username): c.close(); return self.send(page('User','<div class="card"><div class="alert">اسم المستخدم يجب أن يكون حروف عربية/إنجليزية أو أرقام أو ._- من 3 إلى 80 حرفًا.</div></div>',u),400)
                c.execute('INSERT INTO users(username,password_hash,role,full_name,must_change_password,active) VALUES(?,?,?,?,1,1) ON CONFLICT(username) DO UPDATE SET password_hash=excluded.password_hash,role=excluded.role,full_name=excluded.full_name,active=1,must_change_password=1',(username,hashpw(password),role,e['name']))
                c.execute('INSERT INTO employee_user_links(emp_code,username,created_at,created_by) VALUES(?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET username=excluded.username,created_at=excluded.created_at,created_by=excluded.created_by',(emp,username,now(),u['username']))
                c.commit(); c.close(); audit(u['username'],u['role'],'EMPLOYEE_USER_CREATE','Users',username,f'emp={emp};role={role}')
                msg=f'<div class="card"><h2>✅ تم إنشاء/تحديث الحساب</h2><p>Employee: <b>{esc(emp)}</b> · Username: <b>{esc(username)}</b> · Role: <b>{esc(role)}</b></p><div class="alert">كلمة المرور المؤقتة: <b>{esc(password)}</b><br>لن يتم تصدير كلمات المرور إلى Excel.</div><a class="btn" href="/employee/profile/{quote(emp)}">عودة للموظف</a></div>'
                return self.send(page('User Created',msg,u,'employees'))
            if p=='/employees/bulk-users/create': return bulk_users_create(self,u,f)
            if p=='/employee/photo/upload':
                emp=f.get('emp_code','').strip()
                if not emp or not can(u,'employees.edit') or not emp_allowed(u,emp): return self.forbid(u)
                if f.get('delete'):
                    c=db(); c.execute("UPDATE documents SET status='superseded' WHERE emp_code=? AND category='صورة' AND status='current'",(emp,)); c.commit(); c.close(); audit(u['username'],u['role'],'PROFILE_PHOTO_DELETE','Employees',emp); return self.redirect('/employee/profile/'+quote(emp))
                if not files:return self.send(page('Photo','<div class="card"><div class="alert">اختر صورة.</div></div>',u),400)
                _,data,fname=files[0]; ext=os.path.splitext(fname)[1].lower()
                if not data or ext not in ('.jpg','.jpeg','.png','.webp') or len(data)>5*1024*1024:return self.send(page('Photo','<div class="card"><div class="alert">PNG/JPG/WEBP وبحد أقصى 5MB.</div></div>',u),400)
                if not g['validate_file_signature'](fname,data):return self.send(page('Photo','<div class="card"><div class="alert">محتوى الملف لا يطابق امتداده أو الملف تالف.</div></div>',u),400)
                # PHASE 1 PERF FIX: this is the handler that actually runs for
                # /employee/photo/upload (install_admin_flex's post_flex intercepts
                # it before the older post() in install_enterprise ever sees it).
                # The older handler had a Pillow re-encode/thumbnail step with a
                # comment explaining phone-camera photos were slowing every
                # employee-list page down — but that step was dead code, since THIS
                # handler is the one actually storing the file, full-resolution,
                # unresized. Bringing the same fix here so it actually takes effect.
                try:
                    from PIL import Image, ImageOps
                    im=Image.open(io.BytesIO(data)); im=ImageOps.exif_transpose(im); im=im.convert('RGB')
                    im.thumbnail((900,900))
                    buf=io.BytesIO(); im.save(buf,'JPEG',quality=85,optimize=True); data=buf.getvalue(); ext='.jpg'
                except Exception:
                    pass  # Pillow unavailable or unreadable image: keep the validated original as-is.
                rel=g['save_employee_file'](emp,'profile'+ext,data); c=db(); c.execute("UPDATE documents SET status='superseded' WHERE emp_code=? AND category='صورة' AND status='current'",(emp,)); c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,'profile'+ext,ext,u['username'],now(),None,rel,'صورة',1,hashlib.sha256(data).hexdigest(),'current')); c.commit(); c.close(); audit(u['username'],u['role'],'PROFILE_PHOTO_UPDATED','Employees',emp); return self.redirect('/employee/profile/'+quote(emp))
        return old_post(self)
    H.do_GET=get_flex; H.do_POST=post_flex

    old_page=g['page']
    def page_flex(title,body,user,active='dashboard'):
        out=old_page(title,body,user,active)
        if user and can(user,'settings.manage'):
            out=out.replace('</nav>','<details class="nav-group"><summary>إدارة متقدمة</summary><a href="/admin/flex">⚙️ Enterprise Admin Center</a></details></nav>',1)
        return out
    g['page']=page_flex
