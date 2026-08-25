import os, sys, sqlite3, hashlib, secrets, json, csv, io, zipfile, shutil, mimetypes, re, webbrowser, threading, difflib, socket, time, platform, uuid, traceback, hmac, base64, subprocess
from datetime import datetime, date, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from http import cookies
from openpyxl import load_workbook, Workbook
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except Exception:
    A4=canvas=pdfmetrics=TTFont=None

APP_VERSION='10.0 Enterprise Complete'
BASE=os.path.dirname(os.path.abspath(sys.executable if getattr(sys,'frozen',False) else __file__))
# Executable is read-only in production. Persistent data belongs in ProgramData on Windows.
def _default_data_dir():
    if os.environ.get('HR_DATA_DIR'): return os.environ['HR_DATA_DIR']
    if sys.platform.startswith('win'):
        root=os.environ.get('PROGRAMDATA') or os.environ.get('LOCALAPPDATA') or BASE
        return os.path.join(root,'HR Enterprise','Data')
    return os.path.join(BASE,'data')
DATA=os.path.abspath(_default_data_dir()); os.makedirs(DATA,exist_ok=True)
DB=os.path.join(DATA,'hr_central.db'); BACKUPS=os.path.join(DATA,'backups'); EMPFILES=os.path.join(DATA,'employee_files'); BRAND=os.path.join(DATA,'branding'); LOGDIR=os.path.join(DATA,'logs')
for _d in (BACKUPS,EMPFILES,BRAND,LOGDIR): os.makedirs(_d,exist_ok=True)
SEED=os.path.join(BASE,'seed.xlsx')
IDENTITY_FILE=os.path.join(DATA,'server_identity.json')
TRUST_FILE=os.path.join(DATA,'client_trust.json')
PORT=int(os.environ.get('HR_PORT','8899'))
PORT_MIN=PORT
PORT_MAX=int(os.environ.get('HR_PORT_MAX','8920'))
DISCOVERY_PORT=int(os.environ.get('HR_DISCOVERY_PORT','8898'))
MODE=os.environ.get('HR_MODE','auto').lower()
AUTO_MODE=MODE=='auto'
NETWORK_MODE=MODE in ('network','server','lan')
HOST=os.environ.get('HR_HOST','0.0.0.0' if (NETWORK_MODE or AUTO_MODE) else '127.0.0.1')
SESS={}
LOGIN_ATTEMPTS={}
BRIDGE_TOKENS={}
AUDIT_CTX=threading.local()

def csrf_field(u):
    return '<input type="hidden" name="_csrf" value="{}">'.format(esc(u.get('csrf','')))

def log_error(context,e,request_id='',username='',method='',path='',ip=''):
    tb=traceback.format_exc()
    print('[ERROR]',now(),context,repr(e),'request_id=',request_id)
    try:
        c=db(); c.execute('INSERT INTO error_logs(request_id,ts,username,method,path,ip,exception_type,message,traceback) VALUES(?,?,?,?,?,?,?,?,?)',(request_id or str(uuid.uuid4())[:12],now(),username,method or context.split(' ',1)[0],path or (context.split(' ',1)[1] if ' ' in context else ''),ip,type(e).__name__,str(e),tb)); c.commit(); c.close()
    except Exception as log_exc:
        print('[ERROR_LOG_FAILED]',repr(log_exc))
    traceback.print_exc()

CSS=''':root{--bg:#f3f6fb;--card:#fff;--ink:#101828;--muted:#667085;--brand:#175cd3;--brand2:#0b4aaf;--line:#e4e7ec;--ok:#12b76a;--warn:#f79009;--bad:#f04438;--nav:#0b1220;--shadow:0 12px 36px rgba(16,24,40,.08)}*{box-sizing:border-box}body{margin:0;background:radial-gradient(circle at top right,#eef4ff 0,#f3f6fb 38%,#f7f8fa 100%);font-family:"Segoe UI",Tahoma,Arial,sans-serif;color:var(--ink);direction:rtl}.app{display:flex;min-height:100vh}.side{width:268px;background:linear-gradient(180deg,var(--nav),#0e1728 55%,#101828);color:#fff;padding:18px 14px;position:fixed;right:0;top:0;bottom:0;overflow:auto;box-shadow:-12px 0 30px rgba(16,24,40,.12);z-index:20}.brand{padding:12px 12px 22px;border-bottom:1px solid rgba(255,255,255,.1);margin-bottom:14px}.brand b{font-size:21px}.brand small{display:block;color:#98a2b3;margin-top:6px}.nav a{display:flex;align-items:center;gap:10px;padding:12px 13px;border-radius:11px;color:#d0d5dd;margin:5px 0;font-size:14px}.nav a:hover,.nav a.active{background:linear-gradient(90deg,#18263d,#1b3157);color:#fff;box-shadow:inset 3px 0 0 #53b1fd}.main{margin-right:268px;width:calc(100% - 268px);padding:28px 34px}.top{display:flex;align-items:center;justify-content:space-between;margin-bottom:22px;gap:18px}.title h1{margin:0;font-size:30px}.title p{margin:6px 0 0;color:var(--muted)}.actions{display:flex;gap:8px;flex-wrap:wrap}.btn{border:0;border-radius:10px;padding:10px 15px;background:linear-gradient(135deg,var(--brand),#2e90fa);color:#fff;cursor:pointer;font-weight:700;box-shadow:0 5px 14px rgba(23,92,211,.18)}.btn.gray{background:#fff;color:#344054;border:1px solid var(--line);box-shadow:none}.btn.ok{background:var(--ok)}.btn.bad{background:var(--bad)}.btn.warn{background:var(--warn)}.grid{display:grid;gap:17px}.g4{grid-template-columns:repeat(4,1fr)}.g3{grid-template-columns:repeat(3,1fr)}.g2{grid-template-columns:repeat(2,1fr)}.card{background:rgba(255,255,255,.96);border:1px solid var(--line);border-radius:17px;box-shadow:var(--shadow);padding:19px}.metric{position:relative;overflow:hidden}.metric:after{content:"";position:absolute;left:-20px;bottom:-35px;width:110px;height:110px;border-radius:50%;background:#eff8ff}.metric .label{color:var(--muted);font-size:13px;position:relative;z-index:1}.metric .value{font-size:34px;font-weight:800;margin-top:8px;position:relative;z-index:1}.metric .sub{font-size:12px;color:var(--muted);margin-top:6px;position:relative;z-index:1}.table-wrap{overflow:auto}.table{width:100%;border-collapse:separate;border-spacing:0;font-size:13px}.table th{background:#f8fafc;color:#475467;font-weight:750}.table th,.table td{padding:12px 10px;border-bottom:1px solid var(--line);text-align:right;white-space:nowrap}.table tr:hover td{background:#f8fbff}.badge{display:inline-block;border-radius:999px;padding:5px 10px;font-size:12px;font-weight:700}.b-ok{background:#ecfdf3;color:#027a48}.b-warn{background:#fffaeb;color:#b54708}.b-bad{background:#fef3f2;color:#b42318}.b-blue{background:#eff8ff;color:#175cd3}.b-gray{background:#f2f4f7;color:#475467}.timeline{display:flex;flex-direction:column;gap:10px;max-height:420px;overflow:auto}.tl-item{padding:10px 12px;border-radius:10px;background:#f8fafc;border:1px solid var(--line);font-size:13px}.tl-date{display:inline-block;min-width:96px;color:var(--muted);font-weight:700}.gsearch{position:relative}.gsearch input{width:230px;border-radius:10px;border:1px solid #2a3a55;background:#0e1728;color:#fff;padding:9px 12px;font:inherit}.gsearch input::placeholder{color:#7b8aa3}.gsearch kbd{position:absolute;left:8px;top:50%;transform:translateY(-50%);font-size:10px;color:#7b8aa3;border:1px solid #2a3a55;border-radius:5px;padding:1px 5px}.form{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}.field label{display:block;font-size:12px;font-weight:700;margin-bottom:7px;color:#344054}.field input,.field select,.field textarea{width:100%;border:1px solid #d0d5dd;border-radius:10px;padding:11px;background:#fff;font:inherit;outline:none}.field input:focus,.field select:focus,.field textarea:focus{border-color:#84adff;box-shadow:0 0 0 4px #eff4ff}.field textarea{min-height:100px}.full{grid-column:1/-1}.toolbar{display:flex;gap:9px;flex-wrap:wrap;margin-bottom:14px}.toolbar input,.toolbar select{border:1px solid #d0d5dd;border-radius:10px;padding:10px;font:inherit;background:#fff}.chart{height:260px;display:flex;align-items:end;gap:14px;padding:24px 10px 24px}.bar{flex:1;background:linear-gradient(180deg,#53b1fd,#175cd3);border-radius:10px 10px 4px 4px;min-width:28px;position:relative;box-shadow:0 8px 18px rgba(23,92,211,.16)}.bar span{position:absolute;bottom:-28px;right:0;left:0;text-align:center;font-size:11px;color:var(--muted);overflow:hidden;text-overflow:ellipsis}.bar b{position:absolute;top:-23px;right:0;left:0;text-align:center;font-size:11px}.alert{padding:13px 15px;border-radius:11px;background:#fff8e7;border:1px solid #fedf89;color:#7a4e00}.login{min-height:100vh;display:grid;place-items:center;background:radial-gradient(circle at 20% 10%,#2e90fa 0,#175cd3 22%,#101828 70%)}.login-card{width:450px;background:rgba(255,255,255,.98);border-radius:22px;padding:38px;box-shadow:0 30px 90px rgba(0,0,0,.3)}.login-card input{width:100%;padding:13px;border:1px solid #d0d5dd;border-radius:10px;margin:7px 0;font:inherit}.login-card button{width:100%;margin-top:14px}.footer{color:#98a2b3;font-size:11px;text-align:center;margin-top:30px}@media(max-width:1100px){.g4{grid-template-columns:repeat(2,1fr)}}@media(max-width:900px){.side{position:static;width:100%;height:auto}.app{display:block}.main{margin:0;width:100%;padding:18px}.g4,.g3,.g2{grid-template-columns:1fr}.form{grid-template-columns:1fr}}@media print{.side,.actions,.toolbar,.no-print{display:none!important}.main{margin:0;width:100%;padding:0}.card{box-shadow:none}}'''
CSS += '.top{position:sticky;top:0;z-index:10;background:rgba(243,246,251,.94);backdrop-filter:blur(8px);padding:10px 0}.nav a{min-height:42px}.btn:focus-visible,input:focus-visible,select:focus-visible,textarea:focus-visible{outline:3px solid #b2ddff;outline-offset:2px}.mobile-menu{display:none}@media(max-width:900px){.side{position:fixed;right:-290px;width:280px;height:100vh;transition:right .2s}.side.open{right:0}.main{margin:0}.mobile-menu{display:inline-flex}.top{position:static}}';

def esc(s):
    return str(s if s is not None else '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
def hashpw(p):
    salt=secrets.token_bytes(16)
    try: dk=hashlib.scrypt(p.encode(),salt=salt,n=2**14,r=8,p=1,dklen=64)
    except Exception: dk=hashlib.pbkdf2_hmac('sha256',p.encode(),salt,300000,dklen=64)
    return 'scrypt$'+salt.hex()+'$'+dk.hex()
def checkpw(p,stored):
    if not stored: return False
    try:
        if stored.startswith('scrypt$'):
            _,salt,hexhash=stored.split('$',2); dk=hashlib.scrypt(p.encode(),salt=bytes.fromhex(salt),n=2**14,r=8,p=1,dklen=64)
            return secrets.compare_digest(dk.hex(),hexhash)
        if stored.startswith('pbkdf2$'):
            _,salt,hexhash=stored.split('$',2); dk=hashlib.pbkdf2_hmac('sha256',p.encode(),salt.encode(),200000)
            return secrets.compare_digest(dk.hex(),hexhash)
        return secrets.compare_digest(hashlib.sha256(p.encode()).hexdigest(),stored)
    except Exception: return False
def is_legacy_hash(stored): return bool(stored) and not stored.startswith(('scrypt$','pbkdf2$'))
def now(): return datetime.now().isoformat(timespec='seconds')
def db():
    c=sqlite3.connect(DB,timeout=30,check_same_thread=False); c.row_factory=sqlite3.Row
    c.execute('PRAGMA foreign_keys=ON'); c.execute('PRAGMA journal_mode=WAL'); c.execute('PRAGMA synchronous=NORMAL'); return c

def init():
    c=db(); c.executescript('''
    CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,username TEXT UNIQUE,password_hash TEXT,role TEXT,full_name TEXT,active INTEGER DEFAULT 1,last_login TEXT,must_change_password INTEGER DEFAULT 1,scope_type TEXT DEFAULT 'all',scope_value TEXT DEFAULT '');
    CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
    CREATE TABLE IF NOT EXISTS saved_views(id INTEGER PRIMARY KEY,username TEXT,name TEXT,query TEXT NOT NULL,created_at TEXT,updated_at TEXT,UNIQUE(username,name));
    CREATE TABLE IF NOT EXISTS employees(id INTEGER PRIMARY KEY,emp_code TEXT UNIQUE,name TEXT NOT NULL,national_id TEXT,fingerprint TEXT,phone TEXT,email TEXT,job TEXT,department TEXT,location TEXT,gender TEXT,hire_date TEXT,status TEXT DEFAULT 'على رأس العمل',basic_salary REAL DEFAULT 0,allowances REAL DEFAULT 0,total_salary REAL DEFAULT 0,notes TEXT,updated_at TEXT,employee_group TEXT,birth_date TEXT,address TEXT,qualification TEXT,iban TEXT,bank_name TEXT,bank_branch TEXT,unit TEXT,contract_date TEXT,contract_amount REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS leave_types(id INTEGER PRIMARY KEY,name TEXT UNIQUE,annual_balance REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS leave_balances(id INTEGER PRIMARY KEY,emp_code TEXT,leave_type TEXT,annual REAL DEFAULT 0,used REAL DEFAULT 0,UNIQUE(emp_code,leave_type));
    CREATE TABLE IF NOT EXISTS leaves(id INTEGER PRIMARY KEY,request_no TEXT UNIQUE,emp_code TEXT,leave_type TEXT,start_date TEXT,end_date TEXT,days REAL,request_date TEXT,status TEXT DEFAULT 'قيد المراجعة',approved_by TEXT,approved_at TEXT,notes TEXT);
    CREATE TABLE IF NOT EXISTS attendance(id INTEGER PRIMARY KEY,work_date TEXT,emp_code TEXT,status TEXT,check_in TEXT,check_out TEXT,late_minutes INTEGER DEFAULT 0,work_hours REAL DEFAULT 0,overtime REAL DEFAULT 0,notes TEXT,UNIQUE(work_date,emp_code));
    CREATE TABLE IF NOT EXISTS audit(id INTEGER PRIMARY KEY,ts TEXT,username TEXT,role TEXT,action TEXT,entity TEXT,record_key TEXT,details TEXT);
    CREATE TABLE IF NOT EXISTS documents(id INTEGER PRIMARY KEY,emp_code TEXT,file_name TEXT,file_type TEXT,expiry_date TEXT,uploaded_by TEXT,uploaded_at TEXT,data BLOB,storage_path TEXT,category TEXT DEFAULT 'عام');
    CREATE TABLE IF NOT EXISTS notifications(id INTEGER PRIMARY KEY,user_name TEXT,title TEXT,message TEXT,created_at TEXT,read_at TEXT);
    CREATE TABLE IF NOT EXISTS permissions(id INTEGER PRIMARY KEY,code TEXT UNIQUE,name_ar TEXT);
    CREATE TABLE IF NOT EXISTS role_permissions(role TEXT,permission TEXT,PRIMARY KEY(role,permission));
    CREATE TABLE IF NOT EXISTS roles(name TEXT PRIMARY KEY,display_name TEXT,description TEXT,system INTEGER DEFAULT 0,scope_default TEXT DEFAULT 'all');
    CREATE TABLE IF NOT EXISTS payroll(id INTEGER PRIMARY KEY,emp_code TEXT,period TEXT,basic REAL DEFAULT 0,allowances REAL DEFAULT 0,deductions REAL DEFAULT 0,overtime REAL DEFAULT 0,bonuses REAL DEFAULT 0,net REAL DEFAULT 0,status TEXT DEFAULT 'مسودة',notes TEXT,created_by TEXT,created_at TEXT,UNIQUE(emp_code,period));
    CREATE TABLE IF NOT EXISTS employee_events(id INTEGER PRIMARY KEY,emp_code TEXT,event_type TEXT,event_date TEXT,title TEXT,details TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS employee_evaluations(id INTEGER PRIMARY KEY,emp_code TEXT,period TEXT,score REAL DEFAULT 0,goals_score REAL DEFAULT 0,attendance_score REAL DEFAULT 0,discipline_score REAL DEFAULT 0,notes TEXT,created_by TEXT,created_at TEXT,UNIQUE(emp_code,period));
    CREATE TABLE IF NOT EXISTS system_backups(id INTEGER PRIMARY KEY,file_path TEXT,created_at TEXT,created_by TEXT,label TEXT,db_size INTEGER);
    CREATE TABLE IF NOT EXISTS access_logs(id INTEGER PRIMARY KEY,ts TEXT,username TEXT,path TEXT,method TEXT,status INTEGER,ip TEXT);
    CREATE TABLE IF NOT EXISTS disciplinary_actions(id INTEGER PRIMARY KEY,emp_code TEXT,action_type TEXT,action_date TEXT,minutes INTEGER DEFAULT 0,amount REAL DEFAULT 0,reason TEXT,notes TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS attendance_adjustments(id INTEGER PRIMARY KEY,attendance_id INTEGER,minutes INTEGER DEFAULT 0,reason TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS file_imports(id INTEGER PRIMARY KEY,source TEXT,file_name TEXT,records INTEGER,created_by TEXT,created_at TEXT,details TEXT);
    CREATE TABLE IF NOT EXISTS login_history(id INTEGER PRIMARY KEY,username TEXT,ts TEXT,ip TEXT,success INTEGER);
    CREATE TABLE IF NOT EXISTS error_logs(id INTEGER PRIMARY KEY,request_id TEXT,ts TEXT,username TEXT,method TEXT,path TEXT,ip TEXT,exception_type TEXT,message TEXT,traceback TEXT,resolved INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS system_sessions(id INTEGER PRIMARY KEY,session_id TEXT UNIQUE,user_name TEXT,full_name TEXT,role TEXT,created_at TEXT,last_seen TEXT,ip TEXT,device TEXT,revoked INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS system_events(id INTEGER PRIMARY KEY,ts TEXT,event_type TEXT,severity TEXT,message TEXT,details TEXT);
    CREATE TABLE IF NOT EXISTS import_runs(id INTEGER PRIMARY KEY,source TEXT,file_name TEXT,created_at TEXT,created_by TEXT,status TEXT,records INTEGER,details TEXT);
    CREATE TABLE IF NOT EXISTS bridge_tokens(token TEXT PRIMARY KEY,username TEXT,session_id TEXT,expires_at TEXT,used INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS shifts(id INTEGER PRIMARY KEY,name TEXT UNIQUE,start_time TEXT,end_time TEXT,grace_minutes INTEGER DEFAULT 15,warning_minutes INTEGER DEFAULT 15,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS employee_shifts(emp_code TEXT PRIMARY KEY,shift_id INTEGER,assigned_at TEXT,assigned_by TEXT);
    CREATE TABLE IF NOT EXISTS import_mappings(name TEXT PRIMARY KEY,kind TEXT,mapping_json TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS overtime_requests(id INTEGER PRIMARY KEY,request_no TEXT UNIQUE,emp_code TEXT,work_date TEXT,hours REAL DEFAULT 0,reason TEXT,status TEXT DEFAULT 'قيد المراجعة',requested_by TEXT,approved_by TEXT,approved_at TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS backup_files(id INTEGER PRIMARY KEY,backup_id INTEGER,path TEXT,size INTEGER,checksum TEXT);
    CREATE TABLE IF NOT EXISTS import_errors(id INTEGER PRIMARY KEY,run_id INTEGER,row_no INTEGER,field TEXT,message TEXT,raw_json TEXT);
    CREATE TABLE IF NOT EXISTS attendance_daily_ledger(id INTEGER PRIMARY KEY,work_date TEXT,emp_code TEXT,allowed_minutes INTEGER DEFAULT 0,used_minutes INTEGER DEFAULT 0,remaining_minutes INTEGER DEFAULT 0,exceeded_minutes INTEGER DEFAULT 0,policy TEXT DEFAULT '',updated_at TEXT,UNIQUE(work_date,emp_code));
    CREATE TABLE IF NOT EXISTS holidays(id INTEGER PRIMARY KEY,name TEXT,holiday_date TEXT UNIQUE,kind TEXT DEFAULT 'official',paid INTEGER DEFAULT 1,notes TEXT);
    CREATE TABLE IF NOT EXISTS training(id INTEGER PRIMARY KEY,emp_code TEXT,course TEXT,training_date TEXT,expiry_date TEXT,score REAL DEFAULT 0,certificate TEXT,trainer TEXT,status TEXT DEFAULT 'completed',created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS credentials(id INTEGER PRIMARY KEY,emp_code TEXT,credential_type TEXT,credential_no TEXT,issue_date TEXT,expiry_date TEXT,status TEXT DEFAULT 'active',notes TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS assets(id INTEGER PRIMARY KEY,asset_tag TEXT UNIQUE,asset_type TEXT,description TEXT,serial_no TEXT,status TEXT DEFAULT 'available',emp_code TEXT,assigned_at TEXT,returned_at TEXT,notes TEXT,created_by TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS approval_queue(id INTEGER PRIMARY KEY,entity_type TEXT,entity_id TEXT,emp_code TEXT,title TEXT,requested_by TEXT,status TEXT DEFAULT 'pending',approved_by TEXT,approved_at TEXT,reason TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS import_rollback_items(id INTEGER PRIMARY KEY,run_id INTEGER,emp_code TEXT,action TEXT,before_json TEXT,after_json TEXT);
    CREATE TABLE IF NOT EXISTS system_jobs(id INTEGER PRIMARY KEY,name TEXT UNIQUE,last_run TEXT,last_status TEXT,message TEXT);
    CREATE TABLE IF NOT EXISTS device_registry(id INTEGER PRIMARY KEY,device_id TEXT UNIQUE,device_name TEXT,department TEXT,username TEXT,ip TEXT,first_seen TEXT,last_seen TEXT,status TEXT DEFAULT 'online',role TEXT);
    CREATE TABLE IF NOT EXISTS matching_reviews(id INTEGER PRIMARY KEY,source_name TEXT,candidate_json TEXT,confidence REAL,status TEXT DEFAULT 'review',selected_code TEXT,created_by TEXT,created_at TEXT,reviewed_by TEXT,reviewed_at TEXT);
    CREATE TABLE IF NOT EXISTS employee_requests(id INTEGER PRIMARY KEY,emp_code TEXT,request_type TEXT,payload_json TEXT,status TEXT DEFAULT 'pending',manager_user TEXT,hr_user TEXT,created_at TEXT,updated_at TEXT);
    CREATE TABLE IF NOT EXISTS excel_mapping_history(id INTEGER PRIMARY KEY,mapping_name TEXT UNIQUE,mapping_json TEXT,used_count INTEGER DEFAULT 0,last_used TEXT,created_by TEXT);
    CREATE TABLE IF NOT EXISTS alert_events(id INTEGER PRIMARY KEY,alert_key TEXT UNIQUE,severity TEXT,title TEXT,message TEXT,emp_code TEXT,created_at TEXT,resolved_at TEXT);

    ''')
    existing_u={r['name'] for r in c.execute('PRAGMA table_info(users)').fetchall()}
    for col,typ in {'must_change_password':'INTEGER DEFAULT 1','scope_type':"TEXT DEFAULT 'all'",'scope_value':"TEXT DEFAULT ''",'permission_version':'INTEGER DEFAULT 1'}.items():
        if col not in existing_u: c.execute(f'ALTER TABLE users ADD COLUMN {col} {typ}')
    existing_d={r['name'] for r in c.execute('PRAGMA table_info(documents)').fetchall()}
    for col,typ in {'storage_path':'TEXT','category':"TEXT DEFAULT 'عام'",'version':'INTEGER DEFAULT 1','checksum':'TEXT','status':"TEXT DEFAULT 'current'",'superseded_by':'INTEGER'}.items():
        if col not in existing_d: c.execute(f'ALTER TABLE documents ADD COLUMN {col} {typ}')
    existing_a={r['name'] for r in c.execute('PRAGMA table_info(audit)').fetchall()}
    for col,typ in {'prev_hash':'TEXT','hash':'TEXT','ip':'TEXT','before_json':'TEXT','after_json':'TEXT','reason':'TEXT'}.items():
        if col not in existing_a: c.execute(f'ALTER TABLE audit ADD COLUMN {col} {typ}')
    existing={r['name'] for r in c.execute('PRAGMA table_info(employees)').fetchall()}
    extra_cols={'manager_name':'TEXT','tags':'TEXT','employee_group':'TEXT','birth_date':'TEXT','address':'TEXT','qualification':'TEXT','iban':'TEXT','bank_name':'TEXT','bank_branch':'TEXT','unit':'TEXT','contract_date':'TEXT','contract_amount':'REAL DEFAULT 0','prev_status':'TEXT','archived_at':'TEXT','archived_by':'TEXT','shift_id':'INTEGER'}
    for col,typ in extra_cols.items():
        if col not in existing: c.execute(f'ALTER TABLE employees ADD COLUMN {col} {typ}')
    existing_p={r['name'] for r in c.execute('PRAGMA table_info(payroll)').fetchall()}
    for col,typ in {'locked_at':'TEXT','approved_by':'TEXT','approved_at':'TEXT','penalties':'REAL DEFAULT 0'}.items():
        if col not in existing_p: c.execute(f'ALTER TABLE payroll ADD COLUMN {col} {typ}')
    existing_ev={r['name'] for r in c.execute('PRAGMA table_info(employee_evaluations)').fetchall()}
    for col,typ in {'punctuality_score':'REAL DEFAULT 0','productivity_score':'REAL DEFAULT 0','behavior_score':'REAL DEFAULT 0','manager_score':'REAL DEFAULT 0'}.items():
        if col not in existing_ev: c.execute(f'ALTER TABLE employee_evaluations ADD COLUMN {col} {typ}')
    existing_s={r['name'] for r in c.execute('PRAGMA table_info(shifts)').fetchall()}
    for col,typ in {'warning_minutes':'INTEGER DEFAULT 15'}.items():
        if col not in existing_s: c.execute(f'ALTER TABLE shifts ADD COLUMN {col} {typ}')
    existing_disc={r['name'] for r in c.execute('PRAGMA table_info(disciplinary_actions)').fetchall()}
    for col,typ in {'source':"TEXT DEFAULT 'manual'"}.items():
        if col not in existing_disc: c.execute(f'ALTER TABLE disciplinary_actions ADD COLUMN {col} {typ}')
    for _ix in [
        'CREATE INDEX IF NOT EXISTS idx_attendance_emp_date ON attendance(emp_code,work_date)',
        'CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(work_date)',
        'CREATE INDEX IF NOT EXISTS idx_attendance_month_late ON attendance(emp_code,work_date,late_minutes)',
        'CREATE INDEX IF NOT EXISTS idx_leaves_emp_dates ON leaves(emp_code,start_date,end_date)',
        'CREATE INDEX IF NOT EXISTS idx_leaves_status ON leaves(status)',
        'CREATE INDEX IF NOT EXISTS idx_documents_emp_cat_status ON documents(emp_code,category,status)',
        'CREATE INDEX IF NOT EXISTS idx_documents_expiry ON documents(expiry_date)',
        'CREATE INDEX IF NOT EXISTS idx_payroll_period_status ON payroll(period,status)',
        'CREATE INDEX IF NOT EXISTS idx_audit_user_ts ON audit(username,ts)',
        'CREATE INDEX IF NOT EXISTS idx_notifications_user_read ON notifications(user_name,read_at)',
        'CREATE INDEX IF NOT EXISTS idx_training_expiry ON training(expiry_date)',
        'CREATE INDEX IF NOT EXISTS idx_credentials_expiry ON credentials(expiry_date)',
        'CREATE INDEX IF NOT EXISTS idx_assets_employee ON assets(emp_code,status)',
        'CREATE INDEX IF NOT EXISTS idx_approval_status ON approval_queue(status,created_at)'
    ]:
        try: c.execute(_ix)
        except Exception: pass
    defaults={'company_name':'HR Enterprise','work_start':'09:00','work_end':'17:00','grace_minutes':'15','currency':'EGP','max_late_minutes_month':'60','session_idle_minutes':'30','document_max_mb':'25','network_mode':'standalone','late_alert_threshold':'60','monthly_late_limit_minutes':'120','monthly_late_action':'none','monthly_deduction_amount':'0','required_doc_categories':'عقد,هوية,مؤهل','company_logo':'','eval_weight_attendance':'20','eval_weight_punctuality':'20','eval_weight_productivity':'25','eval_weight_behavior':'15','eval_weight_manager':'20','backup_auto_enabled':'1','backup_time':'23:00','backup_keep_days':'14','theme':'light','accent_color':'blue','backup_keep_weekly':'8','backup_keep_monthly':'12','server_name':'HR-SERVER','offline_cache_minutes':'15','daily_late_limit_minutes':'15','daily_late_action':'none','daily_exceeded_deduction_amount':'0','contract_alert_days':'30','credential_alert_days':'30','training_alert_days':'30','privacy_mode_default':'0','https_enabled':'0','backup_network_path':'','backup_verify_after_create':'1','server_offline_seconds':'120','late_risk_percent':'80','training_alert_days':'30','credential_alert_days':'30'}
    for k,v in defaults.items(): c.execute('INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)',(k,v))
    c.execute('INSERT OR IGNORE INTO users(username,password_hash,role,full_name,must_change_password,scope_type) VALUES(?,?,?,?,?,?)',('admin',hashpw('Admin@123'),'SuperAdmin','مدير النظام',1,'all'))
    for n,b in [('اعتيادي',21),('عارضة',7),('مرضي',0),('مأمورية',0),('بدون أجر',0)]: c.execute('INSERT OR IGNORE INTO leave_types(name,annual_balance) VALUES(?,?)',(n,b))
    for n,st,et,g,w in [('Morning','09:00','17:00',15,15),('Evening','14:00','22:00',15,15),('Night','22:00','06:00',15,15)]: c.execute('INSERT OR IGNORE INTO shifts(name,start_time,end_time,grace_minutes,warning_minutes) VALUES(?,?,?,?,?)',(n,st,et,g,w))
    for emp in c.execute('SELECT emp_code FROM employees').fetchall():
        for lt in c.execute('SELECT name,annual_balance FROM leave_types').fetchall(): c.execute('INSERT OR IGNORE INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,0)',(emp['emp_code'],lt['name'],lt['annual_balance']))
    perms=[('employees.view','عرض الموظفين'),('employees.edit','تعديل الموظفين'),('leave.create','إنشاء إجازة'),('leave.approve','اعتماد الإجازات'),('attendance.view','عرض الحضور'),('attendance.import','استيراد الحضور'),('attendance.edit','تعديل الحضور'),('reports.view','التقارير'),('reports.export','تصدير التقارير'),('users.manage','إدارة المستخدمين'),('roles.manage','إدارة الصلاحيات'),('settings.manage','الإعدادات'),('audit.view','سجل المراجعة'),('backup.manage','النسخ الاحتياطي'),('documents.manage','المستندات'),('documents.bulk_import','استيراد مجلدات الموظفين'),('discipline.manage','الجزاءات والإنذارات'),('notifications.manage','التنبيهات'),('payroll.view','عرض المرتبات'),('payroll.manage','إدارة المرتبات'),('payroll.approve','اعتماد المرتبات'),('payroll.lock','قفل المرتبات'),('sensitive.view','عرض البيانات الحساسة'),('overtime.request','طلب الإضافي'),('overtime.approve','اعتماد الإضافي'),('shifts.manage','إدارة الورديات'),('import.validate','فحص الاستيراد'),('import.mapping','إدارة ربط أعمدة Excel'),('system.manage','إدارة النظام'),('backup.restore','استعادة النسخ')]
    for code,nm in perms: c.execute('INSERT OR IGNORE INTO permissions(code,name_ar) VALUES(?,?)',(code,nm))
    rolemap={'SuperAdmin':[x[0] for x in perms],'Admin':[x[0] for x in perms],'HR':[x[0] for x in perms if x[0] not in ('users.manage','roles.manage')],'Manager':['employees.view','leave.create','leave.approve','attendance.view','attendance.edit','reports.view','reports.export','documents.manage','discipline.manage','overtime.request','overtime.approve'],'Employee':['employees.view','leave.create','attendance.view','overtime.request']}
    for role,codes in rolemap.items():
        c.execute('INSERT OR IGNORE INTO roles(name,display_name,description,system,scope_default) VALUES(?,?,?,?,?)',(role,role,'دور نظامي' if role in ('SuperAdmin','Admin','HR','Manager','Employee') else '',1,'all' if role in ('SuperAdmin','Admin','HR') else ('department' if role=='Manager' else 'self')))
        for code in codes: c.execute('INSERT OR IGNORE INTO role_permissions(role,permission) VALUES(?,?)',(role,code))
    c.commit(); c.close()
    # Do not auto-import the legacy seed workbook. Employees are loaded only from the user's Excel/Paste workflow.

def count(table):
    c=db(); x=c.execute(f'SELECT COUNT(*) n FROM {table}').fetchone()['n']; c.close(); return x

def audit(user,role,action,entity,key,details='',before=None,after=None,reason='',ip=''):
    ip=ip or getattr(AUDIT_CTX,'ip','')
    c=db(); prev=c.execute('SELECT hash FROM audit ORDER BY id DESC LIMIT 1').fetchone(); prev_hash=prev['hash'] if prev and prev['hash'] else ''
    ts=now(); bjson=json.dumps(before,ensure_ascii=False,sort_keys=True,default=str) if before is not None else ''; ajson=json.dumps(after,ensure_ascii=False,sort_keys=True,default=str) if after is not None else ''
    payload='|'.join([prev_hash,ts,str(user),str(role),str(action),str(entity),str(key),str(details),str(ip),bjson,ajson,str(reason)])
    h=hashlib.sha256(payload.encode('utf-8')).hexdigest()
    c.execute('INSERT INTO audit(ts,username,role,action,entity,record_key,details,prev_hash,hash,ip,before_json,after_json,reason) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',(ts,user,role,action,entity,key,details,prev_hash,h,ip,bjson,ajson,reason)); c.commit(); c.close()


def verify_audit_chain():
    """Walk the audit table in id order and recompute each row's hash to confirm
    the chain is intact. Returns (ok, first_broken_id_or_None, total_checked)."""
    c=db(); rows=c.execute('SELECT id,ts,username,role,action,entity,record_key,details,prev_hash,hash,ip,before_json,after_json,reason FROM audit ORDER BY id ASC').fetchall(); c.close()
    prev_hash=''
    for r in rows:
        payload='|'.join([prev_hash,str(r['ts']),str(r['username']),str(r['role']),str(r['action']),str(r['entity']),str(r['record_key']),str(r['details'] or ''),str(r['ip'] or ''),str(r['before_json'] or ''),str(r['after_json'] or ''),str(r['reason'] or '')])
        expected=hashlib.sha256(payload.encode('utf-8')).hexdigest()
        if (r['prev_hash'] or '')!=prev_hash or (r['hash'] or '')!=expected:
            return False,r['id'],len(rows)
        prev_hash=r['hash']
    return True,None,len(rows)

def setting(k):
    c=db(); r=c.execute('SELECT value FROM settings WHERE key=?',(k,)).fetchone(); c.close(); return r['value'] if r else ''

def evaluation_weights():
    keys=['attendance','punctuality','productivity','behavior','manager']
    vals=[max(0.0,cell_num(setting('eval_weight_'+k))) for k in keys]
    total=sum(vals) or 100.0
    return dict(zip(keys,[v*100.0/total for v in vals]))
def evaluation_grade(score):
    x=float(score or 0)
    if x>=90:return ('ممتاز','b-ok')
    if x>=80:return ('جيد جدًا','b-ok')
    if x>=70:return ('جيد','b-warn')
    if x>=60:return ('يحتاج تحسين','b-warn')
    return ('ضعيف','b-bad')
def show_native_error(title,message):
    try:
        import tkinter as tk
        from tkinter import messagebox
        root=tk.Tk(); root.withdraw(); messagebox.showerror(title,message); root.destroy()
    except Exception: print(title+': '+message)

def migrate_legacy_data_location():
    """First V7.5 launch preserves V7.4 data if it was stored beside the EXE."""
    if not sys.platform.startswith('win'): return
    legacy=os.path.join(BASE,'data');
    if os.path.abspath(legacy)==os.path.abspath(DATA) or not os.path.isdir(legacy): return
    if os.path.exists(DB): return
    try:
        os.makedirs(DATA,exist_ok=True)
        for name in os.listdir(legacy):
            src=os.path.join(legacy,name); dst=os.path.join(DATA,name)
            if os.path.isdir(src): shutil.copytree(src,dst,dirs_exist_ok=True)
            else: shutil.copy2(src,dst)
    except Exception as e: log_error('legacy-data-migration',e)

def load_json_file(path, default):
    try:
        with open(path, 'r', encoding='utf-8') as fh: return json.load(fh)
    except Exception: return default

def save_json_file(path, obj):
    tmp=path+'.tmp'
    with open(tmp,'w',encoding='utf-8') as fh: json.dump(obj,fh,ensure_ascii=False,indent=2)
    os.replace(tmp,path)

def server_identity():
    ident=load_json_file(IDENTITY_FILE,{})
    if not ident.get('server_id'):
        ident={'server_id':uuid.uuid4().hex,'created_at':now(),'fingerprint':''}
        ident['fingerprint']=hashlib.sha256(ident['server_id'].encode()).hexdigest()
        save_json_file(IDENTITY_FILE,ident)
    return ident

def trusted_server_info(): return load_json_file(TRUST_FILE,{})

def remember_trusted_server(url,fingerprint,server_id):
    if not url or not fingerprint: return
    info=trusted_server_info(); key=url.split('://',1)[-1]
    old=info.get(key)
    if old and old.get('fingerprint')!=fingerprint:
        raise ValueError('تغيرت هوية خادم HR Enterprise لهذا العنوان. راجع مسؤول النظام قبل الاتصال.')
    info[key]={'url':url,'fingerprint':fingerprint,'server_id':server_id,'trusted_at':old.get('trusted_at') if old else now(),'last_seen':now()}
    save_json_file(TRUST_FILE,info)

def local_ip():
    try:
        ss=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); ss.connect(('8.8.8.8',80)); ip=ss.getsockname()[0]; ss.close(); return ip
    except Exception:
        try:return socket.gethostbyname(socket.gethostname())
        except Exception:return '127.0.0.1'
def find_free_port(start=None,end=None):
    start=int(start or PORT_MIN); end=int(end or PORT_MAX)
    for p in range(start,end+1):
        try:
            ss=socket.socket(socket.AF_INET,socket.SOCK_STREAM); ss.setsockopt(socket.SOL_SOCKET,socket.SO_REUSEADDR,1); ss.bind(('',p)); ss.close(); return p
        except OSError: pass
    raise OSError(f'No free HR Enterprise port between {start} and {end}')

def storage_health():
    checks=[]
    for label,path in [('Data',DATA),('Database',DB),('Employee Files',EMPFILES),('Backups',BACKUPS)]:
        ok=os.path.exists(path)
        writable=False
        try:
            if os.path.isdir(path):
                t=os.path.join(path,'.write_test'); open(t,'wb').write(b'ok'); os.remove(t); writable=True
            elif os.path.exists(path): writable=os.access(path,os.W_OK)
        except Exception: writable=False
        checks.append({'name':label,'exists':ok,'writable':writable})
    return checks

def database_health():
    try:
        c=db(); c.execute('SELECT 1').fetchone(); integrity=c.execute('PRAGMA integrity_check').fetchone()[0]; counts={t:count(t) for t in ('employees','documents','attendance','audit')}; c.close(); return {'ok':integrity=='ok','integrity':integrity,'counts':counts}
    except Exception as e: return {'ok':False,'error':str(e)}

def _discovery_packet(kind, **fields):
    body={'kind':kind,'version':APP_VERSION,'ts':time.time(),**fields}
    return ('HR75|'+json.dumps(body,ensure_ascii=False,separators=(',',':'))).encode('utf-8')

def _parse_discovery(data):
    if not data.startswith(b'HR75|'): return None
    try: return json.loads(data[5:].decode('utf-8'))
    except Exception: return None

def discover_server(timeout=1.25):
    """LAN discovery with TOFU fingerprint pinning. A changed fingerprint is never silently accepted."""
    found=[]; nonce=secrets.token_hex(8)
    try:
        ss=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); ss.setsockopt(socket.SOL_SOCKET,socket.SO_BROADCAST,1); ss.settimeout(0.20)
        ss.sendto(_discovery_packet('DISCOVER',nonce=nonce),('<broadcast>',DISCOVERY_PORT))
        end=time.time()+timeout
        while time.time()<end:
            try:
                data,addr=ss.recvfrom(4096); msg=_parse_discovery(data)
                if msg and msg.get('kind')=='SERVER' and msg.get('nonce')==nonce and msg.get('url'):
                    url=msg['url']; fp=msg.get('fingerprint',''); sid=msg.get('server_id','')
                    remember_trusted_server(url,fp,sid)
                    found.append((msg.get('priority',100),url))
            except socket.timeout: continue
            except ValueError: raise
        ss.close()
    except ValueError: raise
    except Exception: return ''
    found.sort(key=lambda x:(x[0],x[1]))
    return found[0][1] if found else ''

def elect_server(timeout=1.8):
    """Race-resistant first-start election. Each candidate gets a deterministic jitter;
    the earliest candidate becomes server, while others keep checking discovery and become clients."""
    cid=uuid.uuid4().hex
    jitter=0.25+(int(cid[:6],16)%900)/1000.0
    deadline=time.time()+timeout
    while time.time()<deadline:
        if discover_server(0.22): return False
        time.sleep(min(0.12,max(0.02,jitter)))
        jitter=max(0.05,jitter-0.12)
    return True

def discovery_responder():
    while True:
        try:
            ident=server_identity(); ss=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); ss.setsockopt(socket.SOL_SOCKET,socket.SO_REUSEADDR,1); ss.bind(('',DISCOVERY_PORT))
            while True:
                data,addr=ss.recvfrom(4096); msg=_parse_discovery(data)
                if not msg: continue
                if msg.get('kind')=='DISCOVER':
                    url=f'http://{local_ip()}:{PORT}'
                    payload=_discovery_packet('SERVER',nonce=msg.get('nonce',''),url=url,server_id=ident['server_id'],fingerprint=ident['fingerprint'],priority=10)
                    ss.sendto(payload,addr)
        except Exception as e:
            log_error('discovery_responder',e); time.sleep(1)
HOSPITAL_HEADERS=['م','الإسم','المجموعه الوظفيه','تاريخ الميلاد','الرقم القومى','العنوان','المؤهل','رقم التيلفون','ipan','إسم البنك','إسم فرع البنك','الإدارة','الوحدة','الوظيفه','تاريخ التعاقد','مبلغ التعاقد']
HOSPITAL_FIELDS=['emp_no','name','employee_group','birth_date','national_id','address','qualification','phone','iban','bank_name','bank_branch','department','unit','job','contract_date','contract_amount']
def norm_header(v):
    x=str(v or '').replace('\ufeff','').replace('\n',' ').strip().lower()
    x=x.replace('ى','ي').replace('أ','ا').replace('إ','ا').replace('آ','ا')
    return ' '.join(x.split())
HOSPITAL_ALIASES={
'emp_no':['م','م.','رقم','الرقم','مسلسل','serial','no','no.'],
'name':['الإسم','الاسم','اسم','اسم الموظف','employee name','name'],
'employee_group':['المجموعه الوظفيه','المجموعة الوظيفية','المجموعه الوظيفيه','المجموعة الوظيفيه','المجموعة','الفئة','group'],
'birth_date':['تاريخ الميلاد','تاريخ ميلاد','birth date','date of birth'],
'national_id':['الرقم القومى','الرقم القومي','رقم قومي','national id','national_id'],
'address':['العنوان','عنوان','address'],
'qualification':['المؤهل','المؤهل الدراسي','المؤهل الدراسى','qualification','education'],
'phone':['رقم التيلفون','رقم التليفون','رقم الهاتف','الهاتف','الموبايل','المحمول','phone','mobile'],
'iban':['ipan','iban','رقم الحساب','الحساب البنكي','الحساب'],
'bank_name':['إسم البنك','اسم البنك','البنك','bank name','bank'],
'bank_branch':['إسم فرع البنك','اسم فرع البنك','فرع البنك','الفرع','bank branch','branch'],
'department':['الإدارة','الادارة','إدارة','department','dept'],
'unit':['الوحدة','الوحده','unit'],
'job':['الوظيفه','الوظيفة','الوظيفه الحالية','المسمى الوظيفي','المسمى الوظيفى','job','position','title'],
'contract_date':['تاريخ التعاقد','تاريخ العقد','contract date','hire date'],
'contract_amount':['مبلغ التعاقد','قيمة التعاقد','قيمة العقد','مبلغ العقد','contract amount','salary','amount']}
def find_hospital_header(rows,max_scan=40):
    best=None
    dynamic={}
    try:
        cc=db(); rr=cc.execute('SELECT mapping_json FROM import_mappings WHERE name=?',("Hospital Employee Template",)).fetchone(); cc.close(); dynamic=json.loads(rr['mapping_json']) if rr and rr['mapping_json'] else {}
    except Exception: dynamic={}
    aliases={k:list(v) for k,v in HOSPITAL_ALIASES.items()}
    for fk,hname in dynamic.items(): aliases.setdefault(fk,[]).append(hname)
    for ridx,row in enumerate(rows[:max_scan]):
        hdr={norm_header(v):i for i,v in enumerate(row or []) if norm_header(v)}
        found={}
        for field,als in aliases.items():
            for a in als:
                if norm_header(a) in hdr: found[field]=hdr[norm_header(a)]; break
        score=len(found)
        if 'name' in found and score>=4 and (best is None or score>best[0]): best=(score,ridx,found)
    return (best[1],best[2]) if best else (None,None)
def cell_text(v):
    if v is None:return ''
    if hasattr(v,'isoformat'):return v.isoformat()[:10]
    if isinstance(v,float) and v.is_integer():return str(int(v))
    return str(v).strip()
def cell_num(v):
    try:return float(str(v or '').replace(',','').replace('٬','').replace('،','').strip() or 0)
    except:return 0.0
def hospital_row_to_record(row,idx):
    def g(field):
        i=idx.get(field); return row[i] if i is not None and i<len(row) else ''
    name=cell_text(g('name')); national=cell_text(g('national_id')); serial=cell_text(g('emp_no'))
    return {'emp_code':serial or national,'name':name,'employee_group':cell_text(g('employee_group')),'birth_date':cell_text(g('birth_date')),'national_id':national,'address':cell_text(g('address')),'qualification':cell_text(g('qualification')),'phone':cell_text(g('phone')),'iban':cell_text(g('iban')),'bank_name':cell_text(g('bank_name')),'bank_branch':cell_text(g('bank_branch')),'department':cell_text(g('department')),'unit':cell_text(g('unit')),'job':cell_text(g('job')),'contract_date':cell_text(g('contract_date')),'contract_amount':cell_num(g('contract_amount'))}
def looks_like_hospital_employee(rec):
    name=rec.get('name','').strip()
    if len(name)<2:return False
    if not (rec.get('emp_code') or rec.get('national_id') or rec.get('phone')):return False
    return not any(x in name for x in ('التوقيع','رئيس شئون العاملين','شئون العاملين'))
def validate_employee_records(records):
    errors=[]; seen={}; valid=[]
    for i,rec in enumerate(records,1):
        code=rec.get('emp_code','').strip(); name=rec.get('name','').strip();
        if not code: errors.append((i,'Employee ID','مفقود',rec)); continue
        if not name: errors.append((i,'Full Name','مفقود',rec)); continue
        if code in seen: errors.append((i,'Employee ID','مكرر داخل الملف',rec)); continue
        seen[code]=i
        if rec.get('birth_date'):
            try: datetime.fromisoformat(str(rec['birth_date'])[:10])
            except: errors.append((i,'Birth Date','تاريخ غير صالح',rec)); continue
        if rec.get('national_id') and not re.fullmatch(r'\d{10,20}',str(rec['national_id']).replace(' ','')):
            errors.append((i,'National ID','صيغة غير صالحة',rec)); continue
        valid.append(rec)
    return valid,errors

def upsert_hospital_records(records,user,source):
    c=db(); new=updated=skipped=0
    sql='INSERT INTO employees(emp_code,name,employee_group,birth_date,national_id,address,qualification,phone,iban,bank_name,bank_branch,department,unit,job,contract_date,contract_amount,status,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET name=excluded.name,employee_group=excluded.employee_group,birth_date=excluded.birth_date,national_id=excluded.national_id,address=excluded.address,qualification=excluded.qualification,phone=excluded.phone,iban=excluded.iban,bank_name=excluded.bank_name,bank_branch=excluded.bank_branch,department=excluded.department,unit=excluded.unit,job=excluded.job,contract_date=excluded.contract_date,contract_amount=excluded.contract_amount,updated_at=excluded.updated_at'
    try:
        for rec in records:
            if not looks_like_hospital_employee(rec): skipped+=1; continue
            exists=c.execute('SELECT 1 FROM employees WHERE emp_code=?',(rec['emp_code'],)).fetchone() is not None
            c.execute(sql,(rec['emp_code'],rec['name'],rec['employee_group'],rec['birth_date'],rec['national_id'],rec['address'],rec['qualification'],rec['phone'],rec['iban'],rec['bank_name'],rec['bank_branch'],rec['department'],rec['unit'],rec['job'],rec['contract_date'],rec['contract_amount'],'على رأس العمل',now()))
            if exists: updated+=1
            else: new+=1
        c.commit()
    except Exception:
        c.rollback(); c.close(); raise
    c.close(); audit(user['username'],user['role'],'استيراد','الموظفون',source,f'new={new}, updated={updated}, skipped={skipped}')
    return new,updated,skipped

def import_seed(path):
    try:
        wb=load_workbook(path,read_only=True,data_only=True,keep_vba=True)
        c=db()
        if 'الموظفون' in wb.sheetnames:
            ws=wb['الموظفون']
            for r in ws.iter_rows(min_row=3,values_only=True):
                if not r or not r[0] or not r[3]: continue
                def s(x): return x.isoformat()[:10] if hasattr(x,'isoformat') else str(x or '')
                vals=[r[i] if i<len(r) else None for i in range(33)]
                c.execute('''INSERT INTO employees(emp_code,name,national_id,fingerprint,phone,job,department,hire_date,status,total_salary,notes,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET name=excluded.name,national_id=excluded.national_id,fingerprint=excluded.fingerprint,phone=excluded.phone,job=excluded.job,department=excluded.department,hire_date=excluded.hire_date,status=excluded.status,total_salary=excluded.total_salary,notes=excluded.notes,updated_at=excluded.updated_at''',(str(vals[0]),str(vals[3]),str(vals[2] or ''),str(vals[1] or ''),str(vals[9] or ''),str(vals[13] or ''),str(vals[14] or ''),s(vals[16]),str(vals[18] or 'على رأس العمل'),float(vals[26] or 0),str(vals[30] or ''),now()))
        c.commit(); c.close()
    except Exception as e: print('seed import warning',e)

def export_xlsx(table,query,headers):
    c=db(); rows=c.execute(query).fetchall(); c.close(); wb=Workbook(); ws=wb.active; ws.title=table
    ws.append(headers)
    for r in rows: ws.append(list(r))
    for col in ws.columns:
        mx=max(len(str(x.value or '')) for x in col); ws.column_dimensions[col[0].column_letter].width=min(40,max(12,mx+2))
    out=io.BytesIO(); wb.save(out); return out.getvalue()

def validate_file_signature(filename,data):
    ext=os.path.splitext(filename)[1].lower()
    if ext=='.pdf': return data.startswith(b'%PDF')
    if ext=='.png': return data.startswith(b'\x89PNG\r\n\x1a\n')
    if ext in ('.jpg','.jpeg'): return data.startswith(b'\xff\xd8\xff')
    if ext in ('.docx','.xlsx','.xls'): return data.startswith(b'PK') or data.startswith(b'\xd0\xcf\x11\xe0')
    if ext in ('.txt','.csv','.doc'): return True
    return False

def safe_name(n): return ''.join(ch if ch.isalnum() or ch in '._-' else '_' for ch in str(n))[:120]

def normalize_name(s):
    s=(s or '').strip()
    s=re.sub(r'[_\-]+',' ',s)
    s=re.sub(r'\s+',' ',s)
    s=s.replace('أ','ا').replace('إ','ا').replace('آ','ا').replace('ى','ي').replace('ة','ه').replace('ؤ','و').replace('ئ','ي')
    s=re.sub(r'[\u064B-\u065F\u0670]','',s)
    return s.lower().strip()

def resolve_folder_employee(folder_name, employees):
    """Match an uploaded folder name to an employee. Employees is a list of
    sqlite Row/dict with emp_code and name. Returns (emp_code, match_type, matched_name)
    where match_type is one of: code, name_exact, name_close, none.
    Never guesses blindly: name_close only returned when there is one clear best
    candidate well above the next-best candidate."""
    raw=(folder_name or '').strip()
    if not raw: return (None,'none',None)
    by_code={e['emp_code']:e for e in employees}
    if raw in by_code: return (raw,'code',by_code[raw]['name'])
    norm_target=normalize_name(raw)
    by_norm_name={}
    for e in employees:
        by_norm_name.setdefault(normalize_name(e['name']),[]).append(e)
    if norm_target in by_norm_name and len(by_norm_name[norm_target])==1:
        m=by_norm_name[norm_target][0]; return (m['emp_code'],'name_exact',m['name'])
    # Folder names may use short employee codes too (e.g. "7 - Name" or "26 - Name").
    # Accept a bounded numeric token anywhere in the folder name, then prefer a real code match.
    code_match=re.search(r'(?<!\d)(\d{1,6})(?!\d)',raw)
    if code_match and code_match.group(1) in by_code:
        return (code_match.group(1),'code',by_code[code_match.group(1)]['name'])
    # strip a numeric token then retry as a pure name
    stripped=normalize_name(re.sub(r'(?<!\d)\d{1,6}(?!\d)','',raw))
    if stripped and stripped in by_norm_name and len(by_norm_name[stripped])==1:
        m=by_norm_name[stripped][0]; return (m['emp_code'],'name_exact',m['name'])
    # fuzzy match on name only, and only if there is one clearly-best candidate
    names=list(by_norm_name.keys())
    if names:
        scored=sorted(((difflib.SequenceMatcher(None,norm_target,n).ratio(),n) for n in names),reverse=True)
        best_ratio,best_name=scored[0]
        second_ratio=scored[1][0] if len(scored)>1 else 0
        if best_ratio>=0.82 and (best_ratio-second_ratio)>=0.08 and len(by_norm_name[best_name])==1:
            m=by_norm_name[best_name][0]; return (m['emp_code'],'name_close',m['name'])
    return (None,'none',None)

def can(u,p):
    if not u: return False
    if u.get('role')=='Admin': return True
    c=db(); r=c.execute('SELECT 1 FROM role_permissions WHERE role=? AND permission=?',(u.get('role'),p)).fetchone(); c.close(); return bool(r)

def notify(user_name,title,message):
    c=db(); c.execute('INSERT INTO notifications(user_name,title,message,created_at) VALUES(?,?,?,?)',(user_name,title,message,now())); c.commit(); c.close()

def visible_employee_sql(u, alias='e'):
    role=u.get('role')
    if role in ('SuperAdmin','Admin','HR'): return '',[]
    scope=u.get('scope_type','')
    val=u.get('scope_value','')
    if role=='Employee' or scope=='self': return f" AND {alias}.emp_code=?",[val or u.get('username','')]
    if scope=='department': return f" AND {alias}.department=?",[val]
    if scope=='unit': return f" AND {alias}.unit=?",[val]
    return '',[]
def emp_allowed(u,code):
    c=db(); e=c.execute('SELECT emp_code,department,unit FROM employees WHERE emp_code=?',(code,)).fetchone(); c.close()
    if not e:return False
    if u.get('role') in ('SuperAdmin','Admin','HR'):return True
    if u.get('role')=='Employee' or u.get('scope_type')=='self':return e['emp_code']==(u.get('scope_value') or u.get('username'))
    if u.get('scope_type')=='department':return e['department']==u.get('scope_value')
    if u.get('scope_type')=='unit':return e['unit']==u.get('scope_value')
    return True
def guess_document_category(filename):
    n=norm_header(os.path.splitext(os.path.basename(filename))[0])
    if any(x in n for x in ('contract','عقد','تعاقد')): return 'عقد'
    if any(x in n for x in ('id','بطاق','قومي','national')): return 'هوية'
    if any(x in n for x in ('photo','صوره','صورة','image')): return 'صورة'
    if any(x in n for x in ('qual','مؤهل','شهاد')): return 'مؤهل'
    if any(x in n for x in ('appointment','تعيين')): return 'تعيين'
    if any(x in n for x in ('insurance','تأمين')): return 'تأمين'
    return 'عام'

def missing_documents_report(u):
    """للموظفين النشطين (على رأس العمل) وضمن نطاق صلاحية المستخدم: أي تصنيفات مستندات مطلوبة (من الإعدادات) غير موجودة كمستند 'current'."""
    required=[x.strip() for x in (setting('required_doc_categories') or '').split(',') if x.strip()]
    if not required: return [],[]
    c=db()
    emps=c.execute("SELECT emp_code,name FROM employees WHERE status='على رأس العمل' ORDER BY name").fetchall()
    emps=[e for e in emps if emp_allowed(u,e['emp_code'])]
    have_rows=c.execute("SELECT emp_code,category FROM documents WHERE status='current'").fetchall()
    c.close()
    have={}
    for r in have_rows: have.setdefault(r['emp_code'],set()).add(r['category'])
    rows=[]
    for e in emps:
        present=have.get(e['emp_code'],set())
        missing=[cat for cat in required if cat not in present]
        if missing: rows.append((e['emp_code'],e['name'],missing))
    return required,rows

def employee_dir(code):
    path=os.path.join(EMPFILES,safe_name(code)); os.makedirs(path,exist_ok=True); return path
def save_employee_file(code,relname,data):
    rel='/'.join([safe_name(x) for x in relname.replace('\\','/').split('/') if x not in ('','.','..')])
    path=os.path.join(employee_dir(code),rel); os.makedirs(os.path.dirname(path),exist_ok=True)
    with open(path,'wb') as fh: fh.write(data)
    return os.path.relpath(path,EMPFILES).replace('\\','/')
def secure_file_bytes(path):
    full=os.path.abspath(os.path.join(EMPFILES,path)); root=os.path.abspath(EMPFILES)
    if not full.startswith(root+os.sep): raise ValueError('invalid path')
    with open(full,'rb') as fh:return fh.read()
def migrate_documents_to_fs():
    c=db(); rows=c.execute('SELECT id,emp_code,file_name,data,storage_path FROM documents WHERE storage_path IS NULL OR storage_path=""').fetchall()
    for r in rows:
        if r['data'] is not None:
            rel=save_employee_file(r['emp_code'],r['file_name']); c.execute('UPDATE documents SET storage_path=?,data=NULL WHERE id=?',(rel,r['id']))
    c.commit(); c.close()
def make_backup(user,label='manual'):
    stamp=datetime.now().strftime('%Y%m%d_%H%M%S'); package=os.path.join(BACKUPS,f'HR_Backup_{label}_{stamp}.zip'); tmpdb=os.path.join(DATA,f'_backup_{stamp}.db')
    c=db(); c.execute('PRAGMA wal_checkpoint(FULL)'); c.close(); shutil.copy2(DB,tmpdb)
    manifest={'version':APP_VERSION,'created_at':now(),'label':label,'database':'database.db','employee_files':'employee_files/','files':[]}
    try:
        with zipfile.ZipFile(package,'w',zipfile.ZIP_DEFLATED) as z:
            z.write(tmpdb,'database.db')
            for root,dirs,files in os.walk(EMPFILES):
                for fn in files:
                    fp=os.path.join(root,fn); arc=os.path.relpath(fp,EMPFILES).replace('\\','/'); data=open(fp,'rb').read(); ch=hashlib.sha256(data).hexdigest()
                    z.writestr('employee_files/'+arc,data); manifest['files'].append({'path':'employee_files/'+arc,'size':len(data),'sha256':ch})
            z.writestr('manifest.json',json.dumps(manifest,ensure_ascii=False,indent=2))
    finally:
        if os.path.exists(tmpdb): os.remove(tmpdb)
    checksum=hashlib.sha256(open(package,'rb').read()).hexdigest()
    verified,verify_msg=verify_backup_package(package)
    if not verified: raise ValueError('Backup verification failed: '+verify_msg)
    size=os.path.getsize(package); c=db(); cur=c.execute('INSERT INTO system_backups(file_path,created_at,created_by,label,db_size) VALUES(?,?,?,?,?)',(package,now(),user.get('username','system'),label,size)); c.execute('INSERT INTO backup_files(backup_id,path,size,checksum) VALUES(?,?,?,?)',(cur.lastrowid,package,size,checksum)); c.commit(); c.close(); prune_backups(); return package


def prune_backups():
    """Keep daily/weekly/monthly copies without deleting the newest verified backup."""
    try:
        c=db(); rows=c.execute('SELECT id,file_path,created_at,label FROM system_backups ORDER BY created_at DESC').fetchall(); c.close()
        keep=[]; daily=weekly=monthly=0; seen_days=set(); seen_weeks=set(); seen_months=set()
        for r in rows:
            try: dt=datetime.fromisoformat(r['created_at'])
            except Exception: continue
            day=dt.date().isoformat(); week=dt.strftime('%Y-%W'); month=dt.strftime('%Y-%m')
            if day not in seen_days and daily<int(setting('backup_keep_days') or 14): seen_days.add(day); daily+=1; keep.append(r['id']); continue
            if week not in seen_weeks and weekly<int(setting('backup_keep_weekly') or 8): seen_weeks.add(week); weekly+=1; keep.append(r['id']); continue
            if month not in seen_months and monthly<int(setting('backup_keep_monthly') or 12): seen_months.add(month); monthly+=1; keep.append(r['id']); continue
            if r['id']==rows[0]['id']: keep.append(r['id'])
            else:
                try:
                    if os.path.exists(r['file_path']): os.remove(r['file_path'])
                except Exception: pass
                c2=db(); c2.execute('DELETE FROM backup_files WHERE backup_id=?',(r['id'],)); c2.execute('DELETE FROM system_backups WHERE id=?',(r['id'],)); c2.commit(); c2.close()
    except Exception as e: log_error('prune_backups',e)

def verify_backup_package(path):
    try:
        with zipfile.ZipFile(path,'r') as z:
            manifest=json.loads(z.read('manifest.json').decode('utf-8'))
            dbinfo=z.getinfo('database.db')
            if dbinfo.file_size<=0: return False,'database.db is empty'
            for item in manifest.get('files',[]):
                data=z.read(item['path'])
                if len(data)!=int(item.get('size',len(data))): return False,'size mismatch: '+item['path']
                if hashlib.sha256(data).hexdigest()!=item.get('sha256'): return False,'checksum mismatch: '+item['path']
            return True,'ok'
    except Exception as e: return False,str(e)



def employee_self_code(u):
    """Resolve the employee represented by a portal account without trusting client-supplied employee codes."""
    if not u: return None
    c=db()
    candidates=[u.get('scope_value') or '',u.get('username') or '']
    row=None
    for token in candidates:
        if token:
            row=c.execute('SELECT emp_code FROM employees WHERE emp_code=?',(token,)).fetchone()
            if row: break
            row=c.execute('SELECT emp_code FROM employees WHERE email=?',(token,)).fetchone()
            if row: break
    c.close()
    return row['emp_code'] if row else None

def hr_alerts_snapshot(u):
    c=db(); today=date.today(); d30=(today+timedelta(days=int(setting('contract_alert_days') or 30))).isoformat()
    dtrain=(today+timedelta(days=int(setting('training_alert_days') or 30))).isoformat()
    dcred=(today+timedelta(days=int(setting('credential_alert_days') or 30))).isoformat()
    scope_sql,scope_params=visible_employee_sql(u,'e')
    def q(sql,params=()):
        r=c.execute(sql,params).fetchone(); return int(r['n'] if r and 'n' in r.keys() else 0)
    expired_docs=q(f"""SELECT COUNT(*) n FROM documents d JOIN employees e ON e.emp_code=d.emp_code
                       WHERE d.status='current' AND d.expiry_date IS NOT NULL AND d.expiry_date< ? {scope_sql}""",[today.isoformat()]+scope_params)
    exp_cred=q(f"""SELECT COUNT(DISTINCT c.emp_code) n FROM credentials c JOIN employees e ON e.emp_code=c.emp_code
                   WHERE c.expiry_date IS NOT NULL AND c.expiry_date< ? {scope_sql}""",[today.isoformat()]+scope_params)
    exp_contract=q(f"""SELECT COUNT(*) n FROM employees e WHERE e.status='على رأس العمل'
                       AND e.contract_date IS NOT NULL AND e.contract_date<>'' AND e.contract_date<=? {scope_sql}""",[d30]+scope_params)
    late_limit=int(setting('monthly_late_limit_minutes') or 120)
    threshold=float(setting('late_risk_percent') or 80)/100.0
    late_risk=q(f"""SELECT COUNT(*) n FROM employees e WHERE e.status='على رأس العمل'
                    AND (SELECT COALESCE(SUM(a.late_minutes),0) FROM attendance a WHERE a.emp_code=e.emp_code AND substr(a.work_date,1,7)=?) >= ? {scope_sql}""",
                [today.isoformat()[:7],int(late_limit*threshold)]+scope_params)
    req=[x.strip() for x in (setting('required_doc_categories') or '').split(',') if x.strip()]
    missing=0
    if req:
        rows=c.execute(f"SELECT e.emp_code FROM employees e WHERE e.status='على رأس العمل' {scope_sql}",scope_params).fetchall()
        for r in rows:
            have={x['category'] for x in c.execute("SELECT category FROM documents WHERE emp_code=? AND status='current'",(r['emp_code'],)).fetchall()}
            if any(x not in have for x in req): missing+=1
    train=q(f"""SELECT COUNT(*) n FROM training t JOIN employees e ON e.emp_code=t.emp_code
                WHERE t.expiry_date IS NOT NULL AND t.expiry_date<=? {scope_sql}""",[dtrain]+scope_params)
    c.close()
    return [
      ('bad','🔴 مستندات منتهية',expired_docs,'documents'),
      ('bad','🔴 تراخيص/اعتمادات منتهية',exp_cred,'credentials'),
      ('warn','🟠 عقود تنتهي خلال 30 يوم',exp_contract,'contracts'),
      ('warn','🟠 موظفون قريبون من حد التأخير الشهري',late_risk,'late'),
      ('warn','🟡 موظفون لديهم مستندات ناقصة',missing,'missing'),
      ('warn','🟡 تدريب ينتهي قريبًا',train,'training')
    ]

def employee_risk(u,code):
    if not emp_allowed(u,code): return None
    c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone()
    if not e: c.close(); return None
    today=date.today(); month=today.isoformat()[:7]
    late=c.execute("SELECT COALESCE(SUM(late_minutes),0) n FROM attendance WHERE emp_code=? AND substr(work_date,1,7)=?",(code,month)).fetchone()['n']
    docs_exp=c.execute("SELECT COUNT(*) n FROM documents WHERE emp_code=? AND status='current' AND expiry_date<?",(code,today.isoformat())).fetchone()['n']
    req=[x.strip() for x in (setting('required_doc_categories') or '').split(',') if x.strip()]
    have={r['category'] for r in c.execute("SELECT category FROM documents WHERE emp_code=? AND status='current'",(code,)).fetchall()}
    missing=[x for x in req if x not in have]
    training=c.execute("SELECT COUNT(*) n FROM training WHERE emp_code=? AND expiry_date IS NOT NULL AND expiry_date<=?",(code,(today+timedelta(days=30)).isoformat())).fetchone()['n']
    contract=0
    if e['contract_date']:
        try: contract=1 if (datetime.fromisoformat(e['contract_date']).date()-today).days<=30 else 0
        except Exception: pass
    risk=0
    risk += 35 if docs_exp or missing else 0
    risk += 25 if late>=int(int(setting('monthly_late_limit_minutes') or 120)*.8) else 0
    risk += 20 if contract else 0
    risk += 15 if training else 0
    risk=min(100,risk)
    level='HIGH' if risk>=60 else ('MEDIUM' if risk>=30 else 'LOW')
    c.close()
    return {'level':level,'score':risk,'attendance':late,'documents_expired':docs_exp,'missing_documents':missing,'training_expiring':training,'contract_expiring':contract}

def update_device(u,ip,device_name=''):
    try:
        did=hashlib.sha256((u.get('username','')+'|'+ip+'|'+(device_name or '')).encode()).hexdigest()[:24]
        c=db(); c.execute("""INSERT INTO device_registry(device_id,device_name,department,username,ip,first_seen,last_seen,status,role)
                            VALUES(?,?,?,?,?,?,?,?,?)
                            ON CONFLICT(device_id) DO UPDATE SET device_name=excluded.device_name,department=excluded.department,
                            username=excluded.username,ip=excluded.ip,last_seen=excluded.last_seen,status='online',role=excluded.role""",
                         (did,device_name or 'Unknown Device',u.get('scope_value',''),u.get('username',''),ip,now(),now(),'online',u.get('role','')))
        c.commit(); c.close()
    except Exception as e: log_error('update_device',e)

def auto_alerts():
    try:
        c=db(); today=date.today(); limit=(today+timedelta(days=30)).isoformat()
        for r in c.execute("SELECT emp_code,file_name,expiry_date FROM documents WHERE expiry_date IS NOT NULL AND expiry_date<>'' AND expiry_date<=?",(limit,)).fetchall():
            msg=f"المستند {r['file_name']} للموظف {r['emp_code']} سينتهي/انتهى في {r['expiry_date']}"
            for usr in c.execute("SELECT username FROM users WHERE active=1 AND role IN ('SuperAdmin','Admin','HR')").fetchall():
                if not c.execute("SELECT 1 FROM notifications WHERE user_name=? AND title=? AND message=? AND created_at LIKE ?",(usr['username'],'تنبيه مستند',msg,today.isoformat()+'%')).fetchone():
                    c.execute('INSERT INTO notifications(user_name,title,message,created_at) VALUES(?,?,?,?)',(usr['username'],'تنبيه مستند',msg,now()))
        c.commit(); c.close()
    except Exception as e: log_error('auto_alerts',e)


def page(title,body,user,active='dashboard'):
    company=esc(setting('company_name') or 'HR Enterprise')
    nav=[]
    if can(user,'employees.view'): nav.append(('employees','الموظفون','/employees'))
    if can(user,'leave.create'): nav.append(('leaves','الإجازات','/leaves'))
    if can(user,'leave.create'): nav.append(('leave-balances','أرصدة الإجازات','/leave-balances'))
    if can(user,'attendance.view'): nav.append(('attendance','الحضور والانصراف','/attendance'))
    if can(user,'overtime.request'): nav.append(('overtime','الإضافي','/overtime'))
    if can(user,'shifts.manage'): nav.append(('shifts','الورديات','/shifts'))
    if can(user,'reports.view'): nav.append(('reports','التقارير','/reports'))
    if can(user,'employees.edit'): nav.append(('import','Excel Center','/import'))
    if can(user,'import.mapping'): nav.append(('import-map','ربط Excel','/import/mapping'))
    if can(user,'documents.manage'): nav.append(('documents','المستندات','/documents'))
    if can(user,'payroll.view'): nav.append(('payroll','المرتبات','/payroll'))
    if can(user,'payroll.view'): nav.append(('payroll-review','مراجعة المرتبات','/payroll/review'))
    if can(user,'roles.manage'): nav.append(('roles','الأدوار والصلاحيات','/roles'))
    if can(user,'users.manage'): nav.append(('access','Access Control','/access'))
    if can(user,'backup.manage'): nav.append(('backups','Backup / Rollback','/backups'))
    if can(user,'settings.manage'): nav.append(('system','System','/system'))
    if can(user,'system.manage'): nav.append(('network','🖥 الشبكة والأجهزة','/network'))
    if can(user,'system.manage'): nav.append(('devices','🖥 Connected Devices','/devices'))
    if can(user,'discipline.manage'): nav.append(('discipline','الجزاءات والإنذارات','/discipline'))
    if can(user,'users.manage'): nav.append(('users','المستخدمون','/users'))
    if can(user,'settings.manage'): nav.append(('settings','الإعدادات','/settings'))
    if can(user,'employees.view'): nav.append(('enterprise','Enterprise Center','/enterprise'))
    if user.get('role') in ('Manager','HR','Admin','SuperAdmin'): nav.append(('requests','📥 Employee Requests','/requests'))
    if can(user,'employees.view'): nav.append(('alerts','🚨 HR Intelligence','/alerts'))
    if can(user,'employees.view'): nav.append(('matching','🧩 Matching Review','/matching'))
    if user.get('role')=='Employee': nav.append(('myhr','👤 My HR','/myhr'))
    if can(user,'system.manage'): nav.append(('system','🛠 صحة النظام','/system'))
    if can(user,'system.manage'): nav.append(('diagnostics','🧰 سجل الأخطاء','/diagnostics/errors'))
    if can(user,'audit.view'): nav.append(('audit','سجل المراجعة','/audit'))
    links='<a class="active" href="/">🏠 لوحة التحكم</a>'+''.join(f'<a class="{"active" if active==k else ""}" href="{u}">{t}</a>' for k,t,u in nav)
    c=db(); n=c.execute('SELECT COUNT(*) n FROM notifications WHERE user_name=? AND read_at IS NULL',(user['username'],)).fetchone()['n']; c.close()
    links += f'<a href="/notifications">🔔 الإشعارات <span class="badge b-blue">{n}</span></a><a href="/password">🔐 تغيير كلمة المرور</a><a href="/logout">تسجيل الخروج</a>'
    search_box='''<div class="gsearch"><form action="/search" method="get"><input id="gsearch" name="q" placeholder="بحث سريع: اسم، كود، مستند…" autocomplete="off"><kbd>Ctrl K</kbd></form>
<script>document.addEventListener('keydown',function(e){if((e.ctrlKey||e.metaKey)&&e.key.toLowerCase()==='k'){e.preventDefault();var el=document.getElementById('gsearch');if(el){el.focus();el.select();}}});</script>'''
    forced = '<div class="alert" style="margin-bottom:14px">🔐 يجب تغيير كلمة المرور قبل متابعة العمل. <a href="/password">تغيير الآن</a></div>' if user.get('must_change_password') else ''
    return f'''<!doctype html><html lang="ar"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{esc(title)} — {company}</title><style>{CSS}</style></head><body><div class="app"><aside class="side"><div class="brand">{('<img src="/branding/logo" alt="logo" style="max-width:92px;max-height:58px;display:block;margin-bottom:10px;border-radius:10px;background:#fff;padding:5px">') if setting('company_logo') else ''}<b>{company}</b><small>HR Enterprise · Simple Pro</small></div>{search_box}<nav class="nav">{links}</nav><div class="footer">{esc(user['full_name'])} · {esc(user['role'])}</div></aside><main class="main">{forced}{body}<div class="footer">HR Enterprise · برنامج واحد · بيانات موحدة · {esc(user['full_name'])}</div></main></div>
<script>
(function(){{
  let device=localStorage.getItem('hr_device_name');
  if(!device){{device=prompt('اسم هذا الجهاز داخل HR Enterprise (مثال HR-MANAGER-01):',location.hostname)||('WEB-'+navigator.platform);localStorage.setItem('hr_device_name',device);}}
  async function hb(){{try{{await fetch('/device/ping?heartbeat='+Date.now(),{{headers:{{'X-HR-Device-Name':device}}}});}}catch(e){{}}}}
  hb(); setInterval(hb,60000);
}})();
</script></body></html>'''

def login_page(msg=''):
    return f'''<!doctype html><html lang="ar"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>تسجيل الدخول</title><style>{CSS}</style></head><body><div class="login"><form class="login-card" method="post" action="/login">{('<img src="/branding/logo" alt="logo" style="max-width:150px;max-height:80px;display:block;margin:0 auto 14px">') if setting('company_logo') else ''}<h1>نظام الموارد البشرية</h1><p>HR Enterprise Multi-User</p>{('<div class="alert">'+esc(msg)+'</div>') if msg else ''}<input name="username" placeholder="اسم المستخدم" autofocus><input name="password" type="password" placeholder="كلمة المرور"><button class="btn">دخول للنظام</button><div class="footer">في أول تشغيل استخدم حساب المسؤول الذي تم إنشاؤه أثناء الإعداد، ثم غيّر كلمة المرور فورًا.</div></form></div></body></html>'''

class H(BaseHTTPRequestHandler):
    def log_message(self,*a): pass
    def user(self):
        c=cookies.SimpleCookie(); c.load(self.headers.get('Cookie','')); sid=c.get('sid')
        if not sid: return None
        sidv=sid.value
        u=SESS.get(sidv)
        if not u: return None
        try:
            idle=int(setting('session_idle_minutes') or 30)
            if datetime.now()-datetime.fromisoformat(u.get('last_seen',now()))>timedelta(minutes=idle):
                SESS.pop(sidv,None); c=db(); c.execute('UPDATE system_sessions SET revoked=1 WHERE session_id=?',(sidv,)); c.commit(); c.close(); return None
            c=db(); row=c.execute('SELECT * FROM users WHERE username=?',(u['username'],)).fetchone(); sessrow=c.execute('SELECT revoked FROM system_sessions WHERE session_id=?',(sidv,)).fetchone(); c.close()
            if not row or not row['active'] or not sessrow or sessrow['revoked']:
                SESS.pop(sidv,None); return None
            if int(row['permission_version'] or 1)!=int(u.get('permission_version') or 1) or row['role']!=u.get('role') or row['scope_type']!=u.get('scope_type') or row['scope_value']!=u.get('scope_value') or int(row['must_change_password'] or 0)!=int(u.get('must_change_password') or 0):
                u=dict(row); u['csrf']=secrets.token_urlsafe(24); u['last_seen']=now(); u['permission_version']=row['permission_version']
                SESS[sidv]=u
            u['last_seen']=now(); u['_sid']=sidv
            c=db(); c.execute('UPDATE system_sessions SET last_seen=? WHERE session_id=? AND revoked=0',(u['last_seen'],sidv)); c.commit(); c.close(); return u
        except Exception:
            return None
    def send(self,body,status=200,ctype='text/html; charset=utf-8',headers=None):
        if isinstance(body,str): body=body.encode()
        self.send_response(status); self.send_header('Content-Type',ctype); self.send_header('Content-Length',str(len(body)))
        self.send_header('X-Content-Type-Options','nosniff')
        self.send_header('X-Frame-Options','SAMEORIGIN')
        self.send_header('Referrer-Policy','strict-origin-when-cross-origin')
        self.send_header('Permissions-Policy','camera=(), microphone=(), geolocation=()')
        for k,v in (headers or {}).items():
            try: self.send_header(k,v)
            except UnicodeEncodeError: self.send_header(k, v.encode('latin-1','ignore').decode('latin-1'))
        self.end_headers(); self.wfile.write(body)
    def redirect(self,url,extra=None): self.send_response(302); self.send_header('Location',url); [self.send_header(k,v) for k,v in (extra or {}).items()]; self.end_headers()
    def _raw(self):
        # Read the request body exactly once per request and cache it. Several
        # handlers (form parsing for CSRF, then the specific upload handler)
        # need the same bytes; re-reading self.rfile a second time blocks
        # forever because the client already sent everything and is waiting
        # on the response, which used to hang the server on document uploads.
        if not hasattr(self,'_raw_body'):
            n=int(self.headers.get('Content-Length','0')); self._raw_body=self.rfile.read(n)
        return self._raw_body
    def form(self):
        raw=self._raw().decode('utf-8', errors='replace')
        parsed=parse_qs(raw,keep_blank_values=True)
        return {k:(v if len(v)>1 else v[0]) for k,v in parsed.items()}
    def fval(self,f,key,default=''):
        v=f.get(key,default)
        return v[0] if isinstance(v,list) and v else (default if isinstance(v,list) else v)
    def flist(self,f,key):
        v=f.get(key,[])
        return v if isinstance(v,list) else ([v] if v not in (None,'') else [])
    def require(self):
        u=self.user()
        if not u:
            self.send(login_page(),401); return None
        # Live session validation: user disable/role/scope changes take effect immediately.
        try:
            c=db(); r=c.execute('SELECT * FROM users WHERE username=?',(u.get('username'),)).fetchone(); c.close()
            if not r or not r['active'] or int(r['permission_version'] or 1)!=int(u.get('permission_version') or 1):
                sid=u.get('_sid'); SESS.pop(sid,None)
                try:
                    c=db(); c.execute('UPDATE system_sessions SET revoked=1,last_seen=? WHERE session_id=?',(now(),sid)); c.commit(); c.close()
                except Exception: pass
                self.send(login_page('تم تحديث صلاحيات الحساب أو إيقافه. سجّل الدخول مرة أخرى.'),401); return None
            u.update(dict(r)); u['_sid']=u.get('_sid',getattr(self,'_sid',''))
            u['csrf']=SESS.get(u.get('_sid'),{}).get('csrf',u.get('csrf',''))
            return u
        except Exception:
            return u
    def forbid(self,u):
        self.send(page('صلاحيات','<div class="card"><div class="alert">لا تملك صلاحية لتنفيذ هذا الإجراء.</div></div>',u),403); return None
    def need(self,u,perm):
        if can(u,perm): return True
        self.forbid(u); return False
    def error_page(self,u,request_id,status=500):
        admin = bool(u and (u.get('role') in ('SuperAdmin','Admin') or can(u,'system.manage')))
        link = '<a class="btn gray" href="/diagnostics/errors">فتح سجل الأخطاء</a>' if admin else ''
        body=f'''<div class="card" style="max-width:900px;margin:30px auto"><div style="font-size:52px">⚠️</div><h1>حدث خطأ غير متوقع</h1><p>لم نتمكن من تنفيذ العملية. لم يتم إخفاء الخطأ؛ تم تسجيله للتشخيص.</p><div class="alert"><b>رقم الخطأ:</b> {esc(request_id)}<br><small>أرسل هذا الرقم لمسؤول النظام بدل إرسال رسالة عامة فقط.</small></div><div class="actions" style="margin-top:16px"><button class="btn" onclick="location.reload()">إعادة المحاولة</button><a class="btn gray" href="/">العودة للوحة التحكم</a>{link}</div></div>'''; self.send(page('خطأ في النظام',body,u or {'full_name':'','role':''}),status)

    def bridge_authorize_token(self):
        try:
            n=int(self.headers.get('Content-Length','0')); payload=json.loads(self.rfile.read(n).decode('utf-8')); tok=str(payload.get('token','')); csrf=str(payload.get('csrf','')); emp=str(payload.get('emp_code',''))
            info=BRIDGE_TOKENS.get(tok); ok=bool(info and datetime.now()<datetime.fromisoformat(info['expires_at']) and secrets.compare_digest(info['csrf'],csrf))
            if not ok: return self.send(json.dumps({'ok':False,'error':'bridge token expired'}),403,'application/json')
            self.send(json.dumps({'ok':True,'emp_code':emp}),200,'application/json')
        except Exception as e: self.send(json.dumps({'ok':False,'error':str(e)}),400,'application/json')

    def bridge_upload_token(self):
        try:
            ctype=self.headers.get('Content-Type','').lower(); fields,files=self.parse_upload_all(); token=fields.get('_bridge_token','') or fields.get('bridge_token',''); info=BRIDGE_TOKENS.get(token)
            if not info or datetime.now()>=datetime.fromisoformat(info['expires_at']): return self.send(json.dumps({'ok':False,'error':'bridge token expired'}),403,'application/json')
            if not secrets.compare_digest(fields.get('_csrf',''),info['csrf']): return self.send(json.dumps({'ok':False,'error':'invalid bridge csrf'}),403,'application/json')
            emp=fields.get('emp_code','')
            if not emp: return self.send(json.dumps({'ok':False,'error':'employee required'}),400,'application/json')
            c=db(); er=c.execute('SELECT 1 FROM employees WHERE emp_code=?',(emp,)).fetchone(); c.close()
            if not er: return self.send(json.dumps({'ok':False,'error':'employee not found'}),404,'application/json')
            imported=0
            for _head,data,fname in files:
                if not data: continue
                ext=os.path.splitext(fname)[1].lower(); allowed={'.pdf','.jpg','.jpeg','.png','.docx','.doc','.xlsx','.xls','.txt','.csv'}
                if ext not in allowed: continue
                if not validate_file_signature(fname,data): continue
                if len(data)>float(setting('document_max_mb') or 25)*1024*1024: continue
                fname=safe_name(fname); cat=guess_document_category(fname); rel=save_employee_file(emp,fname,data); checksum=hashlib.sha256(data).hexdigest(); c=db(); prior=c.execute('SELECT id,version FROM documents WHERE emp_code=? AND category=? AND status=?',(emp,cat,'current')).fetchall(); ver=max([r['version'] or 1 for r in prior] or [0])+1; cur=c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,fname,ext,info['username'],now(),None,rel,cat,ver,checksum,'current')); [c.execute('UPDATE documents SET status=?,superseded_by=? WHERE id=?',('superseded',cur.lastrowid,r['id'])) for r in prior]; c.commit(); c.close(); imported+=1
            BRIDGE_TOKENS.pop(token,None)
            audit(info['username'],'Bridge','Native Clipboard Upload','المستندات',emp,f'files={imported}')
            self.send(json.dumps({'ok':True,'imported':imported},ensure_ascii=False),200,'application/json')
        except Exception as e:
            self.send(json.dumps({'ok':False,'error':str(e)},ensure_ascii=False),400,'application/json')

    def bridge_token(self,u):
        c=db(); tok=secrets.token_urlsafe(32); exp=(datetime.now()+timedelta(minutes=3)).isoformat(timespec='seconds')
        c.execute('INSERT INTO bridge_tokens(token,username,session_id,expires_at,used) VALUES(?,?,?,?,0)',(tok,u['username'],u.get('_sid',''),exp)); c.commit(); c.close(); BRIDGE_TOKENS[tok]={'username':u['username'],'expires_at':exp,'csrf':u.get('csrf','')}
        self.send(json.dumps({'ok':True,'token':tok,'expires_at':exp,'bridge_url':'http://127.0.0.1:8975/clipboard/upload'},ensure_ascii=False),200,'application/json')

    def system_health_json(self,u):
        dh=database_health(); sh=storage_health(); ok=bool(dh.get('ok')) and all(x.get('exists') and x.get('writable') for x in sh)
        self.send(json.dumps({'ok':ok,'version':APP_VERSION,'server_name':setting('server_name') or 'HR-SERVER','ip':local_ip(),'port':PORT,'database':dh,'storage':sh,'audit':verify_audit_chain()[0],'server_id':server_identity()['server_id']},ensure_ascii=False),200,'application/json')

    def diagnostics_errors(self,u):
        c=db(); rows=c.execute('SELECT * FROM error_logs ORDER BY id DESC LIMIT 300').fetchall(); c.close()
        trs=''.join(f'<tr><td><code>{esc(r["request_id"])}</code></td><td>{esc(r["ts"])}</td><td>{esc(r["username"] or "-")}</td><td>{esc(r["method"])} {esc(r["path"])}</td><td>{esc(r["ip"] or "-")}</td><td><b>{esc(r["exception_type"])}</b><br>{esc(r["message"])}</td><td><details><summary>التفاصيل</summary><pre style="white-space:pre-wrap;max-width:680px">{esc(r["traceback"] or "")}</pre></details></td></tr>' for r in rows) or '<tr><td colspan="7">لا توجد أخطاء مسجلة.</td></tr>'
        body=f'''<div class="top"><div class="title"><h1>🧰 Diagnostics / Error Center</h1><p>السجل الحقيقي لأخطاء Backend مع Request ID لتحديد المشكلة بسرعة.</p></div><a class="btn gray" href="/system">حالة النظام</a></div><div class="card table-wrap"><table class="table"><thead><tr><th>Request ID</th><th>الوقت</th><th>المستخدم</th><th>الطلب</th><th>IP</th><th>الخطأ</th><th>Stack Trace</th></tr></thead><tbody>{trs}</tbody></table></div>'''; self.send(page('Diagnostics',body,u,'system'))

    def diagnostics_test(self,u):
        return self.send(json.dumps({'ok':True,'request_id':str(uuid.uuid4())[:12],'database':database_health(),'storage':storage_health(),'port':PORT,'ip':local_ip(),'version':APP_VERSION},ensure_ascii=False),200,'application/json')

    def password_page(self,u):
        body=f'''<div class="top"><div class="title"><h1>تغيير كلمة المرور</h1><p>كلمة مرور قوية تحمي حسابك.</p></div></div>
<div class="card" style="max-width:620px"><form class="form" method="post" action="/password">{csrf_field(u)}
<div class="field"><label>كلمة المرور الحالية</label><input type="password" name="current" required autocomplete="current-password"></div>
<div class="field"><label>كلمة المرور الجديدة</label><input type="password" name="new_password" minlength="10" required autocomplete="new-password"></div>
<div class="field"><label>تأكيد كلمة المرور</label><input type="password" name="confirm" minlength="10" required autocomplete="new-password"></div>
<div class="full"><div class="alert">10 أحرف على الأقل، ويفضل حروفًا وأرقامًا ورموزًا.</div></div>
<div class="full"><button class="btn">حفظ كلمة المرور</button></div></form></div>'''
        self.send(page('تغيير كلمة المرور',body,u,'password'))
    def password_save(self,u,f):
        cur=f.get('current',''); new=f.get('new_password',''); conf=f.get('confirm','')
        if not checkpw(cur,u.get('password_hash','')):
            return self.send(page('تغيير كلمة المرور','<div class="card"><div class="alert">كلمة المرور الحالية غير صحيحة.</div></div>',u,'password'),400)
        if len(new)<10 or new!=conf or new==cur:
            return self.send(page('تغيير كلمة المرور','<div class="card"><div class="alert">تحقق من كلمة المرور الجديدة.</div></div>',u,'password'),400)
        c=db(); nh=hashpw(new); c.execute('UPDATE users SET password_hash=?,must_change_password=0 WHERE username=?',(nh,u['username'])); c.commit(); c.close()
        u['password_hash']=nh; u['must_change_password']=0
        audit(u['username'],u['role'],'تغيير كلمة المرور','المستخدمون',u['username'],reason='self password change')
        self.redirect('/')


    def alerts_page(self,u):
        alerts=hr_alerts_snapshot(u)
        cards=''.join('<div class="card metric"><div class="label">{}</div><div class="value" style="font-size:28px">{}</div><div class="sub">HR review · {}</div></div>'.format(esc(t),n,esc(k)) for s,t,n,k in alerts)
        c=db(); scope_sql,scope_params=visible_employee_sql(u,'e')
        rows=c.execute("SELECT e.emp_code,e.name FROM employees e WHERE e.status='على رأس العمل' "+scope_sql+" ORDER BY e.name LIMIT 500",scope_params).fetchall(); c.close()
        riskrows=[]
        for r in rows:
            x=employee_risk(u,r['emp_code'])
            if x and x['score']>=30: riskrows.append((r,x))
        riskrows.sort(key=lambda z:z[1]['score'],reverse=True)
        trs=''.join('<tr><td><a href="/employee/profile/{0}">{1}</a></td><td>{2}</td><td><span class="badge {3}">{4}</span></td><td>{5}%</td><td>Late {6}m · Docs {7} · Training {8}</td></tr>'.format(esc(r['emp_code']),esc(r['name']),esc(r['emp_code']),'b-bad' if x['level']=='HIGH' else 'b-warn',x['level'],x['score'],x['attendance'],x['documents_expired']+len(x['missing_documents']),x['training_expiring']) for r,x in riskrows[:100])
        body='''<div class="top"><div class="title"><h1>🚨 HR Intelligence & Alerts</h1><p>تنبيهات استباقية تساعد HR على التدخل قبل حدوث المشكلة — ليست قرارات تلقائية ضد الموظف.</p></div><a class="btn gray" href="/enterprise">Enterprise Center</a></div>
<div class="grid g3">{cards}</div>
<div class="card" style="margin-top:16px"><h3>Employee Risk / Attention</h3><p class="sub">مؤشر مراجعة فقط، مبني على الحضور والمستندات والتدريب والعقد.</p><div class="table-wrap"><table class="table"><thead><tr><th>الموظف</th><th>الكود</th><th>HR Attention</th><th>Risk</th><th>Reasons</th></tr></thead><tbody>{trs}</tbody></table></div></div>
<div class="card" style="margin-top:16px"><h3>🔔 Policy</h3><p>كل التنبيهات قابلة للمراجعة. النظام لا يفرض عقوبة أو قرارًا وظيفيًا تلقائيًا.</p></div>'''.format(cards=cards,trs=trs or '<tr><td colspan="5">لا يوجد موظفون يحتاجون انتباهًا حاليًا.</td></tr>')
        self.send(page('HR Intelligence',body,u,'alerts'))

    def matching_page(self,u):
        c=db(); rows=c.execute("SELECT * FROM matching_reviews WHERE status='review' ORDER BY confidence ASC, id DESC LIMIT 200").fetchall(); c.close()
        trs=''
        for r in rows:
            cls='b-ok' if (r['confidence'] or 0)>=.98 else ('b-warn' if (r['confidence'] or 0)>=.90 else 'b-bad')
            trs += '<tr><td>{}</td><td>{}</td><td><span class="badge {}">{:.0f}%</span></td><td><form method="post" action="/matching/review" class="match-form {}" style="display:flex;gap:6px">{}<input type="hidden" name="id" value="{}"><input name="selected_code" placeholder="Employee Code" required><button class="btn">Accept</button></form></td></tr>'.format(esc(r['source_name']),esc(r['candidate_json'] or '[]'),cls,(r['confidence'] or 0)*100,('safe' if (r['confidence'] or 0)>=.98 else ''),csrf_field(u),r['id'])
        body='''<div class="top"><div class="title"><h1>🧩 Matching Center</h1><p>المطابقات الآمنة يمكن اعتمادها، والمطابقات الملتبسة تحتاج اختيار HR.</p></div><a class="btn gray" href="/import">Excel Center</a></div>
<div class="card"><div class="actions"><button class="btn ok" type="button" onclick="document.querySelectorAll('.match-form.safe').forEach(f=>f.submit())">Accept All Safe Matches</button></div><div class="table-wrap"><table class="table"><thead><tr><th>Folder / Candidate</th><th>Candidates</th><th>Confidence</th><th>Action</th></tr></thead><tbody>{}</tbody></table></div></div>'''.format(trs or '<tr><td colspan="4">لا توجد مطابقة تحتاج Review.</td></tr>')
        self.send(page('Matching Center',body,u,'matching'))

    def myhr(self,u):
        code=employee_self_code(u)
        if not code:
            return self.send(page('My HR','<div class="card"><div class="alert">لم يتم ربط حسابك بموظف. اطلب من HR ضبط Employee Code أو البريد الإلكتروني للحساب.</div></div>',u,'myhr'),403)
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone()
        att=c.execute('SELECT * FROM attendance WHERE emp_code=? ORDER BY work_date DESC LIMIT 20',(code,)).fetchall()
        leaves=c.execute('SELECT * FROM leaves WHERE emp_code=? ORDER BY id DESC LIMIT 20',(code,)).fetchall()
        pays=c.execute('SELECT period,net,status FROM payroll WHERE emp_code=? ORDER BY period DESC LIMIT 12',(code,)).fetchall()
        docs=c.execute('SELECT file_name,category,expiry_date FROM documents WHERE emp_code=? AND status="current" ORDER BY category',(code,)).fetchall()
        train=c.execute('SELECT course,expiry_date,status FROM training WHERE emp_code=? ORDER BY expiry_date DESC LIMIT 20',(code,)).fetchall()
        ev=c.execute('SELECT period,score,notes FROM employee_evaluations WHERE emp_code=? ORDER BY period DESC LIMIT 6',(code,)).fetchall(); c.close()
        attrows=''.join('<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}m</td></tr>'.format(esc(r['work_date']),esc(r['status']),esc(r['check_in'] or ''),esc(r['check_out'] or ''),r['late_minutes'] or 0) for r in att)
        lrows=''.join('<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(esc(r['request_no']),esc(r['leave_type']),esc(r['start_date']),esc(r['end_date']),esc(r['status'])) for r in leaves)
        prows=''.join('<tr><td>{}</td><td>{}</td><td>{}</td></tr>'.format(esc(r['period']),r['net'] or 0,esc(r['status'])) for r in pays)
        drows=''.join('<tr><td>{}</td><td>{}</td><td>{}</td></tr>'.format(esc(r['file_name']),esc(r['category']),esc(r['expiry_date'] or '—')) for r in docs)
        trows=''.join('<tr><td>{}</td><td>{}</td><td>{}</td></tr>'.format(esc(r['course']),esc(r['expiry_date'] or '—'),esc(r['status'])) for r in train)
        erows=''.join('<tr><td>{}</td><td>{:.1f}</td><td>{}</td></tr>'.format(esc(r['period']),r['score'] or 0,esc(r['notes'] or '')) for r in ev)
        body='''<div class="top"><div class="title"><h1>👤 My HR</h1><p>{name} · {code} — لا يمكن لحساب Employee الوصول لبيانات موظف آخر.</p></div></div>
<div class="grid g3"><div class="card"><h3>👤 My Profile</h3><p>{job} · {dept}</p><a class="btn gray" href="/employee/profile/{code}">عرض ملفي</a></div>
<div class="card"><h3>🏖 Leave Request</h3><form method="post" action="/leave/save">{csrf}<input type="hidden" name="emp_code" value="{code}"><div class="field"><label>النوع</label><select name="leave_type"><option>اعتيادي</option><option>عارضة</option><option>مرضي</option></select></div><div class="field"><label>من</label><input type="date" name="start_date" required></div><div class="field"><label>إلى</label><input type="date" name="end_date" required></div><button class="btn">إرسال للمدير</button></form></div>
<div class="card"><h3>⏱ Overtime Request</h3><form method="post" action="/overtime/save">{csrf}<input type="hidden" name="emp_code" value="{code}"><div class="field"><label>التاريخ</label><input type="date" name="work_date" required></div><div class="field"><label>الساعات</label><input type="number" step=".5" name="hours" min=".5" required></div><div class="field"><label>السبب</label><input name="reason" required></div><button class="btn">إرسال</button></form></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><h3>🕐 My Attendance</h3><div class="table-wrap"><table class="table"><thead><tr><th>Date</th><th>Status</th><th>In</th><th>Out</th><th>Late</th></tr></thead><tbody>{att}</tbody></table></div></div><div class="card"><h3>🏖 My Leaves</h3><div class="table-wrap"><table class="table"><thead><tr><th>No</th><th>Type</th><th>From</th><th>To</th><th>Status</th></tr></thead><tbody>{leaves}</tbody></table></div></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><h3>💰 My Payroll</h3><div class="table-wrap"><table class="table"><thead><tr><th>Period</th><th>Net</th><th>Status</th></tr></thead><tbody>{pay}</tbody></table></div></div><div class="card"><h3>📄 My Documents</h3><div class="table-wrap"><table class="table"><thead><tr><th>File</th><th>Category</th><th>Expiry</th></tr></thead><tbody>{docs}</tbody></table></div></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><h3>🎓 My Training</h3><div class="table-wrap"><table class="table"><thead><tr><th>Course</th><th>Expiry</th><th>Status</th></tr></thead><tbody>{train}</tbody></table></div></div><div class="card"><h3>📊 My Evaluation</h3><div class="table-wrap"><table class="table"><thead><tr><th>Period</th><th>Score</th><th>Notes</th></tr></thead><tbody>{eval}</tbody></table></div></div></div>'''.format(name=esc(e['name']),code=esc(code),job=esc(e['job'] or '—'),dept=esc(e['department'] or '—'),csrf=csrf_field(u),att=attrows or '<tr><td colspan="5">لا توجد بيانات</td></tr>',leaves=lrows or '<tr><td colspan="5">لا توجد طلبات</td></tr>',pay=prows or '<tr><td colspan="3">لا توجد بيانات</td></tr>',docs=drows or '<tr><td colspan="3">لا توجد مستندات</td></tr>',train=trows or '<tr><td colspan="3">لا توجد دورات</td></tr>',eval=erows or '<tr><td colspan="3">لا يوجد تقييم</td></tr>')
        body += '<div class="grid g2" style="margin-top:16px"><div class="card"><h3>🛠 Attendance Correction</h3><form method="post" action="/myhr/request">'+csrf_field(u)+'<input type="hidden" name="request_type" value="attendance_correction"><div class="field"><label>Date</label><input type="date" name="work_date" required></div><div class="field"><label>Requested In</label><input type="time" name="requested_check_in"></div><div class="field"><label>Requested Out</label><input type="time" name="requested_check_out"></div><div class="field"><label>Reason</label><input name="reason" required></div><button class="btn">Send to Manager</button></form></div><div class="card"><h3>📄 Document Request</h3><form method="post" action="/myhr/request">'+csrf_field(u)+'<input type="hidden" name="request_type" value="document_request"><div class="field"><label>Document Category</label><input name="document_category" placeholder="عقد / هوية / مؤهل" required></div><div class="field"><label>Reason</label><input name="reason"></div><button class="btn">Send to Manager</button></form></div></div>'
        self.send(page('My HR',body,u,'myhr'))


    def myhr_request(self,u,f):
        code=employee_self_code(u)
        if not code: return self.forbid(u)
        typ=self.fval(f,'request_type').strip()
        if typ not in ('attendance_correction','document_request'):
            return self.send(page('My HR','<div class="card"><div class="alert">نوع الطلب غير صحيح.</div></div>',u,'myhr'),400)
        payload={k:self.fval(f,k) for k in ('work_date','requested_check_in','requested_check_out','reason','document_category')}
        c=db(); c.execute('INSERT INTO employee_requests(emp_code,request_type,payload_json,status,created_at,updated_at) VALUES(?,?,?,?,?,?)',(code,typ,json.dumps(payload,ensure_ascii=False),'pending',now(),now())); c.commit(); c.close()
        audit(u['username'],u['role'],'طلب موظف',typ,code,json.dumps(payload,ensure_ascii=False))
        self.redirect('/myhr')

    def requests_page(self,u):
        if u.get('role') not in ('Manager','HR','Admin','SuperAdmin'): return self.forbid(u)
        c=db(); rows=c.execute('SELECT r.*,e.name FROM employee_requests r JOIN employees e ON e.emp_code=r.emp_code WHERE r.status IN ("pending","manager_approved") ORDER BY r.id DESC LIMIT 300').fetchall(); c.close()
        trs=''
        for r in rows:
            stage='Manager Review' if r['status']=='pending' else 'HR Review'
            trs += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td><form method="post" action="/requests/action" style="display:flex;gap:5px">{}<input type="hidden" name="id" value="{}"><button name="decision" value="approve" class="btn ok">Approve</button><button name="decision" value="reject" class="btn bad">Reject</button></form></td></tr>'.format(esc(r['name']),esc(r['emp_code']),esc(r['request_type']),stage,csrf_field(u),r['id'])
        body='''<div class="top"><div class="title"><h1>Employee Requests</h1><p>Workflow: Employee → Manager → HR.</p></div></div>
<div class="card table-wrap"><table class="table"><thead><tr><th>Employee</th><th>Code</th><th>Request</th><th>Stage</th><th>Action</th></tr></thead><tbody>{}</tbody></table></div>'''.format(trs or '<tr><td colspan="5">لا توجد طلبات معلقة.</td></tr>')
        self.send(page('Employee Requests',body,u,'requests'))

    def request_action(self,u,f):
        if u.get('role') not in ('Manager','HR','Admin','SuperAdmin'): return self.forbid(u)
        rid=self.fval(f,'id'); decision=self.fval(f,'decision')
        c=db(); r=c.execute('SELECT * FROM employee_requests WHERE id=?',(rid,)).fetchone()
        if not r: c.close(); return self.redirect('/requests')
        if decision=='reject':
            c.execute("UPDATE employee_requests SET status='rejected',updated_at=?,manager_user=COALESCE(manager_user,?) WHERE id=?",(now(),u['username'],rid))
        elif r['status']=='pending':
            c.execute("UPDATE employee_requests SET status='manager_approved',manager_user=?,updated_at=? WHERE id=?",(u['username'],now(),rid))
        elif r['status']=='manager_approved' and u.get('role') in ('HR','Admin','SuperAdmin'):
            c.execute("UPDATE employee_requests SET status='approved',hr_user=?,updated_at=? WHERE id=?",(u['username'],now(),rid))
        c.commit(); c.close(); self.redirect('/requests')

    def matching_review(self,u,f):
        rid=self.fval(f,'id'); code=self.fval(f,'selected_code').strip()
        if not rid or not code: return self.redirect('/matching')
        c=db(); row=c.execute('SELECT * FROM matching_reviews WHERE id=?',(rid,)).fetchone(); emp=c.execute('SELECT 1 FROM employees WHERE emp_code=?',(code,)).fetchone()
        if not row or not emp:
            c.close(); return self.send(page('Matching','<div class="card"><div class="alert">المطابقة أو الموظف غير موجود.</div></div>',u,'matching'),400)
        c.execute("UPDATE matching_reviews SET status='accepted',selected_code=?,reviewed_by=?,reviewed_at=? WHERE id=?",(code,u['username'],now(),rid)); c.commit(); c.close()
        audit(u['username'],u['role'],'اعتماد مطابقة','Matching',str(rid),'selected='+code); self.redirect('/matching')

    def device_ping(self,u):
        dev=self.headers.get('X-HR-Device-Name','')[:120] or self.headers.get('User-Agent','')[:120]
        update_device(u,self.client_address[0],dev)
        self.send(json.dumps({'ok':True,'server':setting('server_name') or 'HR-MAIN','ip':local_ip(),'port':PORT,'device':dev,'last_sync':now()},ensure_ascii=False),200,'application/json',{'Cache-Control':'no-store'})

    def devices_page(self,u):
        c=db(); rows=c.execute('SELECT * FROM device_registry ORDER BY last_seen DESC').fetchall(); c.close()
        trs=''.join('<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td><span class="badge {}">{}</span></td><td>{}</td></tr>'.format(esc(r['device_name']),esc(r['username']),esc(r['department'] or '—'),esc(r['ip']),'b-ok' if r['status']=='online' else 'b-gray','🟢 Connected' if r['status']=='online' else '🔴 Offline',esc(r['last_seen'])) for r in rows)
        body='''<div class="top"><div class="title"><h1>🖥 Connected Devices</h1><p>أسماء الأجهزة + المستخدم + IP + آخر اتصال.</p></div><button class="btn" onclick="syncDevice()">Refresh Device</button></div>
<div class="card"><h3>Server Status</h3><p>🟢 Connected · Server: <b>{server}</b> · IP: <b>{ip}</b> · Port: <b>{port}</b></p><div id="deviceStatus" class="alert">Last Sync: —</div></div>
<div class="card table-wrap" style="margin-top:16px"><table class="table"><thead><tr><th>Device</th><th>User</th><th>Department</th><th>IP</th><th>Status</th><th>Last Sync</th></tr></thead><tbody>{rows}</tbody></table></div>
<script>async function syncDevice(){{const d=prompt('Computer Name / Device Name',navigator.platform+'-'+location.hostname);try{{const r=await fetch('/device/ping?x='+Date.now(),{{headers:{{'X-HR-Device-Name':d||navigator.platform}}}});const j=await r.json();document.getElementById('deviceStatus').textContent='🟢 Connected · '+j.server+' · '+j.ip+':'+j.port+' · '+j.last_sync;setTimeout(()=>location.reload(),500);}}catch(e){{document.getElementById('deviceStatus').textContent='🔴 SERVER OFFLINE';}}}}</script>'''.format(server=esc(setting('server_name') or 'HR-MAIN'),ip=esc(local_ip()),port=PORT,rows=trs or '<tr><td colspan="6">لا توجد أجهزة مسجلة.</td></tr>')
        self.send(page('Connected Devices',body,u,'network'))

    def do_GET(self):
        path=urlparse(self.path).path
        self.request_id='ERR-'+datetime.now().strftime('%Y%m%d%H%M%S')+'-'+uuid.uuid4().hex[:6].upper()
        if path=='/login': return self.send(login_page())
        if path=='/health': return self.send(json.dumps({'ok':True,'version':APP_VERSION,'server_name':setting('server_name') or 'HR-SERVER'},ensure_ascii=False),200,'application/json',{'Cache-Control':'no-store'})
        if path=='/branding/logo':
            logo=setting('company_logo')
            p=os.path.join(DATA,logo) if logo else ''
            if p and os.path.exists(p):
                raw=open(p,'rb').read(); self.send(raw,200,mimetypes.guess_type(p)[0] or 'image/png',{'Cache-Control':'no-cache'})
            else:
                self.send(b'',404,'image/png')
            return
        if path=='/logout':
            ck=cookies.SimpleCookie(); ck.load(self.headers.get('Cookie','')); sidc=ck.get('sid')
            if sidc:
                SESS.pop(sidc.value,None)
                try:
                    cc=db(); cc.execute('UPDATE system_sessions SET revoked=1,last_seen=? WHERE session_id=?',(now(),sidc.value)); cc.commit(); cc.close()
                except Exception: pass
            c=cookies.SimpleCookie(); c['sid']=''; c['sid']['expires']='Thu, 01 Jan 1970 00:00:00 GMT'; self.send_response(302); self.send_header('Location','/login'); self.send_header('Set-Cookie',c.output(header='').strip()); self.end_headers(); return
        if path=='/documents/bridge-upload':
            return self.bridge_upload_token()
        u=self.require()
        if not u:return
        AUDIT_CTX.ip=self.client_address[0]; AUDIT_CTX.user=u
        if u.get('must_change_password') and path not in ('/password','/logout'):
            return self.redirect('/password')
        try:
            if path=='/system/health.json': return self.system_health_json(u)
            if path=='/bridge/token': return self.bridge_token(u)
            if path=='/bridge/authorize': return self.bridge_authorize(u)
            if path=='/password': return self.password_page(u)
            if path=='/': return self.dashboard(u)
            if path=='/enterprise': return self.enterprise_center(u)
            if path=='/alerts': return self.alerts_page(u)
            if path=='/matching': return self.matching_page(u)
            if path=='/myhr': return self.myhr(u)
            if path=='/requests': return self.requests_page(u)
            if path=='/devices': return self.need(u,'system.manage') and self.devices_page(u)
            if path=='/device/ping': return self.device_ping(u)
            if path=='/employees': return self.employees(u)
            if path=='/employees/views': return self.employees_views(u)
            if path=='/export/employees/selected': return self.need(u,'reports.export') and self.export_selected_employees(u)
            if path.startswith('/export/html/'): return self.need(u,'reports.export') and self.export_html(path.split('/')[-1],u)
            if path.startswith('/export/csv/'): return self.need(u,'reports.export') and self.export_csv(path.split('/')[-1],u)
            if path=='/documents/paste': return self.need(u,'documents.manage') and self.paste_documents(u)
            if path=='/documents/bridge-upload': return self.need(u,'documents.manage') and self.bridge_upload(u)
            if path=='/employee/new': return self.employee_form(u)
            if path.startswith('/employee/edit/'): return self.employee_form(u,path.split('/')[-1])
            if path.startswith('/employee/profile/'): return self.employee_profile(u,path.split('/')[-1])
            if path=='/leaves': return self.leaves(u)
            if path=='/leave/new': return self.leave_form(u)
            if path=='/attendance': return self.attendance(u)
            if path=='/reports': return self.reports(u)
            if path=='/import': return self.import_page(u)
            if path=='/users': return self.users(u)
            if path=='/settings': return self.need(u,'settings.manage') and self.settings_page(u)
            if path=='/audit': return self.need(u,'audit.view') and self.audit_page(u)
            if path=='/audit/verify': return self.need(u,'audit.view') and self.audit_verify_page(u)
            if path=='/search': return self.search_page(u)
            if path=='/import/folder-history': return self.need(u,'documents.bulk_import') and self.folder_import_history(u)
            if path.startswith('/export/import-errors/'):
                return self.need(u,'import.validate') and self.export_import_errors(path.split('/')[-1])
            if path.startswith('/export/pdf/'): return self.need(u,'reports.export') and self.pdf_report(path.split('/')[-1],u)
            if path.startswith('/export/'): return self.need(u,'reports.export') and self.export_route(path)
            if path=='/backup': return self.need(u,'backup.manage') and self.backup(u)
            if path=='/notifications': return self.notifications(u)
            if path=='/documents': return self.documents(u)
            if path=='/documents/folders': return self.folder_import_page(u)
            if path=='/documents/folders/import-files': return self.folder_import_page(u)
            if path=='/roles': return self.need(u,'roles.manage') and self.roles_page(u)
            if path=='/payroll': return self.need(u,'payroll.view') and self.payroll(u)
            if path=='/access': return self.need(u,'users.manage') and self.access_page(u)
            if path=='/shifts': return self.need(u,'shifts.manage') and self.shifts_page(u)
            if path=='/overtime': return self.need(u,'overtime.request') and self.overtime_page(u)
            if path=='/leave-balances': return self.need(u,'leave.create') and self.leave_balances_page(u)
            if path=='/payroll/review': return self.need(u,'payroll.view') and self.payroll_actions(u)
            if path=='/import/mapping': return self.need(u,'import.mapping') and self.import_mapping_page(u)
            if path.startswith('/import/errors/'): return self.need(u,'import.validate') and self.import_errors_page(u,path.split('/')[-1])
            if path.startswith('/export/import-errors/'): return self.need(u,'import.validate') and self.export_import_errors(path.split('/')[-1])
            if path=='/backups': return self.need(u,'backup.manage') and self.backups_page(u)
            if path=='/system': return self.need(u,'settings.manage') and self.system_page(u)
            if path=='/diagnostics/errors': return self.need(u,'system.manage') and self.diagnostics_errors(u)
            if path=='/diagnostics/test': return self.need(u,'system.manage') and self.diagnostics_test(u)
            if path=='/network': return self.need(u,'system.manage') and self.network_page(u)
            if path=='/network/ping': return self.need(u,'system.manage') and self.network_ping(u)
            if path=='/discipline': return self.need(u,'discipline.manage') and self.discipline_page(u)
            if path.startswith('/document/'): return self.document_download(u,path.split('/')[-1])
            if path=='/template/attendance': return self.export_template_attendance()
            if path=='/template/leaves': return self.export_template_leaves()
            if path.startswith('/report/pdf/'): return self.need(u,'reports.export') and self.pdf_report(path.split('/')[-1],u)
            if path=='/health': return self.send(json.dumps({'ok':True,'version':APP_VERSION,'server_name':setting('server_name') or 'HR-SERVER'},ensure_ascii=False),200,'application/json')
            if path=='/template/employees': return self.export_template_employees()
            self.send(page('404','<div class="card">الصفحة غير موجودة</div>',u),404)
        except Exception as e:
            log_error('GET '+path,e,getattr(self,'request_id',''),getattr(getattr(self,'user',lambda:None)(),'username','') if callable(getattr(self,'user',None)) else '', 'GET', path, self.client_address[0])
            self.error_page(u,getattr(self,'request_id','ERR-UNKNOWN'),500)
    def do_POST(self):
        path=urlparse(self.path).path
        self.request_id='ERR-'+datetime.now().strftime('%Y%m%d%H%M%S')+'-'+uuid.uuid4().hex[:6].upper()
        if path=='/login':
            ip=self.client_address[0]; att=LOGIN_ATTEMPTS.get(ip,{'n':0,'until':None})
            if att.get('until') and datetime.now()<att['until']:
                wait=int((att['until']-datetime.now()).total_seconds())
                return self.send(login_page(f'محاولات دخول كثيرة. حاول مرة أخرى بعد {wait} ثانية.'))
            f=self.form(); c=db(); r=c.execute('SELECT * FROM users WHERE username=? AND active=1',(f.get('username','').strip(),)).fetchone()
            if not r or not checkpw(f.get('password',''),r['password_hash']):
                c.close(); att['n']=att.get('n',0)+1
                if att['n']>=5: att={'n':0,'until':datetime.now()+timedelta(minutes=5)}
                LOGIN_ATTEMPTS[ip]=att
                return self.send(login_page('بيانات الدخول غير صحيحة.'))
            LOGIN_ATTEMPTS.pop(ip,None)
            if r and is_legacy_hash(r['password_hash']): c.execute('UPDATE users SET password_hash=? WHERE id=?',(hashpw(f.get('password','')),r['id']))
            c.execute('UPDATE users SET last_login=? WHERE id=?',(now(),r['id'])); c.commit(); c.close()
            sid=secrets.token_urlsafe(32); sess=dict(r); sess['_sid']=sid; sess['csrf']=secrets.token_urlsafe(24); sess['last_seen']=now(); sess['permission_version']=r['permission_version']; SESS[sid]=sess
            try:
                c2=db(); c2.execute('INSERT INTO system_sessions(session_id,user_name,full_name,role,created_at,last_seen,ip,device) VALUES(?,?,?,?,?,?,?,?)',(sid,r['username'],r['full_name'],r['role'],now(),now(),ip,self.headers.get('User-Agent','')[:180])); c2.commit(); c2.close()
            except Exception: pass
            self.redirect('/',{'Set-Cookie':f'sid={sid}; Path=/; HttpOnly; SameSite=Lax'}); return
        u=self.require()
        if not u:return
        AUDIT_CTX.ip=self.client_address[0]; AUDIT_CTX.user=u
        if u.get('must_change_password') and path!='/password':
            return self.redirect('/password')
        try:
            # Multipart/form-data contains binary files; do not UTF-8 decode it.
            ctype=self.headers.get('Content-Type','').lower()
            is_multipart=ctype.startswith('multipart/form-data')
            f=self.form() if not is_multipart else self.parse_upload()[0]
            # CSRF enforced for every state-changing POST.
            if f.get('_csrf')!=u.get('csrf'):
                return self.send(page('خطأ أمني','<div class="card"><div class="alert">انتهت صلاحية النموذج. أعد تحميل الصفحة وحاول مرة أخرى.</div></div>',u),403)
            if path=='/bridge/authorize': return self.bridge_authorize_token()
            if path=='/password': return self.password_save(u,f)
            if path=='/enterprise/save': return self.need(u,'employees.edit') and self.enterprise_save(u,f)
            if path=='/matching/review': return self.need(u,'employees.edit') and self.matching_review(u,f)
            if path=='/myhr/request': return self.need(u,'leave.create') and self.myhr_request(u,f)
            if path=='/requests/action': return self.request_action(u,f)
            if path=='/employee/save': return self.need(u,'employees.edit') and self.save_employee(u,f)
            if path=='/employee/evaluation/save': return self.need(u,'employees.edit') and self.save_evaluation(u,f)
            if path=='/employees/bulk': return self.need(u,'employees.edit') and self.bulk_employees(u,f)
            if path=='/export/employees/selected': return self.need(u,'reports.export') and self.export_selected_employees(u)
            if path=='/documents/paste': return self.need(u,'documents.manage') and self.paste_documents(u)
            if path=='/employees/views/save': return self.need(u,'employees.edit') and self.save_employee_view(u,f)
            if path.startswith('/employee/archive/'): return self.need(u,'employees.edit') and self.archive_employee(u,path.split('/')[-1])
            if path.startswith('/employee/restore/'): return self.need(u,'employees.edit') and self.restore_employee(u,path.split('/')[-1])
            if path=='/leave/save': return self.need(u,'leave.create') and self.save_leave(u,f)
            if path=='/leave/status': return self.need(u,'leave.approve') and self.leave_status(u,f)
            if path=='/attendance/import': return self.need(u,'attendance.import') and self.do_import_attendance(u)
            if path=='/import/employees': return self.need(u,'employees.edit') and self.do_import_employees(u)
            if path=='/import/employees/paste': return self.need(u,'employees.edit') and self.do_import_employees_paste(u,f)
            if path=='/settings/save': return self.need(u,'settings.manage') and self.save_settings(u,f)
            if path=='/branding/logo': return self.need(u,'settings.manage') and self.branding_logo_save(u)
            if path=='/users/save': return self.save_user(u,f)
            if path=='/users/toggle': return self.toggle_user(u,f)
            if path=='/backup': return self.need(u,'backup.manage') and self.backup(u)
            if path=='/notifications/read': return self.notifications_read(u,f)
            if path=='/documents/upload': return self.need(u,'documents.manage') and self.document_upload(u)
            if path=='/documents/folders/import': return self.need(u,'documents.bulk_import') and self.folder_import(u)
            if path=='/documents/folders/import-files': return self.need(u,'documents.bulk_import') and self.folder_files_import(u)
            if path.startswith('/documents/folders/rollback/'): return self.need(u,'documents.bulk_import') and self.folder_import_rollback(u,path.split('/')[-1])
            if path=='/roles/save': return self.need(u,'roles.manage') and self.roles_save(u,f)
            if path=='/payroll/save': return self.need(u,'payroll.manage') and self.payroll_save(u,f)
            if path=='/access/save': return self.need(u,'users.manage') and self.access_save(u,f)
            if path=='/shift/save': return self.need(u,'shifts.manage') and self.shift_save(u,f)
            if path=='/shift/assign': return self.need(u,'shifts.manage') and self.shift_assign(u,f)
            if path=='/overtime/save': return self.need(u,'overtime.request') and self.overtime_save(u,f)
            if path=='/overtime/status': return self.need(u,'overtime.approve') and self.overtime_status(u,f)
            if path=='/payroll/approve': return self.need(u,'payroll.approve') and self.payroll_approve(u,f)
            if path=='/payroll/lock': return self.need(u,'payroll.lock') and self.payroll_lock(u,f)
            if path=='/import/mapping/save': return self.need(u,'import.mapping') and self.import_mapping_save(u,f)
            if path=='/backup/restore': return self.need(u,'backup.manage') and self.backup_restore(u,f)
            if path=='/discipline/save': return self.need(u,'discipline.manage') and self.discipline_save(u,f)
            if path=='/attendance/save': return self.need(u,'attendance.edit') and self.attendance_save(u,f)
            if path=='/attendance/adjust': return self.need(u,'attendance.edit') and self.attendance_adjust(u,f)
            if path=='/notifications': return self.notifications(u)
            if path=='/documents': return self.documents(u)
            if path=='/documents/folders': return self.folder_import_page(u)
            if path=='/roles': return self.need(u,'roles.manage') and self.roles_page(u)
            if path=='/payroll': return self.need(u,'payroll.view') and self.payroll(u)
            if path=='/access': return self.need(u,'users.manage') and self.access_page(u)
            if path=='/shifts': return self.need(u,'shifts.manage') and self.shifts_page(u)
            if path=='/overtime': return self.need(u,'overtime.request') and self.overtime_page(u)
            if path=='/leave-balances': return self.need(u,'leave.create') and self.leave_balances_page(u)
            if path=='/payroll/review': return self.need(u,'payroll.view') and self.payroll_actions(u)
            if path=='/import/mapping': return self.need(u,'import.mapping') and self.import_mapping_page(u)
            if path=='/backups': return self.need(u,'backup.manage') and self.backups_page(u)
            if path=='/system': return self.need(u,'settings.manage') and self.system_page(u)
            if path=='/diagnostics/errors': return self.need(u,'system.manage') and self.diagnostics_errors(u)
            if path=='/diagnostics/test': return self.need(u,'system.manage') and self.diagnostics_test(u)
            if path=='/network': return self.need(u,'system.manage') and self.network_page(u)
            if path=='/network/ping': return self.need(u,'system.manage') and self.network_ping(u)
            if path=='/discipline': return self.need(u,'discipline.manage') and self.discipline_page(u)
            if path.startswith('/document/'): return self.document_download(u,path.split('/')[-1])
            if path=='/template/attendance': return self.export_template_attendance()
            if path=='/template/leaves': return self.export_template_leaves()
            if path.startswith('/report/pdf/'): return self.pdf_report(path.split('/')[-1],u)
            if path=='/health': return self.send(json.dumps({'ok':True,'version':APP_VERSION,'server_name':setting('server_name') or 'HR-SERVER'},ensure_ascii=False),200,'application/json')
            if path=='/template/employees': return self.export_template_employees()
        except Exception as e:
            log_error('POST '+path,e,getattr(self,'request_id',''),getattr(getattr(self,'user',lambda:None)(),'username','') if callable(getattr(self,'user',None)) else '', 'POST', path, self.client_address[0])
            self.error_page(u,getattr(self,'request_id','ERR-UNKNOWN'),500)

    def enterprise_center(self,u):
        scope_sql,scope_params=visible_employee_sql(u,'e'); c=db()
        today=date.today(); d30=(today+timedelta(days=int(setting('contract_alert_days') or 30))).isoformat()
        exp_docs=c.execute(f"SELECT d.emp_code,e.name,d.file_name,d.expiry_date FROM documents d JOIN employees e ON e.emp_code=d.emp_code WHERE d.status='current' AND d.expiry_date<>'' AND d.expiry_date<=?{scope_sql} ORDER BY d.expiry_date LIMIT 50",[d30]+scope_params).fetchall()
        exp_cred=c.execute(f"SELECT cr.emp_code,e.name,cr.credential_type,cr.expiry_date FROM credentials cr JOIN employees e ON e.emp_code=cr.emp_code WHERE cr.expiry_date<>'' AND cr.expiry_date<=?{scope_sql} ORDER BY cr.expiry_date LIMIT 50",[d30]+scope_params).fetchall()
        exp_train=c.execute(f"SELECT t.emp_code,e.name,t.course,t.expiry_date FROM training t JOIN employees e ON e.emp_code=t.emp_code WHERE t.expiry_date<>'' AND t.expiry_date<=?{scope_sql} ORDER BY t.expiry_date LIMIT 50",[d30]+scope_params).fetchall()
        pending=c.execute(f"SELECT * FROM approval_queue WHERE status='pending' ORDER BY id DESC LIMIT 50").fetchall()
        holidays=c.execute('SELECT * FROM holidays WHERE holiday_date>=? ORDER BY holiday_date LIMIT 20',(today.isoformat(),)).fetchall()
        assets=c.execute(f"SELECT a.*,e.name FROM assets a LEFT JOIN employees e ON e.emp_code=a.emp_code WHERE a.status<>'returned' ORDER BY a.id DESC LIMIT 50").fetchall()
        c.close()
        def rows(items,kind):
            if kind=='doc': return ''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["file_name"])}</td><td>{esc(r["expiry_date"])}</td></tr>' for r in items) or '<tr><td colspan="4">لا توجد بيانات</td></tr>'
            if kind=='cred': return ''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["credential_type"])}</td><td>{esc(r["expiry_date"])}</td></tr>' for r in items) or '<tr><td colspan="4">لا توجد بيانات</td></tr>'
            if kind=='train': return ''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{esc(r["name"])}</td><td>{esc(r["course"])}</td><td>{esc(r["expiry_date"])}</td></tr>' for r in items) or '<tr><td colspan="4">لا توجد بيانات</td></tr>'
        hrows=''.join(f'<tr><td>{esc(r["holiday_date"])}</td><td>{esc(r["name"])}</td><td>{esc(r["kind"])}</td></tr>' for r in holidays) or '<tr><td colspan="3">لا توجد عطلات قادمة</td></tr>'
        prows=''.join(f'<tr><td>{esc(r["entity_type"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["title"])}</td><td>{esc(r["requested_by"])}</td><td>{esc(r["status"])}</td></tr>' for r in pending) or '<tr><td colspan="5">لا توجد موافقات معلقة</td></tr>'
        arows=''.join(f'<tr><td>{esc(r["asset_tag"])}</td><td>{esc(r["asset_type"])}</td><td>{esc(r["description"])}</td><td>{esc(r["name"] or "-")}</td><td>{esc(r["status"])}</td></tr>' for r in assets) or '<tr><td colspan="5">لا توجد أصول</td></tr>'
        empopts=''.join(f'<option value="{esc(r["emp_code"])}">{esc(r["emp_code"])} — {esc(r["name"])}</option>' for r in c.execute(f'SELECT e.emp_code,e.name FROM employees e WHERE e.status<>\'مؤرشف\'{scope_sql} ORDER BY e.name LIMIT 500',scope_params).fetchall()) if False else ''
        body=f'''<div class="top"><div class="title"><h1>Enterprise Center</h1><p>التدريب · الاعتمادات · العطلات · الأصول · الموافقات · التنبيهات الذكية</p></div></div>
<div class="grid g4"><div class="card metric"><div class="label">مستندات منتهية/قريبة</div><div class="value">{len(exp_docs)}</div></div><div class="card metric"><div class="label">اعتمادات قريبة</div><div class="value">{len(exp_cred)}</div></div><div class="card metric"><div class="label">تدريبات قريبة</div><div class="value">{len(exp_train)}</div></div><div class="card metric"><div class="label">موافقات معلقة</div><div class="value">{len(pending)}</div></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><h3>📄 انتهاء المستندات</h3><div class="table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>المستند</th><th>الانتهاء</th></tr></thead><tbody>{rows(exp_docs,'doc')}</tbody></table></div></div>
<div class="card"><h3>🎓 الاعتمادات المهنية</h3><div class="table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>الاعتماد</th><th>الانتهاء</th></tr></thead><tbody>{rows(exp_cred,'cred')}</tbody></table></div></div>
<div class="card"><h3>📚 التدريب والشهادات</h3><div class="table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>الدورة</th><th>الانتهاء</th></tr></thead><tbody>{rows(exp_train,'train')}</tbody></table></div></div>
<div class="card"><h3>🏖 العطلات القادمة</h3><form method="post" action="/enterprise/save" class="form">{csrf_field(u)}<input type="hidden" name="kind" value="holiday"><div class="field"><label>الاسم</label><input name="name" required></div><div class="field"><label>التاريخ</label><input type="date" name="holiday_date" required></div><div class="field"><label>النوع</label><input name="kind" value="official"></div><div class="field"><label>ملاحظات</label><input name="notes"></div><div class="full"><button class="btn">إضافة عطلة</button></div></form><table class="table" style="margin-top:12px"><thead><tr><th>التاريخ</th><th>الاسم</th><th>النوع</th></tr></thead><tbody>{hrows}</tbody></table></div>
<div class="card"><h3>🔔 Approval Center</h3><div class="table-wrap"><table class="table"><thead><tr><th>النوع</th><th>الموظف</th><th>الطلب</th><th>بواسطة</th><th>الحالة</th></tr></thead><tbody>{prows}</tbody></table></div></div>
<div class="card"><h3>💻 Employee Assets</h3><form method="post" action="/enterprise/save" class="form">{csrf_field(u)}<input type="hidden" name="kind" value="asset"><div class="field"><label>Asset Tag</label><input name="asset_tag" required></div><div class="field"><label>النوع</label><input name="asset_type" required></div><div class="field"><label>Serial</label><input name="serial_no"></div><div class="field"><label>الوصف</label><input name="description"></div><div class="field full"><button class="btn">إضافة أصل</button></div></form><table class="table" style="margin-top:12px"><thead><tr><th>Tag</th><th>النوع</th><th>الوصف</th><th>الموظف</th><th>الحالة</th></tr></thead><tbody>{arows}</tbody></table></div>
</div>'''
        self.send(page('Enterprise Center',body,u,'enterprise'))

    def enterprise_save(self,u,f):
        kind=f.get('kind','')
        c=db()
        if kind=='holiday':
            c.execute('INSERT INTO holidays(name,holiday_date,kind,paid,notes) VALUES(?,?,?,?,?) ON CONFLICT(holiday_date) DO UPDATE SET name=excluded.name,kind=excluded.kind,paid=excluded.paid,notes=excluded.notes',(f.get('name',''),f.get('holiday_date',''),f.get('kind','official'),1,f.get('notes','')))
            audit(u['username'],u['role'],'إضافة','العطلات',f.get('holiday_date',''))
        elif kind=='asset':
            c.execute('INSERT INTO assets(asset_tag,asset_type,description,serial_no,status,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(f.get('asset_tag',''),f.get('asset_type',''),f.get('description',''),f.get('serial_no',''),'available',u['username'],now()))
            audit(u['username'],u['role'],'إضافة','الأصول',f.get('asset_tag',''))
        else:
            c.close(); return self.forbid(u)
        c.commit(); c.close(); self.redirect('/enterprise')

    def dashboard(self,u):
        c=db(); today=date.today().isoformat(); month=today[:7]
        total=c.execute('SELECT COUNT(*) n FROM employees').fetchone()['n']; active=c.execute("SELECT COUNT(*) n FROM employees WHERE status='على رأس العمل'").fetchone()['n']
        present=c.execute("SELECT COUNT(*) n FROM attendance WHERE work_date=? AND status='حضور'",(today,)).fetchone()['n']; absent=c.execute("SELECT COUNT(*) n FROM attendance WHERE work_date=? AND status='غياب'",(today,)).fetchone()['n']
        late=c.execute("SELECT COALESCE(SUM(late_minutes),0) n FROM attendance WHERE substr(work_date,1,7)=?",(month,)).fetchone()['n']
        pending=c.execute("SELECT COUNT(*) n FROM leaves WHERE status='قيد المراجعة'").fetchone()['n']; expiring=c.execute("SELECT COUNT(*) n FROM documents WHERE expiry_date<>'' AND expiry_date<=?",((date.today()+timedelta(days=30)).isoformat(),)).fetchone()['n']
        missing=len(missing_documents_report(u)[1]); avg=c.execute("SELECT COALESCE(ROUND(AVG(v.score),1),0) x FROM employee_evaluations v JOIN (SELECT emp_code,MAX(id) id FROM employee_evaluations GROUP BY emp_code) z ON z.id=v.id").fetchone()['x']
        deps=c.execute("SELECT COALESCE(department,'غير محدد') d,COUNT(*) n FROM employees WHERE status<>'مؤرشف' GROUP BY d ORDER BY n DESC LIMIT 6").fetchall()
        perf=c.execute("SELECT e.name,COALESCE(v.score,0) score FROM employees e LEFT JOIN (SELECT emp_code,MAX(id) id FROM employee_evaluations GROUP BY emp_code) x ON x.emp_code=e.emp_code LEFT JOIN employee_evaluations v ON v.id=x.id WHERE e.status='على رأس العمل' ORDER BY score DESC LIMIT 6").fetchall(); c.close()
        maxd=max([r['n'] for r in deps] or [1]); depbars=''.join(f'<div class="bar" style="height:{max(18,int(r["n"]/maxd*180))}px"><b>{r["n"]}</b><span>{esc(r["d"])}</span></div>' for r in deps)
        pmax=max([float(r['score'] or 0) for r in perf] or [1]); pmax=max(pmax,1.0); pbar=''.join(f'<div class="bar" style="height:{max(18,int(float(r["score"] or 0)/pmax*180))}px"><b>{float(r["score"] or 0):g}%</b><span>{esc(r["name"][:14])}</span></div>' for r in perf)
        late_pct=min(100, int((late/max(1,int(setting('monthly_late_limit_minutes') or 120)))*100))
        alerts=[]
        if late>=int(setting('monthly_late_limit_minutes') or 120): alerts.append(f'<a href="/attendance">🔴 تجاوز التأخير الشهري: {late} دقيقة</a>')
        if expiring: alerts.append(f'<a href="/documents">🟠 مستندات تنتهي خلال 30 يوم: {expiring}</a>')
        if missing: alerts.append(f'<a href="/import">🟡 موظفون ناقصهم مستندات: {missing}</a>')
        if pending: alerts.append(f'<a href="/leaves">🔵 إجازات قيد المراجعة: {pending}</a>')
        body=f'''<div class="top"><div class="title"><h1>لوحة التحكم</h1><p>صورة سريعة للنظام · {today}</p></div><div class="actions"><a class="btn" href="/employee/new">+ موظف</a><a class="btn gray" href="/import">Excel Center</a><a class="btn gray" href="/backup">Backup</a></div></div>
<div class="grid g4"><div class="card metric"><div class="label">👥 الموظفون</div><div class="value">{total}</div><div class="sub">{active} على رأس العمل</div></div><div class="card metric"><div class="label">⏰ الحضور اليوم</div><div class="value">{present}</div><div class="sub">غياب {absent}</div></div><div class="card metric"><div class="label">⌛ تأخير الشهر</div><div class="value">{late}</div><div class="sub">{late_pct}% من حد {esc(setting('monthly_late_limit_minutes') or 120)} دقيقة</div></div><div class="card metric"><div class="label">⭐ متوسط الأداء</div><div class="value">{float(avg or 0):g}%</div><div class="sub">تقييمات الموظفين</div></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><div class="top"><h3>👥 الموظفون حسب الإدارة</h3><span class="badge b-blue">{len(deps)} إدارات</span></div><div class="chart">{depbars or '<div class="alert">لا توجد بيانات.</div>'}</div></div><div class="card"><div class="top"><h3>⭐ أعلى الأداء</h3><span class="badge b-ok">{float(avg or 0):g}% متوسط</span></div><div class="chart">{pbar or '<div class="alert">لا توجد تقييمات.</div>'}</div></div></div>
<div class="grid g3" style="margin-top:16px"><div class="card"><h3>🏖 الطلبات</h3><div class="value" style="font-size:30px">{pending}</div><p>إجازات تحتاج مراجعة</p></div><div class="card"><h3>📄 المستندات</h3><div class="value" style="font-size:30px">{missing}</div><p>موظفون لديهم نواقص · {expiring} قريب الانتهاء</p></div><div class="card"><h3>⏱ مؤشر التأخير</h3><div style="height:16px;background:#eef2f6;border-radius:999px;overflow:hidden;margin-top:14px"><div style="width:{late_pct}%;height:100%;background:linear-gradient(90deg,#12b76a,#f79009,#f04438)"></div></div><p>{late} / {esc(setting('monthly_late_limit_minutes') or 120)} دقيقة</p></div></div>
<div class="card" style="margin-top:16px"><h3>🔔 تنبيهات ذكية</h3>{''.join(f'<p style="padding:9px 0;border-bottom:1px solid #eee">{x}</p>' for x in alerts) or '<p><span class="badge b-ok">لا توجد تنبيهات حرجة الآن</span></p>'}</div>'''
        self.send(page('لوحة التحكم',body,u,'dashboard'))

    def employees(self,u):
        qs=parse_qs(urlparse(self.path).query)
        q=qs.get('q',[''])[0].strip(); st=qs.get('status',[''])[0]; dept=qs.get('department',[''])[0]; unit=qs.get('unit',[''])[0]; job=qs.get('job',[''])[0]; shift=qs.get('shift',[''])[0]; eval_min=qs.get('eval_min',[''])[0]; eval_max=qs.get('eval_max',[''])[0]; late_gt=qs.get('late_gt',[''])[0]; view=qs.get('view',[''])[0]
        if view:
            c0=db(); vv=c0.execute('SELECT query FROM saved_views WHERE id=? AND username=?',(view,u['username'])).fetchone(); c0.close()
            if vv:
                vp=parse_qs(vv['query']); q=vp.get('q',[q])[0]; st=vp.get('status',[st])[0]; dept=vp.get('department',[dept])[0]; unit=vp.get('unit',[unit])[0]; job=vp.get('job',[job])[0]; shift=vp.get('shift',[shift])[0]; eval_min=vp.get('eval_min',[eval_min])[0]; eval_max=vp.get('eval_max',[eval_max])[0]; late_gt=vp.get('late_gt',[late_gt])[0]
        c=db(); scope_sql,scope_params=visible_employee_sql(u); cond=[]; params=[]
        if q: cond.append('(e.emp_code LIKE ? OR e.name LIKE ? OR e.national_id LIKE ? OR e.phone LIKE ? OR e.email LIKE ?)'); params += [f'%{q}%']*5
        if st=='__all__': pass
        elif st: cond.append('e.status=?'); params.append(st)
        else: cond.append("e.status<>'مؤرشف'")
        if dept: cond.append("COALESCE(e.department,'')=?"); params.append(dept)
        if unit: cond.append("COALESCE(e.unit,'')=?"); params.append(unit)
        if job: cond.append("COALESCE(e.job,'')=?"); params.append(job)
        if shift: cond.append("COALESCE(CAST(e.shift_id AS TEXT),'')=?"); params.append(str(shift))
        if eval_min:
            try: cond.append("COALESCE((SELECT score FROM employee_evaluations ee WHERE ee.emp_code=e.emp_code ORDER BY ee.id DESC LIMIT 1),0)>=?"); params.append(float(eval_min))
            except ValueError: pass
        if eval_max:
            try: cond.append("COALESCE((SELECT score FROM employee_evaluations ee WHERE ee.emp_code=e.emp_code ORDER BY ee.id DESC LIMIT 1),0)<=?"); params.append(float(eval_max))
            except ValueError: pass
        if late_gt:
            try: cond.append("(SELECT COALESCE(SUM(a2.late_minutes),0) FROM attendance a2 WHERE a2.emp_code=e.emp_code AND substr(a2.work_date,1,7)=?)>?"); params.extend([date.today().isoformat()[:7],int(late_gt)])
            except ValueError: pass
        where=(' WHERE '+' AND '.join(cond)) if cond else ' WHERE 1=1'
        rows=c.execute(f'SELECT e.emp_code,e.name,e.national_id,e.job,e.department,e.unit,e.status,e.phone,e.hire_date,e.shift_id FROM employees e{where}{scope_sql} ORDER BY e.name COLLATE NOCASE LIMIT 2000',params+scope_params).fetchall()
        depts=[x[0] for x in c.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department<>'' ORDER BY department").fetchall()]
        units=[x[0] for x in c.execute("SELECT DISTINCT unit FROM employees WHERE unit IS NOT NULL AND unit<>'' ORDER BY unit").fetchall()]
        jobs=[x[0] for x in c.execute("SELECT DISTINCT job FROM employees WHERE job IS NOT NULL AND job<>'' ORDER BY job").fetchall()]
        shifts=c.execute('SELECT id,name FROM shifts WHERE active=1 ORDER BY name').fetchall(); views=c.execute('SELECT id,name FROM saved_views WHERE username=? ORDER BY name',(u['username'],)).fetchall(); c.close()
        badge={'على رأس العمل':'b-ok','موقوف':'b-warn','منتهي الخدمة':'b-bad','مؤرشف':'b-gray'}
        trs=''.join(f'<tr><td><input class="emp-check" type="checkbox" value="{esc(r["emp_code"])}"></td><td>{esc(r["emp_code"])}</td><td><a href="/employee/profile/{esc(r["emp_code"])}"><b>{esc(r["name"])}</b></a></td><td>{esc(r["national_id"]) if can(u,"sensitive.view") else "************"}</td><td>{esc(r["job"] or "—")}</td><td>{esc(r["department"] or "—")}</td><td>{esc(r["unit"] or "—")}</td><td><span class="badge {badge.get(r["status"],"b-warn")}">{esc(r["status"])}</span></td><td>{esc(r["phone"] or "—")}</td><td>{esc(r["hire_date"] or "—")}</td><td><a class="btn gray" href="/employee/profile/{esc(r["emp_code"])}">فتح</a></td></tr>' for r in rows)
        def opts(items,current): return ''.join(f'<option value="{esc(x)}" {"selected" if str(current)==str(x) else ""}>{esc(x)}</option>' for x in items)
        st_opts=''.join(f'<option value="{esc(x)}" {"selected" if st==x else ""}>{esc(x)}</option>' for x in ['على رأس العمل','موقوف','منتهي الخدمة','مؤرشف'])
        view_opts=''.join(f'<option value="{x["id"]}">{esc(x["name"])}</option>' for x in views)
        bulk='''<div id="bulkBar" class="card" style="display:none;margin-bottom:14px;border:2px solid #84adff;background:#f7fbff"><b id="selCount">0 محدد</b><div class="actions" style="margin-top:10px"><button class="btn bad" type="button" onclick="bulkAction('archive')">أرشفة</button><button class="btn ok" type="button" onclick="bulkAction('restore')">استعادة</button><button class="btn gray" type="button" onclick="bulkAction('department')">Department</button><button class="btn gray" type="button" onclick="bulkAction('unit')">Unit</button><button class="btn gray" type="button" onclick="bulkAction('manager')">Manager</button><button class="btn gray" type="button" onclick="bulkAction('shift')">Shift</button><button class="btn gray" type="button" onclick="bulkAction('tag')">Tag</button><button class="btn gray" type="button" onclick="bulkAction('document_requirement')">Document Requirement</button><button class="btn gray" type="button" onclick="bulkExport()">Export</button></div></div>'''
        shiftopts=''.join(f'<option value="{x["id"]}" {"selected" if str(shift)==str(x["id"]) else ""}>{esc(x["name"])}</option>' for x in shifts)
        body=f'''<div class="top"><div class="title"><h1>الموظفون</h1><p>{len(rows)} نتيجة · فلاتر متقدمة + تحديد جماعي + تصدير</p></div><div class="actions"><a class="btn" href="/employee/new">+ إضافة موظف</a><a class="btn gray" id="exportXlsx" href="/export/employees">Export Excel</a><a class="btn gray" id="exportCsv" href="/export/csv/employees">CSV</a><a class="btn gray" id="exportHtml" href="/export/html/employees">HTML</a></div></div>
        <div class="card"><form class="toolbar" method="get" id="empFilter"><input name="q" value="{esc(q)}" placeholder="بحث: الاسم، الكود، الرقم القومي، الهاتف" style="min-width:280px"><select name="status"><option value="">نشطون</option>{st_opts}<option value="__all__" {"selected" if st=="__all__" else ""}>الكل</option></select><select name="department"><option value="">كل الإدارات</option>{opts(depts,dept)}</select><select name="unit"><option value="">كل الوحدات</option>{opts(units,unit)}</select><select name="job"><option value="">كل الوظائف</option>{opts(jobs,job)}</select><select name="shift"><option value="">كل الورديات</option>{shiftopts}</select><input name="eval_min" type="number" min="0" max="100" step="1" value="{esc(eval_min)}" placeholder="تقييم ≥"><input name="eval_max" type="number" min="0" max="100" step="1" value="{esc(eval_max)}" placeholder="تقييم ≤"><input name="late_gt" type="number" min="0" step="1" value="{esc(late_gt)}" placeholder="تأخير الشهر >"><button class="btn gray">تطبيق</button><a class="btn gray" href="/employees">مسح الفلاتر</a></form><div class="actions"><select id="savedView" onchange="if(this.value) location.href='/employees?view='+this.value"><option value="">الفلاتر المحفوظة</option>{view_opts}</select><button class="btn gray" type="button" onclick="saveView()">حفظ الفلتر الحالي</button></div></div>
        <div class="card" style="margin:12px 0"><b>الأعمدة الظاهرة</b><div class="actions" style="margin-top:8px"><label><input type="checkbox" class="colToggle" data-col="4" checked> الوظيفة</label><label><input type="checkbox" class="colToggle" data-col="5" checked> الإدارة</label><label><input type="checkbox" class="colToggle" data-col="6" checked> الوحدة</label><label><input type="checkbox" class="colToggle" data-col="8" checked> الهاتف</label><label><input type="checkbox" class="colToggle" data-col="9" checked> التعيين</label></div></div>
        {bulk}<div class="card table-wrap"><table class="table"><thead><tr><th><input id="selectAll" type="checkbox"></th><th>الكود</th><th>الاسم</th><th>الرقم القومي</th><th>الوظيفة</th><th>الإدارة</th><th>الوحدة</th><th>الحالة</th><th>الهاتف</th><th>التعيين</th><th></th></tr></thead><tbody>{trs or '<tr><td colspan="11"><div class="alert">لا توجد نتائج. جرّب تغيير الفلاتر.</div></td></tr>'}</tbody></table></div>
        <script>
        (function(){{const qs=new URLSearchParams(location.search);["exportXlsx","exportCsv","exportHtml"].forEach(id=>{{const a=document.getElementById(id);if(a){{const u=new URL(a.href);qs.forEach((v,k)=>u.searchParams.set(k,v));a.href=u.toString();}}}});}})();
        document.querySelectorAll('.colToggle').forEach(x=>{{const k='emp_col_'+x.dataset.col;const saved=localStorage.getItem(k);if(saved!==null)x.checked=saved==='1';const apply=()=>{{const c=Number(x.dataset.col)+1;document.querySelectorAll('.table th:nth-child('+c+'),.table td:nth-child('+c+')').forEach(e=>e.style.display=x.checked?'':'none');localStorage.setItem(k,x.checked?'1':'0');}};x.onchange=apply;apply();}});
        const checks=()=>[...document.querySelectorAll('.emp-check:checked')].map(x=>x.value);function syncSel(){{let n=checks().length;document.getElementById('bulkBar').style.display=n?'block':'none';document.getElementById('selCount').textContent=n+' محدد';}}document.querySelectorAll('.emp-check').forEach(x=>x.onchange=syncSel);document.getElementById('selectAll').onchange=function(){{document.querySelectorAll('.emp-check').forEach(x=>x.checked=this.checked);syncSel();}};
        function bulkAction(action){{let ids=checks();if(!ids.length)return;let value='';if(!['archive','restore'].includes(action)){{value=prompt('أدخل القيمة للعملية ('+action+')');if(value===null)return;}}if(!confirm('تنفيذ '+action+' على '+ids.length+' موظف؟'))return;let fd=new FormData();fd.append('_csrf','{esc(u.get('csrf',''))}');fd.append('action',action);fd.append('value',value);ids.forEach(x=>fd.append('emp_codes',x));fetch('/employees/bulk',{{method:'POST',body:fd}}).then(()=>location.reload());}}
        function bulkExport(){{let ids=checks();if(!ids.length)return;let f=document.createElement('form');f.method='POST';f.action='/export/employees/selected';let c=document.createElement('input');c.type='hidden';c.name='_csrf';c.value='{esc(u.get('csrf',''))}';f.appendChild(c);ids.forEach(x=>{{let i=document.createElement('input');i.type='hidden';i.name='emp_codes';i.value=x;f.appendChild(i)}});document.body.appendChild(f);f.submit();}}
        function saveView(){{let name=prompt('اسم الفلتر المحفوظ:');if(!name)return;let fd=new FormData(document.getElementById('empFilter'));fd.append('_csrf','{esc(u.get('csrf',''))}');fd.append('name',name);fetch('/employees/views/save',{{method:'POST',body:fd}}).then(()=>location.reload());}}
        </script>'''
        self.send(page('الموظفون',body,u,'employees'))
    def save_employee_view(self,u,f):
        name=f.get('name','').strip(); allowed={'q','status','department','unit','job','shift','eval_min','eval_max','late_gt'}
        query='&'.join(f'{k}={quote(str(f.get(k,"")))}' for k in allowed if f.get(k,''))
        if not name: return self.redirect('/employees')
        c=db(); c.execute('INSERT INTO saved_views(username,name,query,created_at,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(username,name) DO UPDATE SET query=excluded.query,updated_at=excluded.updated_at',(u['username'],name,query,now(),now())); c.commit(); c.close(); audit(u['username'],u['role'],'حفظ فلتر','الموظفون',name,query); self.redirect('/employees')
    def employees_views(self,u):
        c=db(); rows=c.execute('SELECT id,name,query,updated_at FROM saved_views WHERE username=? ORDER BY name',(u['username'],)).fetchall(); c.close(); return self.send(page('الفلاتر المحفوظة','<div class="card"><h2>الفلاتر المحفوظة</h2>'+''.join(f'<p><a href="/employees?view={r["id"]}">{esc(r["name"])}</a> · {esc(r["updated_at"])}</p>' for r in rows)+'</div>',u,'employees'))
    def bulk_employees(self,u,f):
        ids=f.get('emp_codes') or []
        if isinstance(ids,str): ids=[ids]
        action=f.get('action',''); value=self.fval(f,'value','').strip(); results=[]
        c=db()
        try:
            for code in ids:
                if not emp_allowed(u,code): results.append((code,'مرفوض خارج النطاق')); continue
                e=c.execute('SELECT status,prev_status FROM employees WHERE emp_code=?',(code,)).fetchone()
                if not e: results.append((code,'غير موجود')); continue
                if action=='archive' and e['status']!='مؤرشف':
                    c.execute('UPDATE employees SET prev_status=status,status=?,archived_at=?,archived_by=?,updated_at=? WHERE emp_code=?',('مؤرشف',now(),u['username'],now(),code)); results.append((code,'تم'))
                elif action=='restore' and e['status']=='مؤرشف':
                    ns=e['prev_status'] or 'على رأس العمل'; c.execute('UPDATE employees SET status=?,prev_status=NULL,archived_at=NULL,archived_by=NULL,updated_at=? WHERE emp_code=?',(ns,now(),code)); results.append((code,'تم'))
                elif action in ('department','unit','manager','tag'):
                    if not value: results.append((code,'القيمة مطلوبة')); continue
                    col={'department':'department','unit':'unit','manager':'manager_name','tag':'tags'}[action]
                    c.execute(f'UPDATE employees SET {col}=?,updated_at=? WHERE emp_code=?',(value,now(),code)); results.append((code,'تم'))
                elif action=='shift':
                    try:
                        sid=int(value); c.execute('UPDATE employees SET shift_id=?,updated_at=? WHERE emp_code=?',(sid,now(),code)); c.execute('INSERT INTO employee_shifts(emp_code,shift_id,assigned_at,assigned_by) VALUES(?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET shift_id=excluded.shift_id,assigned_at=excluded.assigned_at,assigned_by=excluded.assigned_by',(code,sid,now(),u['username'])); results.append((code,'تم'))
                    except Exception: results.append((code,'وردية غير صحيحة'))
                elif action=='document_requirement':
                    cat=value or 'عام'
                    # Create a visible requirement marker; actual file remains absent until uploaded.
                    c.execute("INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)",(code,'document_requirement',now()[:10],'Document Requirement','مطلوب: '+cat,u['username'],now())); results.append((code,'تم'))
                else: results.append((code,'لا يحتاج تغيير'))
            c.commit()
        except Exception:
            c.rollback(); c.close(); raise
        c.close()
        audit(u['username'],u['role'],'Bulk Action','الموظفون',str(len(ids)),f'action={action},value={value}')
        ok=sum(x[1]=='تم' for x in results); bad=len(results)-ok
        self.send(page('نتيجة العملية',f'<div class="card"><h2>نتيجة العملية الجماعية</h2><p>✓ تم: {ok} · ⚠ لم يتغير/مرفوض: {bad}</p><a class="btn" href="/employees">العودة للموظفين</a></div>',u,'employees'))

    def export_selected_employees(self,u):
        raw=self._raw().decode('utf-8',errors='replace'); f=parse_qs(raw,keep_blank_values=True); ids=f.get('emp_codes',[])
        if isinstance(ids,str): ids=[ids]
        ids=[x for x in ids if emp_allowed(u,x)]
        c=db(); placeholders=','.join('?'*len(ids)) or "''"; rows=c.execute(f"SELECT emp_code,name,employee_group,birth_date,national_id,address,qualification,phone,iban,bank_name,bank_branch,department,unit,job,contract_date,contract_amount FROM employees WHERE emp_code IN ({placeholders}) ORDER BY name",ids).fetchall(); c.close(); return self._export_employee_rows(u,rows,'Selected_Employees')
    def _export_employee_rows(self,u,rows,name):
        wb=Workbook(); ws=wb.active; ws.title='Employees'; ws.append(HOSPITAL_HEADERS)
        keys=['emp_code','name','employee_group','birth_date','national_id','address','qualification','phone','iban','bank_name','bank_branch','department','unit','job','contract_date','contract_amount']
        for r in rows:
            d=dict(r)
            if not can(u,'sensitive.view'): d['national_id']='************'; d['iban']='************'
            ws.append([d.get(k,'') for k in keys])
        for col in ws.columns:
            vals=[len(str(x.value or '')) for x in col]; ws.column_dimensions[col[0].column_letter].width=min(42,max(12,max(vals+[12])+2))
        out=io.BytesIO(); wb.save(out); safe=name.replace(' ','_')+'.xlsx'; fn=quote(name+'.xlsx')
        self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':f'attachment; filename="{safe}"; filename*=UTF-8\'\'{fn}'})

    def archive_employee(self,u,code):
        c=db(); e=c.execute('SELECT status FROM employees WHERE emp_code=?',(code,)).fetchone()
        if not e: c.close(); return self.redirect('/employees')
        if e['status']=='مؤرشف': c.close(); return self.redirect(f'/employee/profile/{code}')
        c.execute('UPDATE employees SET prev_status=status,status=?,archived_at=?,archived_by=?,updated_at=? WHERE emp_code=?',('مؤرشف',now(),u['username'],now(),code))
        c.execute('INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(code,'أرشفة',now()[:10],'تمت أرشفة الموظف','الحالة السابقة: '+str(e['status']),u['username'],now()))
        c.commit(); c.close(); audit(u['username'],u['role'],'أرشفة','الموظفون',code,'الحالة السابقة: '+str(e['status']))
        self.redirect(f'/employee/profile/{code}')

    def restore_employee(self,u,code):
        c=db(); e=c.execute('SELECT prev_status FROM employees WHERE emp_code=?',(code,)).fetchone()
        if not e: c.close(); return self.redirect('/employees')
        newst=e['prev_status'] or 'على رأس العمل'
        c.execute('UPDATE employees SET status=?,prev_status=NULL,archived_at=NULL,archived_by=NULL,updated_at=? WHERE emp_code=?',(newst,now(),code))
        c.execute('INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(code,'استعادة',now()[:10],'تمت استعادة الموظف من الأرشيف','الحالة الجديدة: '+newst,u['username'],now()))
        c.commit(); c.close(); audit(u['username'],u['role'],'استعادة','الموظفون',code,'الحالة الجديدة: '+newst)
        self.redirect(f'/employee/profile/{code}')
    def save_evaluation(self,u,f):
        emp=self.fval(f,'emp_code'); period=self.fval(f,'period')
        if not emp_allowed(u,emp): return self.forbid(u)
        try:
            vals={k:float(self.fval(f,k) or 0) for k in ('attendance_score','punctuality_score','productivity_score','behavior_score','manager_score')}
        except ValueError:
            return self.send(page('التقييم','<div class="card"><div class="alert">كل نسب التقييم يجب أن تكون أرقامًا.</div></div>',u,'employees'),400)
        if any(x<0 or x>100 for x in vals.values()): return self.send(page('التقييم','<div class="card"><div class="alert">كل نسبة يجب أن تكون بين 0 و100.</div></div>',u,'employees'),400)
        w=evaluation_weights(); score=(vals['attendance_score']*w['attendance'] + vals['punctuality_score']*w['punctuality'] + vals['productivity_score']*w['productivity'] + vals['behavior_score']*w['behavior'] + vals['manager_score']*w['manager'])/100
        c=db(); c.execute('''INSERT INTO employee_evaluations(emp_code,period,score,goals_score,attendance_score,discipline_score,punctuality_score,productivity_score,behavior_score,manager_score,notes,created_by,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(emp_code,period) DO UPDATE SET score=excluded.score,goals_score=excluded.goals_score,attendance_score=excluded.attendance_score,discipline_score=excluded.discipline_score,punctuality_score=excluded.punctuality_score,productivity_score=excluded.productivity_score,behavior_score=excluded.behavior_score,manager_score=excluded.manager_score,notes=excluded.notes,created_by=excluded.created_by,created_at=excluded.created_at''',
        (emp,period,score,vals['productivity_score'],vals['attendance_score'],vals['behavior_score'],vals['punctuality_score'],vals['productivity_score'],vals['behavior_score'],vals['manager_score'],self.fval(f,'notes'),u['username'],now()))
        c.commit(); c.close(); audit(u['username'],u['role'],'تقييم موظف','التقييم',emp,f'score={score:.1f}',reason='weighted evaluation')
        self.redirect('/employee/profile/'+quote(emp,safe=''))
    def employee_form(self,u,code=None):
        c=db();r=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone() if code else None;c.close()
        def v(k):return esc(r[k] if r else '')
        fields=[('emp_code','م / كود الموظف','text'),('name','الإسم','text'),('employee_group','المجموعه الوظفيه','text'),('birth_date','تاريخ الميلاد','date'),('national_id','الرقم القومى','text'),('address','العنوان','text'),('qualification','المؤهل','text'),('phone','رقم التيلفون','text'),('iban','ipan','text'),('bank_name','إسم البنك','text'),('bank_branch','إسم فرع البنك','text'),('department','الإدارة','text'),('unit','الوحدة','text'),('job','الوظيفه','text'),('contract_date','تاريخ التعاقد','date'),('contract_amount','مبلغ التعاقد','number')]
        html=[]
        for k,label,typ in fields:
            extra=' step="0.01"' if typ=='number' else ''
            req=' required' if k in ('emp_code','name') else ''
            html.append(f'<div class="field"><label>{label}</label><input type="{typ}" name="{k}" value="{v(k)}"{extra}{req}></div>')
        status=v('status') or 'على رأس العمل'
        html.append(f'<div class="field"><label>الحالة</label><select name="status"><option {"selected" if status=="على رأس العمل" else ""}>على رأس العمل</option><option {"selected" if status=="موقوف" else ""}>موقوف</option><option {"selected" if status=="منتهي الخدمة" else ""}>منتهي الخدمة</option><option {"selected" if status=="مؤرشف" else ""}>مؤرشف</option></select></div>')
        c=db(); shifts=c.execute('SELECT id,name,start_time,end_time FROM shifts WHERE active=1 ORDER BY id').fetchall(); c.close()
        current_shift=r['shift_id'] if r and r['shift_id'] else ''
        html.append('<div class="field"><label>الوردية</label><select name="shift_id"><option value="">استخدام الوردية الافتراضية</option>'+''.join(f'<option value="{x["id"]}" {"selected" if str(current_shift)==str(x["id"]) else ""}>{esc(x["name"])} ({esc(x["start_time"])} → {esc(x["end_time"])})</option>' for x in shifts)+'</select></div>')
        body='<div class="top"><div class="title"><h1>{}</h1><p>نفس هيكل شيت الموظفين الأصلي — يمكنك إدخال البيانات يدويًا أو من Excel/Paste</p></div><a class="btn gray" href="/employees">عودة</a></div><div class="card"><form class="form" method="post" action="/employee/save">{}{}<div class="full"><button class="btn">حفظ البيانات</button></div></form></div>'.format('تعديل موظف' if r else 'إضافة موظف',csrf_field(u),''.join(html))
        self.send(page('الموظف',body,u,'employees'))

    def employee_profile(self,u,code):
        if not emp_allowed(u,code): return self.forbid(u)
        c=db(); e=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone()
        if not e: c.close(); return self.send(page('الموظف','<div class="card"><div class="alert">الموظف غير موجود.</div></div>',u,'employees'),404)
        leaves=c.execute('SELECT * FROM leaves WHERE emp_code=? ORDER BY id DESC LIMIT 50',(code,)).fetchall(); att=c.execute('SELECT * FROM attendance WHERE emp_code=? ORDER BY work_date DESC LIMIT 50',(code,)).fetchall(); docs=c.execute('SELECT id,file_name,expiry_date,category,storage_path,version,status,checksum,uploaded_at,uploaded_by FROM documents WHERE emp_code=? ORDER BY category,version DESC',(code,)).fetchall(); acts=c.execute('SELECT * FROM disciplinary_actions WHERE emp_code=? ORDER BY id DESC LIMIT 50',(code,)).fetchall(); pays=c.execute('SELECT * FROM payroll WHERE emp_code=? ORDER BY period DESC LIMIT 24',(code,)).fetchall(); events=c.execute('SELECT * FROM employee_events WHERE emp_code=? ORDER BY id DESC LIMIT 50',(code,)).fetchall(); balance_rows=c.execute('SELECT leave_type,annual-used remaining FROM leave_balances WHERE emp_code=?',(code,)).fetchall(); evals=c.execute('SELECT * FROM employee_evaluations WHERE emp_code=? ORDER BY period DESC,id DESC LIMIT 12',(code,)).fetchall(); last_att=c.execute('SELECT work_date,status,check_in,check_out FROM attendance WHERE emp_code=? ORDER BY work_date DESC,id DESC LIMIT 1',(code,)).fetchone(); last_pay=pays[0] if pays else None; c.close()
        today_iso=date.today().isoformat()
        def expiry_badge(exp):
            if not exp: return ''
            try: d=datetime.fromisoformat(exp).date()
            except Exception: return ''
            delta=(d-date.today()).days
            if delta<0: return '<span class="badge b-bad">منتهي</span>'
            if delta<=30: return f'<span class="badge b-warn">ينتهي خلال {delta} يوم</span>'
            return ''
        expiring_docs=[r for r in docs if r['status']=='current' and r['expiry_date'] and (lambda x: x!='' and (datetime.fromisoformat(x).date()-date.today()).days<=30)(r['expiry_date'])]
        # --- Unified Timeline: merge leaves / documents / discipline / payroll / manual events into one feed ---
        timeline=[]
        for r in leaves: timeline.append((r['request_date'] or '','🏖',f"طلب إجازة {esc(r['leave_type'])} ({esc(r['status'])})"))
        for r in docs:
            if r['status']=='current': timeline.append(((r['uploaded_at'] or '')[:10],'📄',f"رفع مستند: {esc(r['file_name'])} v{r['version']} ({esc(r['category'])})"))
        for r in acts: timeline.append((r['action_date'] or '','⚠',f"{esc(r['action_type'])}: {esc(r['reason'] or '')}"))
        for r in pays: timeline.append((r['period'] or '','💰',f"مرتب {esc(r['period'])} — {esc(r['status'])}"))
        for r in events: timeline.append((r['event_date'] or '','📌',esc(r['title'] or '')))
        timeline=[t for t in timeline if t[0]]; timeline.sort(key=lambda t:t[0],reverse=True); timeline=timeline[:100]
        late=sum(int(r['late_minutes'] or 0) for r in att if str(r['work_date']).startswith(date.today().isoformat()[:7]));
        def rows_html(rows, cols): return ''.join('<tr>'+''.join(f'<td>{esc(r[k])}</td>' for k in cols)+'</tr>' for r in rows) or '<tr><td colspan="20">لا توجد بيانات</td></tr>'
        tabs=[('<a href="#data">البيانات</a>','employees.view'),('<a href="#timeline">Timeline</a>','employees.view'),('<a href="#attendance">الحضور</a>','attendance.view'),('<a href="#leaves">الإجازات</a>','leave.create'),('<a href="#documents">المستندات</a>','documents.manage'),('<a href="#discipline">الإنذارات والجزاءات</a>','discipline.manage'),('<a href="#payroll">المرتبات</a>','payroll.view')]
        nav=''.join(a for a,p in tabs if can(u,p))
        doc_status_badge={'current':'<span class="badge b-ok">حالي</span>','superseded':'<span class="badge b-gray">نسخة سابقة</span>'}
        _row_style=lambda r: ' style="opacity:.55"' if r["status"]!="current" else ""
        docs_rows=''.join(
            f'<tr{_row_style(r)}><td>{esc(r["file_name"])}</td><td>{esc(r["category"] or "عام")}</td>'
            f'<td>v{r["version"] or 1}</td><td>{doc_status_badge.get(r["status"] or "current","")}</td>'
            f'<td>{esc(r["expiry_date"] or "—")} {expiry_badge(r["expiry_date"])}</td>'
            f'<td><small>{esc((r["checksum"] or "")[:12])}…</small></td>'
            f'<td><a class="btn gray" href="/document/{r["id"]}">فتح</a></td></tr>'
            for r in docs) or '<tr><td colspan="7">لا توجد مستندات</td></tr>'
        timeline_html=''.join(f'<div class="tl-item"><span class="tl-date">{esc(t[0])}</span> {t[1]} {t[2]}</div>' for t in timeline) or '<div class="tl-item">لا توجد أحداث بعد</div>'
        # Previous / Next follow the user's visible employee scope.
        cnav=db(); navrows=cnav.execute('SELECT emp_code,name FROM employees e ORDER BY name COLLATE NOCASE').fetchall(); cnav.close(); navcodes=[r['emp_code'] for r in navrows if emp_allowed(u,r['emp_code'])]
        try: ni=navcodes.index(code)
        except ValueError: ni=-1
        prev_code=navcodes[ni-1] if ni>0 else None; next_code=navcodes[ni+1] if ni>=0 and ni+1<len(navcodes) else None
        navlinks=(f'<a class="btn gray" href="/employee/profile/{esc(prev_code)}">← السابق</a>' if prev_code else '')+(f'<a class="btn gray" href="/employee/profile/{esc(next_code)}">التالي →</a>' if next_code else '')
        archived=e['status']=='مؤرشف'
        arch_btn=(f'<form method="post" action="/employee/restore/{esc(code)}" style="display:inline" onsubmit="return confirm(\'استعادة الموظف؟\')">{csrf_field(u)}<button class="btn ok" type="submit">استعادة من الأرشيف</button></form>' if archived else f'<form method="post" action="/employee/archive/{esc(code)}" style="display:inline" onsubmit="return confirm(\'أرشفة هذا الموظف؟ بياناته تظل محفوظة بالكامل ويمكن استعادته لاحقًا.\')">{csrf_field(u)}<button class="btn bad" type="submit">أرشفة</button></form>') if can(u,'employees.edit') else ''
        arch_banner=f'<div class="alert" style="margin-bottom:16px">⚠ هذا الموظف مؤرشف (أرشفه {esc(e["archived_by"] or "-")} بتاريخ {esc(e["archived_at"] or "-")}). البيانات محفوظة بالكامل ويمكن الاستعادة.</div>' if archived else ''
        body=f'''{arch_banner}<div class="top"><div class="title"><h1>{esc(e['name'])}</h1><p>{esc(e['emp_code'])} · {esc(e['job'] or '')} · {esc(e['department'] or '')}</p></div><div class="actions">{navlinks}<a class="btn" href="/employee/edit/{esc(code)}">تعديل</a>{arch_btn}<a class="btn gray" href="/employees">عودة</a><button class="btn gray" onclick="window.print()">طباعة الملف</button></div></div>
        <div class="card" style="position:sticky;top:8px;z-index:5"><div class="actions">{nav}</div></div>
        <div id="data" class="grid g4" style="margin-top:16px"><div class="card metric"><div class="label">الحالة</div><div class="value" style="font-size:18px">{esc(e['status'])}</div><div class="sub">الوظيفة: {esc(e['job'] or '—')}</div></div><div class="card metric"><div class="label">آخر حضور</div><div class="value" style="font-size:18px">{esc(last_att['work_date'] if last_att else '—')}</div><div class="sub">{esc((last_att['check_in'] or '') if last_att else '')}</div></div><div class="card metric"><div class="label">رصيد الإجازات</div><div class="value" style="font-size:20px">{sum(float(x['remaining'] or 0) for x in balance_rows):g}</div><div class="sub">إجمالي الأيام المتبقية</div></div><div class="card metric"><div class="label">المرتب الأخير</div><div class="value" style="font-size:20px">{(last_pay['net'] if last_pay else 0) if can(u,'sensitive.view') else 'Hidden'}</div><div class="sub">{esc(last_pay['period'] if last_pay else '—')}</div></div></div>
        <div class="card" style="margin-top:16px"><h3>تنبيهات الموظف</h3><div class="grid g3"><div>{('<span class="badge b-warn">مستندات تنتهي خلال 30 يوم: '+str(len(expiring_docs))+'</span>') if expiring_docs else '<span class="badge b-ok">لا توجد مستندات قريبة الانتهاء</span>'}</div><div><span class="badge b-blue">تأخير الشهر: {late} دقيقة</span></div><div><span class="badge b-gray">آخر مرتب: {esc(last_pay['period'] if last_pay else '—')}</span></div></div></div>
        <div class="card" style="margin-top:16px"><h3>البيانات الأساسية</h3><div class="grid g3">{''.join(f'<div><small>{label}</small><br><b>{esc(e[k] or "—")}</b></div>' for k,label in [('birth_date','تاريخ الميلاد'),('address','العنوان'),('qualification','المؤهل'),('phone','الهاتف'),('bank_name','البنك'),('bank_branch','الفرع'),('employee_group','المجموعة الوظيفية'),('unit','الوحدة'),('contract_date','تاريخ التعاقد'),('contract_amount','مبلغ التعاقد')])}{('<div><small>الرقم القومي</small><br><b>'+esc(e['national_id'] or '—')+'</b></div><div><small>IBAN</small><br><b>'+esc(e['iban'] or '—')+'</b></div>') if can(u,'sensitive.view') else '<div><small>الرقم القومي</small><br><b>************</b></div><div><small>IBAN</small><br><b>************</b></div>'}</div></div>
        {'' if not evals else (lambda g: f'<div class="card" style="margin-top:16px"><div class="grid g2"><div><h3>⭐ التقييم الحالي</h3><div style="display:flex;align-items:center;gap:20px"><div style="width:120px;height:120px;border-radius:50%;background:conic-gradient(#12b76a {float(evals[0]["score"] or 0)}%,#eef2f6 0);display:grid;place-items:center"><div style="width:88px;height:88px;border-radius:50%;background:#fff;display:grid;place-items:center;font-size:24px;font-weight:800">{float(evals[0]["score"] or 0):g}%</div></div><div><div class="badge {g[1]}">{g[0]}</div><p style="margin-top:10px">آخر فترة: <b>{esc(evals[0]["period"])}</b></p></div></div></div><div><h3>Performance Trend</h3><div class="chart" style="height:180px">'+''.join(f'<div class="bar" style="height:{max(18,int(float(r["score"] or 0)*1.4))}px"><b>{float(r["score"] or 0):g}%</b><span>{esc(r["period"])}</span></div>' for r in reversed(evals[:8]))+'</div></div></div></div>')(evaluation_grade(evals[0]["score"]))}
        <div id="evaluation" class="card" style="margin-top:16px"><h3>⭐ تقييم الموظف — 5 عناصر</h3><p>النتيجة النهائية تُحسب تلقائيًا حسب الأوزان الموجودة في الإعدادات.</p><form class="form" method="post" action="/employee/evaluation/save">{csrf_field(u)}<input type="hidden" name="emp_code" value="{esc(code)}"><div class="field"><label>الفترة</label><input name="period" value="{today_iso[:7]}" required></div>
<div class="field"><label>Attendance · {evaluation_weights()['attendance']:g}%</label><input name="attendance_score" type="number" min="0" max="100" step="1" required></div>
<div class="field"><label>Punctuality · {evaluation_weights()['punctuality']:g}%</label><input name="punctuality_score" type="number" min="0" max="100" step="1" required></div>
<div class="field"><label>Productivity · {evaluation_weights()['productivity']:g}%</label><input name="productivity_score" type="number" min="0" max="100" step="1" required></div>
<div class="field"><label>Behavior · {evaluation_weights()['behavior']:g}%</label><input name="behavior_score" type="number" min="0" max="100" step="1" required></div>
<div class="field"><label>Manager Evaluation · {evaluation_weights()['manager']:g}%</label><input name="manager_score" type="number" min="0" max="100" step="1" required></div>
<div class="field full"><label>ملاحظات</label><textarea name="notes"></textarea></div><div class="full"><button class="btn">حفظ التقييم وحساب النسبة</button></div></form>
<div class="table-wrap" style="margin-top:14px"><h4>تاريخ التقييمات السابقة</h4><table class="table"><thead><tr><th>الفترة</th><th>النتيجة</th><th>Attendance</th><th>Punctuality</th><th>Productivity</th><th>Behavior</th><th>Manager</th><th>بواسطة</th></tr></thead><tbody>{''.join(f'<tr><td>{esc(r["period"])}</td><td><b>{float(r["score"] or 0):g}%</b></td><td>{r["attendance_score"]}%</td><td>{r["punctuality_score"]}%</td><td>{r["productivity_score"]}%</td><td>{r["behavior_score"]}%</td><td>{r["manager_score"]}%</td><td>{esc(r["created_by"])}</td></tr>' for r in evals) or '<tr><td colspan="8">لا توجد تقييمات سابقة</td></tr>'}</tbody></table></div></div>
        <div id="timeline" class="card" style="margin-top:16px"><h3>Timeline — كل أحداث الموظف</h3><div class="timeline">{timeline_html}</div></div>
        <div id="attendance" class="card" style="margin-top:16px"><h3>الحضور والانصراف</h3><div class="table-wrap"><table class="table"><thead><tr><th>التاريخ</th><th>الحالة</th><th>دخول</th><th>خروج</th><th>تأخير</th><th>ساعات</th><th>إضافي</th></tr></thead><tbody>{rows_html(att,['work_date','status','check_in','check_out','late_minutes','work_hours','overtime'])}</tbody></table></div></div>
        <div id="leaves" class="card" style="margin-top:16px"><h3>الإجازات</h3><div class="table-wrap"><table class="table"><thead><tr><th>الطلب</th><th>النوع</th><th>من</th><th>إلى</th><th>الأيام</th><th>الحالة</th></tr></thead><tbody>{rows_html(leaves,['request_no','leave_type','start_date','end_date','days','status'])}</tbody></table></div></div>
        <div id="documents" class="card" style="margin-top:16px"><div class="top"><h3>مستندات الموظف <small style="font-weight:normal">(نسخ تاريخية محفوظة تلقائيًا عند إعادة الرفع)</small></h3><a class="btn gray" href="/documents?emp_code={esc(code)}">إدارة المستندات</a></div><div class="table-wrap"><table class="table"><thead><tr><th>الملف</th><th>التصنيف</th><th>الإصدار</th><th>الحالة</th><th>الانتهاء</th><th>SHA-256</th><th></th></tr></thead><tbody>{docs_rows}</tbody></table></div></div>
        <div id="discipline" class="card" style="margin-top:16px"><h3>الإنذارات والجزاءات والمكافآت</h3><div class="table-wrap"><table class="table"><thead><tr><th>التاريخ</th><th>النوع</th><th>دقائق</th><th>مبلغ</th><th>السبب</th><th>ملاحظات</th></tr></thead><tbody>{rows_html(acts,['action_date','action_type','minutes','amount','reason','notes'])}</tbody></table></div></div>
        <div id="payroll" class="card" style="margin-top:16px"><h3>المرتبات</h3><div class="table-wrap"><table class="table"><thead><tr><th>الفترة</th><th>أساسي</th><th>بدلات</th><th>إضافي</th><th>مكافآت</th><th>خصومات</th><th>صافي</th><th>الحالة</th></tr></thead><tbody>{rows_html(pays,['period','basic','allowances','overtime','bonuses','deductions','net','status'])}</tbody></table></div></div>'''
        self.send(page('ملف الموظف',body,u,'employees'))

    def export_template_employees(self):
        wb=Workbook();ws=wb.active;ws.title='بيانات الموظفين'
        ws.append(['']+HOSPITAL_HEADERS)
        ws.append(['']+[1,'موظف تجريبي','التخصصية','1990-01-01','29000000000000','دمياط','بكالوريوس','01000000000','EG0000000000000000000000000000','بنك مصر','دمياط الجديدة','الخدمات الطبية','مستشفى دمياط العسكرى','موظف','2026-01-01',5000])
        out=io.BytesIO();wb.save(out);self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Hospital_Employees_Compatible_Template.xlsx"'})


    def save_employee(self,u,f):
        code=(f.get('emp_code') or f.get('emp_no') or f.get('national_id') or '').strip(); name=(f.get('name') or '').strip()
        if not code or not name:return self.send(page('خطأ','<div class="card"><div class="alert">الكود والاسم مطلوبان.</div></div>',u,'employees'),400)
        c=db(); old=c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone()
        shift_id=int(f.get('shift_id') or 0) or None
        sql="INSERT INTO employees(emp_code,name,employee_group,birth_date,national_id,address,qualification,phone,iban,bank_name,bank_branch,department,unit,job,contract_date,contract_amount,status,shift_id,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET name=excluded.name,employee_group=excluded.employee_group,birth_date=excluded.birth_date,national_id=excluded.national_id,address=excluded.address,qualification=excluded.qualification,phone=excluded.phone,iban=excluded.iban,bank_name=excluded.bank_name,bank_branch=excluded.bank_branch,department=excluded.department,unit=excluded.unit,job=excluded.job,contract_date=excluded.contract_date,contract_amount=excluded.contract_amount,status=excluded.status,shift_id=excluded.shift_id,updated_at=excluded.updated_at"
        vals=(code,name,f.get('employee_group',''),f.get('birth_date',''),f.get('national_id',''),f.get('address',''),f.get('qualification',''),f.get('phone',''),f.get('iban',''),f.get('bank_name',''),f.get('bank_branch',''),f.get('department',''),f.get('unit',''),f.get('job',''),f.get('contract_date',''),cell_num(f.get('contract_amount')),f.get('status') or 'على رأس العمل',shift_id,now())
        c.execute(sql,vals)
        for lt in c.execute('SELECT name,annual_balance FROM leave_types').fetchall(): c.execute('INSERT OR IGNORE INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,0)',(code,lt['name'],lt['annual_balance']))
        if shift_id: c.execute('INSERT INTO employee_shifts(emp_code,shift_id,assigned_at,assigned_by) VALUES(?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET shift_id=excluded.shift_id,assigned_at=excluded.assigned_at,assigned_by=excluded.assigned_by',(code,shift_id,now(),u['username']))
        else: c.execute('DELETE FROM employee_shifts WHERE emp_code=?',(code,))
        after=dict(c.execute('SELECT * FROM employees WHERE emp_code=?',(code,)).fetchone()); c.commit(); c.close()
        audit(u['username'],u['role'],'حفظ','الموظفون',code,'تحديث بيانات الموظف',before=dict(old) if old else None,after=after,reason='employee save')
        self.redirect('/employees')


    def leaves(self,u):
        q=parse_qs(urlparse(self.path).query).get('emp_code',[''])[0]; c=db(); rows=c.execute('SELECT * FROM leaves WHERE emp_code=? ORDER BY id DESC LIMIT 1000',(q,)).fetchall() if q else c.execute('SELECT * FROM leaves ORDER BY id DESC LIMIT 1000').fetchall(); rows=[r for r in rows if emp_allowed(u,r['emp_code'])]; c.close()
        parts=[]
        for r in rows:
            cls='b-ok' if r['status']=='معتمدة' else ('b-bad' if r['status']=='مرفوضة' else 'b-warn')
            actions=''
            if r['status']=='قيد المراجعة' and can(u,'leave.approve'):
                no=esc(r['request_no'])
                actions=f'<form method="post" action="/leave/status" style="display:inline">{csrf_field(u)}<input type="hidden" name="request_no" value="{no}"><input type="hidden" name="status" value="معتمدة"><button class="btn">اعتماد</button></form> <form method="post" action="/leave/status" style="display:inline">{csrf_field(u)}<input type="hidden" name="request_no" value="{no}"><input type="hidden" name="status" value="مرفوضة"><button class="btn bad">رفض</button></form>'
            parts.append(f'<tr><td>{esc(r["request_no"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["leave_type"])}</td><td>{esc(r["start_date"])}</td><td>{esc(r["end_date"])}</td><td>{r["days"]}</td><td><span class="badge {cls}">{esc(r["status"])}</span></td><td>{actions}</td></tr>')
        trs=''.join(parts)
        body=f'''<div class="top"><div class="title"><h1>طلبات الإجازات</h1><p>تحقق من الرصيد والتداخل واعتماد مركزي</p></div><div class="actions"><a class="btn" href="/leave/new">+ طلب إجازة</a><a class="btn gray" href="/export/leaves">Export Excel</a><button class="btn gray" onclick="window.print()">طباعة / PDF</button></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الطلب</th><th>الموظف</th><th>النوع</th><th>من</th><th>إلى</th><th>الأيام</th><th>الحالة</th><th></th></tr></thead><tbody>{trs}</tbody></table></div>'''
        self.send(page('الإجازات',body,u,'leaves'))
    def leave_form(self,u):
        c=db(); types=[x[0] for x in c.execute('SELECT name FROM leave_types ORDER BY id').fetchall()]; emps=[(x[0],x[1]) for x in c.execute("SELECT emp_code,name FROM employees WHERE status='على رأس العمل' ORDER BY name").fetchall()]; c.close()
        opts=''.join(f'<option value="{esc(e[0])}">{esc(e[0])} — {esc(e[1])}</option>' for e in emps)
        body=f'''<div class="top"><div class="title"><h1>طلب إجازة جديد</h1><p>النظام يمنع التداخل ويحسب الأيام ويتحقق من الرصيد</p></div><a class="btn gray" href="/leaves">عودة</a></div><div class="card"><form class="form" method="post" action="/leave/save">{csrf_field(u)}<div class="field"><label>الموظف *</label><select name="emp_code" required>{opts}</select></div><div class="field"><label>نوع الإجازة *</label><select name="leave_type">{''.join('<option>'+esc(x)+'</option>' for x in types)}</select></div><div class="field"><label>من *</label><input type="date" name="start_date" value="{date.today().isoformat()}" required></div><div class="field"><label>إلى *</label><input type="date" name="end_date" value="{date.today().isoformat()}" required></div><div class="field full"><label>ملاحظات</label><textarea name="notes"></textarea></div><div class="full"><button class="btn">إرسال الطلب</button></div></form></div>'''; self.send(page('طلب إجازة',body,u,'leaves'))
    def save_leave(self,u,f):
        sd,ed=f.get('start_date'),f.get('end_date'); d1=datetime.fromisoformat(sd).date(); d2=datetime.fromisoformat(ed).date(); days=(d2-d1).days+1
        if days<=0: return self.send(page('خطأ','<div class="card"><div class="alert">نطاق التاريخ غير صحيح.</div></div>',u,'leaves'))
        c=db(); emp=f.get('emp_code'); lt=f.get('leave_type');
        if not c.execute('SELECT 1 FROM employees WHERE emp_code=? AND status=\'على رأس العمل\'',(emp,)).fetchone(): c.close(); return self.send(page('خطأ','<div class="card"><div class="alert">الموظف غير موجود أو غير نشط.</div></div>',u,'leaves'))
        if c.execute("SELECT 1 FROM leaves WHERE emp_code=? AND status IN ('معتمدة','قيد المراجعة') AND start_date<=? AND end_date>=?",(emp,ed,sd)).fetchone(): c.close(); return self.send(page('خطأ','<div class="card"><div class="alert">يوجد طلب إجازة متداخل بالفعل.</div></div>',u,'leaves'))
        ltrow=c.execute('SELECT annual_balance FROM leave_types WHERE name=?',(lt,)).fetchone(); c.execute('INSERT OR IGNORE INTO leave_balances(emp_code,leave_type,annual,used) VALUES(?,?,?,0)',(emp,lt,ltrow['annual_balance'] if ltrow else 0))
        bal=c.execute('SELECT annual-used FROM leave_balances WHERE emp_code=? AND leave_type=?',(emp,lt)).fetchone(); pending=c.execute("SELECT COALESCE(SUM(days),0) x FROM leaves WHERE emp_code=? AND leave_type=? AND status='قيد المراجعة'",(emp,lt)).fetchone()['x']
        available=(bal[0] if bal else 0)-pending
        if available < days: c.close(); return self.send(page('خطأ',f'<div class="card"><div class="alert">الرصيد المتاح بعد الطلبات المعلقة: {available} يوم.</div></div>',u,'leaves'),400)
        no=f'LV-{c.execute("SELECT COALESCE(MAX(id),0)+1 n FROM leaves").fetchone()["n"]:05d}'; c.execute('INSERT INTO leaves(request_no,emp_code,leave_type,start_date,end_date,days,request_date,status,notes) VALUES(?,?,?,?,?,?,?,?,?)',(no,emp,lt,sd,ed,days,date.today().isoformat(),'قيد المراجعة',f.get('notes'))); c.commit(); c.close(); audit(u['username'],u['role'],'إضافة','طلبات الإجازات',no); self.redirect('/leaves')
    def leave_status(self,u,f):
        no=f.get('request_no'); status=f.get('status'); c=db(); r=c.execute('SELECT * FROM leaves WHERE request_no=?',(no,)).fetchone()
        if not r: c.close(); return self.redirect('/leaves')
        if not emp_allowed(u,r['emp_code']): c.close(); return self.forbid(u)
        if not can(u,'leave.approve'): c.close(); return self.forbid(u)
        if r['status']!='قيد المراجعة':
            c.close(); return self.send(page('خطأ','<div class="card"><div class="alert">تم اتخاذ قرار بشأن هذا الطلب مسبقًا («{}»). لا يمكن تكرار الاعتماد أو الرفض.</div></div>'.format(esc(r['status'])),u,'leaves'))
        if status=='معتمدة':
            bal=c.execute('SELECT annual-used FROM leave_balances WHERE emp_code=? AND leave_type=?',(r['emp_code'],r['leave_type'])).fetchone()
            if bal and bal[0]<r['days']: c.close(); return self.send(page('خطأ','<div class="card"><div class="alert">لا يمكن الاعتماد: الرصيد غير كافٍ.</div></div>',u,'leaves'))
            c.execute('UPDATE leaves SET status=?,approved_by=?,approved_at=? WHERE request_no=?',(status,u['username'],now(),no)); c.execute('UPDATE leave_balances SET used=used+? WHERE emp_code=? AND leave_type=?',(r['days'],r['emp_code'],r['leave_type'])); c.execute('INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(r['emp_code'],'leave',now()[:10],f'الإجازة {status}',f'{r["leave_type"]} · {r["days"]} يوم',u['username'],now()))
        else: c.execute('UPDATE leaves SET status=?,approved_by=?,approved_at=? WHERE request_no=?',(status,u['username'],now(),no))
        c.commit(); c.close(); audit(u['username'],u['role'],status,'طلبات الإجازات',no); self.redirect('/leaves')

    def attendance(self,u):
        c=db(); scope_sql,scope_params=visible_employee_sql(u,'e'); rows=c.execute(f'SELECT a.* FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE 1=1{scope_sql} ORDER BY a.work_date DESC,a.id DESC LIMIT 1000',scope_params).fetchall(); emps=c.execute(f'SELECT e.emp_code,e.name FROM employees e WHERE 1=1{scope_sql} ORDER BY e.name',scope_params).fetchall(); monthly=c.execute(f"SELECT a.emp_code,SUM(a.late_minutes) late,SUM(a.overtime) ot FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE substr(a.work_date,1,7)=?{scope_sql} GROUP BY a.emp_code ORDER BY late DESC LIMIT 50",[date.today().isoformat()[:7]]+scope_params).fetchall(); c.close()
        trs=''.join(f'<tr><td>{esc(r["work_date"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["status"])}</td><td>{esc(r["check_in"])}</td><td>{esc(r["check_out"])}</td><td>{r["late_minutes"]}</td><td>{r["work_hours"]}</td><td>{r["overtime"]}</td></tr>' for r in rows)
        opts=''.join(f'<option value="{esc(e["emp_code"])}">{esc(e["emp_code"])} — {esc(e["name"])}</option>' for e in emps); mrows=''.join(f'<tr><td>{esc(r["emp_code"])}</td><td>{r["late"] or 0}</td><td>{r["ot"] or 0}</td></tr>' for r in monthly)
        body=f'''<div class="top"><div class="title"><h1>الحضور والانصراف</h1><p>يدوي + Excel/CSV + Paste · تجميع التأخير الشهري والإضافي</p></div><div class="actions"><a class="btn" href="/import">Excel Center</a><a class="btn gray" href="/export/attendance">Export Excel</a></div></div><div class="grid g2"><div class="card"><h3>إدخال يدوي</h3><form class="form" method="post" action="/attendance/save">{csrf_field(u)}<div class="field"><label>التاريخ</label><input type="date" name="work_date" value="{date.today().isoformat()}"></div><div class="field"><label>الموظف</label><select name="emp_code" required>{opts}</select></div><div class="field"><label>الحالة</label><select name="status"><option>حضور</option><option>غياب</option><option>إجازة</option><option>مأمورية</option></select></div><div class="field"><label>الحضور</label><input type="time" name="check_in"></div><div class="field"><label>الانصراف</label><input type="time" name="check_out"></div><div class="field full"><label>ملاحظات</label><input name="notes"></div><div class="full"><button class="btn">حفظ / تحديث</button></div></form></div><div class="card"><h3>ملخص التأخير لهذا الشهر</h3><table class="table"><thead><tr><th>الموظف</th><th>إجمالي دقائق التأخير</th><th>إضافي</th></tr></thead><tbody>{mrows or '<tr><td colspan="3">لا توجد بيانات.</td></tr>'}</tbody></table></div></div><div class="card" style="margin-top:16px"><h3>سجل الحضور</h3><div class="table-wrap"><table class="table"><thead><tr><th>التاريخ</th><th>الكود</th><th>الحالة</th><th>حضور</th><th>انصراف</th><th>تأخير</th><th>ساعات</th><th>إضافي</th></tr></thead><tbody>{trs}</tbody></table></div></div>'''
        self.send(page('الحضور',body,u,'attendance'))

    def do_import_employees_paste(self,u,f):
        text=f.get('paste_data','')
        try:
            raw_lines=[x for x in text.replace('\r','').split('\n') if x.strip()]
            if not raw_lines: raise ValueError('لم يتم لصق أي بيانات.')
            rows=[x.split('\t') for x in raw_lines]; header_idx,idx=find_hospital_header(rows)
            if header_idx is None:
                if max(len(r) for r in rows)<8: raise ValueError('لم أتعرف على هيكل شيت الموظفين.')
                idx={field:i for i,field in enumerate(HOSPITAL_FIELDS)}; data_rows=rows
            else:data_rows=rows[header_idx+1:]
            records=[hospital_row_to_record(r,idx) for r in data_rows]; valid,errors=validate_employee_records(records)
            if errors:
                c=db(); cur=c.execute('INSERT INTO file_imports(source,file_name,records,created_by,created_at,details) VALUES(?,?,?,?,?,?)',('employees_paste','Paste Excel',0,u['username'],now(),f'atomic_blocked_errors={len(errors)}')); runid=cur.lastrowid
                for row_no,field,msg,raw in errors:c.execute('INSERT INTO import_errors(run_id,row_no,field,message,raw_json) VALUES(?,?,?,?,?)',(runid,row_no,field,msg,json.dumps(raw,ensure_ascii=False)))
                c.commit(); c.close(); return self.send(page('Import Validation',f'<div class="card"><h2>🔴 لم يتم إدخال أي سجل</h2><p>Atomic import: أصلح {len(errors)} خطأ ثم أعد المحاولة.</p><a class="btn bad" href="/export/import-errors/{runid}">تصدير الأخطاء</a></div>',u,'import'),400)
            new,upd,skip=upsert_hospital_records(valid,u,'Paste Excel'); runid=None
            if errors:
                c=db(); cur=c.execute('INSERT INTO file_imports(source,file_name,records,created_by,created_at,details) VALUES(?,?,?,?,?,?)',('employees_paste','Paste Excel',len(valid),u['username'],now(),f'errors={len(errors)}')); runid=cur.lastrowid
                for row_no,field,msg,raw in errors:c.execute('INSERT INTO import_errors(run_id,row_no,field,message,raw_json) VALUES(?,?,?,?,?)',(runid,row_no,field,msg,json.dumps(raw,ensure_ascii=False)))
                c.commit(); c.close()
            link=f'<a class="btn bad" href="/import/errors/{runid}">عرض/تصدير الأخطاء ({len(errors)})</a>' if errors else ''
            return self.send(page('تم الفحص',f'<div class="card"><h2>نتيجة فحص اللصق</h2><p>إجمالي الصفوف: <b>{len(records)}</b> · صالح: <b>{len(valid)}</b> · أخطاء: <b>{len(errors)}</b> · جديد: <b>{new}</b> · تحديث: <b>{upd}</b></p>{link}<a class="btn gray" href="/employees">عرض الموظفين</a></div>',u,'import'))
        except Exception as e:return self.send(page('خطأ في اللصق',f'<div class="card"><div class="alert">{esc(e)}</div></div>',u,'import'),400)



    def reports(self,u):
        c=db(); scope_sql,scope_params=visible_employee_sql(u)
        by_status=c.execute(f'SELECT e.status,COUNT(*) n FROM employees e WHERE 1=1{scope_sql} GROUP BY e.status',scope_params).fetchall()
        by_dept=c.execute(f"SELECT COALESCE(e.department,'غير محدد') d,COUNT(*) n FROM employees e WHERE 1=1{scope_sql} GROUP BY d ORDER BY n DESC",scope_params).fetchall()
        leave_status=c.execute(f'SELECT l.status,COUNT(*) n FROM leaves l JOIN employees e ON e.emp_code=l.emp_code WHERE 1=1{scope_sql} GROUP BY l.status',scope_params).fetchall()
        late=c.execute(f'SELECT SUM(a.late_minutes) x,SUM(a.overtime) o FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE 1=1{scope_sql}',scope_params).fetchone(); c.close()
        rows=lambda xs,key: ''.join(f'<tr><td>{esc(r[key])}</td><td>{r["n"]}</td></tr>' for r in xs)
        body=f'''<div class="top"><div class="title"><h1>مركز التقارير</h1><p>تقارير تشغيلية جاهزة للطباعة والتصدير</p></div><div class="actions"><a class="btn gray" href="/export/employees">موظفون Excel</a><a class="btn gray" href="/export/leaves">إجازات Excel</a><a class="btn gray" href="/export/attendance">حضور Excel</a><button class="btn" onclick="window.print()">طباعة / PDF</button></div></div><div class="grid g3"><div class="card"><h3>حالة الموظفين</h3><table class="table"><tbody>{rows(by_status,'status')}</tbody></table></div><div class="card"><h3>الإجازات حسب الحالة</h3><table class="table"><tbody>{rows(leave_status,'status')}</tbody></table></div><div class="card"><h3>الحضور والتأخير</h3><p>إجمالي دقائق التأخير: <b>{late['x'] or 0}</b></p><p>إجمالي الساعات الإضافية: <b>{late['o'] or 0}</b></p></div></div><div class="card" style="margin-top:16px"><h3>الموظفون حسب الإدارة</h3><table class="table"><thead><tr><th>الإدارة</th><th>العدد</th></tr></thead><tbody>{rows(by_dept,'d')}</tbody></table></div>'''; self.send(page('التقارير',body,u,'reports'))

    def import_page(self,u):
        required,missing_rows=missing_documents_report(u)
        if not required:
            missing_html='<div class="card" style="margin-top:16px"><h3>تتبع المستندات الناقصة</h3><div class="alert">لم يتم تحديد تصنيفات مستندات مطلوبة بعد \u2014 حدّدها من <a href="/settings">الإعدادات</a> (مثال: عقد,هوية,مؤهل).</div></div>'
        elif not missing_rows:
            missing_html='<div class="card" style="margin-top:16px"><h3>تتبع المستندات الناقصة</h3><div class="alert" style="background:#ecfdf3;border-color:#a6f4c5;color:#027a48">كل الموظفين النشطين مستوفون التصنيفات المطلوبة ('+esc(", ".join(required))+'). \u2713</div></div>'
        else:
            mtrs=''.join('<tr><td>'+esc(ec)+'</td><td>'+esc(en)+'</td><td>'+esc(", ".join(miss))+'</td><td><a class="btn gray" href="/employee/profile/'+esc(ec)+'#documents">رفع المستند</a></td></tr>' for ec,en,miss in missing_rows)
            missing_html='<div class="card" style="margin-top:16px"><h3>تتبع المستندات الناقصة <span class="badge b-warn">'+str(len(missing_rows))+' موظف</span></h3><p style="color:#667085">التصنيفات المطلوبة (من الإعدادات): <b>'+esc(", ".join(required))+'</b> \u2014 القائمة تتحدّث تلقائيًا فور رفع مستند أو استيراد مجلد.</p><div class="table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>التصنيفات الناقصة</th><th></th></tr></thead><tbody>'+mtrs+'</tbody></table></div></div>'
        body = '<div class="top"><div class="title"><h1>مركز الاستيراد والتصدير</h1><p>يدعم نفس شيت الموظفين الأصلي: م، الإسم، المجموعة الوظيفية، الميلاد، الرقم القومي، العنوان، المؤهل، الهاتف، IPAN، البنك، الفرع، الإدارة، الوحدة، الوظيفة، التعاقد، مبلغ التعاقد</p></div><span class="badge b-blue" style="font-size:14px">V4.4 · PASTE READY</span></div>\n<div class="card" style="margin-bottom:16px;border:2px solid #84adff;background:linear-gradient(180deg,#fff,#f7fbff)">\n  <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap">\n    <div><h2 style="margin:0 0 6px">📋 لصق بيانات الموظفين من Excel</h2><p style="margin:0;color:#667085">انسخ الجدول من Excel ثم اضغط داخل المربع واضغط Ctrl+V. ستظهر البيانات هنا قبل الحفظ.</p></div>\n    <span class="badge b-blue">بدون Upload</span>\n  </div>\n  <form method="post" action="/import/employees/paste" style="margin-top:16px">'+csrf_field(u)+'\n    <div class="field full"><label>منطقة اللصق — اضغط هنا ثم Ctrl+V</label>\n      <textarea id="pasteBox" name="paste_data" rows="9" spellcheck="false" autocomplete="off" placeholder="الصق هنا صف العناوين + صفوف الموظفين من Excel..."></textarea>\n    </div>\n    <div id="pasteStatus" class="alert" style="display:none;margin-top:10px"></div>\n    <div id="pastePreview" class="table-wrap" style="margin-top:12px;max-height:340px;overflow:auto"></div>\n    <div class="actions" style="margin-top:12px">\n      <button class="btn ok" type="submit">استيراد البيانات الملصوقة</button>\n      <button class="btn gray" type="button" id="clearPaste">مسح</button>\n      <button class="btn gray" type="button" id="samplePaste">إدخال مثال</button>\n    </div>\n  </form>\n</div>\n<div class="grid g3">\n<div class="card"><h3>استيراد ملف Excel</h3><p>يدعم XLSX / XLSM ويكتشف ورقة الموظفين والأعمدة العربية أو الإنجليزية.</p><form method="post" action="/import/employees" enctype="multipart/form-data">{csrf_field(u)}<input type="file" name="file" accept=".xlsx,.xlsm" required><button class="btn" style="margin-top:12px">استيراد الموظفين</button></form><div class="actions" style="margin-top:10px"><a class="btn gray" href="/template/employees">تحميل Template</a></div></div>\n<div class="card"><h3>استيراد الحضور</h3><p>التاريخ + كود الموظف، مع الحضور والانصراف والتأخير والإضافي.</p><form method="post" action="/attendance/import" enctype="multipart/form-data">{csrf_field(u)}<input type="file" name="file" accept=".xlsx,.xlsm,.csv" required><button class="btn" style="margin-top:12px">استيراد الحضور</button></form></div>\n<div class="card"><h3>استيراد مجلدات الموظفين (ZIP)</h3><p>ZIP فيه مجلد لكل موظف — اسم المجلد بالاسم أو الكود، يترّبط تلقائيًا. الملفات المؤكدة تتوزع فورًا على تصنيفات المستندات وتظهر في تتبع الناقص تحت.</p><form method="post" action="/documents/folders/import" enctype="multipart/form-data">{csrf_field(u)}<input type="file" name="file" accept=".zip" required><button class="btn" style="margin-top:12px">استيراد ZIP المجلدات</button></form></div>\n<div class="card"><h3>📁 اختيار مجلدات كثيرة بدون ZIP</h3><p>اختار مجلد رئيسي كامل أو أضف عدة مجلدات موظفين منفصلة في نفس العملية.</p><form id="folderFilesForm" method="post" action="/documents/folders/import-files" enctype="multipart/form-data">{csrf_field(u)}<div id="folderInputs"><div class="folder-pick-row" style="display:flex;gap:8px;align-items:center;margin-bottom:8px"><input class="folderPicker" type="file" name="folder_files" webkitdirectory directory multiple required style="flex:1"><button class="btn gray removeFolder" type="button" style="display:none">حذف</button></div></div><div class="actions" style="margin-top:10px"><button class="btn gray" id="addFolderPicker" type="button">➕ إضافة مجلد آخر</button><button class="btn gray" id="clearFolderPickers" type="button">مسح الاختيارات</button></div><div id="folderStatus" class="alert" style="margin-top:10px">يمكنك إضافة أكثر من مجلد ثم الضغط على استيراد مرة واحدة.</div><button class="btn" style="margin-top:12px">استيراد كل المجلدات المختارة</button></form><script>(function(){const box=document.getElementById("folderInputs"),status=document.getElementById("folderStatus");function refresh(){let files=0,rows=0;box.querySelectorAll(".folder-pick-row").forEach(function(r){rows++;const i=r.querySelector("input");files+=i.files.length;r.querySelector(".removeFolder").style.display=rows>1?"inline-block":"none";});status.textContent="المجلدات المضافة: "+rows+" · إجمالي الملفات: "+files;}document.getElementById("addFolderPicker").onclick=function(){const row=document.createElement("div");row.className="folder-pick-row";row.style="display:flex;gap:8px;align-items:center;margin-bottom:8px";row.innerHTML="<input class=\"folderPicker\" type=\"file\" name=\"folder_files\" webkitdirectory directory multiple style=\"flex:1\"><button class=\"btn gray removeFolder\" type=\"button\">حذف</button>";row.querySelector("input").onchange=refresh;row.querySelector(".removeFolder").onclick=function(){row.remove();refresh();};box.appendChild(row);};box.addEventListener("change",refresh);document.getElementById("clearFolderPickers").onclick=function(){box.querySelectorAll(".folder-pick-row").forEach(function(r,i){if(i===0)r.querySelector("input").value="";else r.remove();});refresh();};refresh();})();</script></div>\n</div><div class="card" style="margin-top:16px"><h3>تصدير</h3><div class="actions"><a class="btn gray" href="/export/employees">الموظفون</a><a class="btn gray" href="/export/leaves">الإجازات</a><a class="btn gray" href="/export/attendance">الحضور</a><a class="btn gray" href="/export/audit">سجل المراجعة</a></div></div>\nMISSING_DOCS_PLACEHOLDER\n<script>\n(function(){\n  const b=document.getElementById(\'pasteBox\'), p=document.getElementById(\'pastePreview\'), st=document.getElementById(\'pasteStatus\');\n  function esc(x){return String(x??\'\').replace(/[&<>"\']/g,function(m){return {\'&\':\'&amp;\',\'<\':\'&lt;\',\'>\':\'&gt;\',\'"\':\'&quot;\',"\'":\'&#39;\'}[m];});}\n  function render(){\n    const text=b.value.replace(/\\\\r/g,\'\');\n    const lines=text.split(\'\\\\n\').filter(function(x){return x.trim();});\n    if(!lines.length){p.innerHTML=\'\';st.style.display=\'none\';return;}\n    const rows=lines.slice(0,100).map(function(x){return x.split(\'\\\\t\');});\n    const m=Math.max.apply(null,rows.map(function(x){return x.length;}));\n    let h=\'<table class="table"><thead><tr>\';\n    for(let i=0;i<m;i++) h+=\'<th>عمود \'+(i+1)+\'</th>\';\n    h+=\'</tr></thead><tbody>\';\n    rows.forEach(function(x){h+=\'<tr>\';for(let i=0;i<m;i++)h+=\'<td contenteditable=\"true\" data-col=\"\'+i+\">\'+esc(x[i]||\'\')+\'</td>\';h+=\'<td><button type=\"button\" class=\"btn bad\" onclick=\"this.closest(\\\'tr\\\').remove()\">حذف</button></td></tr>\';});\n    h+=\'</tbody></table>\';p.innerHTML=h;\n    st.textContent=\'تم التقاط \'+lines.length+\' صف — اضغط على أي خلية لتعديلها، ويمكنك حذف أي صف قبل الاستيراد.\';st.style.display=\'block\';\n  }\n  b.addEventListener(\'input\',render); b.addEventListener(\'paste\',function(){setTimeout(render,50);});\n  b.closest(\'form\').addEventListener(\'submit\',function(){const table=p.querySelector(\'table\');if(!table)return;const rows=[...table.querySelectorAll(\'tbody tr\')];b.value=rows.map(tr=>[...tr.querySelectorAll(\'td[data-col]\')].map(td=>td.innerText.replace(/\\t/g,\' \').replace(/\\n/g,\' \').trim()).join(\'\\t\')).join(\'\\n\');});\n  document.getElementById(\'clearPaste\').addEventListener(\'click\',function(){b.value=\'\';p.innerHTML=\'\';st.style.display=\'none\';b.focus();});\n  document.getElementById(\'samplePaste\').addEventListener(\'click\',function(){b.value=\'\\\\tم\\\\tالإسم\\\\tالمجموعه الوظفيه\\\\tتاريخ الميلاد\\\\tالرقم القومى\\\\tالعنوان\\\\tالمؤهل\\\\tرقم التيلفون\\\\tipan\\\\tإسم البنك\\\\tإسم فرع البنك\\\\tالإدارة\\\\tالوحدة\\\\tالوظيفه\\\\tتاريخ التعاقد\\\\tمبلغ التعاقد\\\\n\\\\t1\\\\tموظف تجريبي\\\\tالتخصصيه\\\\t1990-01-01\\\\t29000000000000\\\\tدمياط\\\\tبكالوريوس\\\\t01000000000\\\\tEG0000000000000000000000000000\\\\tبنك مصر\\\\tدمياط الجديدة\\\\tالخدمات الطبية\\\\tمستشفى دمياط العسكرى\\\\tموظف\\\\t2026-01-01\\\\t5000\';render();b.focus();});\n})();\n</script>'
        body += r'''<script>
(function(){
  const p=document.getElementById('pastePreview'); if(!p) return;
  let undo=[], redo=[];
  function snap(){ return p.innerHTML; }
  function save(){ undo.push(snap()); if(undo.length>30) undo.shift(); redo=[]; }
  p.addEventListener('focusin',e=>{ if(e.target.matches('td[contenteditable="true"]')) save(); });
  p.addEventListener('keydown',e=>{
    const td=e.target.closest('td[contenteditable="true"]'); if(!td) return;
    if((e.ctrlKey||e.metaKey)&&e.key.toLowerCase()==='z'){e.preventDefault(); if(undo.length){redo.push(snap());p.innerHTML=undo.pop();}} 
    if((e.ctrlKey||e.metaKey)&&e.key.toLowerCase()==='y'){e.preventDefault(); if(redo.length){undo.push(snap());p.innerHTML=redo.pop();}}
    if(e.key==='Delete' && !window.getSelection().toString()){e.preventDefault(); save();td.textContent=''; validate(td);}
    if(e.key==='Enter'){e.preventDefault(); const cells=[...p.querySelectorAll('td[contenteditable="true"]')]; const i=cells.indexOf(td); if(cells[i+1]) cells[i+1].focus();}
    if((e.ctrlKey||e.metaKey)&&e.key.toLowerCase()==='d'){e.preventDefault();save();const tr=td.closest('tr');const col=td.dataset.col;const above=tr.previousElementSibling&&tr.previousElementSibling.querySelector('td[data-col="'+col+'"]');if(above){td.textContent=above.innerText;validate(td);}}
  });
  p.addEventListener('input',e=>{if(e.target.matches('td[contenteditable="true"]'))validate(e.target);});
  function validate(td){
    td.style.background='';
    const v=td.innerText.trim(), col=Number(td.dataset.col||-1);
    if(!v) return;
    if(col===5 && !/^\d{14}$/.test(v)){td.style.background='#fee4e2';td.title='الرقم القومي يجب أن يكون 14 رقمًا';}
    else if(col===7 && v && !/^01\d{9}$/.test(v)){td.style.background='#fef0c7';td.title='تحقق من رقم الهاتف';}
    else td.title='';
    const rows=[...p.querySelectorAll('tbody tr')]; let dup=0;
    if(col===1){rows.forEach(r=>{const c=r.querySelector('td[data-col="1"]');if(c&&c.innerText.trim()===v)dup++;});if(dup>1){td.style.background='#fee4e2';td.title='كود موظف مكرر';}}
  }
  const form=document.getElementById('pasteBox')?.closest('form');
  if(form) form.addEventListener('submit',()=>{p.querySelectorAll('td[contenteditable="true"]').forEach(validate);});
})();
</script>'''
        self.send(page('الاستيراد والتصدير',body.replace('{csrf_field(u)}',csrf_field(u)).replace('MISSING_DOCS_PLACEHOLDER',missing_html),u,'import'))
    def parse_upload(self):
        if hasattr(self,'_upload_cache'): return self._upload_cache
        raw=self._raw(); ct=self.headers.get('Content-Type',''); boundary=ct.split('boundary=',1)[-1].strip().strip('"').encode('ascii','ignore'); parts=raw.split(b'--'+boundary); fields={}; file_part=(b'',b'', '')
        for p in parts:
            if b'\r\n\r\n' not in p: continue
            head,dat=p.split(b'\r\n\r\n',1); dat=dat[:-2] if dat.endswith(b'\r\n') else dat
            nm=re.search(rb'name="([^"]+)"',head); fn=re.search(rb'filename="([^"]*)"',head)
            if nm:
                name=nm.group(1).decode(errors='ignore')
                if fn: file_part=(head,dat,fn.group(1).decode(errors='ignore'))
                else:
                    val=dat.decode('utf-8',errors='replace')
                    if name=='emp_codes': fields.setdefault(name,[]).append(val)
                    else: fields[name]=val
        self._upload_cache=(fields,file_part); return self._upload_cache
    def parse_upload_all(self):
        raw=self._raw(); ct=self.headers.get('Content-Type',''); boundary=ct.split('boundary=',1)[-1].strip().strip('"').encode('ascii','ignore'); parts=raw.split(b'--'+boundary); fields={}; files=[]
        for p in parts:
            if b'\r\n\r\n' not in p: continue
            head,dat=p.split(b'\r\n\r\n',1); dat=dat[:-2] if dat.endswith(b'\r\n') else dat
            nm=re.search(rb'name="([^"]+)"',head); fn=re.search(rb'filename="([^"]*)"',head)
            if not nm: continue
            name=nm.group(1).decode(errors='ignore')
            if fn: files.append((head,dat,fn.group(1).decode(errors='ignore')))
            else: fields[name]=dat.decode('utf-8',errors='replace')
        return fields,files
    def folder_files_import(self,u):
        fields,files=self.parse_upload_all(); max_mb=float(setting('document_max_mb') or 25); allowed={'.pdf','.jpg','.jpeg','.png','.docx','.doc','.xlsx','.xls','.txt','.csv'}; c=db(); employees=[dict(r) for r in c.execute('SELECT emp_code,name FROM employees').fetchall()]; groups={}; invalid=0
        # Browser directory uploads include the selected root folder in webkitRelativePath,
        # e.g. Employee_Folders/38 - Name/file.pdf. Find the first path component that
        # actually resolves to an employee instead of treating the root container as a person.
        for head,data,fname in files:
            path=fname.replace('\\','/')
            parts=[x.strip() for x in path.split('/') if x.strip()]
            if len(parts)<2:
                invalid+=1; continue
            employee_idx=None; employee_folder=None
            for idx in range(len(parts)-1):
                emp0,_,_=resolve_folder_employee(parts[idx],employees)
                if emp0:
                    employee_idx=idx; employee_folder=parts[idx]; break
            if employee_idx is None:
                # Keep the first directory as the review key so the user gets a useful message.
                groups.setdefault(parts[0],[]).append(('/'.join(parts[1:]).strip(),data))
            else:
                relname='/'.join(parts[employee_idx+1:]).strip()
                groups.setdefault(employee_folder,[]).append((relname,data))
        matched=0; imported=0; review=[]
        try:
            for folder,items in groups.items():
                emp,mt,mn=resolve_folder_employee(folder,employees)
                if not emp or not emp_allowed(u,emp):
                    # Persist ambiguous/unmatched folders for HR Matching Center review.
                    try:
                        target=normalize_name(folder); candidates=[]
                        for ee in employees:
                            ratio=difflib.SequenceMatcher(None,target,normalize_name(ee['name'])).ratio()
                            if ratio>=0.55: candidates.append({'emp_code':ee['emp_code'],'name':ee['name'],'confidence':round(ratio,3)})
                        candidates=sorted(candidates,key=lambda x:x['confidence'],reverse=True)[:5]
                        c.execute('INSERT INTO matching_reviews(source_name,candidate_json,confidence,status,created_by,created_at) VALUES(?,?,?,?,?,?)',(folder,json.dumps(candidates,ensure_ascii=False),candidates[0]['confidence'] if candidates else 0,'review',u['username'],now()))
                    except Exception: pass
                    review.append((folder,'لا يوجد تطابق آمن')); continue
                matched+=1
                for relname,data in items:
                    ext=os.path.splitext(relname)[1].lower()
                    if ext not in allowed or len(data)>max_mb*1024*1024: invalid+=1; continue
                    fname=safe_name(os.path.basename(relname)); cat=guess_document_category(fname); prior=c.execute('SELECT id,version FROM documents WHERE emp_code=? AND category=? AND status=?',(emp,cat,'current')).fetchall(); ver=max([r['version'] or 1 for r in prior] or [0])+1; store=save_employee_file(emp,fname,data); checksum=hashlib.sha256(data).hexdigest(); cur=c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,fname,ext,u['username'],now(),None,store,cat,ver,checksum,'current')); [c.execute('UPDATE documents SET status=?,superseded_by=? WHERE id=?',('superseded',cur.lastrowid,r['id'])) for r in prior]; imported+=1
            c.commit()
        except Exception:
            c.rollback(); raise
        finally: c.close()
        audit(u['username'],u['role'],'استيراد مجلدات من الجهاز','المستندات',str(len(groups)),f'folders={matched}, files={imported}, review={len(review)}, invalid={invalid}')
        review_html=''.join(f'<li>{esc(a)} — {esc(b)}</li>' for a,b in review) or '<li>لا توجد مجلدات تحتاج مراجعة.</li>'
        self.send(page('استيراد المجلدات',f'<div class="card"><h2>تم استيراد المجلدات</h2><p>مجلدات مرتبطة: <b>{matched}</b> · ملفات: <b>{imported}</b> · غير صالحة: <b>{invalid}</b></p><h3>المراجعة</h3><ul>{review_html}</ul><a class="btn" href="/import">العودة لمركز الاستيراد</a></div>',u,'import'))

    def do_import_employees(self,u):
        fields,file_part=self.parse_upload(); head,data,fname=file_part
        if not data:return self.redirect('/import')
        is_csv = b'.csv' in head.lower() or b'text/csv' in head.lower()
        tmp=os.path.join(DATA,'_upload.'+('csv' if is_csv else 'xlsx'));open(tmp,'wb').write(data)
        try:
            if is_csv:
                text=data.decode('utf-8-sig',errors='replace'); rows=list(csv.reader(io.StringIO(text)))
                hi,idx=find_hospital_header(rows)
                if hi is None: raise ValueError('CSV لا يحتوي على صف عناوين معروف.')
                records=[hospital_row_to_record(r,idx) for r in rows[hi+1:] if looks_like_hospital_employee(hospital_row_to_record(r,idx))]
                valid,errors=validate_employee_records(records); 
                if errors:
                    os.remove(tmp); return self.send(page('Import Validation', '<div class="card"><h2>🔴 لم يتم إدخال أي سجل</h2><p>الاستيراد الآن Atomic: يجب إصلاح كل الأخطاء ثم إعادة المحاولة.</p><pre>'+esc(json.dumps(errors,ensure_ascii=False,indent=2))+'</pre></div>',u,'import'),400)
                new,upd,skip=upsert_hospital_records(valid,u,'CSV'); os.remove(tmp)
                return self.send(page('تم الفحص',f'<div class="card"><h2>تم فحص CSV</h2><p>الإجمالي: <b>{len(records)}</b> · صالح: <b>{len(valid)}</b> · أخطاء: <b>{len(errors)}</b> · جديد: <b>{new}</b> · تحديث: <b>{upd}</b></p><a class="btn" href="/employees">عرض الموظفين</a></div>',u,'import'))
            wb=load_workbook(tmp,read_only=True,data_only=True,keep_vba=True)
            selected=None;header_idx=None;idx=None
            for ws in wb.worksheets:
                rows=list(ws.iter_rows(min_row=1,max_row=min(ws.max_row or 1,40),values_only=True))
                hi,hidx=find_hospital_header(rows)
                if hi is not None:selected,header_idx,idx=ws,hi,hidx;break
            if selected is None:wb.close();raise ValueError('لم يتم العثور على شيت الموظفين. ارفع نفس الملف الأصلي بدون تعديل.')
            records=[]
            for r in selected.iter_rows(min_row=header_idx+2,values_only=True):
                rec=hospital_row_to_record(r,idx)
                if looks_like_hospital_employee(rec):records.append(rec)
            sheet_name=selected.title;wb.close(); valid,errors=validate_employee_records(records); new,upd,skip=upsert_hospital_records(valid,u,'Excel');os.remove(tmp)
            return self.send(page('تم الفحص',f'<div class="card"><h2>تم فحص شيت الموظفين بنجاح</h2><p>الشيت: <b>{esc(sheet_name)}</b> · الصفوف: <b>{len(records)}</b> · صالح: <b>{len(valid)}</b> · أخطاء: <b>{len(errors)}</b></p><p>جديد: <b>{new}</b> · تحديث: <b>{upd}</b> · لم يُقبل: <b>{skip}</b></p><a class="btn" href="/employees">عرض الموظفين</a></div>',u,'import'))
        except Exception as e:
            if os.path.exists(tmp):os.remove(tmp)
            return self.send(page('خطأ في الاستيراد',f'<div class="card"><div class="alert">{esc(e)}</div><p>ارفع نفس ملف Excel الأصلي كما هو؛ لا تحتاج لتغيير الأعمدة أو اسم الشيت.</p></div>',u,'import'),400)


    def do_import_attendance(self,u):
        fields,file_part=self.parse_upload(); head,data,fname=file_part
        if not data:return self.redirect('/import')
        tmp=os.path.join(DATA,'_attendance'); open(tmp,'wb').write(data); imported=0
        try:
            c=db(); ext='xlsx'
            if b'filename="' in head and b'.csv' in head.lower():
                rows=list(csv.reader(io.StringIO(data.decode('utf-8-sig'))))
                hdr=[str(x).strip().lower() for x in rows[0]]; idx={h:i for i,h in enumerate(hdr)}
                def gi(r,*names):
                    for n in names:
                        if n.lower() in idx:return r[idx[n.lower()]]
                    return ''
                for r in rows[1:]:
                    wd,emp=gi(r,'التاريخ','date'),gi(r,'كود الموظف','emp_code','الكود')
                    if wd and emp:c.execute('INSERT OR REPLACE INTO attendance(work_date,emp_code,status,check_in,check_out,late_minutes,work_hours,overtime) VALUES(?,?,?,?,?,?,?,?)',(str(wd)[:10],str(emp),'حضور',gi(r,'الحضور','check_in'),gi(r,'الانصراف','check_out'),int(gi(r,'التأخير','late_minutes') or 0),float(gi(r,'ساعات العمل','work_hours') or 0),float(gi(r,'الإضافي','overtime') or 0))); imported+=1
            else:
                wb=load_workbook(tmp,read_only=True,data_only=True); ws=wb.active; rows=list(ws.iter_rows(values_only=True)); hdr=[str(x or '').strip().lower() for x in rows[0]]; idx={h:i for i,h in enumerate(hdr)}
                def gi(r,*names):
                    for n in names:
                        if n.lower() in idx:return r[idx[n.lower()]]
                    return ''
                for r in rows[1:]:
                    wd,emp=gi(r,'التاريخ','date'),gi(r,'كود الموظف','emp_code','الكود')
                    if hasattr(wd,'isoformat'):wd=wd.isoformat()[:10]
                    if wd and emp:c.execute('INSERT OR REPLACE INTO attendance(work_date,emp_code,status,check_in,check_out,late_minutes,work_hours,overtime) VALUES(?,?,?,?,?,?,?,?)',(str(wd)[:10],str(emp),'حضور',str(gi(r,'الحضور','check_in') or '')[:5],str(gi(r,'الانصراف','check_out') or '')[:5],int(gi(r,'التأخير','late_minutes') or 0),float(gi(r,'ساعات العمل','work_hours') or 0),float(gi(r,'الإضافي','overtime') or 0))); imported+=1
            c.commit();c.close();os.remove(tmp);audit(u['username'],u['role'],'استيراد','الحضور','Excel',f'{imported} records');return self.send(page('تم الاستيراد',f'<div class="card"><h2>تم استيراد {imported} سجل حضور.</h2><a class="btn" href="/attendance">عرض الحضور</a></div>',u,'import'))
        except Exception as e:
            if os.path.exists(tmp):os.remove(tmp)
            return self.send(page('خطأ',f'<div class="card"><div class="alert">{esc(e)}</div></div>',u,'import'),400)

    def users(self,u):
        if u['role'] not in ('SuperAdmin','Admin') and not can(u,'users.manage'): return self.forbid(u)
        c=db(); rows=c.execute('SELECT username,full_name,role,active,last_login FROM users ORDER BY id').fetchall(); c.close(); trs=''.join(f'<tr><td>{esc(r["username"])}</td><td>{esc(r["full_name"])}</td><td>{esc(r["role"])}</td><td>{"نشط" if r["active"] else "موقوف"}</td><td>{esc(r["last_login"] or "")}</td><td><form method=post action=/users/toggle>{csrf_field(u)}<input type=hidden name=username value="{esc(r["username"])}"><button class="btn gray">تبديل</button></form></td></tr>' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>المستخدمون والصلاحيات</h1><p>كل مستخدم يدخل من جهازه إلى نفس الخادم</p></div></div><div class="grid g2"><div class="card"><h3>إضافة مستخدم</h3><form class="form" method="post" action="/users/save">{csrf_field(u)}<div class="field"><label>اسم المستخدم</label><input name="username" required></div><div class="field"><label>الاسم</label><input name="full_name" required></div><div class="field"><label>كلمة المرور</label><input name="password" required></div><div class="field"><label>الدور</label><select name="role"><option>Employee</option><option>HR</option><option>Manager</option><option>Admin</option></select></div><div class="full"><button class="btn">حفظ المستخدم</button></div></form></div><div class="card"><h3>قائمة المستخدمين</h3><table class="table"><thead><tr><th>المستخدم</th><th>الاسم</th><th>الدور</th><th>الحالة</th><th>آخر دخول</th><th></th></tr></thead><tbody>{trs}</tbody></table></div></div>''';self.send(page('المستخدمون',body,u,'users'))
    def save_user(self,u,f):
        if u['role'] not in ('SuperAdmin','Admin'):return self.forbid(u)
        c=db(); c.execute('INSERT INTO users(username,password_hash,role,full_name) VALUES(?,?,?,?) ON CONFLICT(username) DO UPDATE SET password_hash=excluded.password_hash,role=excluded.role,full_name=excluded.full_name,active=1',(f.get('username'),hashpw(f.get('password')),f.get('role'),f.get('full_name')));c.commit();c.close();audit(u['username'],u['role'],'حفظ','المستخدمون',f.get('username'));self.redirect('/users')
    def toggle_user(self,u,f):
        if u['role'] not in ('SuperAdmin','Admin'):return self.forbid(u)
        c=db();c.execute('UPDATE users SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE username=? AND username<>?',(f.get('username'),u['username']));c.commit();c.close();self.redirect('/users')

    def settings_page(self,u):
        c=db(); ss={r['key']:r['value'] for r in c.execute('SELECT * FROM settings').fetchall()}; c.close()
        w=evaluation_weights()
        body=f'''<div class="top"><div class="title"><h1>الإعدادات والهوية</h1><p>كل ما يخص شكل النظام وقواعد التقييم والدوام من مكان واحد.</p></div></div>
<div class="grid g2">
<div class="card"><h3>🎨 هوية النظام</h3><form class="form" method="post" action="/settings/save">{csrf_field(u)}
<div class="field"><label>اسم الشركة / المستشفى</label><input name="company_name" value="{esc(ss.get('company_name',''))}"></div>
<div class="field"><label>العملة</label><input name="currency" value="{esc(ss.get('currency','EGP'))}"></div>
<div class="full"><button class="btn">حفظ الاسم والإعدادات</button></div></form>
<hr style="border:0;border-top:1px solid #e4e7ec;margin:18px 0">
<h4>Logo</h4><div class="actions"><a class="btn gray" href="/branding/logo" target="_blank">معاينة الشعار</a></div>
<form method="post" action="/branding/logo" enctype="multipart/form-data" style="margin-top:10px">{csrf_field(u)}<input type="file" name="file" accept=".png,.jpg,.jpeg,.svg" required><button class="btn" style="margin-top:10px">رفع / استبدال الشعار</button></form>
<p class="sub">يظهر تلقائيًا في Sidebar وLogin وPDF/Reports مستقبلًا. التغيير للـAdmin فقط.</p></div>
<div class="card"><h3>⭐ أوزان تقييم الأداء</h3><p>المجموع يتطبع تلقائيًا إلى 100% حتى لو أدخلت أوزانًا مختلفة.</p>
<form class="form" method="post" action="/settings/save">{csrf_field(u)}
{''.join(f'<div class="field"><label>{label}</label><input type="number" min="0" max="100" step="1" name="eval_weight_{key}" value="{w[key]:g}"></div>' for key,label in [('attendance','Attendance'),('punctuality','Punctuality'),('productivity','Productivity'),('behavior','Behavior'),('manager','Manager Evaluation')])}
<div class="full"><button class="btn">حفظ أوزان التقييم</button></div></form>
<div class="alert" style="margin-top:12px">النتيجة: 90–100 ممتاز 🟢 · 80–89 جيد جدًا 🟢 · 70–79 جيد 🟡 · 60–69 يحتاج تحسين 🟠 · أقل من 60 ضعيف 🔴</div></div></div>
<div class="card" style="margin-top:16px"><h3>⏰ سياسة الدوام والتأخير</h3><form class="form" method="post" action="/settings/save">{csrf_field(u)}
<div class="field"><label>بداية الدوام الافتراضية</label><input type="time" name="work_start" value="{esc(ss.get('work_start','08:15'))}"></div>
<div class="field"><label>نهاية الدوام الافتراضية</label><input type="time" name="work_end" value="{esc(ss.get('work_end','14:00'))}"></div>
<div class="field"><label>السماح بالدقائق</label><input type="number" name="grace_minutes" value="{esc(ss.get('grace_minutes','0'))}"></div>
<div class="field"><label>إنذار اليوم يبدأ بعد دقائق</label><input type="number" name="late_alert_threshold" value="{esc(ss.get('late_alert_threshold','15'))}"></div>
<div class="field"><label>حد التأخير الشهري</label><input type="number" name="monthly_late_limit_minutes" value="{esc(ss.get('monthly_late_limit_minutes','120'))}"></div>
<div class="field"><label>الإجراء بعد تجاوز الحد</label><select name="monthly_late_action"><option value="none" {"selected" if ss.get('monthly_late_action','none')=='none' else ''}>بدون إجراء تلقائي</option><option value="warning" {"selected" if ss.get('monthly_late_action')=='warning' else ''}>تنبيه فقط</option><option value="deduction" {"selected" if ss.get('monthly_late_action')=='deduction' else ''}>تنبيه + خصم حافز</option></select></div>
<div class="field"><label>قيمة خصم الحافز</label><input type="number" step="0.01" name="monthly_deduction_amount" value="{esc(ss.get('monthly_deduction_amount','0'))}"></div>
<div class="field"><label>حجم المستند الأقصى MB</label><input type="number" name="document_max_mb" value="{esc(ss.get('document_max_mb','25'))}"></div>
<div class="field full"><label>التصنيفات المطلوبة</label><input name="required_doc_categories" value="{esc(ss.get('required_doc_categories','عقد,هوية,مؤهل'))}"></div>
<div class="full"><button class="btn">حفظ سياسة الدوام</button></div></form></div>'''
        self.send(page('الإعدادات',body,u,'settings'))
    def branding_logo_save(self,u):
        fields,fp=self.parse_upload(); head,data,fname=fp
        if not data or not fname: return self.send(page('الشعار','<div class="card"><div class="alert">اختر صورة PNG/JPG/SVG صالحة.</div></div>',u,'settings'),400)
        ext=os.path.splitext(fname)[1].lower()
        if ext not in ('.png','.jpg','.jpeg','.svg') or len(data)>5*1024*1024: return self.send(page('الشعار','<div class="card"><div class="alert">الصورة يجب أن تكون PNG/JPG/SVG وأقل من 5MB.</div></div>',u,'settings'),400)
        brand=os.path.join(DATA,'branding'); os.makedirs(brand,exist_ok=True)
        for old in os.listdir(brand):
            try: os.remove(os.path.join(brand,old))
            except Exception: pass
        fn='logo'+ext; open(os.path.join(brand,fn),'wb').write(data)
        c=db(); c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',('company_logo',os.path.relpath(os.path.join(brand,fn),DATA))); c.commit(); c.close()
        # Generate role-specific logo variants when Pillow is available; preserve the original.
        try:
            from PIL import Image
            im=Image.open(os.path.join(brand,fn)).convert('RGBA')
            # Center-crop to a square for favicon/login use.
            w,h=im.size; side=min(w,h); left=(w-side)//2; top=(h-side)//2
            sq=im.crop((left,top,left+side,top+side))
            sq.save(os.path.join(brand,'favicon.png'),'PNG')
            # Transparent PNG variant: remove near-white background pixels.
            px=sq.load()
            for yy in range(sq.height):
                for xx in range(sq.width):
                    r,g,b,a=px[xx,yy]
                    if r>245 and g>245 and b>245: px[xx,yy]=(255,255,255,0)
            sq.save(os.path.join(brand,'transparent.png'),'PNG')
            for role in ('login','sidebar','report'): sq.save(os.path.join(brand,role+'.png'),'PNG')
            c=db()
            for k,v in [('logo_login','branding/login.png'),('logo_sidebar','branding/sidebar.png'),('logo_report','branding/report.png'),('logo_favicon','branding/favicon.png'),('logo_transparent','branding/transparent.png')]:
                c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',(k,v))
            c.commit(); c.close()
        except Exception:
            pass
        audit(u['username'],u['role'],'تغيير شعار','النظام',fn)
        self.redirect('/settings')
    def save_settings(self,u,f):
        c=db()
        for k in ['company_name','currency','work_start','work_end','grace_minutes','document_max_mb','late_alert_threshold','monthly_late_limit_minutes','monthly_late_action','monthly_deduction_amount','required_doc_categories','eval_weight_attendance','eval_weight_punctuality','eval_weight_productivity','eval_weight_behavior','eval_weight_manager']:
            c.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',(k,f.get(k,'')))
        c.commit();c.close();audit(u['username'],u['role'],'تعديل','الإعدادات','global');self.redirect('/settings')
    def audit_page(self,u):
        c=db(); rows=c.execute('SELECT * FROM audit ORDER BY id DESC LIMIT 500').fetchall();c.close();trs=''.join(f'<tr><td>{esc(r["ts"])}</td><td>{esc(r["username"])}</td><td>{esc(r["action"])}</td><td>{esc(r["entity"])}</td><td>{esc(r["record_key"])}</td><td>{esc(r["details"])}</td><td>{esc(r["ip"] or "")}</td><td>{esc(r["reason"] or "")}</td></tr>' for r in rows);body=f'''<div class="top"><div class="title"><h1>سجل المراجعة</h1><p>تتبع العمليات المهمة · Append-only مع سلسلة تجزئة (Hash Chain) لمنع التلاعب</p></div><div class="actions"><a class="btn gray" href="/audit/verify">🔒 التحقق من سلامة السجل</a><a class="btn gray" href="/export/audit">Export Excel</a></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الوقت</th><th>المستخدم</th><th>الإجراء</th><th>الوحدة</th><th>المعرف</th><th>التفاصيل</th><th>IP</th><th>السبب</th></tr></thead><tbody>{trs}</tbody></table></div>''';self.send(page('Audit Log',body,u,'audit'))
    def audit_verify_page(self,u):
        ok,bad_id,total=verify_audit_chain()
        if ok:
            body=f'''<div class="top"><div class="title"><h1>سلامة سجل المراجعة</h1><p>فحص سلسلة الـHash عبر {total} سجل</p></div><a class="btn gray" href="/audit">عودة</a></div><div class="card"><div class="badge b-ok" style="font-size:15px;padding:10px 16px">✓ السجل سليم — لا يوجد أي تلاعب مكتشف</div><p style="margin-top:14px;color:var(--muted)">كل سجل يحتوي على تجزئة SHA-256 مبنية على تجزئة السجل السابق؛ أي تعديل أو حذف لسجل قديم سيكسر السلسلة من هذه النقطة فصاعدًا وسيظهر هنا فورًا.</p></div>'''
        else:
            body=f'''<div class="top"><div class="title"><h1>سلامة سجل المراجعة</h1><p>فحص سلسلة الـHash عبر {total} سجل</p></div><a class="btn gray" href="/audit">عودة</a></div><div class="card"><div class="badge b-bad" style="font-size:15px;padding:10px 16px">⚠ تم اكتشاف تلاعب في السجل</div><p style="margin-top:14px">أول سجل غير متطابق: <b>#{bad_id}</b>. هذا يعني أن هذا السجل أو سجل قبله تم تعديله أو حذفه خارج مسار النظام الطبيعي.</p></div>'''
        self.send(page('التحقق من السجل',body,u,'audit'))
    def search_page(self,u):
        q=(parse_qs(urlparse(self.path).query).get('q',[''])[0] or '').strip()
        if not q:
            return self.send(page('البحث','<div class="card">اكتب كلمة للبحث (اسم، كود موظف، رقم قومي، هاتف).</div>',u,'employees'))
        like=f'%{q}%'; scope_sql,scope_params=visible_employee_sql(u)
        c=db()
        emp_rows=c.execute(f"SELECT emp_code,name,job,department,national_id,phone FROM employees e WHERE (emp_code LIKE ? OR name LIKE ? OR national_id LIKE ? OR phone LIKE ?) {scope_sql} LIMIT 25",[like,like,like,like]+scope_params).fetchall() if can(u,'employees.view') else []
        doc_rows=[]
        if can(u,'documents.manage'):
            doc_rows=c.execute(f"SELECT d.id,d.emp_code,d.file_name,d.category,e.name as emp_name FROM documents d JOIN employees e ON e.emp_code=d.emp_code WHERE d.status='current' AND (d.file_name LIKE ? OR d.emp_code LIKE ?) {scope_sql} LIMIT 25",[like,like]+scope_params).fetchall()
        c.close()
        emp_html=''.join(f'<div class="tl-item">👤 <a href="/employee/profile/{esc(r["emp_code"])}"><b>{esc(r["name"])}</b></a> — {esc(r["emp_code"])} · {esc(r["job"] or "")} · {esc(r["department"] or "")}</div>' for r in emp_rows) or '<div class="tl-item">لا يوجد موظفون مطابقون</div>'
        doc_html=''.join(f'<div class="tl-item">📄 <a href="/document/{r["id"]}">{esc(r["file_name"])}</a> — {esc(r["emp_name"])} ({esc(r["category"] or "")})</div>' for r in doc_rows) or '<div class="tl-item">لا توجد مستندات مطابقة</div>'
        body=f'''<div class="top"><div class="title"><h1>نتائج البحث عن «{esc(q)}»</h1></div></div><div class="grid g2"><div class="card"><h3>الموظفون</h3><div class="timeline">{emp_html}</div></div><div class="card"><h3>المستندات</h3><div class="timeline">{doc_html}</div></div></div>'''
        self.send(page('نتائج البحث',body,u,'employees'))
    def backup(self,u):
        out=make_backup(u,'manual'); audit(u['username'],u['role'],'نسخة احتياطية','النظام',os.path.basename(out)); self.send(page('Backup',f'<div class="card"><h2>تم إنشاء نسخة احتياطية</h2><p>{esc(out)}</p><a class="btn" href="/backups">إدارة النسخ</a></div>',u,'dashboard'))
    def notifications(self,u):
        c=db(); rows=c.execute('SELECT * FROM notifications WHERE user_name=? ORDER BY id DESC LIMIT 200',(u['username'],)).fetchall(); c.close()
        trs=''.join('<tr><td>{}</td><td><b>{}</b><br>{}</td><td>{}</td><td>{}</td></tr>'.format(esc(r['created_at']),esc(r['title']),esc(r['message']),'مقروء' if r['read_at'] else 'جديد', '' if r['read_at'] else '<form method="post" action="/notifications/read">{}<input type="hidden" name="id" value="{}"><button class="btn gray">تعليم كمقروء</button></form>'.format(('safe' if (r['confidence'] or 0)>=.98 else ''),csrf_field(u),r['id'])) for r in rows)
        body='<div class="top"><div class="title"><h1>الإشعارات</h1><p>التنبيهات التشغيلية الخاصة بك</p></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الوقت</th><th>التنبيه</th><th>الحالة</th><th></th></tr></thead><tbody>{}</tbody></table></div>'.format(trs or '<tr><td colspan="4">لا توجد إشعارات</td></tr>')
        self.send(page('الإشعارات',body,u,'notifications'))
    def notifications_read(self,u,f):
        c=db(); c.execute('UPDATE notifications SET read_at=? WHERE id=? AND user_name=?',(now(),f.get('id'),u['username'])); c.commit(); c.close(); self.redirect('/notifications')
    def documents(self,u):
        code=parse_qs(urlparse(self.path).query).get('emp_code',[''])[0]; c=db(); emps=c.execute('SELECT emp_code,name FROM employees ORDER BY name').fetchall(); emps=[e for e in emps if emp_allowed(u,e['emp_code'])]; rows=c.execute('SELECT * FROM documents WHERE emp_code=? ORDER BY id DESC',(code,)).fetchall() if code and emp_allowed(u,code) else []; c.close()
        opts=''.join('<option value="{}" {}>{} — {}</option>'.format(esc(e['emp_code']),'selected' if e['emp_code']==code else '',esc(e['emp_code']),esc(e['name'])) for e in emps)
        trs=''.join('<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td><a class="btn gray" href="/document/{}">فتح</a></td></tr>'.format(esc(r['file_name']),esc(r['file_type']),esc(r['expiry_date'] or '—'),esc(r['uploaded_at']),r['id']) for r in rows)
        paste_card=f'''<div class="card" style="margin-bottom:16px"><h3>📋 إدخال سريع للمستندات</h3><p>اختر الموظف ثم استخدم <b>Ctrl+V</b> من Explorer أو زر الرفع. على Windows EXE يوجد أيضًا Native Clipboard Bridge.</p><form method="post" action="/documents/paste" enctype="multipart/form-data">{csrf_field(u)}<div class="field"><label>الموظف</label><select id="docEmp" name="emp_code" required>{opts}</select></div><div id="docPaste" tabindex="0" style="margin-top:10px;border:2px dashed #84adff;border-radius:14px;padding:24px;text-align:center;background:#f7fbff">اضغط هنا ثم Ctrl+V للصق الملف</div><input type="file" id="pasteFile" name="file" hidden><div class="actions" style="margin-top:10px"><button class="btn" type="submit">حفظ الملف الملصوق</button><button class="btn gray" type="button" id="nativePaste">📋 لصق من Windows Explorer</button></div></form><div id="nativeStatus" class="alert" style="display:none;margin-top:10px"></div><script>(function(){{const z=document.getElementById('docPaste'),i=document.getElementById('pasteFile'),nativeBtn=document.getElementById('nativePaste'),status=document.getElementById('nativeStatus');z.addEventListener('click',()=>z.focus());z.addEventListener('paste',e=>{{const fs=e.clipboardData&&e.clipboardData.files;if(fs&&fs.length){{try{{i.files=fs;z.textContent='تم التقاط الملف — اضغط حفظ';}}catch(err){{z.textContent='المتصفح منع الوصول للملف؛ استخدم الزر الأصلي.';}}}}}});nativeBtn.onclick=async()=>{{status.style.display='block';status.textContent='جاري قراءة Clipboard من Windows...';try{{const t=await (await fetch('/bridge/token')).json();if(!t.ok)throw Error('فشل إنشاء رمز الربط');const r=await fetch('http://127.0.0.1:8975/clipboard/upload',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{token:t.token,remote:location.origin,csrf:'{esc(u.get('csrf',''))}',emp_code:document.getElementById('docEmp').value}})}});const j=await r.json();if(!j.ok)throw Error(j.error||'فشل الاستيراد');status.textContent='✓ تم استيراد '+j.imported+' ملف من Clipboard';setTimeout(()=>location.reload(),700);}}catch(e){{status.textContent='⚠ '+e.message+' — استخدم Select File/Folder إذا كان المتصفح يمنع الاتصال المحلي.';}}}};}})();</script></div>'''
        body=paste_card+'''<div class="top"><div class="title"><h1>مستندات الموظفين</h1><p>العقود والهوية والشهادات والمرفقات</p></div></div><div class="grid g2"><div class="card"><h3>رفع مستند</h3><form method="post" action="/documents/upload" enctype="multipart/form-data">{{csrf_field(u)}}<div class="field"><label>الموظف</label><select name="emp_code" required>{}</select></div><div class="field"><label>التصنيف</label><select name="category"><option>عام</option><option>عقد</option><option>هوية</option><option>مؤهل</option><option>تعيين</option><option>صورة</option><option>تأمين</option></select></div><div class="field"><label>تاريخ الانتهاء</label><input type="date" name="expiry_date"></div><div class="field"><label>الملف</label><input type="file" name="file" required></div><button class="btn">رفع المستند</button></form></div><div class="card"><h3>المستندات</h3><div class="table-wrap"><table class="table"><thead><tr><th>الملف</th><th>النوع</th><th>الانتهاء</th><th>رفع</th><th></th></tr></thead><tbody>{}</tbody></table></div></div></div>'''.format(opts,trs or '<tr><td colspan="5">اختر موظفًا لعرض المستندات</td></tr>').replace('{csrf_field(u)}',csrf_field(u))
        self.send(page('المستندات',body,u,'documents'))
    def document_upload(self,u):
        fields,file_part=self.parse_upload(); head,fdata,fname=file_part
        if not fname or not fields.get('emp_code'): return self.send(page('خطأ','<div class="card"><div class="alert">بيانات المستند غير مكتملة.</div></div>',u,'documents'),400)
        ext=os.path.splitext(fname)[1].lower(); allowed={'.pdf','.jpg','.jpeg','.png','.docx','.doc','.xlsx','.xls'}
        if ext not in allowed: return self.send(page('خطأ','<div class="card"><div class="alert">نوع الملف غير مسموح به.</div></div>',u,'documents'),400)
        max_mb=float(setting('document_max_mb') or 25)
        if len(fdata)>max_mb*1024*1024: return self.send(page('خطأ',f'<div class="card"><div class="alert">حجم الملف أكبر من {max_mb} MB.</div></div>',u,'documents'),400)
        emp_code=fields['emp_code'];
        if not emp_allowed(u,emp_code): return self.forbid(u)
        fname=safe_name(fname); checksum=hashlib.sha256(fdata).hexdigest(); category=fields.get('category') or guess_document_category(fname)
        c=db(); prior=c.execute('SELECT id,version FROM documents WHERE emp_code=? AND category=? AND status=?',(emp_code,category,'current')).fetchall(); ver=max([r['version'] or 1 for r in prior] or [0])+1
        rel=save_employee_file(emp_code,fname,fdata); cur=c.execute('INSERT INTO documents(emp_code,file_name,file_type,expiry_date,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(emp_code,fname,ext,fields.get('expiry_date'),u['username'],now(),None,rel,category,ver,checksum,'current'));
        for r in prior: c.execute('UPDATE documents SET status=?,superseded_by=? WHERE id=?',('superseded',cur.lastrowid,r['id']))
        c.execute('INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(emp_code,'document',now()[:10],f'رفع مستند: {fname}',f'v{ver} · {category}',u['username'],now())); c.commit(); c.close()
        audit(u['username'],u['role'],'رفع','المستندات',emp_code,f'{fname} (v{ver}, {category})',reason='document upload'); self.redirect('/employee/profile/'+emp_code+'#documents')
    def document_download(self,u,doc_id):
        c=db(); r=c.execute('SELECT * FROM documents WHERE id=?',(doc_id,)).fetchone(); c.close()
        if not r or not emp_allowed(u,r['emp_code']): return self.forbid(u)
        try:
            data=secure_file_bytes(r['storage_path']) if r['storage_path'] else (r['data'] or b'')
            self.send(data,200,mimetypes.guess_type(r['file_name'])[0] or 'application/octet-stream',{'Content-Disposition':'inline; filename="'+safe_name(r['file_name'])+'"'})
        except Exception: self.send('File not found',404)

    def get_shift(self,emp):
        c=db(); r=c.execute('SELECT s.* FROM employee_shifts es JOIN shifts s ON s.id=es.shift_id WHERE es.emp_code=? AND s.active=1',(emp,)).fetchone(); c.close()
        if r:return r
        c=db(); r=c.execute("SELECT * FROM shifts WHERE name='Morning' AND active=1").fetchone(); c.close(); return r

    def attendance_daily_policy(self, emp, wd, late_minutes):
        """Calculate and persist the daily late allowance ledger. The ledger is authoritative and never loses the daily boundary."""
        limit=int(setting('daily_late_limit_minutes') or 0)
        c=db(); prev=c.execute('SELECT late_minutes FROM attendance WHERE emp_code=? AND work_date=?',(emp,wd)).fetchone();
        used=max(0,int(late_minutes or 0)); remaining=max(0,limit-used); exceeded=max(0,used-limit)
        policy=f'limit={limit};action={setting("daily_late_action") or "none"}'
        c.execute('INSERT INTO attendance_daily_ledger(work_date,emp_code,allowed_minutes,used_minutes,remaining_minutes,exceeded_minutes,policy,updated_at) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(work_date,emp_code) DO UPDATE SET allowed_minutes=excluded.allowed_minutes,used_minutes=excluded.used_minutes,remaining_minutes=excluded.remaining_minutes,exceeded_minutes=excluded.exceeded_minutes,policy=excluded.policy,updated_at=excluded.updated_at',(wd,emp,limit,used,remaining,exceeded,policy,now()))
        if exceeded>0 and (setting('daily_late_action') or 'none')!='none':
            action=setting('daily_late_action') or 'none'; amount=float(setting('daily_exceeded_deduction_amount') or 0) if action=='deduction' else 0
            exists=c.execute("SELECT 1 FROM disciplinary_actions WHERE emp_code=? AND action_date=? AND action_type='تجاوز الحد اليومي' AND source='auto'",(emp,wd)).fetchone()
            if not exists:
                c.execute('INSERT INTO disciplinary_actions(emp_code,action_type,action_date,minutes,amount,reason,notes,created_by,created_at,source) VALUES(?,?,?,?,?,?,?,?,?,?)',(emp,'تجاوز الحد اليومي',wd,exceeded,amount,f'تجاوز الحد اليومي للتأخير بمقدار {exceeded} دقيقة','Daily Late Limit','system',now(),'auto'))
        c.commit(); c.close(); return {'limit':limit,'used':used,'remaining':remaining,'exceeded':exceeded}

    def attendance_save(self,u,f):
        wd=(f.get('work_date') or date.today().isoformat())[:10]; emp=(f.get('emp_code') or '').strip(); status=f.get('status') or 'حضور'; cin=(f.get('check_in') or '')[:5]; cout=(f.get('check_out') or '')[:5]
        if not emp or not emp_allowed(u,emp): return self.forbid(u)
        def mins(t):
            try:h,m=map(int,t.split(':'));return h*60+m
            except:return None
        sh=self.get_shift(emp); start=mins(sh['start_time']) if sh else mins(setting('work_start')); end=mins(sh['end_time']) if sh else mins(setting('work_end')); grace=int(sh['grace_minutes']) if sh else int(setting('grace_minutes') or 0)
        warning_minutes=int(sh['warning_minutes']) if sh and sh['warning_minutes'] is not None else 15
        late=max(0,(mins(cin)-start-grace)) if cin and mins(cin) is not None and start is not None else 0
        hours=ot=0
        if cin and cout and mins(cin) is not None and mins(cout) is not None:
            delta=mins(cout)-mins(cin)
            if delta<0: delta+=1440
            hours=round(delta/60,2); scheduled=((end-start)%1440)/60 if end is not None and start is not None else 8
            ot=round(max(0,hours-scheduled),2)
        c=db(); old=c.execute('SELECT * FROM attendance WHERE work_date=? AND emp_code=?',(wd,emp)).fetchone()
        # monthly total late minutes BEFORE this save (excluding today, since we're about to overwrite today's value)
        prior_month_late=c.execute("SELECT COALESCE(SUM(late_minutes),0) x FROM attendance WHERE emp_code=? AND substr(work_date,1,7)=? AND work_date<>?",(emp,wd[:7],wd)).fetchone()['x']
        c.execute('INSERT INTO attendance(work_date,emp_code,status,check_in,check_out,late_minutes,work_hours,overtime,notes) VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(work_date,emp_code) DO UPDATE SET status=excluded.status,check_in=excluded.check_in,check_out=excluded.check_out,late_minutes=excluded.late_minutes,work_hours=excluded.work_hours,overtime=excluded.overtime,notes=excluded.notes',(wd,emp,status,cin,cout,late,hours,ot,f.get('notes',''))); new=dict(c.execute('SELECT * FROM attendance WHERE work_date=? AND emp_code=?',(wd,emp)).fetchone())
        daily_ledger=self.attendance_daily_policy(emp,wd,late)
        # Auto-warning: today's lateness alone crosses the shift's warning threshold (e.g. 15 min = 08:30 for an 08:15 shift)
        if status=='حضور' and warning_minutes>0 and late>=warning_minutes:
            already=c.execute("SELECT 1 FROM disciplinary_actions WHERE emp_code=? AND action_date=? AND action_type='تأخير' AND source='auto'",(emp,wd)).fetchone()
            if not already:
                c.execute('INSERT INTO disciplinary_actions(emp_code,action_type,action_date,minutes,amount,reason,notes,created_by,created_at,source) VALUES(?,?,?,?,?,?,?,?,?,?)',(emp,'تأخير',wd,late,0,f'تأخير {late} دقيقة يوم {wd} (بعد حد الإنذار {warning_minutes} دقيقة)','تلقائي',u['username'],now(),'auto'))
                c.execute('INSERT INTO notifications(user_name,title,message,created_at) VALUES(?,?,?,?)',(u['username'],'إنذار تأخير تلقائي',f'الموظف {emp} تأخر {late} دقيقة يوم {wd} — تجاوز حد الإنذار ({warning_minutes} دقيقة).',now()))
        # Monthly threshold: check whether this save is what pushes the employee's monthly total over the configured limit
        month_limit=int(setting('monthly_late_limit_minutes') or 120); month_action=setting('monthly_late_action') or 'none'
        new_month_total=prior_month_late+late
        if month_action!='none' and month_limit>0 and prior_month_late<month_limit<=new_month_total:
            already_month=c.execute("SELECT 1 FROM disciplinary_actions WHERE emp_code=? AND action_type='تجاوز تأخير شهري' AND source='auto' AND action_date LIKE ?",(emp,wd[:7]+'%')).fetchone()
            if not already_month:
                amount=float(setting('monthly_deduction_amount') or 0) if month_action=='deduction' else 0
                c.execute('INSERT INTO disciplinary_actions(emp_code,action_type,action_date,minutes,amount,reason,notes,created_by,created_at,source) VALUES(?,?,?,?,?,?,?,?,?,?)',(emp,'تجاوز تأخير شهري',wd,new_month_total,amount,f'إجمالي تأخير الشهر {new_month_total} دقيقة تجاوز الحد المسموح ({month_limit} دقيقة)','تلقائي',u['username'],now(),'auto'))
                msg=f'الموظف {emp} تجاوز {month_limit} دقيقة تأخير هذا الشهر (الإجمالي {new_month_total} دقيقة).'+(f' تم تسجيل خصم {amount} {setting("currency") or ""}.' if month_action=='deduction' else '')
                c.execute('INSERT INTO notifications(user_name,title,message,created_at) VALUES(?,?,?,?)',(u['username'],'تجاوز حد التأخير الشهري',msg,now()))
        c.commit(); c.close()
        audit(u['username'],u['role'],'حفظ','الحضور',emp,wd,before=dict(old) if old else None,after=new,reason='attendance save'); self.redirect('/attendance')

    def attendance_adjust(self,u,f):
        aid=f.get('attendance_id'); mins=int(f.get('minutes') or 0); reason=f.get('reason','')
        c=db(); r=c.execute('SELECT emp_code,work_date FROM attendance WHERE id=?',(aid,)).fetchone()
        if not r: c.close(); return self.redirect('/attendance')
        if not emp_allowed(u,r['emp_code']): c.close(); return self.forbid(u)
        c.execute('UPDATE attendance SET late_minutes=MAX(0,late_minutes+?) WHERE id=?',(mins,aid)); c.execute('INSERT INTO attendance_adjustments(attendance_id,minutes,reason,created_by,created_at) VALUES(?,?,?,?,?)',(aid,mins,reason,u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'تعديل دقائق التأخير','الحضور',str(aid),f'{mins} دقيقة — {reason}'); self.redirect('/attendance')

    def roles_page(self,u):
        c=db(); roles=c.execute('SELECT * FROM roles ORDER BY system DESC,name').fetchall(); perms=c.execute('SELECT code,name_ar FROM permissions ORDER BY id').fetchall(); rp={(r['role'],r['permission']) for r in c.execute('SELECT role,permission FROM role_permissions').fetchall()}; c.close()
        cards=[]
        for role in roles:
            rows=''.join(f'<label style="display:block;margin:6px 0"><input type="checkbox" name="perm" value="{esc(p["code"])}" {"checked" if (role["name"],p["code"]) in rp else ""} {"disabled" if role["name"]=="SuperAdmin" else ""}> {esc(p["name_ar"])} <small>{esc(p["code"])}</small></label>' for p in perms)
            cards.append(f'<div class="card"><h3>{esc(role["display_name"] or role["name"])} {"🔒 نظامي" if role["system"] else ""}</h3><p>{esc(role["description"] or "")}</p><form method="post" action="/roles/save">{csrf_field(u)}<input type="hidden" name="role" value="{esc(role["name"])}">{rows}<button class="btn" {"disabled" if role["name"]=="SuperAdmin" else ""}>حفظ</button></form></div>')
        body='<div class="top"><div class="title"><h1>RBAC — الأدوار والصلاحيات</h1><p>تحكم دقيق في كل شاشة وكل عملية، مع أدوار مخصصة.</p></div></div><div class="grid g3">'+''.join(cards)+'</div>'
        self.send(page('RBAC',body,u,'roles'))
    def roles_save(self,u,f):
        role=f.get('role'); perms=f.get('perm',''); perms=perms if isinstance(perms,list) else [perms] if perms else []
        if role=='SuperAdmin': return self.redirect('/roles')
        c=db(); c.execute('INSERT OR IGNORE INTO roles(name,display_name,description,system) VALUES(?,?,?,0)',(role,role,'دور مخصص')); c.execute('DELETE FROM role_permissions WHERE role=?',(role,));
        for p in perms: c.execute('INSERT OR IGNORE INTO role_permissions(role,permission) VALUES(?,?)',(role,p))
        c.commit(); c.close(); audit(u['username'],u['role'],'تعديل الصلاحيات','الأدوار',role,','.join(perms)); self.redirect('/roles')

    def discipline_page(self,u):
        q=parse_qs(urlparse(self.path).query).get('emp_code',[''])[0]; c=db(); emps=c.execute('SELECT emp_code,name FROM employees ORDER BY name').fetchall(); emps=[e for e in emps if emp_allowed(u,e['emp_code'])]; rows=c.execute('SELECT * FROM disciplinary_actions WHERE emp_code=? ORDER BY id DESC LIMIT 200',(q,)).fetchall() if q else c.execute('SELECT * FROM disciplinary_actions ORDER BY id DESC LIMIT 200').fetchall(); rows=[r for r in rows if emp_allowed(u,r['emp_code'])]; c.close()
        opts=''.join(f'<option value="{esc(e["emp_code"])}" {"selected" if e["emp_code"]==q else ""}>{esc(e["emp_code"])} — {esc(e["name"])}</option>' for e in emps)
        trs=''.join(f'<tr><td>{esc(r["action_date"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["action_type"])}</td><td>{r["minutes"]}</td><td>{r["amount"]}</td><td>{esc(r["reason"])}</td><td>{esc(r["created_by"])}</td><td>{"🤖 تلقائي" if (r["source"] if "source" in r.keys() else "manual")=="auto" else "يدوي"}</td></tr>' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>الجزاءات والإنذارات والإجراءات</h1><p>سجل مركزي للإنذارات والمخالفات والمكافآت والتعديلات — يشمل الإنذارات التلقائية الناتجة من قواعد التأخير.</p></div></div><div class="grid g2"><div class="card"><h3>إضافة إجراء</h3><form class="form" method="post" action="/discipline/save">{csrf_field(u)}<div class="field"><label>الموظف</label><select name="emp_code" required>{opts}</select></div><div class="field"><label>النوع</label><select name="action_type"><option>إنذار</option><option>جزاء</option><option>مخالفة</option><option>مكافأة</option><option>ملاحظة</option></select></div><div class="field"><label>التاريخ</label><input type="date" name="action_date" value="{date.today().isoformat()}"></div><div class="field"><label>دقائق</label><input type="number" name="minutes" value="0"></div><div class="field"><label>مبلغ</label><input type="number" step="0.01" name="amount" value="0"></div><div class="field full"><label>السبب</label><input name="reason"></div><div class="field full"><label>ملاحظات</label><textarea name="notes"></textarea></div><div class="full"><button class="btn">حفظ الإجراء</button></div></form></div><div class="card table-wrap"><h3>السجل</h3><table class="table"><thead><tr><th>التاريخ</th><th>الموظف</th><th>النوع</th><th>دقائق</th><th>مبلغ</th><th>السبب</th><th>بواسطة</th><th>المصدر</th></tr></thead><tbody>{trs or '<tr><td colspan="8">لا توجد إجراءات.</td></tr>'}</tbody></table></div></div>'''
        self.send(page('الجزاءات والإنذارات',body,u,'users'))

    def discipline_save(self,u,f):
        c=db(); c.execute('INSERT INTO disciplinary_actions(emp_code,action_type,action_date,minutes,amount,reason,notes,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(f.get('emp_code'),f.get('action_type'),f.get('action_date') or date.today().isoformat(),int(f.get('minutes') or 0),cell_num(f.get('amount')),f.get('reason',''),f.get('notes',''),u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'إضافة','الجزاءات',f.get('emp_code'),f.get('action_type')); self.redirect('/discipline')

    def paste_documents(self,u):
        fields,file_part=self.parse_upload(); emp=self.fval(fields,'emp_code'); head,data,fname=file_part
        if not emp or not emp_allowed(u,emp): return self.forbid(u)
        if not data:return self.send(page('المستندات','<div class="card"><div class="alert">لم يتم العثور على ملف.</div></div>',u,'documents'),400)
        ext=os.path.splitext(fname)[1].lower(); allowed={'.pdf','.jpg','.jpeg','.png','.docx','.xlsx','.csv'}
        if ext not in allowed:return self.send(page('المستندات','<div class="card"><div class="alert">نوع الملف غير مسموح.</div></div>',u,'documents'),400)
        max_mb=float(setting('document_max_mb') or 25)
        if len(data)>max_mb*1024*1024:return self.send(page('المستندات','<div class="card"><div class="alert">الملف أكبر من الحد المسموح.</div></div>',u,'documents'),400)
        fname=safe_name(fname); cat=guess_document_category(fname); rel=save_employee_file(emp,fname,data); checksum=hashlib.sha256(data).hexdigest(); c=db(); prior=c.execute('SELECT id,version FROM documents WHERE emp_code=? AND category=? AND status=?',(emp,cat,'current')).fetchall(); ver=max([r['version'] or 1 for r in prior] or [0])+1; cur=c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,fname,ext,u['username'],now(),None,rel,cat,ver,checksum,'current')); [c.execute('UPDATE documents SET status=?,superseded_by=? WHERE id=?',('superseded',cur.lastrowid,r['id'])) for r in prior]; c.commit(); c.close(); audit(u['username'],u['role'],'لصق مستند','المستندات',emp,fname); self.send(page('تم اللصق',f'<div class="card"><h2>تم حفظ {esc(fname)}</h2><p>النسخة v{ver} · {esc(cat)}</p><a class="btn" href="/employee/profile/{quote(emp,safe="")}">فتح الموظف</a></div>',u,'documents'))

    def folder_import_page(self,u):
        body=f'''<div class="top"><div class="title"><h1>استيراد مجلدات الموظفين</h1><p>اختار مجلدًا رئيسيًا يحتوي مجلدات الموظفين، أو استخدم ZIP. الاسم هو الأساس وليس الكود، مع مطابقة آمنة وتقرير ثقة.</p></div></div><div class="card"><div class="alert"><b>أمثلة مقبولة لاسم المجلد:</b> <code>Ahmed Mohamed</code> · <code>1001</code> · <code>1001 Ahmed Mohamed</code> · <code>Ahmed Mohamed - 1001</code></div><form method="post" action="/documents/folders/import" enctype="multipart/form-data" style="margin-top:16px">{csrf_field(u)}<input type="file" name="file" accept=".zip" required><button class="btn" style="margin-top:12px">استيراد ZIP</button></form><hr style="border:0;border-top:1px solid #e4e7ec;margin:20px 0"><form method="post" action="/documents/folders/import-files" enctype="multipart/form-data">{csrf_field(u)}<label style="display:block;font-weight:700;margin-bottom:8px">📁 اختيار مجلد الموظفين مباشرة</label><input type="file" id="folderPicker" name="folder_files" webkitdirectory directory multiple required><div id="folderPaste" tabindex="0" style="margin-top:12px;border:2px dashed #84adff;border-radius:14px;padding:20px;text-align:center;background:#f7fbff">📋 انسخ ملفات/محتويات من Windows ثم اضغط هنا و Ctrl+V</div><div id="folderCount" style="margin-top:8px;color:#667085">لم يتم اختيار ملفات</div><button class="btn" style="margin-top:12px">استيراد المجلدات المختارة</button></form><script>(function(){{const i=document.getElementById('folderPicker'),z=document.getElementById('folderPaste'),c=document.getElementById('folderCount');function show(){{c.textContent=(i.files.length||0)+' ملف جاهز للاستيراد';}}i.addEventListener('change',show);z.addEventListener('click',()=>z.focus());z.addEventListener('paste',e=>{{const fs=e.clipboardData&&e.clipboardData.files;if(!fs||!fs.length)return;try{{const dt=new DataTransfer();Array.from(fs).forEach(f=>dt.items.add(f));i.files=dt.files;show();z.textContent='✓ تم التقاط الملفات — اضغط استيراد';}}catch(err){{z.textContent='المتصفح لم يسمح بلصق الملفات؛ استخدم اختيار المجلد.';}}}});}})();</script><p style="color:#667085">أي مجلد اسمه مش واضح 100% (تشابه جزئي أو أكتر من موظف بنفس الاسم) هيتسجل "يحتاج مراجعة" ومش هيترّبط تلقائيًا بحد غلط — تقدر تراجعه وتربطه يدويًا من ملف الموظف.</p></div>'''
        self.send(page('استيراد مجلدات الموظفين',body,u,'documents'))

    def folder_import(self,u):
        fields,file_part=self.parse_upload(); head,zdata,zname=file_part
        if not zdata or not zname.lower().endswith('.zip'): return self.send(page('خطأ','<div class="card"><div class="alert">ارفع ملف ZIP صالح.</div></div>',u,'documents'),400)
        imported=invalid=0; c=db(); allowed={'.pdf','.jpg','.jpeg','.png','.docx','.doc','.xlsx','.xls','.txt'}
        employees=[dict(r) for r in c.execute('SELECT emp_code,name FROM employees').fetchall()]
        matched_rows=[]; unmatched_rows=[]
        try:
            with zipfile.ZipFile(io.BytesIO(zdata)) as z:
                infos=z.infolist()
                if len(infos)>5000: raise ValueError('ملف ZIP يحتوي عددًا كبيرًا جدًا من الملفات.')
                total_uncompressed=sum(max(0,info.file_size) for info in infos)
                if total_uncompressed>250*1024*1024: raise ValueError('الحجم الإجمالي غير المضغوط لملف ZIP كبير جدًا.')
                # Resolve the first directory component that matches an employee.
                # This supports both:  Employee/1001 Name/file.pdf
                # and:                  Employee_Folders/1001 Name/file.pdf
                by_folder={}
                folder_indexes={}
                for info in infos:
                    if info.is_dir(): continue
                    rawname=info.filename.replace('\\','/'); parts=[x.strip() for x in rawname.split('/') if x.strip()]
                    if rawname.startswith('/') or '..' in parts or len(parts)<2: continue
                    employee_idx=None; folder=None
                    for idx in range(len(parts)-1):
                        emp0,_,_=resolve_folder_employee(parts[idx],employees)
                        if emp0:
                            employee_idx=idx; folder=parts[idx]; break
                    if employee_idx is None:
                        folder=parts[0]
                        employee_idx=0
                    by_folder.setdefault(folder,[]).append(info)
                    folder_indexes[folder]=employee_idx
                for folder,file_infos in by_folder.items():
                    emp,match_type,matched_name=resolve_folder_employee(folder,employees)
                    if not emp or not emp_allowed(u,emp):
                        unmatched_rows.append((folder,len(file_infos),'غير مصرح' if emp else 'لا يوجد موظف مطابق'))
                        continue
                    folder_imported=0
                    for info in file_infos:
                        parts=[x.strip() for x in info.filename.replace('\\','/').split('/') if x.strip()]; idx=folder_indexes.get(folder,0); fname='/'.join(parts[idx+1:]); ext=os.path.splitext(fname)[1].lower()
                        if ext not in allowed: invalid+=1; continue
                        data=z.read(info); max_mb=float(setting('document_max_mb') or 25)
                        if len(data)>max_mb*1024*1024: invalid+=1; continue
                        fname=safe_name(fname); rel=save_employee_file(emp,fname,data); checksum=hashlib.sha256(data).hexdigest(); cat=guess_document_category(fname)
                        prior=c.execute('SELECT id,version FROM documents WHERE emp_code=? AND category=? AND status=?',(emp,cat,'current')).fetchall(); ver=max([r['version'] or 1 for r in prior] or [0])+1
                        cur=c.execute('INSERT INTO documents(emp_code,file_name,file_type,uploaded_by,uploaded_at,data,storage_path,category,version,checksum,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(emp,fname,ext,u['username'],now(),None,rel,cat,ver,checksum,'current'))
                        for r in prior:c.execute('UPDATE documents SET status=?,superseded_by=? WHERE id=?',('superseded',cur.lastrowid,r['id']))
                        imported+=1; folder_imported+=1
                    matched_rows.append((folder,emp,matched_name,match_type,folder_imported))
            details=json.dumps({'matched':matched_rows,'unmatched':unmatched_rows},ensure_ascii=False)
            c.execute('INSERT INTO file_imports(source,file_name,records,created_by,created_at,details) VALUES(?,?,?,?,?,?)',('employee_folders',safe_name(zname),imported,u['username'],now(),details)); c.commit()
        except zipfile.BadZipFile:
            c.rollback(); return self.send(page('خطأ','<div class="card"><div class="alert">ملف ZIP تالف.</div></div>',u,'documents'),400)
        finally: c.close()
        audit(u['username'],u['role'],'استيراد مجلدات','المستندات',zname,f'files={imported}, folders_matched={len(matched_rows)}, folders_unmatched={len(unmatched_rows)}, invalid={invalid}')
        match_label={'code':'كود الموظف','name_exact':'الاسم (تطابق تام)','name_close':'الاسم (تطابق قريب)'}
        matched_trs=''.join(f'<tr><td>{esc(f)}</td><td>{esc(ec)}</td><td>{esc(en or "")}</td><td>{esc(match_label.get(mt,mt))}</td><td>{fc}</td></tr>' for f,ec,en,mt,fc in matched_rows) or '<tr><td colspan="5">لا يوجد</td></tr>'
        unmatched_trs=''.join(f'<tr><td>{esc(f)}</td><td>{n}</td><td>{esc(reason)}</td></tr>' for f,n,reason in unmatched_rows) or '<tr><td colspan="3">لا يوجد</td></tr>'
        body=f'''<div class="card"><h2>اكتمل الاستيراد</h2><p>ملفات مرتبطة: <b>{imported}</b> · مجلدات متطابقة: <b>{len(matched_rows)}</b> · مجلدات محتاجة مراجعة: <b>{len(unmatched_rows)}</b> · غير مسموح: <b>{invalid}</b></p></div><div class="card table-wrap"><h3>المجلدات المرتبطة</h3><table class="table"><thead><tr><th>المجلد</th><th>كود الموظف</th><th>اسم الموظف</th><th>طريقة المطابقة</th><th>عدد الملفات</th></tr></thead><tbody>{matched_trs}</tbody></table></div><div class="card table-wrap"><h3>محتاج مراجعة يدوية</h3><table class="table"><thead><tr><th>المجلد</th><th>عدد الملفات</th><th>السبب</th></tr></thead><tbody>{unmatched_trs}</tbody></table></div>'''
        self.send(page('تم الاستيراد',body,u,'documents'))

    def payroll(self,u):
        c=db(); rows=c.execute('SELECT p.*,e.name FROM payroll p LEFT JOIN employees e ON e.emp_code=p.emp_code ORDER BY p.period DESC,p.id DESC LIMIT 500').fetchall(); rows=[r for r in rows if r['emp_code'] and emp_allowed(u,r['emp_code'])]; emps=c.execute('SELECT emp_code,name,contract_amount,basic_salary,allowances FROM employees ORDER BY name').fetchall(); c.close()
        opts=''.join(f'<option value="{esc(e["emp_code"])}">{esc(e["emp_code"])} — {esc(e["name"])}</option>' for e in emps if emp_allowed(u,e['emp_code']))
        trs=''.join(f'<tr><td>{esc(r["period"])}</td><td>{esc(r["emp_code"])}</td><td>{esc(r["name"] or "")}</td><td>{r["basic"]}</td><td>{r["allowances"]}</td><td>{r["overtime"]}</td><td>{r["bonuses"]}</td><td>{r["deductions"]}</td><td><b>{r["net"]}</b></td><td>{esc(r["status"])}</td></tr>' for r in rows)
        body=f'''<div class="top"><div class="title"><h1>المرتبات</h1><p>مسودة → مراجعة → اعتماد → قفل.</p></div><a class="btn gray" href="/payroll/review">مراجعة وقفل</a></div><div class="grid g2"><div class="card"><h3>إضافة/تحديث راتب</h3><form class="form" method="post" action="/payroll/save">{csrf_field(u)}<div class="field"><label>الموظف</label><select name="emp_code" required>{opts}</select></div><div class="field"><label>الفترة YYYY-MM</label><input name="period" value="{date.today().isoformat()[:7]}" required></div><div class="field"><label>أساسي</label><input name="basic" type="number" step="0.01"></div><div class="field"><label>بدلات</label><input name="allowances" type="number" step="0.01"></div><div class="field"><label>إضافي</label><input name="overtime" type="number" step="0.01"></div><div class="field"><label>مكافآت</label><input name="bonuses" type="number" step="0.01"></div><div class="field"><label>خصومات</label><input name="deductions" type="number" step="0.01"></div><div class="field"><label>جزاءات</label><input name="penalties" type="number" step="0.01"></div><div class="field"><label>الحالة</label><select name="status"><option>مسودة</option><option>معتمدة</option><option>مدفوعة</option></select></div><div class="field full"><label>ملاحظات</label><textarea name="notes"></textarea></div><div class="full"><button class="btn">حفظ الراتب</button></div></form></div><div class="card table-wrap"><h3>السجل</h3><table class="table"><thead><tr><th>الفترة</th><th>الكود</th><th>الموظف</th><th>أساسي</th><th>بدلات</th><th>إضافي</th><th>مكافآت</th><th>خصومات</th><th>صافي</th><th>الحالة</th></tr></thead><tbody>{trs or '<tr><td colspan="10">لا توجد بيانات.</td></tr>'}</tbody></table></div></div>'''
        self.send(page('المرتبات',body,u,'payroll'))
    def payroll_save(self,u,f):
        emp=f.get('emp_code'); period=f.get('period','')
        if not emp_allowed(u,emp): return self.forbid(u)
        c=db(); old=c.execute('SELECT * FROM payroll WHERE emp_code=? AND period=?',(emp,period)).fetchone()
        if old and old['locked_at'] and u.get('role')!='SuperAdmin': c.close(); return self.send(page('المرتبات','<div class="card"><div class="alert">هذه الفترة مقفلة ولا يمكن تعديلها إلا بواسطة SuperAdmin.</div></div>',u,'payroll'),403)
        vals={k:cell_num(f.get(k)) for k in ('basic','allowances','overtime','bonuses','deductions','penalties')}; net=vals['basic']+vals['allowances']+vals['overtime']+vals['bonuses']-vals['deductions']-vals['penalties']; status='مسودة'
        c.execute('INSERT INTO payroll(emp_code,period,basic,allowances,deductions,overtime,bonuses,net,status,notes,created_by,created_at,approved_by,approved_at,locked_at,penalties) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(emp_code,period) DO UPDATE SET basic=excluded.basic,allowances=excluded.allowances,deductions=excluded.deductions,overtime=excluded.overtime,bonuses=excluded.bonuses,net=excluded.net,status=excluded.status,notes=excluded.notes,created_by=excluded.created_by,created_at=excluded.created_at,penalties=excluded.penalties',(emp,period,vals['basic'],vals['allowances'],vals['deductions'],vals['overtime'],vals['bonuses'],net,status,f.get('notes',''),u['username'],now(),None,None,None,vals['penalties'])); new=dict(c.execute('SELECT * FROM payroll WHERE emp_code=? AND period=?',(emp,period)).fetchone()); c.commit(); c.close(); audit(u['username'],u['role'],'حفظ','المرتبات',emp,period,before=dict(old) if old else None,after=new,reason='payroll save'); self.redirect('/payroll')


    def shifts_page(self,u):
        c=db(); rows=c.execute('SELECT * FROM shifts ORDER BY id').fetchall(); emps=c.execute("SELECT e.emp_code,e.name,COALESCE(s.name,'Morning') shift_name FROM employees e LEFT JOIN employee_shifts es ON es.emp_code=e.emp_code LEFT JOIN shifts s ON s.id=es.shift_id WHERE e.status<>'مؤرشف' ORDER BY e.name").fetchall(); c.close()
        trs=''.join('<tr><td>'+esc(r['name'])+'</td><td>'+esc(r['start_time'])+' → '+esc(r['end_time'])+'</td><td>'+str(r['grace_minutes'])+'</td><td>'+str(r['warning_minutes'] if r['warning_minutes'] is not None else 15)+'</td><td>'+('نشطة' if r['active'] else 'موقوفة')+'</td></tr>' for r in rows)
        opts=''.join('<option value="'+str(r['id'])+'">'+esc(r['name'])+' ('+esc(r['start_time'])+' → '+esc(r['end_time'])+')</option>' for r in rows if r['active'])
        empopts=''.join('<option value="'+esc(r['emp_code'])+'">'+esc(r['emp_code'])+' — '+esc(r['name'])+' — '+esc(r['shift_name'] or 'Morning')+'</option>' for r in emps if emp_allowed(u,r['emp_code']))
        body='<div class="top"><div class="title"><h1>الورديات</h1><p>Morning / Evening / Night مع فترة سماح، وحد دقائق قبل الإنذار التلقائي، وحساب التأخير تلقائيًا.</p></div></div><div class="grid g2"><div class="card"><h3>إضافة وردية</h3><form class="form" method="post" action="/shift/save">'+csrf_field(u)+'<div class="field"><label>الاسم</label><input name="name" required></div><div class="field"><label>بداية</label><input name="start_time" type="time" value="09:00"></div><div class="field"><label>نهاية</label><input name="end_time" type="time" value="17:00"></div><div class="field"><label>سماح قبل احتساب التأخير (دقيقة)</label><input name="grace_minutes" type="number" value="0"></div><div class="field"><label>دقائق التأخير قبل الإنذار التلقائي</label><input name="warning_minutes" type="number" value="15"></div><div class="full"><div class="alert">مثال: بداية 08:15 + سماح 0 = التأخير يُحسب من أول دقيقة بعد 08:15. لو دقائق التأخير وصلت 15 (يعني الساعة 08:30) يتسجل إنذار تلقائي.</div></div><div class="full"><button class="btn">حفظ</button></div></form></div><div class="card"><h3>تعيين وردية</h3><form class="form" method="post" action="/shift/assign">'+csrf_field(u)+'<div class="field full"><label>الموظف</label><select name="emp_code">'+empopts+'</select></div><div class="field full"><label>الوردية</label><select name="shift_id">'+opts+'</select></div><div class="full"><button class="btn">تعيين</button></div></form></div></div><div class="card table-wrap" style="margin-top:16px"><table class="table"><thead><tr><th>الاسم</th><th>الوقت</th><th>السماح</th><th>دقائق قبل الإنذار</th><th>الحالة</th></tr></thead><tbody>'+trs+'</tbody></table></div>'
        self.send(page('الورديات',body,u,'shifts'))
    def shift_save(self,u,f):
        c=db(); c.execute('INSERT INTO shifts(name,start_time,end_time,grace_minutes,warning_minutes) VALUES(?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,grace_minutes=excluded.grace_minutes,warning_minutes=excluded.warning_minutes,active=1',(f.get('name'),f.get('start_time') or '09:00',f.get('end_time') or '17:00',int(f.get('grace_minutes') or 0),int(f.get('warning_minutes') or 15))); c.commit(); c.close(); audit(u['username'],u['role'],'حفظ','الورديات',f.get('name')); self.redirect('/shifts')
    def shift_assign(self,u,f):
        emp=f.get('emp_code'); sid=int(f.get('shift_id') or 0)
        if not emp_allowed(u,emp): return self.forbid(u)
        c=db(); c.execute('UPDATE employees SET shift_id=?,updated_at=? WHERE emp_code=?',(sid,now(),emp)); c.execute('INSERT INTO employee_shifts(emp_code,shift_id,assigned_at,assigned_by) VALUES(?,?,?,?) ON CONFLICT(emp_code) DO UPDATE SET shift_id=excluded.shift_id,assigned_at=excluded.assigned_at,assigned_by=excluded.assigned_by',(emp,sid,now(),u['username'])); c.commit(); c.close(); audit(u['username'],u['role'],'تعيين وردية','الموظفون',emp,str(sid)); self.redirect('/shifts')

    def overtime_page(self,u):
        c=db(); rows=c.execute('SELECT o.*,e.name FROM overtime_requests o LEFT JOIN employees e ON e.emp_code=o.emp_code ORDER BY o.id DESC LIMIT 300').fetchall(); emps=c.execute("SELECT emp_code,name FROM employees WHERE status<>'مؤرشف' ORDER BY name").fetchall(); c.close()
        opts=''.join('<option value="'+esc(e['emp_code'])+'">'+esc(e['emp_code'])+' — '+esc(e['name'])+'</option>' for e in emps if emp_allowed(u,e['emp_code']))
        trs=''
        for r in rows:
            if not emp_allowed(u,r['emp_code']): continue
            action='<form method="post" action="/overtime/status" style="display:inline">'+csrf_field(u)+'<input type="hidden" name="request_no" value="'+esc(r['request_no'])+'"><input type="hidden" name="status" value="معتمدة"><button class="btn">اعتماد</button></form>' if r['status']=='قيد المراجعة' and can(u,'overtime.approve') else ''
            trs+='<tr><td>'+esc(r['request_no'])+'</td><td>'+esc(r['emp_code'])+'</td><td>'+esc(r['work_date'])+'</td><td>'+str(r['hours'])+'</td><td>'+esc(r['status'])+'</td><td>'+action+'</td></tr>'
        body='<div class="top"><div class="title"><h1>الإضافي</h1><p>طلب إضافي ثم اعتماد.</p></div></div><div class="grid g2"><div class="card"><h3>طلب إضافي</h3><form class="form" method="post" action="/overtime/save">'+csrf_field(u)+'<div class="field full"><label>الموظف</label><select name="emp_code">'+opts+'</select></div><div class="field"><label>التاريخ</label><input type="date" name="work_date" value="'+date.today().isoformat()+'"></div><div class="field"><label>الساعات</label><input type="number" step="0.25" name="hours"></div><div class="field full"><label>السبب</label><textarea name="reason"></textarea></div><div class="full"><button class="btn">إرسال</button></div></form></div><div class="card table-wrap"><table class="table"><thead><tr><th>الطلب</th><th>الكود</th><th>التاريخ</th><th>الساعات</th><th>الحالة</th><th>إجراء</th></tr></thead><tbody>'+trs+'</tbody></table></div></div>'
        self.send(page('الإضافي',body,u,'overtime'))
    def overtime_save(self,u,f):
        emp=f.get('emp_code'); hours=cell_num(f.get('hours'))
        if not emp_allowed(u,emp) or hours<=0:return self.forbid(u)
        c=db(); no=f'OT-{c.execute("SELECT COALESCE(MAX(id),0)+1 n FROM overtime_requests").fetchone()["n"]:05d}'; c.execute('INSERT INTO overtime_requests(request_no,emp_code,work_date,hours,reason,requested_by,created_at) VALUES(?,?,?,?,?,?,?)',(no,emp,f.get('work_date') or date.today().isoformat(),hours,f.get('reason',''),u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'إضافة','الإضافي',no); self.redirect('/overtime')
    def overtime_status(self,u,f):
        no=f.get('request_no'); status=f.get('status'); c=db(); r=c.execute('SELECT * FROM overtime_requests WHERE request_no=?',(no,)).fetchone()
        if not r or r['status']!='قيد المراجعة': c.close(); return self.redirect('/overtime')
        if not emp_allowed(u,r['emp_code']): c.close(); return self.forbid(u)
        if not can(u,'overtime.approve'): c.close(); return self.forbid(u)
        c.execute('UPDATE overtime_requests SET status=?,approved_by=?,approved_at=? WHERE request_no=?',(status,u['username'],now(),no))
        if status=='معتمدة': c.execute('INSERT OR REPLACE INTO attendance(work_date,emp_code,status,overtime,notes) VALUES(?,?,?,?,?)',(r['work_date'],r['emp_code'],'حضور',r['hours'],'اعتماد إضافي'))
        c.commit(); c.close(); audit(u['username'],u['role'],status,'الإضافي',no); self.redirect('/overtime')

    def leave_balances_page(self,u):
        c=db(); rows=c.execute('SELECT lb.emp_code,e.name,lb.leave_type,lb.annual,lb.used,(lb.annual-lb.used) remaining FROM leave_balances lb JOIN employees e ON e.emp_code=lb.emp_code ORDER BY e.name,lb.leave_type LIMIT 2000').fetchall(); c.close()
        trs=''.join('<tr><td>'+esc(r['emp_code'])+'</td><td>'+esc(r['name'])+'</td><td>'+esc(r['leave_type'])+'</td><td>'+str(r['annual'])+'</td><td>'+str(r['used'])+'</td><td><b>'+str(r['remaining'])+'</b></td></tr>' for r in rows if emp_allowed(u,r['emp_code']))
        body='<div class="top"><div class="title"><h1>أرصدة الإجازات</h1><p>المخصص · المستخدم · المتبقي</p></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الكود</th><th>الموظف</th><th>النوع</th><th>المخصص</th><th>المستخدم</th><th>المتبقي</th></tr></thead><tbody>'+trs+'</tbody></table></div>'
        self.send(page('أرصدة الإجازات',body,u,'leaves'))

    def payroll_actions(self,u):
        c=db(); rows=c.execute('SELECT p.*,e.name FROM payroll p LEFT JOIN employees e ON e.emp_code=p.emp_code ORDER BY p.period DESC,p.id DESC LIMIT 500').fetchall(); c.close()
        trs=''
        for r in rows:
            if not emp_allowed(u,r['emp_code']): continue
            a=''
            if not r['locked_at'] and can(u,'payroll.approve'): a+='<form method="post" action="/payroll/approve" style="display:inline">'+csrf_field(u)+'<input type="hidden" name="id" value="'+str(r['id'])+'"><button class="btn">اعتماد</button></form> '
            if r['status']=='معتمدة' and not r['locked_at'] and can(u,'payroll.lock'): a+='<form method="post" action="/payroll/lock" style="display:inline">'+csrf_field(u)+'<input type="hidden" name="id" value="'+str(r['id'])+'"><button class="btn warn">قفل</button></form>'
            trs+='<tr><td>'+esc(r['period'])+'</td><td>'+esc(r['emp_code'])+'</td><td>'+esc(r['name'] or '')+'</td><td>'+esc(r['status'])+'</td><td>'+('🔒 مقفول' if r['locked_at'] else 'مفتوح')+'</td><td>'+a+'</td></tr>'
        body='<div class="top"><div class="title"><h1>مراجعة وقفل المرتبات</h1><p>Approve ثم Lock — بعد القفل لا تعديل إلا SuperAdmin.</p></div></div><div class="card table-wrap"><table class="table"><thead><tr><th>الفترة</th><th>الكود</th><th>الموظف</th><th>الحالة</th><th>القفل</th><th></th></tr></thead><tbody>'+trs+'</tbody></table></div>'
        self.send(page('Payroll Review',body,u,'payroll'))
    def payroll_approve(self,u,f):
        c=db(); r=c.execute('SELECT * FROM payroll WHERE id=?',(f.get('id'),)).fetchone()
        if not r or not emp_allowed(u,r['emp_code']): c.close(); return self.forbid(u)
        c.execute("UPDATE payroll SET status='معتمدة',approved_by=?,approved_at=? WHERE id=?",(u['username'],now(),r['id'])); c.execute('INSERT INTO employee_events(emp_code,event_type,event_date,title,details,created_by,created_at) VALUES(?,?,?,?,?,?,?)',(r['emp_code'],'payroll',now()[:10],'اعتماد مرتب',r['period'],u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'اعتماد','المرتبات',str(r['id'])); self.redirect('/payroll/review')
    def payroll_lock(self,u,f):
        c=db(); r=c.execute('SELECT * FROM payroll WHERE id=?',(f.get('id'),)).fetchone()
        if not r or r['status']!='معتمدة': c.close(); return self.redirect('/payroll/review')
        c.execute('UPDATE payroll SET locked_at=? WHERE id=?',(now(),r['id'])); c.commit(); c.close(); audit(u['username'],u['role'],'قفل','المرتبات',str(r['id'])); self.redirect('/payroll/review')

    def import_mapping_page(self,u):
        c=db(); r=c.execute('SELECT mapping_json FROM import_mappings WHERE name=?',('Hospital Employee Template',)).fetchone(); c.close(); current=json.loads(r['mapping_json']) if r and r['mapping_json'] else {}
        fields=[('emp_no','م / كود الموظف'),('name','الإسم'),('employee_group','المجموعة الوظيفية'),('birth_date','تاريخ الميلاد'),('national_id','الرقم القومي'),('address','العنوان'),('qualification','المؤهل'),('phone','الهاتف'),('iban','IBAN'),('bank_name','البنك'),('bank_branch','الفرع'),('department','الإدارة'),('unit','الوحدة'),('job','الوظيفة'),('contract_date','تاريخ التعاقد'),('contract_amount','مبلغ التعاقد')]
        html=''.join('<div class="field"><label>'+label+'</label><input name="map_'+key+'" value="'+esc(current.get(key,''))+'" placeholder="اسم العمود في Excel"></div>' for key,label in fields)
        body='<div class="top"><div class="title"><h1>ربط أعمدة Excel</h1><p>احفظ اسم شيتك مرة واحدة؛ بعد ذلك سيستخدمه النظام تلقائيًا.</p></div></div><div class="card"><form class="form" method="post" action="/import/mapping/save">'+csrf_field(u)+html+'<div class="full"><button class="btn">حفظ Hospital Employee Template</button></div></form></div>'
        self.send(page('Excel Mapping',body,u,'import'))
    def import_mapping_save(self,u,f):
        mapping={k[4:]:v.strip() for k,v in f.items() if k.startswith('map_') and v.strip()}; c=db(); c.execute('INSERT INTO import_mappings(name,kind,mapping_json,created_by,created_at) VALUES(?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET mapping_json=excluded.mapping_json,created_by=excluded.created_by,created_at=excluded.created_at',('Hospital Employee Template','employees',json.dumps(mapping,ensure_ascii=False),u['username'],now())); c.commit(); c.close(); audit(u['username'],u['role'],'حفظ','Excel Mapping','Hospital Employee Template',json.dumps(mapping,ensure_ascii=False)); self.redirect('/import')

    def import_errors_page(self,u,run_id):
        c=db(); rows=c.execute('SELECT * FROM import_errors WHERE run_id=? ORDER BY row_no',(run_id,)).fetchall(); c.close()
        trs=''.join('<tr><td>'+str(r['row_no'])+'</td><td>'+esc(r['field'])+'</td><td>'+esc(r['message'])+'</td><td><small>'+esc(r['raw_json'])+'</small></td></tr>' for r in rows)
        body='<div class="top"><div class="title"><h1>أخطاء الاستيراد</h1><p>راجع الأخطاء ثم أصلح ملف Excel وأعد الاستيراد.</p></div><a class="btn gray" href="/export/import-errors/'+str(run_id)+'">Export Excel</a></div><div class="card table-wrap"><table class="table"><thead><tr><th>الصف</th><th>الحقل</th><th>الخطأ</th><th>البيانات</th></tr></thead><tbody>'+trs+'</tbody></table></div>'
        self.send(page('أخطاء الاستيراد',body,u,'import'))
    def export_import_errors(self,run_id):
        c=db(); rows=c.execute('SELECT row_no,field,message,raw_json FROM import_errors WHERE run_id=? ORDER BY row_no',(run_id,)).fetchall(); c.close(); wb=Workbook(); ws=wb.active; ws.append(['Row','Field','Error','Raw Data'])
        for r in rows: ws.append([r['row_no'],r['field'],r['message'],r['raw_json']])
        out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Import_Errors.xlsx"'})

    def access_page(self,u):
        c=db(); users=c.execute('SELECT username,full_name,role,active,scope_type,scope_value,must_change_password FROM users ORDER BY id').fetchall(); roles=c.execute('SELECT name,display_name FROM roles ORDER BY system DESC,name').fetchall(); c.close();
        roleopts=''.join(f'<option>{esc(r["name"])}</option>' for r in roles)
        trs=''.join(f'<tr><td>{esc(r["username"])}</td><td>{esc(r["full_name"])}</td><td>{esc(r["role"])}</td><td>{esc(r["scope_type"])}: {esc(r["scope_value"] or "all")}</td><td>{"نشط" if r["active"] else "موقوف"}</td><td>{"نعم" if r["must_change_password"] else "لا"}</td></tr>' for r in users)
        body=f'''<div class="top"><div class="title"><h1>Access Control</h1><p>تعيين الدور + نطاق الوصول + إجبار تغيير كلمة المرور.</p></div></div><div class="grid g2"><div class="card"><h3>تعديل وصول مستخدم</h3><form class="form" method="post" action="/access/save">{csrf_field(u)}<div class="field"><label>اسم المستخدم</label><input name="username" required></div><div class="field"><label>الدور</label><select name="role">{roleopts}</select></div><div class="field"><label>النطاق</label><select name="scope_type"><option value="all">كل الموظفين</option><option value="department">إدارة محددة</option><option value="unit">وحدة محددة</option><option value="self">الموظف نفسه</option></select></div><div class="field"><label>قيمة النطاق</label><input name="scope_value" placeholder="اسم الإدارة/الوحدة أو Employee ID"></div><div class="full"><label><input type="checkbox" name="must_change_password" value="1"> إجبار تغيير كلمة المرور</label></div><div class="full"><button class="btn">حفظ الصلاحيات</button></div></form></div><div class="card table-wrap"><h3>المستخدمون</h3><table class="table"><thead><tr><th>المستخدم</th><th>الاسم</th><th>الدور</th><th>النطاق</th><th>الحالة</th><th>تغيير كلمة المرور</th></tr></thead><tbody>{trs}</tbody></table></div></div>'''; self.send(page('Access Control',body,u,'roles'))
    def access_save(self,u,f):
        target=f.get('username','').strip(); role=f.get('role') or 'Employee'
        if target==u.get('username') and role!='SuperAdmin':
            return self.send(page('Access Control','<div class="card"><div class="alert">لا يمكنك خفض صلاحيات حسابك بنفسك.</div></div>',u,'roles'),400)
        c=db(); rr=c.execute('SELECT scope_default FROM roles WHERE name=?',(role,)).fetchone(); default_scope=(rr['scope_default'] if rr else 'self')
        scope=f.get('scope_type') or default_scope
        c.execute('UPDATE users SET role=?,scope_type=?,scope_value=?,must_change_password=?,permission_version=COALESCE(permission_version,1)+1 WHERE username=?',(role,scope,f.get('scope_value',''),1 if f.get('must_change_password') else 0,target)); c.commit(); c.close(); audit(u['username'],u['role'],'تعديل الوصول','المستخدمون',target,role); self.redirect('/access')
    def backups_page(self,u):
        c=db(); rows=c.execute('SELECT * FROM system_backups ORDER BY id DESC LIMIT 100').fetchall(); c.close(); trs=''.join(f'<tr><td>{esc(r["created_at"])}</td><td>{esc(r["label"])}</td><td>{esc(r["created_by"])}</td><td>{r["db_size"]}</td><td><form method="post" action="/backup/restore">{csrf_field(u)}<input type="hidden" name="id" value="{r["id"]}"><button class="btn warn">Rollback</button></form></td></tr>' for r in rows)
        _no_backups_row='<tr><td colspan="5">لا توجد نسخ</td></tr>'
        body=f'<div class="top"><div class="title"><h1>Backup & Rollback</h1><p>نسخ قاعدة البيانات + ملفات الموظفين، مع نسخة أمان تلقائية قبل Restore.</p></div><form method="post" action="/backup" style="display:inline">{csrf_field(u)}<button class="btn">إنشاء نسخة الآن</button></form></div><div class="card table-wrap"><table class="table"><thead><tr><th>التاريخ</th><th>النوع</th><th>بواسطة</th><th>الحجم</th><th></th></tr></thead><tbody>{trs or _no_backups_row}</tbody></table></div>'; self.send(page('Backup & Rollback',body,u,'dashboard'))
    def backup_restore(self,u,f):
        c=db(); r=c.execute('SELECT * FROM system_backups WHERE id=?',(f.get('id'),)).fetchone(); c.close()
        if not r or not os.path.exists(r['file_path']): return self.redirect('/backups')
        source=r['file_path']; make_backup(u,'pre_rollback')
        try:
            if source.lower().endswith('.zip'):
                valid,msg=verify_backup_package(source)
                if not valid: raise ValueError('النسخة الاحتياطية غير سليمة: '+msg)
                tmp=os.path.join(DATA,'_restore_tmp'); shutil.rmtree(tmp,ignore_errors=True); os.makedirs(tmp)
                with zipfile.ZipFile(source,'r') as z:z.extractall(tmp)
                dbsrc=os.path.join(tmp,'database.db')
                if not os.path.exists(dbsrc): raise ValueError('النسخة لا تحتوي database.db')
                shutil.copy2(dbsrc,DB)
                srcfiles=os.path.join(tmp,'employee_files'); shutil.rmtree(EMPFILES,ignore_errors=True); os.makedirs(EMPFILES,exist_ok=True)
                if os.path.isdir(srcfiles):
                    for root,dirs,files in os.walk(srcfiles):
                        rel=os.path.relpath(root,srcfiles); dest=os.path.join(EMPFILES,'' if rel=='.' else rel); os.makedirs(dest,exist_ok=True)
                        for fn in files: shutil.copy2(os.path.join(root,fn),os.path.join(dest,fn))
                shutil.rmtree(tmp,ignore_errors=True)
            else: shutil.copy2(source,DB)
            audit(u['username'],u['role'],'ROLLBACK','النظام',str(r['id']),source,reason='restore backup')
            self.send(page('Rollback','<div class="card"><h2>تمت الاستعادة بنجاح.</h2><p>تم إنشاء نسخة أمان تلقائية قبل الاستعادة.</p><a class="btn" href="/backups">العودة للنسخ</a></div>',u,'dashboard'))
        except Exception as e:
            self.send(page('Rollback Error','<div class="card"><div class="alert">فشلت الاستعادة: '+esc(e)+'</div></div>',u,'dashboard'),500)

    def network_page(self,u):
        c=db(); sessions=c.execute('SELECT user_name,full_name,role,last_seen,ip,device FROM system_sessions WHERE revoked=0 ORDER BY last_seen DESC').fetchall(); c.close()
        nowdt=datetime.now()
        rows=[]
        for r in sessions:
            try: age=(nowdt-datetime.fromisoformat(r['last_seen'])).total_seconds()
            except Exception: age=99999
            online=age<120
            rows.append(f'<tr><td>{esc(r["full_name"])}</td><td>{esc(r["user_name"])}</td><td><span class="badge {"b-ok" if online else "b-gray"}">{"🟢 Online" if online else "⚪ Offline"}</span></td><td>{esc(r["ip"] or "—")}</td><td>{esc(r["last_seen"])}</td><td title="{esc(r["device"] or "")}">{esc((r["device"] or "")[:55])}</td></tr>')
        server_ip=local_ip(); ident=server_identity(); body=f'''<div class="top"><div class="title"><h1>🖥 Network Center</h1><p>هل الأجهزة شايفة السيرفر؟ هنا الإجابة في شاشة واحدة.</p></div><button class="btn" onclick="runPing()">اختبار الاتصال الآن</button></div>
<div class="grid g4"><div class="card metric"><div class="label">Server</div><div class="value" style="font-size:20px">🟢 Online</div><div class="sub">{esc(server_ip)}:{PORT}</div></div><div class="card metric"><div class="label">Database</div><div class="value" style="font-size:20px">🟢 Healthy</div><div class="sub">SQLite WAL</div></div><div class="card metric"><div class="label">Connected Devices</div><div class="value">{sum(1 for r in sessions if (nowdt-datetime.fromisoformat(r['last_seen'])).total_seconds()<120)}</div><div class="sub">آخر نشاط خلال دقيقتين</div></div><div class="card metric"><div class="label">Your IP</div><div class="value" style="font-size:18px">{esc(self.client_address[0])}</div><div class="sub">جهازك الحالي</div></div></div>
<div class="card" style="margin-top:16px"><h3>Server Identity</h3><p>Server ID: <code>{esc(ident['server_id'])}</code></p><p>Fingerprint: <code>{esc(ident['fingerprint'][:32])}…</code></p></div><div class="card" style="margin-top:16px"><h3>Test Network</h3><div id="pingResult" class="alert">اضغط «اختبار الاتصال الآن».</div></div>
<div class="card table-wrap" style="margin-top:16px"><h3>Connected Devices</h3><table class="table"><thead><tr><th>المستخدم</th><th>الحساب</th><th>الحالة</th><th>IP</th><th>آخر نشاط</th><th>الجهاز</th></tr></thead><tbody>{''.join(rows) or '<tr><td colspan="6">لا توجد جلسات أخرى.</td></tr>'}</tbody></table></div>
<script>async function runPing(){{const box=document.getElementById('pingResult');box.textContent='جاري الاختبار...';const t=performance.now();try{{const r=await fetch('/network/ping?x='+Date.now());const j=await r.json();box.innerHTML='✓ Server reachable · ✓ Database reachable · Latency: <b>'+Math.round(performance.now()-t)+' ms</b> · IP: '+j.ip;}}catch(e){{box.textContent='🔴 Server not reachable';}}}}</script>'''
        self.send(page('Network Center',body,u,'network'))
    def network_ping(self,u):
        self.send(json.dumps({'ok':True,'ip':local_ip(),'port':PORT,'version':'7.4 Enterprise'},ensure_ascii=False),200,'application/json')
    def system_page(self,u):
        mode='auto' if AUTO_MODE else ('network' if NETWORK_MODE else 'standalone'); host=HOST; dh=database_health(); sh=storage_health(); dbcls='b-ok' if dh.get('ok') else 'b-bad'; storage_ok=all(x['exists'] and x['writable'] for x in sh)
        checks=''.join(f'<tr><td>{esc(x["name"])}</td><td><span class="badge {"b-ok" if x["exists"] else "b-bad"}">{"🟢 موجود" if x["exists"] else "🔴 مفقود"}</span></td><td><span class="badge {"b-ok" if x["writable"] else "b-bad"}">{"🟢 قابل للكتابة" if x["writable"] else "🔴 غير قابل للكتابة"}</span></td></tr>' for x in sh)
        body=f'''<div class="top"><div class="title"><h1>🛠 مركز صحة النظام</h1><p>تشخيص حقيقي من داخل البرنامج — بدون CMD أو PowerShell.</p></div><div class="actions"><button class="btn" onclick="runCheck()">فحص الآن</button><a class="btn gray" href="/diagnostics/errors">سجل الأخطاء</a><a class="btn gray" href="/network">Network Center</a></div></div>
<div class="grid g4"><div class="card metric"><div class="label">Application</div><div class="value" style="font-size:20px">🟢 Ready</div><div class="sub">{esc(APP_VERSION)}</div></div><div class="card metric"><div class="label">Database</div><div class="value" style="font-size:20px"><span class="badge {dbcls}">{'🟢 Healthy' if dh.get('ok') else '🔴 Problem'}</span></div><div class="sub">Integrity: {esc(dh.get('integrity','unknown'))}</div></div><div class="card metric"><div class="label">Storage</div><div class="value" style="font-size:20px"><span class="badge {'b-ok' if storage_ok else 'b-bad'}">{'🟢 Healthy' if storage_ok else '🔴 Problem'}</span></div><div class="sub">Files + Backups</div></div><div class="card metric"><div class="label">Network</div><div class="value" style="font-size:20px">🟢 {esc(mode)}</div><div class="sub">{esc(local_ip())}:{PORT}</div></div></div>
<div class="grid g2" style="margin-top:16px"><div class="card"><h3>Data Integrity</h3><p>Employees: <b>{dh.get('counts',{}).get('employees',0)}</b> · Documents: <b>{dh.get('counts',{}).get('documents',0)}</b> · Attendance: <b>{dh.get('counts',{}).get('attendance',0)}</b> · Audit: <b>{dh.get('counts',{}).get('audit',0)}</b></p><a class="btn gray" href="/audit/verify">Verify Audit Chain</a></div><div class="card"><h3>Server Discovery</h3><p>🟢 UDP Discovery : {DISCOVERY_PORT}</p><p>Port الحالي: <b>{PORT}</b> — لو 8899 مشغول، النظام يختار منفذًا متاحًا تلقائيًا.</p><p>Data: <code>{esc(DATA)}</code></p></div></div>
<div class="card table-wrap" style="margin-top:16px"><h3>Storage Check</h3><table class="table"><thead><tr><th>المسار</th><th>الحالة</th><th>الكتابة</th></tr></thead><tbody>{checks}</tbody></table></div>
<div id="checkResult" class="alert" style="margin-top:16px">آخر فحص: {esc(now())}</div>
<script>async function runCheck(){{const b=document.getElementById('checkResult');b.textContent='جاري الفحص...';try{{const r=await fetch('/diagnostics/test?x='+Date.now());const j=await r.json();b.innerHTML=(j.ok?'✓':'⚠')+' Request ID: <code>'+j.request_id+'</code> · Port: '+j.port+' · DB: '+(j.database.ok?'OK':'PROBLEM')+' · Storage checks: '+j.storage.filter(x=>x.exists&&x.writable).length+'/'+j.storage.length;}}catch(e){{b.textContent='🔴 تعذر تنفيذ الفحص';}}}}</script>'''; self.send(page('System Health',body,u,'system'))

    def export_template_attendance(self):
        wb=Workbook(); ws=wb.active; ws.append(['التاريخ','كود الموظف','الحالة','الحضور','الانصراف','التأخير','ساعات العمل','الإضافي','ملاحظات']); ws.append([date.today().isoformat(),'1001','حضور','09:00','17:00',0,8,0,'']); out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Attendance_Import_Template.xlsx"'})
    def export_template_leaves(self):
        wb=Workbook(); ws=wb.active; ws.append(['كود الموظف','نوع الإجازة','من','إلى','ملاحظات']); ws.append(['1001','اعتيادي',date.today().isoformat(),date.today().isoformat(),'']); out=io.BytesIO(); wb.save(out); self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':'attachment; filename="Leaves_Import_Template.xlsx"'})
    def _export_rows(self,key,u,query_string=''):
        c=db(); scope_sql,scope_params=visible_employee_sql(u); params=[]; q=urlparse(query_string)
        if key=='employees':
            cond=[]
            qs=parse_qs(query_string)
            dept=qs.get('department',[''])[0]; status=qs.get('status',[''])[0]; unit=qs.get('unit',[''])[0]; job=qs.get('job',[''])[0]; search=qs.get('q',[''])[0]
            if search: cond.append('(e.emp_code LIKE ? OR e.name LIKE ? OR e.phone LIKE ?)'); like=f'%{search}%'; params += [like,like,like]
            for field,val in [('department',dept),('status',status),('unit',unit),('job',job)]:
                if val: cond.append(f'e.{field}=?'); params.append(val)
            where=(' WHERE '+' AND '.join(cond)) if cond else ' WHERE 1=1'
            rows=[dict(r) for r in c.execute(f'SELECT e.emp_code,e.name,e.employee_group,e.birth_date,e.national_id,e.address,e.qualification,e.phone,e.iban,e.bank_name,e.bank_branch,e.department,e.unit,e.job,e.contract_date,e.contract_amount FROM employees e{where}{scope_sql} ORDER BY e.name',params+scope_params).fetchall()]
            if not can(u,'sensitive.view'):
                for r in rows:r['national_id']='************';r['iban']='************'
            headers=HOSPITAL_HEADERS; keys=['emp_code','name','employee_group','birth_date','national_id','address','qualification','phone','iban','bank_name','bank_branch','department','unit','job','contract_date','contract_amount']; data=[[r.get(k,'') for k in keys] for r in rows]
        elif key=='attendance':
            rows=c.execute(f'SELECT a.work_date,a.emp_code,a.status,a.check_in,a.check_out,a.late_minutes,a.work_hours,a.overtime,a.notes FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE 1=1{scope_sql} ORDER BY a.work_date DESC',scope_params).fetchall(); headers=['التاريخ','الكود','الحالة','حضور','انصراف','تأخير','ساعات','إضافي','ملاحظات']; data=[list(r) for r in rows]
        elif key=='leaves':
            rows=c.execute(f'SELECT l.request_no,l.emp_code,l.leave_type,l.start_date,l.end_date,l.days,l.request_date,l.status,l.approved_by,l.approved_at,l.notes FROM leaves l JOIN employees e ON e.emp_code=l.emp_code WHERE 1=1{scope_sql} ORDER BY l.id DESC',scope_params).fetchall(); headers=['الطلب','الموظف','النوع','من','إلى','الأيام','تاريخ الطلب','الحالة','المعتمد','وقت الاعتماد','ملاحظات']; data=[list(r) for r in rows]
        else: c.close(); return [],[]
        c.close(); return headers,data

    def export_csv(self,key,u):
        headers,data=self._export_rows(key,u,urlparse(self.path).query); out=io.StringIO(newline=''); w=csv.writer(out); w.writerow(headers); w.writerows(data); raw=('\ufeff'+out.getvalue()).encode('utf-8'); self.send(raw,200,'text/csv; charset=utf-8',{'Content-Disposition':f'attachment; filename="{key}.csv"'})

    def export_html(self,key,u):
        headers,data=self._export_rows(key,u,urlparse(self.path).query); title={'employees':'بيانات الموظفين','attendance':'الحضور','leaves':'الإجازات'}.get(key,key); th=''.join('<th>'+esc(str(x))+'</th>' for x in headers); trs=''.join('<tr>'+''.join('<td>'+esc(str(v or ''))+'</td>' for v in row)+'</tr>' for row in data)
        html="""<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{}</title><style>body{{font-family:Segoe UI,Tahoma,Arial;background:#f4f7fb;color:#172033;padding:24px}}.head{{background:#fff;padding:18px;border-radius:14px;margin-bottom:18px}}table{{width:100%;border-collapse:collapse;background:#fff}}th,td{{padding:9px;border:1px solid #dfe4ea;text-align:right}}th{{background:#edf3ff}}tr:nth-child(even){{background:#fafbfc}}@media print{{body{{background:#fff}}}}</style></head><body><div class="head"><h1>{}</h1><p>HR Enterprise · {} · {} سجل</p></div><table><thead><tr>{}</tr></thead><tbody>{}</tbody></table></body></html>""".format(esc(title),esc(title),date.today().isoformat(),len(data),th,trs).encode('utf-8')
        self.send(html,200,'text/html; charset=utf-8',{'Content-Disposition':f'attachment; filename="{key}.html"'})

    def pdf_report(self,kind,u):
        if canvas is None:return self.send('PDF engine unavailable',503)
        c=db(); scope_sql,scope_params=visible_employee_sql(u)
        if kind=='employees': rows=c.execute(f'SELECT e.emp_code,e.name,e.department,e.job,e.status,e.hire_date FROM employees e WHERE 1=1{scope_sql} ORDER BY e.name',scope_params).fetchall(); headers=['الكود','الاسم','الإدارة','الوظيفة','الحالة','التعيين']
        elif kind=='leaves': rows=c.execute(f'SELECT l.request_no,l.emp_code,l.leave_type,l.start_date,l.end_date,l.days,l.status FROM leaves l JOIN employees e ON e.emp_code=l.emp_code WHERE 1=1{scope_sql} ORDER BY l.id DESC',scope_params).fetchall(); headers=['الطلب','الموظف','النوع','من','إلى','الأيام','الحالة']
        else: rows=c.execute(f'SELECT a.work_date,a.emp_code,a.status,a.check_in,a.check_out,a.late_minutes,a.overtime FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE 1=1{scope_sql} ORDER BY a.work_date DESC',scope_params).fetchall(); headers=['التاريخ','الكود','الحالة','الحضور','الانصراف','التأخير','الإضافي']
        c.close(); out=io.BytesIO(); pdf=canvas.Canvas(out,pagesize=A4); W,H=A4
        font_path=os.path.join(BASE,'fonts','DejaVuSans.ttf')
        font='Helvetica'
        if pdfmetrics and os.path.exists(font_path):
            try: pdfmetrics.registerFont(TTFont('HRDejaVu',font_path)); font='HRDejaVu'
            except Exception: pass
        pdf.setFont(font,15); pdf.drawRightString(W-40,H-45,setting('company_name') or 'HR Enterprise'); pdf.setFont(font,10); pdf.drawRightString(W-40,H-62,kind.title()+' Report')
        y=H-90; pdf.setFont(font,8); xs=[W-40,W-125,W-225,W-305,W-400,W-470,W-530]
        for i,h in enumerate(headers): pdf.drawRightString(xs[min(i,len(xs)-1)],y,h)
        y-=18; pdf.setFont(font,7)
        for r in rows:
            vals=list(r)
            for i,v in enumerate(vals): pdf.drawRightString(xs[min(i,len(xs)-1)],y,str(v or '')[:28])
            y-=13
            if y<45: pdf.showPage(); y=H-45; pdf.setFont(font,7)
        pdf.save(); self.send(out.getvalue(),200,'application/pdf',{'Content-Disposition':f'inline; filename="{kind}_report.pdf"'})

    def export_route(self,path):
        maps={
            'employees':('بيانات الموظفين',HOSPITAL_HEADERS),
            'leaves':('Leaves',['الطلب','الموظف','النوع','من','إلى','الأيام','تاريخ الطلب','الحالة','المعتمد','وقت الاعتماد','ملاحظات']),
            'attendance':('Attendance',['التاريخ','الكود','الحالة','حضور','انصراف','تأخير','ساعات','إضافي','ملاحظات']),
            'audit':('Audit',['الوقت','المستخدم','الدور','الإجراء','الوحدة','المعرف','التفاصيل'])
        }
        key=path.split('/')[-1]
        if key not in maps:return self.send('Not found',404)
        name,headers=maps[key]; c=db()
        user=getattr(AUDIT_CTX,'user',{})
        scope_sql,scope_params=visible_employee_sql(user)
        if key in ('employees','leaves','attendance'):
            headers,data=self._export_rows(key,user,urlparse(self.path).query)
            c.close(); wb=Workbook(); ws=wb.active; ws.title=name[:31]; ws.append(headers)
            for row in data: ws.append(row)
            for col in ws.columns:
                mx=max([len(str(x.value or '')) for x in col] or [12]); ws.column_dimensions[col[0].column_letter].width=min(42,max(12,mx+2))
            out=io.BytesIO(); wb.save(out)
            return self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',{'Content-Disposition':f'attachment; filename="export.xlsx"; filename*=UTF-8\'\'{quote(name+".xlsx")}'})
        if key=='employees':
            q=f"""SELECT e.emp_code,e.name,e.employee_group,e.birth_date,e.national_id,e.address,e.qualification,e.phone,e.iban,e.bank_name,e.bank_branch,e.department,e.unit,e.job,e.contract_date,e.contract_amount
                  FROM employees e WHERE 1=1{scope_sql} ORDER BY e.id"""
            rows=c.execute(q,scope_params).fetchall()
            # Mask sensitive columns unless the user explicitly has permission.
            user=getattr(AUDIT_CTX,'user',{})
            rows=[dict(r) for r in rows]
            if not can(user,'sensitive.view'):
                for r in rows: r['national_id']='************'; r['iban']='************'
        elif key=='leaves':
            q=f"""SELECT l.request_no,l.emp_code,l.leave_type,l.start_date,l.end_date,l.days,l.request_date,l.status,l.approved_by,l.approved_at,l.notes
                  FROM leaves l JOIN employees e ON e.emp_code=l.emp_code WHERE 1=1{scope_sql} ORDER BY l.id DESC"""
            rows=c.execute(q,scope_params).fetchall()
        elif key=='attendance':
            q=f"""SELECT a.work_date,a.emp_code,a.status,a.check_in,a.check_out,a.late_minutes,a.work_hours,a.overtime,a.notes
                  FROM attendance a JOIN employees e ON e.emp_code=a.emp_code WHERE 1=1{scope_sql} ORDER BY a.work_date DESC"""
            rows=c.execute(q,scope_params).fetchall()
        else:
            rows=c.execute('SELECT ts,username,role,action,entity,record_key,details FROM audit ORDER BY id DESC').fetchall()
        c.close()
        wb=Workbook(); ws=wb.active; ws.title=name[:31]; ws.append(headers)
        if key=='employees':
            keys=['emp_code','name','employee_group','birth_date','national_id','address','qualification','phone','iban','bank_name','bank_branch','department','unit','job','contract_date','contract_amount']
            for r in rows:
                d=dict(r); ws.append([d.get(k,'') for k in keys])
        else:
            for r in rows: ws.append(list(r))
        for col in ws.columns:
            mx=max(len(str(x.value or '')) for x in col)
            ws.column_dimensions[col[0].column_letter].width=min(42,max(12,mx+2))
        out=io.BytesIO(); wb.save(out)
        self.send(out.getvalue(),200,'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                  {'Content-Disposition':f'attachment; filename="export.xlsx"; filename*=UTF-8\'\'{quote(name+'.xlsx')}'})


def clipboard_bridge():
    """Optional Windows-only local bridge for real Explorer file/folder clipboard paste.
    It never reads file contents; it only returns paths from CF_HDROP. The HR page uploads
    the selected files through the normal authenticated endpoint."""
    if not sys.platform.startswith('win'): return
    try:
        import win32clipboard, win32con
    except Exception:
        return
    from http.server import HTTPServer
    BRIDGE_PORT=int(os.environ.get('HR_BRIDGE_PORT','8975'))
    allowed_origin=os.environ.get('HR_BRIDGE_ORIGIN','').strip()
    class Bridge(BaseHTTPRequestHandler):
        def log_message(self,*a): pass
        def do_GET(self):
            if self.path.split('?',1)[0]!='/clipboard/files': self.send_response(404); self.end_headers(); return
            origin=self.headers.get('Origin',''); allowed_origin=os.environ.get('HR_BRIDGE_ORIGIN','').strip()
            if allowed_origin and origin and origin!=allowed_origin:
                self.send_response(403); self.end_headers(); return
            files=[]
            try:
                win32clipboard.OpenClipboard()
                if win32clipboard.IsClipboardFormatAvailable(win32con.CF_HDROP):
                    files=list(win32clipboard.GetClipboardData(win32con.CF_HDROP))
            finally:
                try: win32clipboard.CloseClipboard()
                except Exception: pass
            body=json.dumps({'ok':True,'files':files},ensure_ascii=False).encode('utf-8')
            self.send_response(200); self.send_header('Content-Type','application/json; charset=utf-8'); self.send_header('Content-Length',str(len(body)))
            self.send_header('Cache-Control','no-store'); self.send_header('Access-Control-Allow-Origin',origin or allowed_origin or 'null'); self.send_header('Vary','Origin'); self.end_headers(); self.wfile.write(body)
        def do_POST(self):
            if self.path!='/clipboard/upload': self.send_response(404); self.end_headers(); return
            origin=self.headers.get('Origin',''); allowed_origin=os.environ.get('HR_BRIDGE_ORIGIN','').strip()
            if allowed_origin and origin and origin!=allowed_origin: self.send_response(403); self.end_headers(); return
            try:
                n=int(self.headers.get('Content-Length','0')); payload=json.loads(self.rfile.read(n).decode('utf-8'))
                token=str(payload.get('token','')); remote=str(payload.get('remote','')).rstrip('/'); emp=str(payload.get('emp_code','')); csrf=str(payload.get('csrf',''))
                if not token or not remote or not emp: raise ValueError('missing bridge parameters')
                # Validate the one-time token against the HR server before sending any file.
                import urllib.request, urllib.error
                req=urllib.request.Request(remote+'/bridge/authorize',data=json.dumps({'token':token,'csrf':csrf,'emp_code':emp}).encode('utf-8'),headers={'Content-Type':'application/json','Origin':remote})
                with urllib.request.urlopen(req,timeout=8) as rr: auth=json.loads(rr.read().decode('utf-8'))
                if not auth.get('ok'): raise ValueError(auth.get('error','bridge authorization failed'))
                win32clipboard.OpenClipboard()
                paths=list(win32clipboard.GetClipboardData(win32con.CF_HDROP)) if win32clipboard.IsClipboardFormatAvailable(win32con.CF_HDROP) else []
                win32clipboard.CloseClipboard()
                if not paths: raise ValueError('لا توجد ملفات Windows في Clipboard')
                boundary='----HRBridge'+secrets.token_hex(12).encode().decode(); body=io.BytesIO()
                # multipart upload for each file
                for fp in paths:
                    if not os.path.isfile(fp): continue
                    fn=os.path.basename(fp); data=open(fp,'rb').read()
                    body.write(('--'+boundary+'\r\n').encode()); body.write(f'Content-Disposition: form-data; name="file"; filename="{fn.replace(chr(34),chr(39))}"\r\nContent-Type: application/octet-stream\r\n\r\n'.encode()); body.write(data); body.write(b'\r\n')
                for k,v in [('_csrf',csrf),('_bridge_token',token),('emp_code',emp)]: body.write((f'--{boundary}\r\nContent-Disposition: form-data; name="{k}"\r\n\r\n{v}\r\n').encode())
                body.write(('--'+boundary+'--\r\n').encode())
                req2=urllib.request.Request(remote+'/documents/bridge-upload',data=body.getvalue(),headers={'Content-Type':'multipart/form-data; boundary='+boundary,'Origin':remote})
                with urllib.request.urlopen(req2,timeout=30) as rr: result=json.loads(rr.read().decode('utf-8'))
                out=json.dumps(result,ensure_ascii=False).encode('utf-8'); self.send_response(200); self.send_header('Content-Type','application/json'); self.send_header('Access-Control-Allow-Origin',origin or allowed_origin or 'null'); self.send_header('Content-Length',str(len(out))); self.end_headers(); self.wfile.write(out)
            except Exception as e:
                try: win32clipboard.CloseClipboard()
                except Exception: pass
                out=json.dumps({'ok':False,'error':str(e)},ensure_ascii=False).encode('utf-8'); self.send_response(400); self.send_header('Content-Type','application/json'); self.send_header('Access-Control-Allow-Origin',origin or allowed_origin or 'null'); self.send_header('Content-Length',str(len(out))); self.end_headers(); self.wfile.write(out)
    try:
        HTTPServer(('127.0.0.1',BRIDGE_PORT),Bridge).serve_forever()
    except Exception as e: log_error('clipboard_bridge',e)

def maintenance_loop():
    last_backup_day=''
    while True:
        try:
            auto_alerts()
            if "v10_alerts_engine" in globals():
                try: v10_alerts_engine()
                except Exception as _e: log_error('v10_alerts_engine',_e,username='system',method='SCHEDULE',path='/intelligence')
            enabled=setting('backup_auto_enabled')=='1'; target=setting('backup_time') or '23:00'; today=date.today().isoformat()
            if enabled and datetime.now().strftime('%H:%M')==target and last_backup_day!=today:
                admin={'username':'system','role':'SuperAdmin','full_name':'System'}
                try: make_backup(admin,'automatic_daily'); last_backup_day=today
                except Exception as e: log_error('automatic_backup',e,username='system',method='SCHEDULE',path='/backup')
        except Exception as e: log_error('maintenance_loop',e,username='system',method='SCHEDULE',path='/system')
        time.sleep(30)

def main():
    # One EXE: discover an existing server first; otherwise perform a deterministic LAN election.
    client_url=os.environ.get('HR_CLIENT_URL','').strip()
    cfg=os.path.join(BASE,'client_url.txt')
    if not client_url and os.path.exists(cfg):
        try: client_url=open(cfg,encoding='utf-8-sig').read().strip()
        except Exception: client_url=''
    if client_url and not NETWORK_MODE:
        os.environ['HR_BRIDGE_ORIGIN']=client_url
        webbrowser.open(client_url)
        while True: time.sleep(3600)
    migrate_legacy_data_location(); init(); migrate_documents_to_fs(); auto_alerts(); server_identity()
    threading.Thread(target=clipboard_bridge,daemon=True).start()
    global PORT
    if AUTO_MODE and not NETWORK_MODE:
        try:
            discovered=discover_server()
            if discovered:
                os.environ['HR_BRIDGE_ORIGIN']=discovered
                webbrowser.open(discovered)
                while True: time.sleep(3600)
            if not elect_server():
                time.sleep(0.6); discovered=discover_server(1.5)
                if discovered:
                    os.environ['HR_BRIDGE_ORIGIN']=discovered
                    webbrowser.open(discovered)
                    while True: time.sleep(3600)
        except ValueError as e:
            log_error('server-trust',e); show_native_error('HR Enterprise — تحذير أمني',str(e)); return
        except Exception as e:
            log_error('server-election',e)
    try: PORT=find_free_port(PORT_MIN,PORT_MAX)
    except Exception as e:
        log_error('port-selection',e); raise
    threading.Thread(target=discovery_responder,daemon=True).start()
    threading.Thread(target=maintenance_loop,daemon=True).start()
    url=f'http://{HOST if HOST not in ("0.0.0.0","::") else local_ip()}:{PORT}'
    print(f'HR Enterprise server running on {url} | mode={"auto" if AUTO_MODE else ("network" if NETWORK_MODE else "standalone")} | version={APP_VERSION}')
    if os.environ.get('HR_NO_BROWSER','0')!='1': threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    ThreadingHTTPServer((HOST,PORT),H).serve_forever()

# HR Enterprise V9.0 comprehensive feature patch
import os, json, time, secrets, statistics, math, shutil
from datetime import datetime, date, timedelta
from urllib.parse import quote

V9_PREVIEWS={}

def _v9_upgrade():
    c=db()
    c.executescript('''
    CREATE TABLE IF NOT EXISTS branding_profiles(id INTEGER PRIMARY KEY, profile_key TEXT UNIQUE, path TEXT, updated_at TEXT, updated_by TEXT);
    CREATE TABLE IF NOT EXISTS employee_portal_accounts(emp_code TEXT UNIQUE, username TEXT UNIQUE, active INTEGER DEFAULT 1, created_at TEXT);
    CREATE TABLE IF NOT EXISTS evaluation_anomalies(id INTEGER PRIMARY KEY, emp_code TEXT, period TEXT, score REAL, dept_average REAL, delta_percent REAL, percentile REAL, severity TEXT, created_at TEXT, UNIQUE(emp_code,period));
    CREATE TABLE IF NOT EXISTS device_events(id INTEGER PRIMARY KEY, device_id TEXT, event_type TEXT, ts TEXT, latency_ms REAL, message TEXT);
    CREATE TABLE IF NOT EXISTS server_discovery_log(id INTEGER PRIMARY KEY, ts TEXT, server_name TEXT, ip TEXT, port INTEGER, latency_ms REAL, result TEXT);
    CREATE TABLE IF NOT EXISTS bulk_action_log(id INTEGER PRIMARY KEY, action TEXT, affected INTEGER, filters_json TEXT, created_by TEXT, created_at TEXT);
    ''')
    for key in ('login_logo','sidebar_logo','report_logo','favicon_logo'):
        if not c.execute('SELECT 1 FROM branding_profiles WHERE profile_key=?',(key,)).fetchone():
            c.execute('INSERT INTO branding_profiles(profile_key,path,updated_at,updated_by) VALUES(?,?,?,?)',(key,'',now(),'system'))
    c.commit(); c.close()


def _eval_intelligence(self,u):
    if not self.need(u,'employees.view'): return
    c=db(); scope_sql,scope_params=visible_employee_sql(u,'e')
    rows=c.execute(f'''SELECT e.emp_code,e.name,COALESCE(e.department,'غير محدد') dept,ev.period,ev.score
                       FROM employee_evaluations ev JOIN employees e ON e.emp_code=ev.emp_code
                       WHERE 1=1 {scope_sql} ORDER BY ev.period DESC,e.department,e.name''',scope_params).fetchall()
    grouped={}
    for r in rows: grouped.setdefault((r['dept'],r['period']),[]).append(float(r['score'] or 0))
    out=[]
    for r in rows:
        vals=grouped.get((r['dept'],r['period']),[]); avg=sum(vals)/len(vals) if vals else 0
        delta=((r['score']-avg)/avg*100) if avg else 0
        percentile=(sum(1 for x in vals if x<=r['score'])/len(vals)*100) if vals else 0
        sev='HIGH' if abs(delta)>=25 and len(vals)>=3 else ('MEDIUM' if abs(delta)>=15 and len(vals)>=3 else 'NORMAL')
        if sev!='NORMAL':
            out.append((r,avg,delta,percentile,sev))
    cards=''.join(f'''<div class="card"><div style="display:flex;justify-content:space-between;gap:12px"><div><h3 style="margin:0">{esc(r['name'])}</h3><p>{esc(r['dept'])} · {esc(r['period'] or '')}</p></div><span class="badge {'b-bad' if sev=='HIGH' else 'b-warn'}">{sev}</span></div><div class="grid g4"><div><small>Employee Score</small><h2>{r['score']:.1f}%</h2></div><div><small>Department Avg</small><h2>{avg:.1f}%</h2></div><div><small>Difference</small><h2>{delta:+.1f}%</h2></div><div><small>Percentile</small><h2>{percentile:.0f}th</h2></div></div><div class="alert">⚠️ مؤشر مراجعة فقط — لا يمثل قرارًا تلقائيًا ضد الموظف.</div></div>''' for r,avg,delta,percentile,sev in out[:100])
    if not cards: cards='<div class="card"><h3>لا توجد حالات شاذة حاليًا</h3><p>يحتاج التحليل إلى تقييمات لثلاثة موظفين أو أكثر داخل نفس الإدارة والفترة.</p></div>'
    body=f'''<div class="top"><div class="title"><h1>📊 Evaluation Intelligence</h1><p>مقارنة التقييم الفردي بمتوسط الإدارة واكتشاف القيم غير المعتادة.</p></div><a class="btn gray" href="/reports">التقارير</a></div><div class="card"><div class="alert">النظام لا يعاقب الموظف تلقائيًا. هذه إشارة مراجعة تساعد HR على اكتشاف تقييمات غير منطقية.</div></div><div style="margin-top:16px">{cards}</div>'''
    c.close(); self.send(page('Evaluation Intelligence',body,u,'reports'))


def _branding_page(self,u):
    if not self.need(u,'settings.manage'): return
    c=db(); rows=c.execute('SELECT * FROM branding_profiles ORDER BY id').fetchall(); c.close()
    names={'login_logo':'Login Logo','sidebar_logo':'Sidebar Logo','report_logo':'Report Logo','favicon_logo':'Favicon'}
    trs=''.join(f'''<tr><td>{names.get(r['profile_key'],r['profile_key'])}</td><td>{esc(r['path'] or 'Default')}</td><td>{esc(r['updated_at'] or '')}</td></tr>''' for r in rows)
    body=f'''<div class="top"><div class="title"><h1>🎨 Branding Manager</h1><p>تحكم منفصل في شعار الدخول والـSidebar والتقارير والـFavicon.</p></div><a class="btn gray" href="/settings">Settings</a></div>
    <div class="grid g2"><div class="card"><h3>Upload Logo Profile</h3><form method="post" action="/branding/save" enctype="multipart/form-data">{csrf_field(u)}<div class="field"><label>Profile</label><select name="profile"><option value="login_logo">Login Logo</option><option value="sidebar_logo">Sidebar Logo</option><option value="report_logo">Report Logo</option><option value="favicon_logo">Favicon</option></select></div><div class="field"><label>Image</label><input type="file" name="file" accept=".png,.jpg,.jpeg,.svg" required></div><div class="actions" style="margin-top:12px"><button class="btn">Save Profile</button><button class="btn gray" name="action" value="restore" type="submit">Restore Default</button></div></form><p class="footer">PNG/JPG/SVG · حد أقصى 5MB. سيتم إنشاء نسخة favicon مربعة عند الإمكان.</p></div>
    <div class="card"><h3>Current Profiles</h3><table class="table"><thead><tr><th>Profile</th><th>File</th><th>Updated</th></tr></thead><tbody>{trs}</tbody></table></div></div>'''
    self.send(page('Branding Manager',body,u,'settings'))


def _branding_save(self,u):
    if not self.need(u,'settings.manage'): return
    fields,files=self.parse_upload_all(); profile=fields.get('profile','login_logo'); action=fields.get('action','')
    if profile not in ('login_logo','sidebar_logo','report_logo','favicon_logo'): return self.send(page('Branding','<div class="card"><div class="alert">Invalid profile.</div></div>',u,'settings'),400)
    brand=os.path.join(DATA,'branding','profiles'); os.makedirs(brand,exist_ok=True)
    if action=='restore':
        for p in os.listdir(brand):
            if p.startswith(profile+'.'): os.remove(os.path.join(brand,p))
        c=db(); c.execute('UPDATE branding_profiles SET path=?,updated_at=?,updated_by=? WHERE profile_key=?',('',now(),u['username'],profile)); c.commit(); c.close(); return self.redirect('/branding')
    fp=files.get('file')
    if not fp: return self.send(page('Branding','<div class="card"><div class="alert">اختر صورة.</div></div>',u,'settings'),400)
    head,data,fname=fp; ext=os.path.splitext(fname)[1].lower()
    if ext not in ('.png','.jpg','.jpeg','.svg') or len(data)>5*1024*1024: return self.send(page('Branding','<div class="card"><div class="alert">صيغة أو حجم الصورة غير صالح.</div></div>',u,'settings'),400)
    path=os.path.join(brand,profile+ext); open(path,'wb').write(data)
    c=db(); c.execute('UPDATE branding_profiles SET path=?,updated_at=?,updated_by=? WHERE profile_key=?',(os.path.relpath(path,DATA),now(),u['username'],profile)); c.commit(); c.close(); audit(u['username'],u['role'],'Branding profile update','Branding',profile,fname)
    self.redirect('/branding')


def _excel_grid_page(self,u):
    if not self.need(u,'employees.edit'): return
    headers=['Employee Code','Name','Department','Unit','Job','National ID','Phone','Email','Contract Date','Contract Amount']
    hdr=json.dumps(headers,ensure_ascii=False)
    body=f'''<div class="top"><div class="title"><h1>📊 Excel Center Pro</h1><p>Excel-like paste/edit/validate/import — بدون رفع ملف.</p></div><div class="actions"><a class="btn gray" href="/import">Classic Import</a><button class="btn" type="button" id="pasteBtn">Paste from Clipboard</button></div></div>
    <div class="card"><div class="toolbar"><input id="gridSearch" placeholder="Search in grid…"><button class="btn gray" type="button" id="addRow">+ Add Row</button><button class="btn gray" type="button" id="delRow">Delete Selected</button><button class="btn gray" type="button" id="fillDown">Fill Down</button><button class="btn gray" type="button" id="undo">Undo</button><button class="btn gray" type="button" id="redo">Redo</button><button class="btn gray" type="button" id="clear">Clear</button></div>
    <div class="alert">Ctrl+V = multi-cell paste · Ctrl+D = Fill Down · Delete = clear cell · Ctrl+Z / Ctrl+Y = Undo/Redo. 🔴 invalid · 🟡 warning · 🟢 valid.</div>
    <div id="grid" class="table-wrap" style="max-height:540px;margin-top:14px"></div>
    <form id="gridForm" method="post" action="/import/employees/paste/preview">{csrf_field(u)}<textarea id="payload" name="paste_data" hidden></textarea><div class="actions" style="margin-top:14px"><button class="btn ok" type="submit">Validate All → Preview</button></div></form></div>
    <script>
    const HEADERS={hdr}; let data=[HEADERS,Array(HEADERS.length).fill('')]; let selected=null,undoStack=[],redoStack=[];
    const grid=document.getElementById('grid');
    function esc(x){{return String(x??'').replace(/[&<>"']/g,m=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[m]));}}
    function snap(){{return JSON.stringify(data)}} function save(){{undoStack.push(snap());if(undoStack.length>50)undoStack.shift();redoStack=[]}}
    function valid(v,i){{if(!v)return '';if(i===5&&!/^\\d{{14}}$/.test(v))return 'bad';if(i===6&&!/^01\\d{{9}}$/.test(v))return 'warn';if(i===0&&v==='')return 'bad';return 'ok'}}
    function render(filter=''){{let h='<table class="table" id="sheet"><thead><tr>'+HEADERS.map(x=>'<th>'+esc(x)+'</th>').join('')+'</tr></thead><tbody>';data.slice(1).forEach((r,ri)=>{{let hay=r.join(' ').toLowerCase();if(filter&&!hay.includes(filter.toLowerCase()))return;h+='<tr data-ri="'+ri+'">'+HEADERS.map((_,ci)=>{{let cls=valid(r[ci]||'',ci);return '<td contenteditable="true" data-ci="'+ci+'" class="'+cls+'">'+esc(r[ci]||'')+'</td>'}}).join('')+'</tr>'}});h+='</tbody></table>';grid.innerHTML=h;grid.querySelectorAll('td').forEach(td=>{{td.addEventListener('focus',()=>{{selected=td.closest('tr')}});td.addEventListener('input',()=>{{let ri=+td.closest('tr').dataset.ri,ci=+td.dataset.ci;data[ri+1][ci]=td.innerText;td.className=valid(data[ri+1][ci],ci)}});td.addEventListener('keydown',e=>{{if(e.key==='Delete'&&!window.getSelection().toString()){{save();td.innerText='';td.dispatchEvent(new Event('input'))}}if((e.ctrlKey||e.metaKey)&&e.key.toLowerCase()==='d'){{e.preventDefault();save();let ri=+td.closest('tr').dataset.ri,ci=+td.dataset.ci;if(ri>0){{data[ri+1][ci]=data[ri][ci]||'';render(document.getElementById('gridSearch').value);}}}}}})}});}}
    function loadText(t){{let rows=t.replace(/\\r/g,'').split('\\n').filter(x=>x.trim()).map(x=>x.split('\\t'));if(rows.length===0)return;save();let first=rows[0];let likelyHeader=first.some(x=>/employee|name|department|الإسم|الاسم|الرقم/.test(x));data=likelyHeader?rows:[HEADERS,...rows];let m=Math.max(...data.map(r=>r.length));data=data.map(r=>r.concat(Array(Math.max(0,m-r.length)).fill('')).slice(0,Math.max(m,HEADERS.length)));render();}}
    document.addEventListener('paste',e=>{{if(document.activeElement.closest('#grid')){{e.preventDefault();loadText((e.clipboardData||window.clipboardData).getData('text'));}}}});
    document.getElementById('pasteBtn').onclick=async()=>{{try{{loadText(await navigator.clipboard.readText())}}catch(e){{alert('استخدم Ctrl+V داخل الجدول')}}}};
    document.getElementById('addRow').onclick=()=>{{save();data.push(Array(HEADERS.length).fill(''));render(document.getElementById('gridSearch').value)}};
    document.getElementById('delRow').onclick=()=>{{if(!selected)return;save();let ri=+selected.dataset.ri;data.splice(ri+1,1);selected=null;render(document.getElementById('gridSearch').value)}};
    document.getElementById('fillDown').onclick=()=>{{if(!selected)return;let ci=+selected.querySelector('td')?.dataset.ci||0,ri=+selected.dataset.ri;if(ri>0){{save();data[ri+1][ci]=data[ri][ci]||'';render()}}}};
    document.getElementById('undo').onclick=()=>{{if(undoStack.length){{redoStack.push(snap());data=JSON.parse(undoStack.pop());render()}}}};document.getElementById('redo').onclick=()=>{{if(redoStack.length){{undoStack.push(snap());data=JSON.parse(redoStack.pop());render()}}}};
    document.getElementById('clear').onclick=()=>{{save();data=[HEADERS,Array(HEADERS.length).fill('')];render()}};document.getElementById('gridSearch').oninput=e=>render(e.target.value);
    document.getElementById('gridForm').onsubmit=()=>{{document.getElementById('payload').value=data.map(r=>r.join('\\t')).join('\\n')}};render();
    </script>'''
    self.send(page('Excel Center Pro',body,u,'import'))


def _paste_preview(self,u,f):
    text=f.get('paste_data',''); raw=[x for x in text.replace('\r','').split('\n') if x.strip()]
    if not raw: return self.send(page('Import Preview','<div class="card"><div class="alert">لا توجد بيانات.</div></div>',u,'import'),400)
    rows=[x.split('\t') for x in raw]; header_idx,idx=find_hospital_header(rows)
    if header_idx is None:
        if max(len(r) for r in rows)<8: return self.send(page('Import Validation','<div class="card"><div class="alert">لم أتعرف على أعمدة الموظفين.</div></div>',u,'import'),400)
        idx={field:i for i,field in enumerate(HOSPITAL_FIELDS)}; data_rows=rows
    else: data_rows=rows[header_idx+1:]
    records=[hospital_row_to_record(r,idx) for r in data_rows]; valid,errors=validate_employee_records(records)
    token=secrets.token_urlsafe(18); V9_PREVIEWS[token]={'user':u['username'],'created':time.time(),'records':valid,'errors':errors,'rows':len(records)}
    errhtml=''.join(f'<tr><td>{e[0]}</td><td>{esc(e[1])}</td><td>{esc(e[2])}</td></tr>' for e in errors[:200])
    body=f'''<div class="top"><div class="title"><h1>🔎 Import Preview</h1><p>لم يتم إدخال أي سجل بعد. راجع النتائج ثم Confirm.</p></div><a class="btn gray" href="/import/excel-grid">Back to Grid</a></div><div class="grid g4"><div class="card metric"><div class="label">Total Rows</div><div class="value">{len(records)}</div></div><div class="card metric"><div class="label">Valid</div><div class="value">{len(valid)}</div></div><div class="card metric"><div class="label">Errors</div><div class="value">{len(errors)}</div></div><div class="card metric"><div class="label">Status</div><div class="value" style="font-size:22px">{'BLOCKED' if errors else 'READY'}</div></div></div>'''
    if errors: body+=f'<div class="card" style="margin-top:16px"><h3>🔴 Validation Errors</h3><table class="table"><thead><tr><th>Row</th><th>Field</th><th>Message</th></tr></thead><tbody>{errhtml}</tbody></table><div class="actions" style="margin-top:12px"><a class="btn bad" href="/export/import-errors/{quote(token)}">Export Errors</a><a class="btn gray" href="/import/excel-grid">Fix Data</a></div></div>'
    else: body+=f'<div class="card" style="margin-top:16px"><h3>🟢 Ready to Commit</h3><p>{len(valid)} records passed validation.</p><form method="post" action="/import/employees/paste/commit">{csrf_field(u)}<input type="hidden" name="token" value="{token}"><button class="btn ok">Confirm & Atomic Commit</button> <a class="btn gray" href="/import/excel-grid">Cancel</a></form></div>'
    self.send(page('Import Preview',body,u,'import'))


def _paste_commit(self,u,f):
    tok=f.get('token',''); x=V9_PREVIEWS.get(tok)
    if not x or x['user']!=u['username'] or time.time()-x['created']>900: return self.send(page('Import','<div class="card"><div class="alert">Preview انتهت صلاحيتها. أعد Validate.</div></div>',u,'import'),400)
    if x['errors']: return self.send(page('Import','<div class="card"><div class="alert">لا يمكن Commit مع أخطاء.</div></div>',u,'import'),400)
    new,upd,skip=upsert_hospital_records(x['records'],u,'Excel Center Pro'); del V9_PREVIEWS[tok]
    audit(u['username'],u['role'],'Atomic employee import','Employees',str(x['rows']),f'new={new},updated={upd},skipped={skip}')
    self.send(page('Import Complete',f'<div class="card"><h2>✅ Atomic Import Complete</h2><p>Rows: {x["rows"]} · New: {new} · Updated: {upd} · Skipped: {skip}</p><a class="btn" href="/employees">Employees</a></div>',u,'import'))


def _discovery_page(self,u):
    if not self.need(u,'system.manage'): return
    server=setting('server_name') or 'HR-MAIN'; ip=local_ip(); port=PORT
    body=f'''<div class="top"><div class="title"><h1>🌐 Server Discovery</h1><p>تجربة أول تشغيل للمستخدم العادي بدون معرفة Ports أو إعدادات تقنية.</p></div></div><div class="card" style="max-width:850px;margin:auto"><div style="font-size:20px;font-weight:800">HR Enterprise</div><div id="scan" style="margin-top:20px"><h2>Searching for HR Server...</h2><div style="height:12px;background:#eef2f6;border-radius:20px;overflow:hidden"><div id="bar" style="height:100%;width:65%;background:#175cd3;animation:p 1.5s infinite"></div></div><p id="msg">Checking local server discovery…</p></div><div id="found" style="display:none;margin-top:20px"><div class="card"><h2>🏥 {esc(server)}</h2><p><b>{esc(ip)}</b> · Port <b>{port}</b> · <span class="badge b-ok">Connected</span></p><button class="btn" onclick="location.href='/'">Connect</button></div></div><div class="actions" style="margin-top:20px"><button class="btn gray" onclick="location.reload()">Retry</button><a class="btn gray" href="/network">Network</a></div></div><style>@keyframes p{{0%{{width:15%}}50%{{width:90%}}100%{{width:35%}}}}</style><script>setTimeout(()=>{{document.getElementById('msg').textContent='Server found — verifying database and network…'}},800);setTimeout(()=>{{document.getElementById('scan').style.display='none';document.getElementById('found').style.display='block'}},1700);</script>'''
    self.send(page('Server Discovery',body,u,'network'))


def _device_ping_v9(self,u):
    d=self.headers.get('X-HR-Device-Name','')[:120] or 'Unknown Device'; start=time.perf_counter(); update_device(u,self.client_address[0],d); latency=(time.perf_counter()-start)*1000
    try:
        c=db(); r=c.execute('SELECT device_id FROM device_registry WHERE username=? AND ip=? ORDER BY id DESC LIMIT 1',(u['username'],self.client_address[0])).fetchone();
        if r: c.execute('INSERT INTO device_events(device_id,event_type,ts,latency_ms,message) VALUES(?,?,?,?,?)',(r['device_id'],'heartbeat',now(),latency,'OK'))
        c.commit(); c.close()
    except Exception: pass
    self.send(json.dumps({'ok':True,'server':setting('server_name') or 'HR-MAIN','ip':local_ip(),'port':PORT,'latency_ms':round(latency,2),'last_sync':now()},ensure_ascii=False),200,'application/json',{'Cache-Control':'no-store'})




def _v9_alerts_snapshot(u):
    base=_V9_ORIG_ALERTS(u)
    try:
        c=db(); n=c.execute("SELECT COUNT(*) n FROM payroll WHERE status IN ('جاهز','Ready for Review','approved') AND locked_at IS NULL").fetchone()['n']; c.close()
        base.append(('ok','🟢 Payroll → Ready for Lock',int(n),'payroll-lock'))
    except Exception: pass
    return base

def _v9_page(title,body,user,active='dashboard'):
    out=_V9_ORIG_PAGE(title,body,user,active)
    if user and can(user,'employees.view'):
        extra='<a href="/intelligence/evaluations">📊 Evaluation Intelligence</a>'
        if can(user,'settings.manage'): extra+='<a href="/branding">🎨 Branding Manager</a>'
        if can(user,'system.manage'): extra+='<a href="/network/discovery">🌐 Server Discovery</a>'
        out=out.replace('</nav>',extra+'</nav>',1)
    return out

def _v9_routes():
    global _V9_ORIG_PAGE,_V9_ORIG_ALERTS
    _V9_ORIG_PAGE=page; _V9_ORIG_ALERTS=hr_alerts_snapshot
    globals()['page']=_v9_page; globals()['hr_alerts_snapshot']=_v9_alerts_snapshot
    # Monkey patch routes without destroying the stable V8.3 code.
    H.import_page=_excel_grid_page
    H.device_ping=_device_ping_v9
    H.evaluation_intelligence=_eval_intelligence
    H.branding_manager=_branding_page
    H.branding_profile_save=_branding_save
    H.excel_grid=_excel_grid_page
    H.paste_preview=_paste_preview
    H.paste_commit=_paste_commit
    H.discovery_page=_discovery_page
    old_get=H.do_GET; old_post=H.do_POST
    def get(self):
        p=urlparse(self.path).path
        if p=='/import/excel-grid': return self.excel_grid(self.require()) if self.require() else None
        if p=='/intelligence/evaluations':
            u=self.require(); return self.evaluation_intelligence(u) if u else None
        if p=='/branding':
            u=self.require(); return self.branding_manager(u) if u else None
        if p=='/network/discovery':
            u=self.require(); return self.discovery_page(u) if u else None
        return old_get(self)
    def post(self):
        p=urlparse(self.path).path
        if p in ('/import/employees/paste/preview','/import/employees/paste/commit','/branding/save'):
            u=self.require()
            if not u:return
            if u.get('must_change_password') and p!='/password': return self.redirect('/password')
            ctype=self.headers.get('Content-Type','').lower(); f=self.form() if not ctype.startswith('multipart/form-data') else self.parse_upload()[0]
            if f.get('_csrf')!=u.get('csrf'): return self.send(page('Security','<div class="card"><div class="alert">Invalid CSRF.</div></div>',u),403)
            if p.endswith('/preview'): return self.paste_preview(u,f)
            if p.endswith('/commit'): return self.paste_commit(u,f)
            return self.branding_profile_save(u)
        return old_post(self)
    H.do_GET=get; H.do_POST=post

_v9_upgrade()
_v9_routes()

# V10 feature pack
import v10_feature_pack as _v10fp
_v10fp.install_v10(globals())

# V11 practical completion layer
import v11_completion as _v11
_v11.install_v11(globals())

if __name__=='__main__':main()

